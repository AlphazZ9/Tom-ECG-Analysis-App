# -*- coding: utf-8 -*-
"""
ecg.ui.app
----------
ECGApp -- the main application window (CustomTkinter CTk).
Imports all other ecg.* modules and wires them together.
"""
from __future__ import annotations

import copy
import dataclasses
import io
import logging
import os
import threading
import traceback
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional, cast

import customtkinter as ctk  # type: ignore[import-untyped]
import tkinter as tk
from tkinter import filedialog, messagebox

import matplotlib
import matplotlib.figure
import matplotlib.ticker
from matplotlib.ticker import MultipleLocator
from matplotlib.figure import Figure
if __name__ == "__main__":
    matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import numpy as np
# numpy 2.x may expose only trapezoid; ensure np.trapz alias exists.
if not hasattr(np, "trapz"):
    if hasattr(np, "trapezoid"):
        np.trapz = np.trapezoid  # type: ignore[attr-defined]
    else:
        def _trapz(y, x=None, dx=1.0, axis=-1):
            y_arr = np.asanyarray(y)
            if x is None:
                n = y_arr.shape[axis]
                x = np.arange(n, dtype=float) * dx
            else:
                x = np.asanyarray(x)
            # Broadcasting rules apply; use numpy.diff and slicing
            slice_all = [slice(None)] * y_arr.ndim
            slice1 = slice_all.copy(); slice2 = slice_all.copy()
            slice1[axis] = slice(1, None)
            slice2[axis] = slice(0, -1)
            dx_arr = np.diff(x)
            y_avg = (y_arr[tuple(slice1)] + y_arr[tuple(slice2)]) / 2.0
            return np.sum(y_avg * dx_arr, axis=axis)
        np.trapz = _trapz  # type: ignore[attr-defined]
import pandas as pd
from PIL import Image, ImageTk
from scipy.interpolate import CubicSpline
from scipy.signal import welch as _scipy_welch
from openpyxl import Workbook

# ── ecg.core ──────────────────────────────────────────────────────────────────
from models import (
    ArrhythmiaEvent, MouseECG, ContextRanges, EXPERIMENTAL_CONTEXTS, _CONTEXT_FIELD_MAP,
    FilterParams, AnalysisResults,
)
from filtering import (
    bandpass, notch, normalize,
    downsample_for_display, downsample_pair, envelope_for_display,
)
from detection import (
    fix_polarity, apply_threshold,
    detect_peaks_sg_derivative, detect_peaks_wavelet,
    detect_peaks_envelope_max,
    detect_rr_artifacts, apply_artifact_decisions, correct_rr_artifacts,
    recover_missed_beats, classify_arrhythmias,
)
from analysis import (
    analyse_core, analyse_hrv_freq, analyse_hrv_nonlinear, analyse_intervals,
)
from wave_template import WaveTemplate, detect_waves_on_beat
from state import SignalState, DetectionState, AnalysisState, UIState, SessionState
from navigation_controller import NavigationController
from export_controller import ExportController

# ── ecg.io ────────────────────────────────────────────────────────────────────
from loaders import (
    load_mat_signal, list_channels,
    _serialise_results, _deserialise_results,
)
from session import (
    save_session, load_session, delete_session,
)
from db import (
    _DB_AVAILABLE, get_notes, set_notes,
    recent_recordings, upsert_recording,
)
from export import ExcelExporter, PrismExporter

# ── ecg.ui ────────────────────────────────────────────────────────────────────
from theme import (
    THEME, ThemeConfig, apply_theme_config, apply_plot_theme, make_font,
    NK_AVAILABLE, H5_AVAILABLE, APP_ICON_PATH,
    BG, PANEL, CARD, BORDER, BORDER2, TEXT, MUTED, LIGHT, PLOT,
    RED, BLUE, GREEN, ORANGE,
    BLUE_DARK, BLUE_HOVER, BLUE_MID, BLUE_DEEP,
    PURPLE, PURPLE_DARK, PINK, TEAL, CYAN, CYAN_BRIGHT,
    GREEN_DARK, GREEN_MID, ORANGE_DARK, ORANGE_DEEP,
    AMBER, AMBER_DARK, RED_DARK, RED_MID, RED_LIGHT,
    CORAL, NAVY, GRAY, GRAY_LIGHT,
    FONT_TITLE, FONT_SECTION_HDR, FONT_LABEL, FONT_SMALL, FONT_BODY, FONT_MONO,
    FONT_KPI_VALUE, FONT_KPI_LABEL, FONT_BTN_PRIMARY, FONT_BTN_SEC, FONT_SIDEBAR_HDR,
    FONT_MICRO, FONT_HINT, FONT_BADGE, FONT_SUBSECTION, FONT_CARD_TITLE,
    nk,
)
from plots import CanvasSlot, style_axes
from dialogs import (
    ThemeDialog, ArtifactReviewDialog,
    AnnotationDialog, AnnotationManagerDialog,
)
from wave_editor import WaveTemplateMiniEditor, WaveTemplateEditor
from sidebar import _SidebarSection, IntervalVerifierPanel

log = logging.getLogger("ecg")

# ════════════════════════════════════════════════════════════
#  SPACING SCALE — single source of truth for padx/pady values
# ════════════════════════════════════════════════════════════
# Avant : 30+ couples (a, b) différents et arbitraires dispersés dans le
# fichier (pady=(0,4), (8,4), (6,2), (18,4), (5,1)...) → rythme visuel
# incohérent entre cartes/sections, sans logique apparente.
#
# Après : 4 paliers nommés couvrant tous les usages réels du fichier.
# Toute nouvelle valeur doit être l'un de ces 4 paliers plutôt qu'un
# nombre choisi au hasard. La valeur 0 (pas d'espacement, volontaire)
# reste un littéral explicite — elle n'est jamais arrondie à SPACE_XS.
SPACE_XS = 2   # éléments très liés : label + son champ, icône + texte
SPACE_S  = 4   # lignes à l'intérieur d'une même carte
SPACE_M  = 8   # entre cartes / sous-sections, entre graphiques empilés
SPACE_L  = 12  # entre blocs majeurs, séparateurs horizontaux, top-level

# ════════════════════════════════════════════════════════════
#  SPACING SCALE — single source of truth for padx/pady values
# ════════════════════════════════════════════════════════════
# Avant : 30+ couples (a, b) différents et arbitraires dispersés dans le
# fichier (pady=(0, SPACE_S), (8,4), (6,2), (18,4), (5,1)...) → rythme visuel
# incohérent entre cartes/sections.
#
# Après : 4 paliers nommés couvrant tous les usages réels du fichier.
# Toute nouvelle valeur doit être l'un de ces 4 paliers (ou une combinaison
# explicite SPACE_X/SPACE_Y) plutôt qu'un nombre choisi au hasard.
SPACE_XS = 2   # éléments très liés : label + son champ, icône + texte
SPACE_S  = 4   # lignes à l'intérieur d'une même carte
SPACE_M  = 8   # entre cartes / sous-sections, entre graphiques empilés
SPACE_L  = 12  # entre blocs majeurs, séparateurs horizontaux, top-level

class ECGApp(ctk.CTk):

    def __init__(self) -> None:
        super().__init__()
        self.title("ECG Analysis")
        self.geometry("1920x1080")
        self.minsize(1200, 750)
        # Open in fullscreen/maximized state
        self.after(100, lambda: self.state("zoomed"))  # Windows fullscreen
        self.configure(fg_color=BG)

        # Forward declarations for widgets built during _build() — lets Pylance
        # know these attributes exist even though they're assigned later.
        # (declaration moved to _init_state)
        self.ent_subject:  ctk.CTkEntry
        self.ent_fs:       ctk.CTkEntry
        self.ent_epoch:    ctk.CTkEntry
        self.ent_overlap:  ctk.CTkEntry
        self.txt_td:       ctk.CTkTextbox
        self.txt_fd:       ctk.CTkTextbox

        self._load_icon()
        self._init_state()
        self._build()
        self.after(200, self._setup_dnd)
        self._bind_keyboard_shortcuts()

    # ─── Startup helpers ──────────────────────────────────────

    def _on_window_resize(self, event) -> None:
        """Adjust sidebar width proportionally when the window is resized.

        Debounced to 80 ms so rapid drag-resize events don't flood the layout
        engine with redundant configure calls.
        """
        if event.widget is not self:
            return
        if self._thr_debounce_id is not None:
            try:
                self.after_cancel(self._thr_debounce_id)
            except Exception:
                pass
        new_w = max(220, min(340, int(event.width * 0.20)))
        self._thr_debounce_id = self.after(
            80, lambda w=new_w: self._apply_resize(w)
        )

    def _apply_resize(self, new_w: int) -> None:
        """Apply the debounced sidebar width update."""
        self._thr_debounce_id = None
        try:
            self.sidebar.configure(width=new_w)
        except Exception as e:
            log.debug("sidebar.configure width failed: %s", e)

    def _load_icon(self) -> None:
        if not APP_ICON_PATH:
            return
        try:
            # Pillow 9+ uses Image.Resampling.LANCZOS; older versions use Image.LANCZOS
            _resample = (getattr(getattr(Image, "Resampling", None), "LANCZOS", None)
                         or getattr(Image, "LANCZOS", Image.BICUBIC))  # type: ignore[attr-defined]
            img = Image.open(APP_ICON_PATH).resize((256, 256), _resample)
            # PhotoImage is accepted at runtime; Tkinter stubs are incomplete
            self.iconphoto(True, ImageTk.PhotoImage(img))  # type: ignore[arg-type]
        except Exception as exc:
            log.debug("App icon not loaded: %s", exc)

    def _init_state(self) -> None:
        """Initialise all instance variables that hold application state.

        Non-widget state lives in typed dataclasses from state.py (`self.signal`,
        `self.detection`, `self.analysis`, `self.ui`, `self.session`). The
        "Legacy state shims" property block right after this method keeps every
        existing `self._xxx` access (inside this class and in dialogs.py /
        wave_editor.py / models.py) working unchanged. New code should read/write
        the dataclasses directly.
        """
        self.signal = SignalState()
        self.detection = DetectionState()
        self.analysis = AnalysisState()
        self.ui = UIState()
        self.session = SessionState()

        self.nav_ctrl = NavigationController(self)
        self.export_ctrl = ExportController(self)

        self._batch_bc_outdir: ctk.CTkEntry
        self._batch_bc_channel: ctk.CTkEntry
        self._batch_bc_workers: ctk.CTkEntry

        # Widget registries (populated in _build)
        self._slots:       dict[str, CanvasSlot]   = {}
        self._kpi:         dict[str, ctk.CTkLabel] = {}

        # ── Forward declarations for widgets created in _build ────────────────
        # Declared here so that methods called before the UI is fully constructed
        # (e.g. _collect_session_state called during startup restore) get a
        # predictable None rather than raising AttributeError, and so that
        # hasattr(self, ...) guards can be replaced with "is not None" checks.
        #
        # Sidebar / session controls
        self.btn_save_session:  "Optional[ctk.CTkButton]"  = None
        self.lbl_session_info:  "Optional[ctk.CTkLabel]"   = None
        self.lbl_template_info: "Optional[ctk.CTkLabel]"   = None
        # Per-module on-demand analysis buttons + status labels
        self.btn_run_freq:      "Optional[ctk.CTkButton]"  = None
        self.btn_run_nonlin:    "Optional[ctk.CTkButton]"  = None
        self.btn_run_ivl:       "Optional[ctk.CTkButton]"  = None
        self.btn_run_arrhythmia:"Optional[ctk.CTkButton]"  = None
        self.frm_ivl_nav:       "Optional[tk.Frame]"       = None
        self._ivl_verifier:     "Optional[IntervalVerifierPanel]" = None
        self.lbl_freq_status:   "Optional[ctk.CTkLabel]"   = None
        self.lbl_nonlin_status: "Optional[ctk.CTkLabel]"   = None
        self.lbl_ivl_status:    "Optional[ctk.CTkLabel]"   = None
        # Filter / signal controls (widgets used in _collect_session_state)
        self.ent_channel:       "Optional[ctk.CTkEntry]"   = None
        self.ent_t_start:       "Optional[ctk.CTkEntry]"   = None
        self.ent_t_end:         "Optional[ctk.CTkEntry]"   = None
        self.ent_analysis_t0:   "Optional[ctk.CTkEntry]"   = None
        self.ent_analysis_t1:   "Optional[ctk.CTkEntry]"   = None
        self.lbl_analysis_window: "Optional[ctk.CTkLabel]" = None
        self.ent_lp:            "Optional[ctk.CTkEntry]"   = None
        self.ent_hp:            "Optional[ctk.CTkEntry]"   = None
        self.ent_minrr:         "Optional[ctk.CTkEntry]"   = None
        self.sw_notch:          "Optional[ctk.CTkSwitch]"  = None
        self.sw_invert_signal:  "Optional[ctk.CTkSwitch]"  = None
        self.sw_filter_preview: "Optional[ctk.CTkSwitch]"  = None
        self.cb_clean:          "Optional[ctk.CTkComboBox]" = None
        self.sl_thr:            "Optional[ctk.CTkSlider]"  = None
        self.ent_thr:           "Optional[ctk.CTkEntry]"   = None
        self.sw_permissive:     "Optional[ctk.CTkSwitch]"  = None
        self.btn_annotations:   "Optional[ctk.CTkButton]"  = None
        self.lbl_ann_count:     "Optional[ctk.CTkLabel]"   = None
        self.btn_copy_rr:       "Optional[ctk.CTkButton]"  = None
        self.btn_copy_ivl:      "Optional[ctk.CTkButton]"  = None
        self.btn_copy_epochs:   "Optional[ctk.CTkButton]"  = None
        self.cb_det_method:     "Optional[ctk.CTkComboBox]"  = None
        self.ent_sg_target_fs:  "Optional[ctk.CTkEntry]"     = None
        self.ent_sg_window_ms:  "Optional[ctk.CTkEntry]"     = None
        self._sg_frame:         "Optional[ctk.CTkFrame]"     = None
        self.btn_lang:          "Optional[ctk.CTkButton]"    = None
        self.cb_qtc_formula:    "Optional[ctk.CTkComboBox]"  = None
        self.cb_freq_band:      "Optional[ctk.CTkComboBox]"  = None   # HRV band preset
        # Sidebar / detection tab widgets (forward-declared)
        self.lbl_npeaks:           "Optional[ctk.CTkLabel]"              = None
        self.btn_review_art:       "Optional[ctk.CTkButton]"             = None
        self.lbl_file:             "Optional[ctk.CTkLabel]"              = None
        self.lbl_context_subtitle: "Optional[ctk.CTkLabel]"              = None
        self.lbl_arrhythmia_status:"Optional[ctk.CTkLabel]"              = None
        self.lbl_roll_status:      "Optional[ctk.CTkLabel]"              = None
        self.lbl_arr_event_title:  "Optional[ctk.CTkLabel]"              = None
        self._arr_card_widgets:    "list[ctk.CTkBaseClass]"              = []
        self._adv_filters_frame:   "Optional[ctk.CTkFrame]"              = None
        self._interp_scroll:       "Optional[ctk.CTkScrollableFrame]"    = None
        # Navigation bar widgets (forward-declared so _reset_for_new_file can update them)
        self.ent_nav_pos:       "Optional[ctk.CTkEntry]"    = None
        self.lbl_sig_duration:  "Optional[ctk.CTkLabel]"    = None
        # Intervals tab widgets
        self.ent_max_beats:     "Optional[ctk.CTkEntry]"    = None  # removed from UI but kept for compat
        # Interpretation removed
        self._interp_cards:      "dict[str, tuple[ctk.CTkLabel, ctk.CTkLabel, str]]" = {}
        self._interp_ref_labels: "dict[str, ctk.CTkLabel]" = {}
        # HRV unified tab internal state
        self._hrv_subframes:     "dict[str, ctk.CTkFrame]" = {}
        self._hrv_seg:           "Optional[ctk.CTkSegmentedButton]" = None
        self._hrv_content_area:  "Optional[ctk.CTkFrame]"           = None
        # Summary KPI label widgets (populated in _build_tab_summary)
        self._sum_kpi_vals:      "dict[str, ctk.CTkLabel]" = {}

    # ─── Legacy state shims ───────────────────────────────────────────────
    # Backward-compatible properties for the old flat self._xxx attributes,
    # now held in state.py dataclasses. Every existing method body in this
    # class, plus the sibling modules that reach into an ECGApp instance
    # (dialogs.py: _rebuild_ui/_annotations/_time/_session_dirty/_draw_detail/
    # _update_ann_count; wave_editor.py: _wave_template; models.py:
    # _peak_distance_ms/_safe_float), keep working unchanged through these.
    # New/extracted code should read and write self.signal/.detection/
    # .analysis/.ui/.session directly instead of these shims.

    # -- SignalState --
    @property
    def _filepath(self) -> "Optional[str]":
        return self.signal.filepath
    @_filepath.setter
    def _filepath(self, value: "Optional[str]") -> None:
        self.signal.filepath = value

    @property
    def _signal_raw(self) -> "Optional[np.ndarray]":
        return self.signal.raw
    @_signal_raw.setter
    def _signal_raw(self, value: "Optional[np.ndarray]") -> None:
        self.signal.raw = value

    @property
    def _signal_raw_norm(self) -> "Optional[np.ndarray]":
        return self.signal.raw_norm
    @_signal_raw_norm.setter
    def _signal_raw_norm(self, value: "Optional[np.ndarray]") -> None:
        self.signal.raw_norm = value

    @property
    def _signal_flt(self) -> "Optional[np.ndarray]":
        return self.signal.filtered
    @_signal_flt.setter
    def _signal_flt(self, value: "Optional[np.ndarray]") -> None:
        self.signal.filtered = value

    @property
    def _time(self) -> "Optional[np.ndarray]":
        return self.signal.time
    @_time.setter
    def _time(self, value: "Optional[np.ndarray]") -> None:
        self.signal.time = value

    @property
    def _fs(self) -> int:
        return self.signal.fs
    @_fs.setter
    def _fs(self, value: int) -> None:
        self.signal.fs = value

    @property
    def _peak_distance_ms(self) -> float:
        return self.signal.peak_distance_ms
    @_peak_distance_ms.setter
    def _peak_distance_ms(self, value: float) -> None:
        self.signal.peak_distance_ms = value

    @property
    def _raw_only_loaded(self) -> bool:
        return self.signal.raw_only_loaded
    @_raw_only_loaded.setter
    def _raw_only_loaded(self, value: bool) -> None:
        self.signal.raw_only_loaded = value

    @property
    def _no_filter_mode(self) -> bool:
        return self.signal.no_filter_mode
    @_no_filter_mode.setter
    def _no_filter_mode(self, value: bool) -> None:
        self.signal.no_filter_mode = value

    @property
    def _signal_inverted(self) -> bool:
        return self.signal.inverted
    @_signal_inverted.setter
    def _signal_inverted(self, value: bool) -> None:
        self.signal.inverted = value

    # -- DetectionState --
    @property
    def _rpeaks_ok(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_ok
    @_rpeaks_ok.setter
    def _rpeaks_ok(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_ok = value

    @property
    def _rpeaks_rej(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_rej
    @_rpeaks_rej.setter
    def _rpeaks_rej(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_rej = value

    @property
    def _all_cands(self) -> "Optional[np.ndarray]":
        return self.detection.all_candidates
    @_all_cands.setter
    def _all_cands(self, value: "Optional[np.ndarray]") -> None:
        self.detection.all_candidates = value

    @property
    def _all_proms(self) -> "Optional[np.ndarray]":
        return self.detection.all_prominences
    @_all_proms.setter
    def _all_proms(self, value: "Optional[np.ndarray]") -> None:
        self.detection.all_prominences = value

    @property
    def _thresh_amp(self) -> float:
        return self.detection.thresh_amp
    @_thresh_amp.setter
    def _thresh_amp(self, value: float) -> None:
        self.detection.thresh_amp = value

    @property
    def _sig_quality(self) -> "Optional[int]":
        return self.detection.sig_quality
    @_sig_quality.setter
    def _sig_quality(self, value: "Optional[int]") -> None:
        self.detection.sig_quality = value

    @property
    def _manual_excluded(self) -> "set[int]":
        return self.detection.manual_excluded
    @_manual_excluded.setter
    def _manual_excluded(self, value: "set[int]") -> None:
        self.detection.manual_excluded = value

    @property
    def _rpeaks_manual_excl(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_manual_excl
    @_rpeaks_manual_excl.setter
    def _rpeaks_manual_excl(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_manual_excl = value

    @property
    def _manual_added(self) -> "set[int]":
        return self.detection.manual_added
    @_manual_added.setter
    def _manual_added(self, value: "set[int]") -> None:
        self.detection.manual_added = value

    @property
    def _rpeaks_manual_added(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_manual_added
    @_rpeaks_manual_added.setter
    def _rpeaks_manual_added(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_manual_added = value

    @property
    def _edit_mode(self) -> bool:
        return self.detection.edit_mode
    @_edit_mode.setter
    def _edit_mode(self, value: bool) -> None:
        self.detection.edit_mode = value

    @property
    def _edit_free_placement(self) -> bool:
        return self.detection.edit_free_placement
    @_edit_free_placement.setter
    def _edit_free_placement(self, value: bool) -> None:
        self.detection.edit_free_placement = value

    @property
    def _edit_undo(self) -> "list[tuple[frozenset, frozenset]]":
        return self.detection.edit_undo
    @_edit_undo.setter
    def _edit_undo(self, value: "list[tuple[frozenset, frozenset]]") -> None:
        self.detection.edit_undo = value

    @property
    def _edit_redo(self) -> "list[tuple[frozenset, frozenset]]":
        return self.detection.edit_redo
    @_edit_redo.setter
    def _edit_redo(self, value: "list[tuple[frozenset, frozenset]]") -> None:
        self.detection.edit_redo = value

    @property
    def _EDIT_UNDO_LIMIT(self) -> int:
        return DetectionState.EDIT_UNDO_LIMIT

    @property
    def _hover_samp(self) -> "Optional[int]":
        return self.detection.hover_samp
    @_hover_samp.setter
    def _hover_samp(self, value: "Optional[int]") -> None:
        self.detection.hover_samp = value

    @property
    def _hover_samp_near(self) -> bool:
        return self.detection.hover_samp_near
    @_hover_samp_near.setter
    def _hover_samp_near(self, value: bool) -> None:
        self.detection.hover_samp_near = value

    # -- AnalysisState --
    @property
    def _results(self) -> "Optional[dict]":
        return self.analysis.results
    @_results.setter
    def _results(self, value: "Optional[dict]") -> None:
        self.analysis.results = value

    @property
    def _epoch_df(self) -> "Optional[pd.DataFrame]":
        return self.analysis.epoch_df
    @_epoch_df.setter
    def _epoch_df(self, value: "Optional[pd.DataFrame]") -> None:
        self.analysis.epoch_df = value

    @property
    def _rolling_hrv_df(self) -> "Optional[pd.DataFrame]":
        return self.analysis.rolling_hrv_df
    @_rolling_hrv_df.setter
    def _rolling_hrv_df(self, value: "Optional[pd.DataFrame]") -> None:
        self.analysis.rolling_hrv_df = value

    @property
    def _arrhythmia_events(self) -> "list":
        return self.analysis.arrhythmia_events
    @_arrhythmia_events.setter
    def _arrhythmia_events(self, value: "list") -> None:
        self.analysis.arrhythmia_events = value

    @property
    def _arrhythmia_tsv(self) -> str:
        return self.analysis.arrhythmia_tsv
    @_arrhythmia_tsv.setter
    def _arrhythmia_tsv(self, value: str) -> None:
        self.analysis.arrhythmia_tsv = value

    @property
    def _arr_selected_idx(self) -> int:
        return self.analysis.arr_selected_idx
    @_arr_selected_idx.setter
    def _arr_selected_idx(self, value: int) -> None:
        self.analysis.arr_selected_idx = value

    @property
    def _arr_nav_pos(self) -> float:
        return self.analysis.arr_nav_pos
    @_arr_nav_pos.setter
    def _arr_nav_pos(self, value: float) -> None:
        self.analysis.arr_nav_pos = value

    @property
    def _arr_win(self) -> float:
        return self.analysis.arr_win
    @_arr_win.setter
    def _arr_win(self, value: float) -> None:
        self.analysis.arr_win = value

    @property
    def _arr_edit_mode(self) -> bool:
        return self.analysis.arr_edit_mode
    @_arr_edit_mode.setter
    def _arr_edit_mode(self, value: bool) -> None:
        self.analysis.arr_edit_mode = value

    @property
    def _last_seg_a(self) -> "Optional[dict]":
        return self.analysis.last_seg_a
    @_last_seg_a.setter
    def _last_seg_a(self, value: "Optional[dict]") -> None:
        self.analysis.last_seg_a = value

    @property
    def _last_seg_b(self) -> "Optional[dict]":
        return self.analysis.last_seg_b
    @_last_seg_b.setter
    def _last_seg_b(self, value: "Optional[dict]") -> None:
        self.analysis.last_seg_b = value

    @property
    def _artifact_report(self) -> "Optional[dict]":
        return self.analysis.artifact_report
    @_artifact_report.setter
    def _artifact_report(self, value: "Optional[dict]") -> None:
        self.analysis.artifact_report = value

    @property
    def _exp_context(self) -> str:
        return self.analysis.exp_context
    @_exp_context.setter
    def _exp_context(self, value: str) -> None:
        self.analysis.exp_context = value

    @property
    def _analysis_t_start(self) -> float:
        return self.analysis.t_start
    @_analysis_t_start.setter
    def _analysis_t_start(self, value: float) -> None:
        self.analysis.t_start = value

    @property
    def _analysis_t_end(self) -> float:
        return self.analysis.t_end
    @_analysis_t_end.setter
    def _analysis_t_end(self, value: float) -> None:
        self.analysis.t_end = value

    @property
    def _annotations(self) -> "list[dict]":
        return self.analysis.annotations
    @_annotations.setter
    def _annotations(self, value: "list[dict]") -> None:
        self.analysis.annotations = value

    @property
    def _wave_template(self) -> "Optional[WaveTemplate]":
        return self.analysis.wave_template
    @_wave_template.setter
    def _wave_template(self, value: "Optional[WaveTemplate]") -> None:
        self.analysis.wave_template = value

    # -- UIState --
    @property
    def _nav_pos(self) -> float:
        return self.ui.nav_pos
    @_nav_pos.setter
    def _nav_pos(self, value: float) -> None:
        self.ui.nav_pos = value

    @property
    def _dark_mode(self) -> bool:
        return self.ui.dark_mode
    @_dark_mode.setter
    def _dark_mode(self, value: bool) -> None:
        self.ui.dark_mode = value

    @property
    def _show_raw(self) -> bool:
        return self.ui.show_raw
    @_show_raw.setter
    def _show_raw(self, value: bool) -> None:
        self.ui.show_raw = value

    @property
    def _filter_preview_on(self) -> bool:
        return self.ui.filter_preview_on
    @_filter_preview_on.setter
    def _filter_preview_on(self, value: bool) -> None:
        self.ui.filter_preview_on = value

    @property
    def _thr_debounce_id(self) -> "str | None":
        return self.ui.thr_debounce_id
    @_thr_debounce_id.setter
    def _thr_debounce_id(self, value: "str | None") -> None:
        self.ui.thr_debounce_id = value

    @property
    def _hover_motion_cid(self) -> "Optional[int]":
        return self.ui.hover_motion_cid
    @_hover_motion_cid.setter
    def _hover_motion_cid(self, value: "Optional[int]") -> None:
        self.ui.hover_motion_cid = value

    @property
    def _hover_after_id(self) -> "str | None":
        return self.ui.hover_after_id
    @_hover_after_id.setter
    def _hover_after_id(self, value: "str | None") -> None:
        self.ui.hover_after_id = value

    @property
    def _beat_nav_cid(self) -> "Optional[int]":
        return self.ui.beat_nav_cid
    @_beat_nav_cid.setter
    def _beat_nav_cid(self, value: "Optional[int]") -> None:
        self.ui.beat_nav_cid = value

    @property
    def _rr_click_cid(self) -> "Optional[int]":
        return self.ui.rr_click_cid
    @_rr_click_cid.setter
    def _rr_click_cid(self, value: "Optional[int]") -> None:
        self.ui.rr_click_cid = value

    @property
    def _ov_ylim(self) -> "Optional[tuple]":
        return self.ui.ov_ylim
    @_ov_ylim.setter
    def _ov_ylim(self, value: "Optional[tuple]") -> None:
        self.ui.ov_ylim = value

    @property
    def _hrv_current_view(self) -> str:
        return self.ui.hrv_current_view
    @_hrv_current_view.setter
    def _hrv_current_view(self, value: str) -> None:
        self.ui.hrv_current_view = value

    @property
    def _ctx_keys(self) -> "list":
        return self.ui.ctx_keys
    @_ctx_keys.setter
    def _ctx_keys(self, value: "list") -> None:
        self.ui.ctx_keys = value

    @property
    def _operation_start_time(self) -> "Optional[float]":
        return self.ui.operation_start_time
    @_operation_start_time.setter
    def _operation_start_time(self, value: "Optional[float]") -> None:
        self.ui.operation_start_time = value

    @property
    def _ui_update_batch(self) -> "list[tuple]":
        return self.ui.ui_update_batch
    @_ui_update_batch.setter
    def _ui_update_batch(self, value: "list[tuple]") -> None:
        self.ui.ui_update_batch = value

    @property
    def _figure_cache(self) -> "dict":
        return self.ui.figure_cache
    @_figure_cache.setter
    def _figure_cache(self, value: "dict") -> None:
        self.ui.figure_cache = value

    @property
    def _tsv_store(self) -> "dict[int, str]":
        return self.ui.tsv_store
    @_tsv_store.setter
    def _tsv_store(self, value: "dict[int, str]") -> None:
        self.ui.tsv_store = value

    @property
    def _ds_time(self) -> "Optional[np.ndarray]":
        return self.ui.ds_time
    @_ds_time.setter
    def _ds_time(self, value: "Optional[np.ndarray]") -> None:
        self.ui.ds_time = value

    @property
    def _ds_sig(self) -> "Optional[np.ndarray]":
        return self.ui.ds_sig
    @_ds_sig.setter
    def _ds_sig(self, value: "Optional[np.ndarray]") -> None:
        self.ui.ds_sig = value

    @property
    def _ds_sig_max(self) -> "Optional[np.ndarray]":
        return self.ui.ds_sig_max
    @_ds_sig_max.setter
    def _ds_sig_max(self, value: "Optional[np.ndarray]") -> None:
        self.ui.ds_sig_max = value

    @property
    def _ds_sig_mid(self) -> "Optional[np.ndarray]":
        return self.ui.ds_sig_mid
    @_ds_sig_mid.setter
    def _ds_sig_mid(self, value: "Optional[np.ndarray]") -> None:
        self.ui.ds_sig_mid = value

    @property
    def _ds_raw_sig(self) -> "Optional[np.ndarray]":
        return self.ui.ds_raw_sig
    @_ds_raw_sig.setter
    def _ds_raw_sig(self, value: "Optional[np.ndarray]") -> None:
        self.ui.ds_raw_sig = value

    @property
    def _ds_raw_sig_max(self) -> "Optional[np.ndarray]":
        return self.ui.ds_raw_sig_max
    @_ds_raw_sig_max.setter
    def _ds_raw_sig_max(self, value: "Optional[np.ndarray]") -> None:
        self.ui.ds_raw_sig_max = value

    @property
    def _ds_raw_sig_mid(self) -> "Optional[np.ndarray]":
        return self.ui.ds_raw_sig_mid
    @_ds_raw_sig_mid.setter
    def _ds_raw_sig_mid(self, value: "Optional[np.ndarray]") -> None:
        self.ui.ds_raw_sig_mid = value

    # -- SessionState --
    @property
    def _recent(self) -> "list[str]":
        return self.session.recent_files
    @_recent.setter
    def _recent(self, value: "list[str]") -> None:
        self.session.recent_files = value

    @property
    def _session_dirty(self) -> bool:
        return self.session.dirty
    @_session_dirty.setter
    def _session_dirty(self, value: bool) -> None:
        self.session.dirty = value

    @property
    def _recording_notes(self) -> str:
        return self.session.recording_notes
    @_recording_notes.setter
    def _recording_notes(self, value: str) -> None:
        self.session.recording_notes = value

    def _batch_ui_update(self, widget: ctk.CTkBaseClass, **kwargs) -> None:
        """Batch UI updates to reduce flicker during session restore."""
        self._ui_update_batch.append((widget, kwargs))

    def _flush_ui_updates(self) -> None:
        """Apply all batched UI updates at once."""
        for widget, kwargs in self._ui_update_batch:
            try:
                widget.configure(**kwargs)
            except Exception as exc:
                log.debug("Batch UI update failed: %s", exc)
        self._ui_update_batch.clear()

    def _current_ref(self, key: str) -> "tuple[float, float]":
        """Return (lo, hi) reference range for the given metric key from current experimental context."""
        ctx = EXPERIMENTAL_CONTEXTS.get(self._exp_context)
        if ctx is None:
            # Fallback defaults
            if "HR" in key:
                return MouseECG.HR_MIN_BPM, MouseECG.HR_MAX_BPM
            elif "RR" in key:
                return MouseECG.RR_MIN_MS, MouseECG.RR_MAX_MS
            else:
                return 0.0, 1000.0
        lo_field, hi_field = _CONTEXT_FIELD_MAP.get(key, ("hr_lo", "hr_hi"))
        return getattr(ctx, lo_field), getattr(ctx, hi_field)

    # ════════════════════════════════════════════════════════
    #  UI CONSTRUCTION
    # ════════════════════════════════════════════════════════

    def _build(self) -> None:
        self.update_idletasks()
        w = self.winfo_width() or 1480
        _sidebar_w = max(220, min(340, int(w * 0.20)))
        self.sidebar = ctk.CTkFrame(self, width=_sidebar_w, fg_color=PANEL, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self.bind("<Configure>", self._on_window_resize)
        # Keyboard navigation shortcuts (active when focus is on the main window)
        self.bind("<Left>",  lambda e: self._kb_navigate(-1))
        self.bind("<Right>", lambda e: self._kb_navigate(+1))
        self._build_sidebar()
        # Apply the default no-filter state immediately so filter widgets are
        # greyed out at startup (no_filter is ON by default).
        self.after(50, self._on_no_filter_toggle)

        main = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        main.pack(side="left", fill="both", expand=True)
        self._build_kpi_bar(main)
        self._build_tabs(main)
        # Global keyboard shortcuts
        self.bind("<Control-z>", self._undo_edit)
        self.bind("<Control-Z>", self._undo_edit)
        self.bind("<Control-y>", self._redo_edit)
        self.bind("<Control-Y>", self._redo_edit)

    # ─── Sidebar ──────────────────────────────────────────────

    def _build_sidebar(self) -> None:
        """Compact sidebar: minimal always-visible workflow + scrollable settings."""
        px = dict(padx=SPACE_L)

        # ══════════════════════════════════════════════════════
        #  FIXED TOP — title, file, workflow actions, threshold
        # ══════════════════════════════════════════════════════
        top = ctk.CTkFrame(self.sidebar, fg_color=PANEL, corner_radius=0)
        top.pack(fill="x", side="top")

        # Title + version
        title_row = ctk.CTkFrame(top, fg_color="transparent")
        title_row.pack(fill="x", padx=SPACE_L, pady=(SPACE_M, SPACE_XS))
        ctk.CTkLabel(title_row, text="ECG Analysis",
                     font=FONT_TITLE, text_color=TEXT,
                     anchor="w").pack(side="left")
        ctk.CTkLabel(title_row, text="v6",
                     font=FONT_HINT, text_color=LIGHT).pack(side="right")

        # File label
        self.lbl_file = ctk.CTkLabel(
            top, text="No file loaded", font=FONT_SMALL,
            text_color=MUTED, wraplength=260, anchor="w", justify="left")
        self.lbl_file.pack(**px, pady=(0, SPACE_XS), fill="x")

        # Open button + Channels / Recent on same row
        open_row = ctk.CTkFrame(top, fg_color="transparent")
        open_row.pack(fill="x", expand=True, **px, pady=(0, SPACE_XS))
        open_row.rowconfigure(0, weight=1)
        open_row.columnconfigure(0, weight=3)
        open_row.columnconfigure(1, weight=1)
        open_row.columnconfigure(2, weight=1)
        ctk.CTkButton(open_row, text="Open .mat file",
                      fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
                      font=FONT_BTN_PRIMARY, height=30, corner_radius=8,
                      command=self._open_file).grid(row=0, column=0, sticky="ewns", padx=(0, SPACE_XS))
        ctk.CTkButton(open_row, text="Channel", height=10,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, corner_radius=6,
                      command=self._show_channels).grid(row=0, column=1, sticky="ewns", padx=(0, SPACE_XS))
        ctk.CTkButton(open_row, text="Recent", height=10,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, corner_radius=6,
                      command=self._open_recent).grid(row=0, column=2, sticky="ewns")

        ctk.CTkFrame(top, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_L, pady=(SPACE_XS, SPACE_XS))

        # Peak count status
        self.lbl_npeaks = ctk.CTkLabel(
            top, text="Peaks detected: —",
            font=FONT_CARD_TITLE, text_color=BLUE, anchor="w")
        self.lbl_npeaks.pack(**px, fill="x", pady=(0, SPACE_XS))

        # ── WORKFLOW: 3 inline action rows ───────────────────
        # Row 1: Preview Detection
        self.btn_preview = ctk.CTkButton(
            top, text="1 ▶  Preview Detection",
            command=self._preview,
            fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
            font=FONT_BTN_PRIMARY, height=34, corner_radius=8)
        self.btn_preview.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))

        # Method selector — always visible, directly under Preview
        det_card = ctk.CTkFrame(top, fg_color=CARD, corner_radius=8,
                                border_width=1, border_color=BORDER)
        det_card.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        det_card_row = ctk.CTkFrame(det_card, fg_color="transparent")
        det_card_row.pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        det_card_row.columnconfigure(1, weight=1)
        ctk.CTkLabel(det_card_row, text="Method",
                     font=FONT_SIDEBAR_HDR, text_color=TEXT,
                     anchor="w").grid(row=0, column=0, sticky="w", padx=(0, SPACE_M))
        self.cb_det_method = ctk.CTkComboBox(
            det_card_row, font=FONT_LABEL, height=26,
            fg_color=BG, border_color=BLUE, button_color=BLUE,
            text_color=TEXT, dropdown_fg_color=BG, dropdown_text_color=TEXT,
            values=["Auto (NeuroKit2)", "SG + Derivative (10 kHz)", "Wavelet (CWT)", "Envelope Max"],
            command=self._on_det_method_change)
        self.cb_det_method.set("SG + Derivative (10 kHz)")
        self.cb_det_method.grid(row=0, column=1, sticky="ew")

        # Row 2: Analysis window (inline, compact)
        aw_frame = ctk.CTkFrame(top, fg_color=CARD, corner_radius=8,
                                border_width=1, border_color=BORDER)
        aw_frame.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        aw_hdr = ctk.CTkFrame(aw_frame, fg_color="transparent")
        aw_hdr.pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        ctk.CTkLabel(aw_hdr, text="2  Analysis window", font=FONT_SUBSECTION,
                     text_color=TEXT, anchor="w").pack(side="left")
        ctk.CTkLabel(aw_hdr, text="optional", font=FONT_KPI_LABEL,
                     text_color=LIGHT).pack(side="right")
        aw_entries = ctk.CTkFrame(aw_frame, fg_color="transparent")
        aw_entries.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        aw_entries.columnconfigure(1, weight=1)
        aw_entries.columnconfigure(3, weight=1)
        ctk.CTkLabel(aw_entries, text="From", font=FONT_SMALL, text_color=MUTED,
                     width=36, anchor="w").grid(row=0, column=0)
        self.ent_analysis_t0 = ctk.CTkEntry(
            aw_entries, height=24, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT,
            corner_radius=5, placeholder_text="0 s")
        self.ent_analysis_t0.grid(row=0, column=1, sticky="ew", padx=(SPACE_XS, SPACE_M))
        ctk.CTkLabel(aw_entries, text="To", font=FONT_SMALL, text_color=MUTED,
                     width=24, anchor="w").grid(row=0, column=2)
        self.ent_analysis_t1 = ctk.CTkEntry(
            aw_entries, height=24, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT,
            corner_radius=5, placeholder_text="end")
        self.ent_analysis_t1.grid(row=0, column=3, sticky="ew", padx=(SPACE_XS, 0))
        aw_btns = ctk.CTkFrame(aw_frame, fg_color="transparent")
        aw_btns.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkButton(aw_btns, text="Apply", width=70, height=26,
                      fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
                      font=FONT_SMALL, corner_radius=5,
                      command=self._apply_analysis_window).pack(side="left", padx=(0, SPACE_S))
        ctk.CTkButton(aw_btns, text="Full", width=56, height=26,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_SMALL, corner_radius=5,
                      command=self._reset_analysis_window).pack(side="left")
        self.lbl_analysis_window = ctk.CTkLabel(
            aw_btns, text="", font=FONT_KPI_LABEL, text_color=MUTED, anchor="e")
        self.lbl_analysis_window.pack(side="right", fill="x", expand=True)

        # Row 3: Run Full Analysis
        run_color = GREEN if NK_AVAILABLE else BORDER2
        self.btn_run = ctk.CTkButton(
            top, text="3 ▶▶  Run Full Analysis",
            command=self._run_analysis,
            fg_color=run_color, hover_color=GREEN_DARK if NK_AVAILABLE else BORDER2,
            text_color="white", font=FONT_BTN_PRIMARY, height=34, corner_radius=8,
            state="normal" if NK_AVAILABLE else "disabled")
        self.btn_run.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))

        if not NK_AVAILABLE:
            ctk.CTkLabel(top, text="⚠  pip install neurokit2",
                         font=FONT_HINT, text_color=ORANGE, anchor="w").pack(**px, fill="x")

        ctk.CTkFrame(top, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))

        # Status
        self.lbl_status = ctk.CTkLabel(
            top, text="Ready", font=FONT_SMALL, text_color=MUTED,
            anchor="w", wraplength=260, justify="left")
        self.lbl_status.pack(**px, pady=(SPACE_XS, SPACE_XS), fill="x")

        # ── THRESHOLD — always visible, prominent ─────────────
        thr_card = ctk.CTkFrame(top, fg_color=CARD, corner_radius=8,
                                border_width=1, border_color=BORDER)
        thr_card.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
        thr_top_row = ctk.CTkFrame(thr_card, fg_color="transparent")
        thr_top_row.pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        self.lbl_thr = ctk.CTkLabel(
            thr_top_row, text="Threshold:  0.50",
            font=FONT_CARD_TITLE, text_color=TEXT, anchor="w")
        self.lbl_thr.pack(side="left")
        ctk.CTkLabel(thr_top_row, text="strict ↑  /  sensitive ↓",
                     font=FONT_KPI_LABEL, text_color=LIGHT, anchor="e").pack(side="right")
        self.sl_thr = ctk.CTkSlider(
            thr_card, from_=0.01, to=2.0,  # type: ignore
            progress_color=RED, button_color=RED,
            button_hover_color="#FF5252", fg_color=BORDER,
            height=20, command=self._on_threshold_slide)
        self.sl_thr.set(0.50)
        self.sl_thr.pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        thr_bot = ctk.CTkFrame(thr_card, fg_color="transparent")
        thr_bot.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
        thr_bot.columnconfigure(1, weight=1)
        ctk.CTkLabel(thr_bot, text="Exact:", font=FONT_SMALL,
                     text_color=MUTED).grid(row=0, column=0, padx=(0, SPACE_S))
        self.ent_thr = ctk.CTkEntry(thr_bot, height=24, font=FONT_LABEL,
                                     fg_color=PANEL, border_color=BORDER2, text_color=TEXT)
        self.ent_thr.insert(0, "0.50")
        self.ent_thr.grid(row=0, column=1, sticky="ew")
        self.ent_thr.bind("<Return>",   self._on_threshold_entry)
        self.ent_thr.bind("<FocusOut>", self._on_threshold_entry)

        # ══════════════════════════════════════════════════════
        #  SCROLLABLE SETTINGS
        # ══════════════════════════════════════════════════════
        scroll = ctk.CTkScrollableFrame(
            self.sidebar, fg_color=PANEL,
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=BORDER2)
        scroll.pack(fill="both", expand=True, side="top")
        s = scroll
        fpx = dict(padx=SPACE_L)

        # ── SIGNAL ────────────────────────────────────────────
        sec_sig = _SidebarSection(s, "SIGNAL", initially_open=False)
        f = sec_sig.frame
        self._sidebar_entry_row(f, fpx, [
            ("Channel", "channel", "ECG"),
            ("Subject ID", "subject", "subject_01"),
        ])
        self._sidebar_entry(f, "Sampling rate (Hz)", "fs",
                            str(MouseECG.FS_DEFAULT), fpx)
        self.lbl_fs_source = ctk.CTkLabel(
            f, text="", font=FONT_KPI_LABEL, text_color=MUTED,
            anchor="w", wraplength=230)
        self.lbl_fs_source.pack(**fpx, fill="x", pady=(0, SPACE_XS))
        self._sidebar_entry_row(f, fpx, [
            ("Crop start (s)", "t_start", "0"),
            ("Crop end (s)",   "t_end",   "0"),
        ])
        self.sw_show_raw = self._switch(
            f, "Show raw signal (vs filtered)", fpx, default_on=True)
        self.sw_show_raw.configure(command=self._on_show_raw_toggle)

        # ── FILTERS ───────────────────────────────────────────
        sec_flt = _SidebarSection(s, "FILTERS", initially_open=False)
        f = sec_flt.frame
        self.sw_no_filter = self._switch(
            f, "Raw signal — no DSP filters", fpx, default_on=True)
        self.sw_no_filter.configure(command=self._on_no_filter_toggle)
        ctk.CTkLabel(f, text="Enable filters below for advanced processing",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.sw_invert_signal = self._switch(
            f, "⟳  Invert signal (polarity)", fpx, default_on=False)
        self.sw_invert_signal.configure(command=lambda: self._preview())
        ctk.CTkLabel(f, text="Useful if R peaks appear negative",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.sw_filter_preview = self._switch(
            f, "👁  Preview filter effect (before/after)", fpx, default_on=False)
        self.sw_filter_preview.configure(command=self._on_filter_preview_toggle)
        ctk.CTkLabel(f, text="Overlays the filtered signal on the visible window,\n"
                             "using the settings below — no need to run Preview Detection.",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", justify="left", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))
        self._filter_widgets_frame = ctk.CTkFrame(f, fg_color="transparent")
        self._filter_widgets_frame.pack(**fpx, fill="x")
        fw = self._filter_widgets_frame
        self.sw_notch = self._switch(fw, "Notch 50 Hz", dict(padx=0), default_on=False)
        self.sw_notch.configure(command=self._refresh_filter_preview)
        self._adv_filters_open = False
        self._adv_wrapper      = ctk.CTkFrame(fw, fg_color="transparent")
        self._adv_wrapper.pack(fill="x", pady=(SPACE_XS, 0))
        self._adv_filters_frame = ctk.CTkFrame(self._adv_wrapper, fg_color="transparent")
        adv_hdr = ctk.CTkFrame(self._adv_wrapper, fg_color="transparent")
        adv_hdr.pack(fill="x")
        def _toggle_adv():
            self._adv_filters_open = not self._adv_filters_open
            arrow = "▼" if self._adv_filters_open else "▶"
            self._btn_adv_flt.configure(text=f"{arrow}  Advanced filters")
            if self._adv_filters_open:
                self._adv_filters_frame.pack(fill="x", pady=(SPACE_XS, 0))  # type: ignore
            else:
                self._adv_filters_frame.pack_forget()  # type: ignore
        self._btn_adv_flt = ctk.CTkButton(
            adv_hdr, text="▶  Advanced filters",
            font=FONT_HINT, text_color=MUTED,
            fg_color="transparent", hover_color=BORDER,
            anchor="w", height=22, corner_radius=4,
            command=_toggle_adv)
        self._btn_adv_flt.pack(fill="x")
        af = self._adv_filters_frame
        self._sidebar_entry_row(af, dict(padx=0), [
            ("HP cut (Hz)", "lp", str(MouseECG.BP_LO_HZ)),
            ("LP cut (Hz)", "hp", str(int(MouseECG.BP_HI_HZ))),
        ])
        for _ent in (self.ent_lp, self.ent_hp):
            if _ent is not None:
                _ent.bind("<Return>",   lambda _e: self._refresh_filter_preview())
                _ent.bind("<FocusOut>", lambda _e: self._refresh_filter_preview())
        ctk.CTkLabel(af, text="Clean method:", font=FONT_SMALL,
                     text_color=MUTED, anchor="w").pack(anchor="w", pady=(SPACE_XS, 0))
        self.cb_clean = ctk.CTkComboBox(
            af, font=FONT_LABEL, height=28, fg_color=BG, border_color=BORDER2,
            button_color=BORDER2, text_color=TEXT, dropdown_fg_color=BG,
            dropdown_text_color=TEXT,
            values=["neurokit", "pantompkins1985", "elgendi2010", "hamilton2002", "biosppy"],
            command=lambda _v: self._refresh_filter_preview())
        self.cb_clean.set("neurokit")
        self.cb_clean.pack(fill="x", pady=(SPACE_XS, SPACE_S))
        self._filter_control_widgets: "list" = []
        def _collect_fw(frame) -> None:
            for w in frame.winfo_children():
                self._filter_control_widgets.append(w)
                _collect_fw(w)
        self.after(100, lambda: (
            self._filter_control_widgets.clear() or
            _collect_fw(self._filter_widgets_frame) or
            (self._adv_filters_frame is not None and
             _collect_fw(self._adv_filters_frame))
        ))

        # ── DETECTION ─────────────────────────────────────────
        sec_det = _SidebarSection(s, "DETECTION", initially_open=False)
        f = sec_det.frame
        self._sidebar_entry(f, "Min R-R physio (ms)", "minrr",
                            str(int(MouseECG.MIN_RR_MS)), fpx)
        self._sg_frame = ctk.CTkFrame(f, fg_color="transparent")
        self._sidebar_entry_row(self._sg_frame, fpx, [
            ("Target fs (Hz)", "sg_target_fs", "10000"),
            ("SG window (ms)", "sg_window_ms",  "20"),
        ])
        ctk.CTkLabel(f, text="SG+Deriv: downsample → Savitzky-Golay derivative\n"
                             "Wavelet: CWT bruit/QRS/J-wave séparés (pywt requis)\n"
                             "Envelope Max: maximum local — idéal signaux saturés (clipping ADC)",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))

        # ── ARTIFACTS ─────────────────────────────────────────
        sec_art = _SidebarSection(s, "ARTIFACTS", initially_open=False)
        f = sec_art.frame
        self.btn_review_art = ctk.CTkButton(
            f, text="🔍  Review Artifacts",
            command=self._open_artifact_review,
            fg_color=ORANGE, hover_color=ORANGE_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(30, int(34 * THEME.font_scale)),
            corner_radius=8, state="disabled")
        self.btn_review_art.pack(**fpx, fill="x", pady=(SPACE_S, SPACE_XS))
        ctk.CTkLabel(f, text="Detect + review every artifact. Run Preview first.",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230, justify="left").pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.sw_artifact = self._switch(
            f, "Auto-correct on Full Analysis", fpx, default_on=False)
        ctk.CTkLabel(f, text="OFF by default — use Review for full control",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))

        # ── SESSION & EXPORT ──────────────────────────────────
        sec_ses = _SidebarSection(s, "SESSION & EXPORT", initially_open=False)
        f = sec_ses.frame
        self.btn_save_session = ctk.CTkButton(
            f, text="💾  Save Session", command=self._save_session,
            fg_color=GREEN, hover_color=GREEN_DARK, text_color="white",
            font=FONT_CARD_TITLE, height=34, corner_radius=8,
            state="disabled")
        self.btn_save_session.pack(**fpx, fill="x", pady=(SPACE_S, SPACE_S))
        self.lbl_session_info = ctk.CTkLabel(
            f, text="No session saved for this file", font=FONT_HINT,
            text_color=MUTED, anchor="w", wraplength=230, justify="left")
        self.lbl_session_info.pack(**fpx, pady=(0, SPACE_S), fill="x")
        self._btn(f, "🗑  Clear Session Cache", self._delete_session, fpx, fg=BORDER, h=26)
        ctk.CTkFrame(f, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_S))
        self._btn(f, "📊  Export Excel",              self._export_excel,      fpx, fg=BORDER, h=28)
        self._btn(f, "📄  Export RR CSV  (Ctrl+W)",   self._export_rr_csv,     fpx, fg=BORDER, h=28)
        self._btn(f, "🖼  Export Figures  (PNG)",      self._export_figures,    fpx, fg=BLUE_DARK, h=28)
        self._btn(f, "📦  Export ZIP  (Excel+Figs)",  self._export_zip,        fpx, fg=BORDER, h=28)
        self._btn(f, "📄  PDF Report  (1 page)",      self._export_pdf_report, fpx, fg=BORDER, h=28)
        self._btn(f, "🔬  Export Arrhythmia PDF",     self._export_arrhythmia_pdf, fpx, fg=BORDER, h=28)
        self._btn(f, "🔬  Export GraphPad Prism",     self._export_prism,      fpx, fg=BORDER, h=28)
        ctk.CTkFrame(f, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_S))
        self._btn(f, "📝  Notes (this recording)",    self._open_notes_dialog, fpx, fg=BORDER, h=28)
        self._btn(f, "⏱  Event annotations",          self._open_annotation_dialog, fpx, fg=BORDER, h=28)

        # ── BOTTOM BUTTONS ────────────────────────────────────
        ctk.CTkFrame(s, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        ctk.CTkButton(
            s, text="⚖  Compare Segments",
            command=self._open_compare_segments,
            fg_color=TEAL, hover_color="#00695C", text_color="white",
            font=FONT_SIDEBAR_HDR, height=28, corner_radius=8,
        ).pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkButton(
            s, text="⚡  Batch Processing",
            command=self._open_batch_dialog,
            fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
            font=FONT_SIDEBAR_HDR, height=28, corner_radius=8,
        ).pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkButton(
            s, text="⚙  Parameters",
            command=self._open_params_dialog,
            fg_color=BLUE_DARK, hover_color=BLUE, text_color="white",
            font=FONT_SIDEBAR_HDR, height=28, corner_radius=8,
        ).pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkButton(
            s, text="↺  Reset to Mouse ECG Defaults",
            command=self._reset_params,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            font=FONT_HINT, height=24, corner_radius=5,
        ).pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkLabel(s, text="F1 — keyboard shortcuts",
                     font=FONT_MICRO, text_color=LIGHT, anchor="center",
                     cursor="hand2").pack(pady=(0, SPACE_S))




    # ════════════════════════════════════════════════════════
    #  FREQUENCY BAND HELPER
    # ════════════════════════════════════════════════════════

    def _get_freq_bands(self) -> "tuple[tuple, tuple, tuple]":
        """Return (vlf, lf, hf) tuples from the selected band preset.

        Reads ``cb_freq_band`` if it exists; falls back to Mouse Thireau defaults.
        """
        try:
            sel = self.cb_freq_band.get() if self.cb_freq_band is not None else ""
            for name, vlf, lf, hf in MouseECG.FREQ_BAND_PRESETS:
                if name == sel:
                    return vlf, lf, hf
        except Exception:
            pass
        return (0.0, 0.4), (0.4, 1.5), (1.5, 5.0)

    # ════════════════════════════════════════════════════════
    #  EXPORT FIGURES (clean PNG, no spike markers)
    # ════════════════════════════════════════════════════════

    def _export_figures(self) -> None:
        """Export curated publication-ready PNG figures to a chosen folder."""
        self.export_ctrl.export_figures()

    # ════════════════════════════════════════════════════════
    #  COMPARE SEGMENTS
    # ════════════════════════════════════════════════════════

    def _open_compare_segments(self) -> None:
        """Open the Segment Comparison window.

        Two configurable time windows (A / B) → side-by-side metric table
        + superimposed RR tachogram.  All stats run on a background thread.
        """
        if self._signal_flt is None or self._rpeaks_ok is None or self._fs is None:
            messagebox.showwarning("No data", "Load a file and run Core Analysis first.")
            return

        sig    = self._signal_flt
        rp     = self._rpeaks_ok.copy()
        fs     = self._fs
        dur    = float(len(sig)) / fs
        vlf, lf, hf = self._get_freq_bands()
        sub    = self.ent_subject.get().strip() if self.ent_subject else "subject"

        win = ctk.CTkToplevel(self)
        win.title("Compare Segments")
        win.geometry("1060x720")
        win.configure(fg_color=BG)
        win.resizable(True, True)

        hdr = ctk.CTkFrame(win, fg_color=PANEL, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="⚖  Segment Comparison",
                     font=FONT_KPI_VALUE, text_color=TEXT,
                     anchor="w").pack(side="left", padx=SPACE_L, pady=SPACE_M)
        ctk.CTkLabel(hdr,
                     text=f"{sub}  ·  {dur:.1f} s  ·  {len(rp)} peaks",
                     font=FONT_SMALL, text_color=MUTED).pack(side="left", padx=(0, SPACE_L))
        ctk.CTkButton(hdr, text="✗  Close", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=28, corner_radius=6).pack(
            side="right", padx=SPACE_L, pady=SPACE_M)

        seg_frame = ctk.CTkFrame(win, fg_color="transparent")
        seg_frame.pack(fill="x", padx=SPACE_L, pady=(SPACE_M, SPACE_S))

        seg_entries: "list[dict]" = []
        _colors = [BLUE, ORANGE_DARK]
        _names  = ["Segment A", "Segment B"]

        for i, (name, col) in enumerate(zip(_names, _colors)):
            card = ctk.CTkFrame(seg_frame, fg_color=CARD, corner_radius=8,
                                border_width=2, border_color=col)
            card.pack(side="left", fill="x", expand=True,
                      padx=(0, 8 if i == 0 else 0))
            ctk.CTkLabel(card, text=name, font=FONT_SIDEBAR_HDR,
                         text_color=col, anchor="w").pack(padx=SPACE_L, pady=(SPACE_M, SPACE_S), fill="x")
            row = ctk.CTkFrame(card, fg_color="transparent")
            row.pack(fill="x", padx=SPACE_L, pady=(0, SPACE_M))
            ctk.CTkLabel(row, text="From (s):", font=FONT_SMALL,
                         text_color=MUTED, width=62).pack(side="left")
            lo_e = ctk.CTkEntry(row, width=80, height=28, font=FONT_LABEL,
                                fg_color=BG, border_color=col, text_color=TEXT)
            lo_e.insert(0, str(int(dur * i / 2)))
            lo_e.pack(side="left", padx=(0, SPACE_M))
            ctk.CTkLabel(row, text="To (s):", font=FONT_SMALL,
                         text_color=MUTED, width=46).pack(side="left")
            hi_e = ctk.CTkEntry(row, width=80, height=28, font=FONT_LABEL,
                                fg_color=BG, border_color=col, text_color=TEXT)
            hi_e.insert(0, str(int(dur * (i + 1) / 2)))
            hi_e.pack(side="left", padx=(0, SPACE_M))
            lbl_e = ctk.CTkEntry(row, width=120, height=28, font=FONT_LABEL,
                                 fg_color=BG, border_color=BORDER2, text_color=TEXT,
                                 placeholder_text="Label (optional)")
            lbl_e.insert(0, name)
            lbl_e.pack(side="left")
            seg_entries.append({"lo": lo_e, "hi": hi_e, "lbl": lbl_e, "color": col})

        ctrl = ctk.CTkFrame(win, fg_color="transparent")
        ctrl.pack(fill="x", padx=SPACE_L, pady=(0, SPACE_S))
        lbl_status = ctk.CTkLabel(ctrl, text="", font=FONT_SMALL, text_color=MUTED)
        lbl_status.pack(side="right", padx=(SPACE_M, 0))
        progress_bar = ctk.CTkProgressBar(ctrl, height=8, mode="indeterminate",
                                          progress_color=TEAL)
        progress_bar.pack(side="right", padx=(SPACE_M, 0))
        progress_bar.pack_forget()

        run_btn = ctk.CTkButton(ctrl, text="▶  Compare",
                                fg_color=TEAL, hover_color="#00695C",
                                text_color="white",
                                font=FONT_BTN_PRIMARY, height=34, corner_radius=8)
        run_btn.pack(side="left")

        results_frame = ctk.CTkFrame(win, fg_color="transparent")
        results_frame.pack(fill="both", expand=True, padx=SPACE_L, pady=(0, SPACE_M))

        table_card = ctk.CTkFrame(results_frame, fg_color=CARD, corner_radius=8)
        table_card.pack(side="left", fill="both", expand=True, padx=(0, SPACE_M))
        ctk.CTkLabel(table_card, text="Metric comparison",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").pack(padx=SPACE_M, pady=(SPACE_M, SPACE_S), fill="x")
        tbl = ctk.CTkScrollableFrame(table_card, fg_color=BG, height=380,
                                     scrollbar_button_color=BORDER)
        tbl.pack(fill="both", expand=True, padx=SPACE_S, pady=(0, SPACE_M))
        for ci in range(4):
            tbl.grid_columnconfigure(ci, weight=1)

        plot_card = ctk.CTkFrame(results_frame, fg_color=CARD, corner_radius=8)
        plot_card.pack(side="right", fill="both", expand=True)
        ctk.CTkLabel(plot_card, text="RR tachograms (superimposed)",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").pack(padx=SPACE_M, pady=(SPACE_M, SPACE_XS), fill="x")
        from plots import CanvasSlot as _CS
        plot_slot = _CS(plot_card, 8, 4, toolbar=False, yscale_bar=True)

        _METRICS: "list[tuple[str, str, str]]" = [
            ("n_beats",   "Beats",           ""),
            ("duration_s","Duration",         "s"),
            ("hr_mean",   "Mean HR",          "bpm"),
            ("hr_sd",     "HR SD",            "bpm"),
            ("hr_min",    "Min HR",           "bpm"),
            ("hr_max",    "Max HR",           "bpm"),
            ("rr_mean",   "Mean RR",          "ms"),
            ("rr_sd",     "RR SD",            "ms"),
            ("rr_cv",     "RR CV",            "%"),
            ("sdnn",      "SDNN",             "ms"),
            ("rmssd",     "RMSSD",            "ms"),
            ("pnn6",      "pNN6",             "%"),
            ("lf_nu",     "LF (n.u.)",        "%"),
            ("hf_nu",     "HF (n.u.)",        "%"),
            ("lf_hf",     "LF/HF",            ""),
            ("sd1",       "Poincaré SD1",      "ms"),
            ("sd2",       "Poincaré SD2",      "ms"),
            ("sd_ratio",  "SD1/SD2",           ""),
        ]

        def _fmt(v: "Any") -> str:
            try:
                fv = float(v)
                if not np.isfinite(fv):
                    return "—"
                if abs(fv) >= 100: return f"{fv:.1f}"
                if abs(fv) >= 10:  return f"{fv:.2f}"
                return f"{fv:.3f}"
            except Exception:
                return str(v) if v else "—"

        def _populate_table(sa: dict, sb: dict) -> None:
            for w in tbl.winfo_children():
                w.destroy()
            for ci, (txt, col) in enumerate([
                ("Metric", TEXT), (sa["label"], _colors[0]),
                (sb["label"], _colors[1]), ("Δ%", MUTED),
            ]):
                ctk.CTkLabel(tbl, text=txt, font=FONT_BADGE,
                             text_color=col, anchor="w").grid(
                    row=0, column=ci, sticky="ew", padx=SPACE_S, pady=(SPACE_XS, SPACE_S))
            for ri, (key, name, unit) in enumerate(_METRICS, start=1):
                va = float(sa.get(key, float("nan")))
                vb = float(sb.get(key, float("nan")))
                bg = CARD if ri % 2 == 0 else BG
                lbl_txt = f"{name}  {unit}" if unit else name
                ctk.CTkLabel(tbl, text=lbl_txt, font=FONT_KPI_LABEL,
                             text_color=MUTED, fg_color=bg,
                             anchor="w").grid(row=ri, column=0, sticky="ew", padx=SPACE_S, pady=SPACE_XS)
                ctk.CTkLabel(tbl, text=_fmt(va), font=FONT_KPI_LABEL,
                             text_color=TEXT, fg_color=bg,
                             anchor="e").grid(row=ri, column=1, sticky="ew", padx=SPACE_S, pady=SPACE_XS)
                ctk.CTkLabel(tbl, text=_fmt(vb), font=FONT_KPI_LABEL,
                             text_color=TEXT, fg_color=bg,
                             anchor="e").grid(row=ri, column=2, sticky="ew", padx=SPACE_S, pady=SPACE_XS)
                if np.isfinite(va) and np.isfinite(vb) and abs(va) > 1e-9:
                    delta = 100.0 * (vb - va) / abs(va)
                    ctk.CTkLabel(tbl, text=f"{'▲' if delta>0 else '▼'} {abs(delta):.1f}%",
                                 font=FONT_MICRO, text_color=MUTED, fg_color=bg,
                                 anchor="e").grid(row=ri, column=3, sticky="ew", padx=SPACE_S, pady=SPACE_XS)

        def _populate_plot(sa: dict, sb: dict) -> None:
            _t_a, _r_a, _la, _ca = sa["t_rr"], sa["rr_ms"], sa["label"], _colors[0]
            _t_b, _r_b, _lb, _cb = sb["t_rr"], sb["rr_ms"], sb["label"], _colors[1]
            def _draw(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                if len(_t_a) and len(_r_a):
                    ax.plot(_t_a - _t_a[0], _r_a, color=_ca, lw=1.0, alpha=0.85, label=_la)
                    ax.axhline(float(np.mean(_r_a)), color=_ca, lw=1.0, ls="--", alpha=0.5)
                if len(_t_b) and len(_r_b):
                    ax.plot(_t_b - _t_b[0], _r_b, color=_cb, lw=1.0, alpha=0.85, label=_lb)
                    ax.axhline(float(np.mean(_r_b)), color=_cb, lw=1.0, ls="--", alpha=0.5)
                ax.set_xlabel("Time within segment (s)", fontsize=9)
                ax.set_ylabel("RR (ms)", fontsize=9)
                ax.set_title("RR intervals — segments superimposed", loc="left", fontsize=9)
                ax.legend(framealpha=0, fontsize=8)
            plot_slot.update(_draw)

        def _run_compare() -> None:
            try:
                lo_a = float(seg_entries[0]["lo"].get()); hi_a = float(seg_entries[0]["hi"].get())
                la   = seg_entries[0]["lbl"].get().strip() or "Segment A"
                lo_b = float(seg_entries[1]["lo"].get()); hi_b = float(seg_entries[1]["hi"].get())
                lb   = seg_entries[1]["lbl"].get().strip() or "Segment B"
            except ValueError:
                lbl_status.configure(text="⚠  Invalid time values", text_color=RED)
                return
            for lo, hi, lbl_s in [(lo_a, hi_a, "A"), (lo_b, hi_b, "B")]:
                if lo >= hi:
                    lbl_status.configure(text=f"⚠  Segment {lbl_s}: start ≥ end", text_color=RED)
                    return
                if lo < 0 or hi > dur:
                    lbl_status.configure(
                        text=f"⚠  Segment {lbl_s}: outside recording ({dur:.1f} s)",
                        text_color=RED)
                    return
            run_btn.configure(state="disabled")
            progress_bar.pack(side="right", padx=(SPACE_M, 0))
            progress_bar.start()
            lbl_status.configure(text="  Computing…", text_color=ORANGE)

            import threading as _th
            def _worker():
                from analysis import compute_segment_stats
                sa = compute_segment_stats(sig, rp, fs, lo_a, hi_a, la,
                                            lf_band=lf, hf_band=hf)
                sb = compute_segment_stats(sig, rp, fs, lo_b, hi_b, lb,
                                            lf_band=lf, hf_band=hf)
                win.after(0, lambda: _on_done(sa, sb))

            def _on_done(sa: dict, sb: dict) -> None:
                run_btn.configure(state="normal")
                progress_bar.stop(); progress_bar.pack_forget()
                # Persist for Prism export
                self._last_seg_a = sa
                self._last_seg_b = sb
                errs = [s["error"] for s in (sa, sb) if s.get("error")]
                if errs:
                    lbl_status.configure(text=f"⚠  {errs[0]}", text_color=RED)
                else:
                    lbl_status.configure(
                        text=f"✓  {sa['label']}: {sa['n_beats']} beats  ·  "
                             f"{sb['label']}: {sb['n_beats']} beats",
                        text_color=GREEN)
                    export_btn.configure(state="normal")   # unlock Export button
                _populate_table(sa, sb)
                _populate_plot(sa, sb)

            _th.Thread(target=_worker, daemon=True).start()

        # ── Export function ───────────────────────────────────────────────
        _METRICS_EXPORT: "list[tuple[str, str, str]]" = [
            ("n_beats",   "Beats",             ""),
            ("duration_s","Duration",           "s"),
            ("hr_mean",   "Mean HR",            "bpm"),
            ("hr_sd",     "HR SD",              "bpm"),
            ("hr_min",    "Min HR",             "bpm"),
            ("hr_max",    "Max HR",             "bpm"),
            ("rr_mean",   "Mean RR",            "ms"),
            ("rr_sd",     "RR SD",              "ms"),
            ("rr_cv",     "RR CV",              "%"),
            ("sdnn",      "SDNN",               "ms"),
            ("rmssd",     "RMSSD",              "ms"),
            ("pnn6",      "pNN6",               "%"),
            ("lf_nu",     "LF n.u.",            "%"),
            ("hf_nu",     "HF n.u.",            "%"),
            ("lf_hf",     "LF/HF",              ""),
            ("sd1",       "Poincaré SD1",       "ms"),
            ("sd2",       "Poincaré SD2",       "ms"),
            ("sd_ratio",  "SD1/SD2",            ""),
        ]

        def _export_comparison() -> None:
            sa = self._last_seg_a
            sb = self._last_seg_b
            if sa is None or sb is None:
                messagebox.showwarning("No results",
                                       "Run the comparison first.", parent=win)
                return

            sub  = self.ent_subject.get().strip() if self.ent_subject else "subject"
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"{sub}_comparison_{ts}"

            path = filedialog.asksaveasfilename(
                parent=win,
                title="Export segment comparison",
                defaultextension=".xlsx",
                initialfile=f"{default_name}.xlsx",
                filetypes=[
                    ("Excel workbook", "*.xlsx"),
                    ("CSV",            "*.csv"),
                    ("All files",      "*.*"),
                ],
            )
            if not path:
                return

            la = sa["label"]
            lb = sb["label"]

            # ── Build metric comparison DataFrame ─────────────────────────
            rows = []
            for key, name, unit in _METRICS_EXPORT:
                va = sa.get(key, float("nan"))
                vb = sb.get(key, float("nan"))
                try:
                    va = float(va) if va is not None else float("nan")
                    vb = float(vb) if vb is not None else float("nan")
                except (TypeError, ValueError):
                    va = vb = float("nan")
                delta = (100.0 * (vb - va) / abs(va)
                         if np.isfinite(va) and np.isfinite(vb) and abs(va) > 1e-9
                         else float("nan"))
                rows.append({
                    "Metric":      f"{name}  ({unit})" if unit else name,
                    la:            va,
                    lb:            vb,
                    "Δ% (A→B)":    delta,
                })
            metrics_df = pd.DataFrame(rows)

            # ── RR series per segment ─────────────────────────────────────
            rr_a = pd.DataFrame({
                "Time (s)": np.asarray(sa["t_rr"], dtype=float),
                "RR (ms)":  np.asarray(sa["rr_ms"], dtype=float),
            })
            rr_b = pd.DataFrame({
                "Time (s)": np.asarray(sb["t_rr"], dtype=float),
                "RR (ms)":  np.asarray(sb["rr_ms"], dtype=float),
            })

            # ── Write ─────────────────────────────────────────────────────
            if path.lower().endswith(".csv"):
                metrics_df.to_csv(path, index=False, float_format="%.4f")
                lbl_status.configure(
                    text=f"✓  CSV saved → {os.path.basename(path)}",
                    text_color=GREEN)
            else:
                try:
                    import openpyxl
                    from openpyxl.styles import (Font, PatternFill,
                                                  Alignment, Border, Side)
                    from openpyxl.utils import get_column_letter
                except ImportError:
                    messagebox.showerror("Missing dependency",
                                         "openpyxl required: pip install openpyxl",
                                         parent=win)
                    return

                wb = openpyxl.Workbook()

                # ── Sheet 1: Metric comparison ────────────────────────────
                ws = wb.active
                assert ws is not None
                ws.title = "Comparison"

                # Header style
                hdr_fill = PatternFill("solid", fgColor="1A1A2E")
                col_fills = {
                    la: PatternFill("solid", fgColor="1565C0"),   # blue
                    lb: PatternFill("solid", fgColor="E65100"),   # orange
                    "Δ% (A→B)": PatternFill("solid", fgColor="2E2E2E"),
                }
                hdr_font = Font(bold=True, color="FFFFFF", size=10)
                thin = Side(style="thin", color="333333")
                border = Border(left=thin, right=thin,
                                top=thin, bottom=thin)

                headers = list(metrics_df.columns)
                for ci, h in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=ci, value=h)
                    cell.font = hdr_font
                    cell.fill = col_fills.get(h, hdr_fill)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                for ri, row_d in enumerate(metrics_df.itertuples(index=False), start=2):
                    for ci, val in enumerate(row_d, start=1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        cell.border = border
                        cell.alignment = Alignment(
                            horizontal="right" if ci > 1 else "left")
                        # Shade alternating rows
                        if ri % 2 == 0:
                            cell.fill = PatternFill("solid", fgColor="1E1E2E")
                        # Colour-code Δ% column
                        if headers[ci-1] == "Δ% (A→B)" and isinstance(val, float):
                            if np.isfinite(val):
                                cell.font = Font(
                                    color="66BB6A" if val > 0 else "EF5350",
                                    bold=True)

                # Auto-fit column widths
                for col in ws.columns:
                    if not col:
                        continue
                    first_cell = col[0]
                    if first_cell.column is None:
                        continue
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value), default=8)
                    ws.column_dimensions[
                        get_column_letter(first_cell.column)].width = min(
                        max_len + 3, 30)

                # ── Sheet 2: RR series ────────────────────────────────────
                def _rr_sheet(name: str, rr_df: pd.DataFrame,
                              fill_color: str) -> None:
                    ws2 = wb.create_sheet(name)
                    hf2 = PatternFill("solid", fgColor=fill_color)
                    for ci, col in enumerate(rr_df.columns, start=1):
                        c = ws2.cell(row=1, column=ci, value=col)
                        c.font = Font(bold=True, color="FFFFFF")
                        c.fill = hf2
                        c.border = border
                    for ri, row_vals in enumerate(rr_df.itertuples(index=False),
                                                   start=2):
                        for ci, v in enumerate(row_vals, start=1):
                            ws2.cell(row=ri, column=ci, value=round(float(v), 4))
                    ws2.column_dimensions["A"].width = 14
                    ws2.column_dimensions["B"].width = 12

                _rr_sheet(f"RR — {la}", rr_a, "1565C0")
                _rr_sheet(f"RR — {lb}", rr_b, "E65100")

                # ── Sheet 3: Summary stats ────────────────────────────────
                ws3 = wb.create_sheet("Summary")
                ws3.cell(row=1, column=1, value="Subject").font = Font(bold=True)
                ws3.cell(row=1, column=2, value=sub)
                ws3.cell(row=2, column=1, value="Export date").font = Font(bold=True)
                ws3.cell(row=2, column=2,
                         value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                ws3.cell(row=3, column=1, value="LF band (Hz)").font = Font(bold=True)
                ws3.cell(row=3, column=2, value=f"{lf[0]:.2f}–{lf[1]:.2f}")
                ws3.cell(row=4, column=1, value="HF band (Hz)").font = Font(bold=True)
                ws3.cell(row=4, column=2, value=f"{hf[0]:.2f}–{hf[1]:.2f}")
                for col in ws3.columns:
                    ws3.column_dimensions[
                        get_column_letter(col[0].column)].width = 22

                wb.save(path)
                lbl_status.configure(
                    text=f"✓  Exported → {os.path.basename(path)}",
                    text_color=GREEN)

            # ── Also save the plot as PNG next to the xlsx ─────────────────
            png_path = os.path.splitext(path)[0] + "_plot.png"
            try:
                import matplotlib
                export_fig = Figure(
                    figsize=(10, 4), dpi=200,
                    facecolor=PLOT.get("bg", "#1A1A2E"))
                if hasattr(export_fig, "set_constrained_layout_pads"):
                    getattr(export_fig, "set_constrained_layout_pads")(
                        left=0.11, right=0.98, top=0.96, bottom=0.08)
                if plot_slot._draw_fn is not None:
                    plot_slot._draw_fn(export_fig)
                    export_fig.savefig(
                        png_path, dpi=200,
                        facecolor=export_fig.get_facecolor(),
                        bbox_inches="tight")
                    plt.close(export_fig)
                    lbl_status.configure(
                        text=lbl_status.cget("text") + f"  +  {os.path.basename(png_path)}",
                        text_color=GREEN)
            except Exception as exc:
                log.debug("comparison PNG export: %s", exc)

        run_btn.configure(command=_run_compare)

        # Export button — disabled until comparison has run
        export_btn = ctk.CTkButton(
            ctrl, text="📊  Export",
            command=_export_comparison,
            fg_color=GREEN_DARK, hover_color=GREEN, text_color="white",
            font=FONT_BTN_PRIMARY, height=34, corner_radius=8,
            state="disabled",
        )
        export_btn.pack(side="left", padx=(SPACE_M, 0))

        ctk.CTkLabel(win,
                     text=f"Tip: equal durations give a fair comparison  ·  "
                          f"LF {lf[0]:.2f}–{lf[1]:.2f} Hz  HF {hf[0]:.2f}–{hf[1]:.2f} Hz",
                     font=FONT_MICRO, text_color=LIGHT, anchor="w").pack(
            fill="x", padx=SPACE_L, pady=(0, SPACE_S))

    # ─── KPI bar ──────────────────────────────────────────────

    def _build_kpi_bar(self, parent) -> None:
        bar = ctk.CTkFrame(parent, fg_color=PANEL, corner_radius=0)
        bar.pack(fill="x")
        ctk.CTkFrame(bar, height=1, fg_color=BORDER).pack(side="bottom", fill="x")

        # ── Progress bar row (shown only during analysis) ──────────────────
        self._prog_row = ctk.CTkFrame(bar, fg_color="transparent", height=20)
        self.progress = ctk.CTkProgressBar(
            self._prog_row, height=4, mode="determinate",
            progress_color=BLUE, fg_color=BORDER, corner_radius=2,
        )
        self.progress.set(0)
        self.progress.pack(side="left", fill="x", expand=True, padx=(0, SPACE_M))
        self.lbl_progress = ctk.CTkLabel(
            self._prog_row, text="", font=FONT_HINT,
            text_color=MUTED, width=260, anchor="w",
        )
        self.lbl_progress.pack(side="left")

        # ── KPI grid: 2 rows × 4 columns — adapts to any window width ─────
        kpi_grid = ctk.CTkFrame(bar, fg_color="transparent")
        kpi_grid.pack(side="left", fill="both", expand=True, padx=SPACE_S, pady=SPACE_S)
        for col in range(4):
            kpi_grid.columnconfigure(col, weight=1, uniform="kpi")

        kpi_defs = [
            ("HR Mean",  "hr_mean"),  ("HR Range", "hr_range"),
            ("Mean RR",  "rr_mean"),  ("N Beats",  "n_beats"),
            ("SDNN",     "sdnn"),     ("RMSSD",    "rmssd"),
            ("pNN6",     "pnn50"),    ("Duration", "dur"),
        ]
        for i, (label, key) in enumerate(kpi_defs):
            r, c = divmod(i, 4)
            cell = ctk.CTkFrame(kpi_grid, fg_color="transparent")
            cell.grid(row=r, column=c, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS) if r == 0 else (0, 4))
            ctk.CTkLabel(cell, text=label, font=FONT_KPI_LABEL,
                         text_color=MUTED).pack(anchor="w")
            value_lbl = ctk.CTkLabel(cell, text="—",
                                     font=FONT_KPI_VALUE, text_color=TEXT)
            value_lbl.pack(anchor="w")
            self._kpi[key] = value_lbl

        # ── Right: quality indicator + Theme button + Language toggle ─────────
        right = ctk.CTkFrame(bar, fg_color="transparent")
        right.pack(side="right", padx=SPACE_L)
        self.lbl_quality = ctk.CTkLabel(right, text="", font=FONT_BTN_SEC,
                                         text_color=MUTED)
        self.lbl_quality.pack(anchor="e", pady=(SPACE_S, SPACE_XS))
        btn_row_right = ctk.CTkFrame(right, fg_color="transparent")
        btn_row_right.pack(anchor="e")
        ctk.CTkButton(btn_row_right, text="Theme", width=76, height=28,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      font=FONT_BTN_SEC, command=self._open_theme_dialog,
                      corner_radius=8).pack(side="left", padx=(0, SPACE_S))
        ctk.CTkButton(btn_row_right, text="☀/☾", width=52, height=28,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      font=FONT_BTN_SEC, corner_radius=8,
                      command=self._toggle_dark_live,
                      ).pack(side="left", padx=(0, SPACE_S))
        self.btn_lang = ctk.CTkButton(
            btn_row_right, text="FR", width=50, height=28,
            fg_color=BORDER2, hover_color=BLUE, text_color=TEXT,
            font=FONT_BTN_SEC, corner_radius=8,
            command=self._toggle_language,
        )
        self.btn_lang.pack(side="left", padx=(0, SPACE_S))
        # Quality badge — updated by _update_quality_badge() after analysis
        self._lbl_quality_badge = ctk.CTkLabel(
            btn_row_right, text="", font=FONT_BADGE,
            text_color="white", fg_color="transparent",
            corner_radius=6, width=120, height=22, anchor="center")
        self._lbl_quality_badge.pack(side="left")

    # ─── Tabs ─────────────────────────────────────────────────

    def _build_tabs(self, parent) -> None:
        self.tabs = ctk.CTkTabview(
            parent, fg_color=BG,
            segmented_button_fg_color=PANEL,
            segmented_button_selected_color=BLUE,
            segmented_button_selected_hover_color=BLUE_HOVER,
            segmented_button_unselected_color=PANEL,
            segmented_button_unselected_hover_color=BORDER,
            text_color=TEXT, text_color_disabled=MUTED,
        )
        self.tabs.pack(fill="both", expand=True)
        for name in ["Detection", "HRV", "Intervals",
                     "Beat Template", "Arrhythmias", "Summary"]:
            self.tabs.add(name)

        self._build_tab_detection()
        self._build_tab_hrv_unified()
        self._build_tab_intervals()
        self._build_tab_beat_template()
        self._build_tab_arrhythmias()
        self._build_tab_summary()

    def _build_tab_detection(self) -> None:
        t = self.tabs.tab("Detection")

        # Layout (top → bottom, all pack):
        #   nav bar     (fill=x, no expand) — time navigation controls
        #   thin separator
        #   detail toolbar (fill=x, no expand)
        #   detail plot    (fill=both, expand=True)

        # ── Navigation bar ────────────────────────────────────────────────
        nav = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0, height=38)
        nav.pack(side="top", fill="x")
        nav.pack_propagate(False)

        # Jump-to-start / jump-to-end
        ctk.CTkButton(nav, text="⏮", width=36, height=28, font=FONT_LABEL,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      corner_radius=8,
                      command=self._nav_reset).pack(side="left", padx=(SPACE_M, SPACE_XS), pady=SPACE_S)
        ctk.CTkButton(nav, text="◀◀", width=40, height=28, font=FONT_SMALL,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      corner_radius=8,
                      command=lambda: self._navigate_big(-1)).pack(side="left", padx=SPACE_XS, pady=SPACE_S)
        ctk.CTkButton(nav, text="◀", width=36, height=28, font=FONT_LABEL,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      corner_radius=8,
                      command=lambda: self._navigate(-1)).pack(side="left", padx=SPACE_XS, pady=SPACE_S)

        # Current position entry
        ctk.CTkLabel(nav, text="t =", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(SPACE_L, SPACE_XS))
        self.ent_nav_pos = ctk.CTkEntry(
            nav, width=72, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT,
            corner_radius=6,
            placeholder_text="0.000")
        self.ent_nav_pos.pack(side="left", padx=(0, SPACE_XS))  # type: ignore[union-attr]
        ctk.CTkLabel(nav, text="s", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        ctk.CTkButton(nav, text="Go", width=48, height=28, font=FONT_SMALL,
                      fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
                      corner_radius=8,
                      command=self._nav_goto).pack(side="left", padx=(SPACE_S, SPACE_XS))
        # Bind Enter key on the position field
        self.ent_nav_pos.bind("<Return>", lambda _e: self._nav_goto())  # type: ignore[union-attr]

        ctk.CTkButton(nav, text="▶", width=36, height=28, font=FONT_LABEL,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      corner_radius=8,
                      command=lambda: self._navigate(+1)).pack(side="left", padx=SPACE_XS, pady=SPACE_S)
        ctk.CTkButton(nav, text="▶▶", width=40, height=28, font=FONT_SMALL,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      corner_radius=8,
                      command=lambda: self._navigate_big(+1)).pack(side="left", padx=SPACE_XS, pady=SPACE_S)
        ctk.CTkButton(nav, text="⏭", width=36, height=28, font=FONT_LABEL,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      corner_radius=8,
                      command=self._nav_end).pack(side="left", padx=(SPACE_XS, SPACE_XS), pady=SPACE_S)

        # Separator + window size
        ctk.CTkFrame(nav, width=1, fg_color=BORDER).pack(
            side="left", fill="y", padx=(SPACE_L, SPACE_S), pady=SPACE_S)
        ctk.CTkLabel(nav, text="Window:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(0, SPACE_S))
        self.ent_window = ctk.CTkEntry(nav, width=48, height=28, font=FONT_LABEL,
                                       fg_color=BG, border_color=BORDER2, text_color=TEXT,
                                       corner_radius=6)
        self.ent_window.insert(0, "2")
        self.ent_window.pack(side="left")
        ctk.CTkLabel(nav, text="s", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(SPACE_XS, 0))

        # Duration label (filled after signal load)
        self.lbl_sig_duration = ctk.CTkLabel(
            nav, text="", font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_sig_duration.pack(side="left", padx=(SPACE_L, 0))  # type: ignore[union-attr]

        # Separator
        tk.Frame(t, height=1, bg=BORDER).pack(side="top", fill="x", padx=SPACE_M, pady=SPACE_XS)

        # Detail toolbar (fixed height)
        hdr = ctk.CTkFrame(t, fg_color="transparent", height=36)
        hdr.pack(side="top", fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        hdr.pack_propagate(False)

        ctk.CTkLabel(hdr, text="SIGNAL ECG", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left", anchor="w")

        ctk.CTkFrame(hdr, width=1, fg_color=BORDER).pack(side="left", fill="y", padx=(SPACE_L, SPACE_S), pady=SPACE_XS)
        self.btn_edit_mode = ctk.CTkButton(
            hdr, text="Edit Peaks", width=96, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._toggle_edit_mode,
        )
        self.btn_edit_mode.pack(side="left", padx=SPACE_XS)
        self.lbl_edit_hint = ctk.CTkLabel(
            hdr,
            text="L-click: exclude/restore   R-click: add   Ctrl+Z: undo",
            font=FONT_HINT, text_color=ORANGE,
        )
        self.btn_undo_edit = ctk.CTkButton(
            hdr, text="↩ Undo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._undo_edit,
        )
        self.btn_undo_edit.pack(side="left", padx=(SPACE_XS, 0))
        self.btn_redo_edit = ctk.CTkButton(
            hdr, text="↪ Redo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._redo_edit,
        )
        self.btn_redo_edit.pack(side="left", padx=(SPACE_XS, SPACE_S))
        self.btn_clear_excl = ctk.CTkButton(
            hdr, text="Clear Edits", width=100, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._clear_manual_exclusions,
        )
        self.btn_clear_excl.pack(side="left", padx=SPACE_XS)

        # ── Free Placement toggle (bypass proximity guard) ────────────────
        ctk.CTkFrame(hdr, width=1, fg_color=BORDER).pack(side="left", fill="y", padx=(SPACE_M, SPACE_S), pady=SPACE_XS)
        self.btn_free_placement = ctk.CTkButton(
            hdr, text="Free Placement", width=118, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._toggle_free_placement,
        )
        self.btn_free_placement.pack(side="left", padx=SPACE_XS)
        _fp_tip = ctk.CTkLabel(
            hdr, text="?", width=28, height=28,
            font=FONT_CARD_TITLE, text_color=MUTED,
            fg_color=BORDER, corner_radius=14,
        )
        _fp_tip.pack(side="left", padx=(SPACE_XS, SPACE_S))
        _fp_tip.bind("<Enter>", lambda e: self._set_status(
            "Free Placement: right-click adds a peak at the exact clicked position — "
            "no snapping, no proximity guard, works even on top of existing peaks.", MUTED))
        _fp_tip.bind("<Leave>", lambda e: self._set_status(""))

        ctk.CTkFrame(hdr, width=1, fg_color=BORDER).pack(side="left", fill="y", padx=(SPACE_M, SPACE_S), pady=SPACE_XS)
        self.btn_annotations = ctk.CTkButton(
            hdr, text="Annotations", width=110, height=28, font=FONT_SMALL,
            fg_color=PURPLE_DARK, hover_color=PURPLE, text_color="white",
            corner_radius=8,
            command=self._open_annotations,
        )
        self.btn_annotations.pack(side="left", padx=SPACE_XS)  # type: ignore[union-attr]
        self.lbl_ann_count = ctk.CTkLabel(
            hdr, text="", font=FONT_HINT, text_color=MUTED, anchor="w")
        self.lbl_ann_count.pack(side="left", padx=(SPACE_XS, SPACE_S))  # type: ignore[union-attr]

        # Detail plot — expands to fill all remaining space
        det_frame = tk.Frame(t, bg=PLOT["bg"], bd=0, highlightthickness=0)
        det_frame.pack(side="top", fill="both", expand=True, padx=SPACE_XS, pady=(0, SPACE_XS))
        self._slots["detail"] = CanvasSlot(det_frame, 12, 6.5, toolbar=True)

        self._slots["detail"].canvas.mpl_connect(
            "scroll_event", self._on_detail_scroll)
        self._slots["detail"].canvas.mpl_connect(
            "button_press_event", self._on_detail_click)
        self._hover_motion_cid = self._slots["detail"].canvas.mpl_connect(
            "motion_notify_event", self._on_detail_motion)

    def _build_tab_hrv_unified(self) -> None:
        """Unified HRV tab with internal segmented navigation.

        Merges: RR / HR  |  Temporel  |  Fréquentiel  |  Non-linéaire  |  Epochs  |  Glissant
        """
        t = self.tabs.tab("HRV")
        t.grid_rowconfigure(1, weight=1)
        t.grid_columnconfigure(0, weight=1)

        # ── Top: segmented sub-tab selector ───────────────────────────────
        _HRV_VIEWS = ["RR / HR", "Time Domain", "Frequency", "Non-linear", "Epochs", "Rolling"]
        self._hrv_subframes: "dict[str, ctk.CTkFrame]" = {}

        seg_bar = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0)
        seg_bar.grid(row=0, column=0, sticky="ew")
        self._hrv_seg = ctk.CTkSegmentedButton(
            seg_bar,
            values=_HRV_VIEWS,
            font=FONT_SMALL,
            fg_color=BORDER,
            selected_color=BLUE,
            selected_hover_color=BLUE_HOVER,
            unselected_color=BORDER,
            unselected_hover_color=BORDER2,
            text_color=TEXT,
            text_color_disabled=MUTED,
            command=self._on_hrv_view_change,
        )
        self._hrv_seg.pack(padx=SPACE_M, pady=SPACE_S, anchor="w")  # type: ignore[union-attr]
        self._hrv_seg.set("RR / HR")  # type: ignore[union-attr]

        # ── Content area — one frame per view, pack/pack_forget ───────────
        content_area = ctk.CTkFrame(t, fg_color="transparent")
        content_area.grid(row=1, column=0, sticky="nsew")
        content_area.grid_rowconfigure(0, weight=1)
        content_area.grid_columnconfigure(0, weight=1)
        self._hrv_content_area = content_area

        for view in _HRV_VIEWS:
            f = ctk.CTkFrame(content_area, fg_color="transparent")
            f.grid_rowconfigure(0, weight=1)
            f.grid_columnconfigure(0, weight=1)
            self._hrv_subframes[view] = f

        # ── Build each sub-view content ───────────────────────────────────
        self._build_hrv_view_rr(self._hrv_subframes["RR / HR"])
        self._build_hrv_view_temporel(self._hrv_subframes["Time Domain"])
        self._build_hrv_view_freq(self._hrv_subframes["Frequency"])
        self._build_hrv_view_nonlin(self._hrv_subframes["Non-linear"])
        self._build_hrv_view_epochs(self._hrv_subframes["Epochs"])
        self._build_hrv_view_rolling(self._hrv_subframes["Rolling"])

        # Show initial view
        self._hrv_subframes["RR / HR"].pack(fill="both", expand=True)
        self._hrv_current_view = "RR / HR"

    def _on_hrv_view_change(self, view: str) -> None:
        """Switch the visible HRV sub-view."""
        if not hasattr(self, "_hrv_subframes"):
            return
        for name, frame in self._hrv_subframes.items():
            if name == view:
                frame.pack(fill="both", expand=True)
            else:
                frame.pack_forget()
        self._hrv_current_view = view

    def _build_hrv_view_rr(self, parent: ctk.CTkFrame) -> None:
        """RR tachogram + histogram — formerly the 'RR / HR' tab."""
        # Ratio 4:1 (était 3:2) — les deux tachogrammes empilés (RR + HR)
        # sont le contenu principal de l'onglet et bénéficient d'un maximum
        # de hauteur ; le tableau de stats + histogramme du bas est compact
        # et n'a pas besoin d'autant d'espace que ce que 2/5 lui donnait.
        parent.grid_rowconfigure(0, weight=4)
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        rr_frame = ctk.CTkFrame(parent, fg_color="transparent")
        rr_frame.grid(row=0, column=0, sticky="nsew", padx=SPACE_S, pady=(SPACE_S, SPACE_XS))
        rr_frame.grid_rowconfigure(0, weight=1)
        rr_frame.grid_columnconfigure(0, weight=1)
        self._slots["rr"] = CanvasSlot(rr_frame, 14, 5.0, toolbar=False)

        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(SPACE_XS, SPACE_S))
        row.grid_rowconfigure(0, weight=1)
        row.grid_columnconfigure(0, weight=1)
        row.grid_columnconfigure(1, weight=2)

        # Stats textbox (narrow)
        stats_card = ctk.CTkFrame(row, fg_color=CARD, corner_radius=0)
        stats_card.grid(row=0, column=0, sticky="nsew", padx=(0, SPACE_S))
        stats_card.grid_rowconfigure(0, weight=0)
        stats_card.grid_rowconfigure(1, weight=1)
        stats_card.grid_columnconfigure(0, weight=1)
        _rr_btn = ctk.CTkButton(stats_card, text="📋 Copy for Excel", height=22,
                                font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                                text_color="white",
                                command=lambda: self._copy_tsv(self.txt_rr))
        _rr_btn.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, 0))
        self.txt_rr = ctk.CTkTextbox(stats_card, font=FONT_MONO, fg_color=CARD,
                                     text_color=TEXT, border_width=0)
        self.txt_rr.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(SPACE_XS, SPACE_S))

        # Histogram
        hist_card = ctk.CTkFrame(row, fg_color=CARD, corner_radius=0)
        hist_card.grid(row=0, column=1, sticky="nsew")
        hist_card.grid_rowconfigure(0, weight=1)
        hist_card.grid_columnconfigure(0, weight=1)
        self._slots["rr_hist"] = CanvasSlot(hist_card, 7, 4.5, toolbar=False)

    def _build_hrv_view_temporel(self, parent: ctk.CTkFrame) -> None:
        """Time-domain HRV — text tables (computed in Core Analysis)."""
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        inner = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=6)
        inner.grid(row=0, column=0, sticky="nsew", padx=SPACE_M, pady=SPACE_M)
        inner.grid_rowconfigure(1, weight=1)
        inner.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(inner, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_XS))
        ctk.CTkLabel(hdr, text="HRV TIME DOMAIN", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left")
        ctk.CTkButton(hdr, text="📋 Copy for Excel", height=22,
                      font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                      text_color="white", width=140,
                      command=lambda: self._copy_tsv(self.txt_td)
                      ).pack(side="right")
        self.txt_td = ctk.CTkTextbox(inner, font=FONT_MONO, fg_color=CARD,
                                     text_color=TEXT, border_width=0)
        self.txt_td.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_M))

    def _build_hrv_view_freq(self, parent: ctk.CTkFrame) -> None:
        """Frequency-domain HRV — PSD + radar + tables."""
        parent.grid_rowconfigure(0, weight=0)
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        # Action bar
        bar = ctk.CTkFrame(parent, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        self.btn_run_freq = ctk.CTkButton(
            bar, text="⚡  Freq HRV",
            command=self._run_freq,
            fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=5,
        )
        self.btn_run_freq.pack(side="left")  # type: ignore[union-attr]
        self.lbl_freq_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_freq_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        # Content: text left, PSD+radar right
        content = ctk.CTkFrame(parent, fg_color="transparent")
        content.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(SPACE_XS, SPACE_S))
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=2)

        left = ctk.CTkFrame(content, fg_color=CARD, corner_radius=6)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, SPACE_S))
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)
        _hdr = ctk.CTkFrame(left, fg_color="transparent")
        _hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        ctk.CTkLabel(_hdr, text="FREQUENCY DOMAIN", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left")
        ctk.CTkButton(_hdr, text="📋 Copy for Excel", height=20,
                      font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                      text_color="white", width=140,
                      command=lambda: self._copy_tsv(self.txt_fd)
                      ).pack(side="right")
        self.txt_fd = ctk.CTkTextbox(left, font=FONT_MONO, fg_color=CARD,
                                     text_color=TEXT, border_width=0)
        self.txt_fd.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_S))

        right = ctk.CTkFrame(content, fg_color="transparent")
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)
        for row_i, slot_key in enumerate(["psd", "radar"]):
            card = ctk.CTkFrame(right, fg_color=CARD, corner_radius=0)
            # Espacement symétrique entre les deux graphiques empilés (PSD/radar) :
            # avant, le bas avait pady=0 → les deux cartes semblaient "collées"
            # côté inférieur alors qu'il y avait de l'air côté supérieur.
            card.grid(row=row_i, column=0, sticky="nsew",
                      pady=(0, SPACE_S) if row_i == 0 else (SPACE_S, 0))
            card.grid_rowconfigure(0, weight=1)
            card.grid_columnconfigure(0, weight=1)
            self._slots[slot_key] = CanvasSlot(card, 8, 4.5, toolbar=False)

    def _build_hrv_view_nonlin(self, parent: ctk.CTkFrame) -> None:
        """Non-linear HRV — Poincaré + SampEn/DFA table."""
        parent.grid_rowconfigure(0, weight=0)
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        bar = ctk.CTkFrame(parent, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        self.btn_run_nonlin = ctk.CTkButton(
            bar, text="⚡  Non-linear HRV",
            command=self._run_nonlinear,
            fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=5,
        )
        self.btn_run_nonlin.pack(side="left")  # type: ignore[union-attr]
        self.lbl_nonlin_status = ctk.CTkLabel(
            bar, text="  SampEn + DFA peuvent prendre 30 s+ sur longs enregistrements",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_nonlin_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        content = ctk.CTkFrame(parent, fg_color="transparent")
        content.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(SPACE_XS, SPACE_S))
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=2)

        left = ctk.CTkFrame(content, fg_color=CARD, corner_radius=6)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, SPACE_S))
        left.grid_rowconfigure(0, weight=0)
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)
        _nl_hdr = ctk.CTkFrame(left, fg_color="transparent")
        _nl_hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        ctk.CTkLabel(_nl_hdr, text="NON-LINEAR", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left")
        ctk.CTkButton(_nl_hdr, text="📋 Copy for Excel", height=20,
                      font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                      text_color="white", width=140,
                      command=lambda: self._copy_tsv(self.txt_nl)
                      ).pack(side="right")
        self.txt_nl = ctk.CTkTextbox(left, font=FONT_MONO, fg_color=CARD,
                                     text_color=TEXT, border_width=0)
        self.txt_nl.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_S))

        right = ctk.CTkFrame(content, fg_color=CARD, corner_radius=0)
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(0, weight=1)
        right.grid_columnconfigure(0, weight=1)
        self._slots["poincare"] = CanvasSlot(right, 8, 8, toolbar=False)

    def _build_hrv_view_epochs(self, parent: ctk.CTkFrame) -> None:
        """Epoch HRV analysis — formerly the 'Epochs' tab."""
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(parent, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))
        ctk.CTkLabel(hdr, text="HRV PER EPOCH",
                     font=FONT_SECTION_HDR, text_color=MUTED).pack(side="left")
        self.btn_compute_epochs = ctk.CTkButton(
            hdr, text="⟳  Compute epochs", command=self._compute_epochs,
            fg_color=BLUE, hover_color=BLUE_DEEP, text_color="white",
            font=FONT_SMALL, height=28, corner_radius=5)
        self.btn_compute_epochs.pack(side="right")
        self.lbl_epoch_count = ctk.CTkLabel(
            hdr, text="", font=FONT_SMALL, text_color=BLUE)
        self.lbl_epoch_count.pack(side="right", padx=(0, SPACE_M))

        # Settings row
        settings_row = ctk.CTkFrame(parent, fg_color="transparent")
        settings_row.grid(row=1, column=0, sticky="ew", padx=SPACE_M, pady=(0, SPACE_S))
        for lbl_t, attr_n, default in [
            ("Epoch (s)", "epoch", str(int(MouseECG.EPOCH_DEFAULT_S))),
            ("Overlap (s)", "overlap", "0"),
        ]:
            ctk.CTkLabel(settings_row, text=lbl_t, font=FONT_SMALL,
                         text_color=MUTED).pack(side="left", padx=(0, SPACE_S))
            ent = ctk.CTkEntry(settings_row, width=58, height=26,
                               font=FONT_LABEL, fg_color=BG,
                               border_color=BORDER2, text_color=TEXT)
            ent.insert(0, default)
            ent.pack(side="left", padx=(0, SPACE_L))
            setattr(self, f"ent_{attr_n}", ent)
        self.sw_epoch = self._switch(settings_row, "Auto-run après analyse",
                                     dict(padx=0))
        self.sw_epoch.pack(side="left")
        self.lbl_epoch_info = ctk.CTkLabel(
            settings_row, text="", font=FONT_HINT, text_color=MUTED)
        self.lbl_epoch_info.pack(side="right", padx=(0, SPACE_S))

        parent.grid_rowconfigure(1, weight=0)
        parent.grid_rowconfigure(2, weight=0)
        parent.grid_rowconfigure(3, weight=0)
        parent.grid_rowconfigure(4, weight=1)

        # copy bar
        ep_bar = ctk.CTkFrame(parent, fg_color="transparent")
        ep_bar.grid(row=2, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, 0))
        ctk.CTkLabel(ep_bar, text="Epoch table:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self.btn_copy_epochs = ctk.CTkButton(
            ep_bar, text="📋  Copy for Excel", height=24,
            font=FONT_SMALL, fg_color=BLUE, hover_color=BLUE_HOVER,
            text_color="white", width=160,
            command=lambda: self._copy_tsv(self.txt_epochs))
        self.btn_copy_epochs.pack(side="right", padx=(0, SPACE_XS))  # type: ignore[union-attr]

        ep_sf = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=0)
        ep_sf.grid(row=3, column=0, sticky="nsew", padx=SPACE_M, pady=(SPACE_S, SPACE_S))
        ep_sf.grid_rowconfigure(0, weight=1)
        ep_sf.grid_columnconfigure(0, weight=1)
        self._slots["epochs"] = CanvasSlot(ep_sf, 14, 5.5, toolbar=False)

        epochs_tb = ctk.CTkTextbox(parent, font=FONT_BODY, fg_color="transparent",
                                   text_color=TEXT, border_width=0,
                                   scrollbar_button_color=BORDER,
                                   scrollbar_button_hover_color=BORDER2,
                                   height=160)
        epochs_tb.grid(row=4, column=0, sticky="ew", padx=SPACE_M, pady=(0, SPACE_S))
        self.txt_epochs = epochs_tb

    def _build_hrv_view_rolling(self, parent: ctk.CTkFrame) -> None:
        """Rolling / sliding-window HRV — formerly the 'Rolling HRV' tab."""
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        bar = ctk.CTkFrame(parent, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))

        ctk.CTkLabel(bar, text="Window (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self.ent_roll_win = ctk.CTkEntry(
            bar, width=58, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_roll_win.insert(0, "30")
        self.ent_roll_win.pack(side="left", padx=(SPACE_S, SPACE_L))

        ctk.CTkLabel(bar, text="Step (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self.ent_roll_step = ctk.CTkEntry(
            bar, width=48, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_roll_step.insert(0, "5")
        self.ent_roll_step.pack(side="left", padx=(SPACE_S, SPACE_L))

        ctk.CTkLabel(bar, text="Metrics:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self._roll_metrics: "dict[str, ctk.CTkCheckBox]" = {}
        for metric, default in [("HR", True), ("SDNN", True),
                                 ("RMSSD", True), ("pNN6", False)]:
            cb = ctk.CTkCheckBox(bar, text=metric, font=FONT_SMALL,
                                 text_color=MUTED, fg_color=BLUE,
                                 checkmark_color="white",
                                 border_color=BORDER2, width=16)
            if default:
                cb.select()
            cb.pack(side="left", padx=(SPACE_S, 0))
            self._roll_metrics[metric] = cb

        self.btn_roll_compute = ctk.CTkButton(
            bar, text="⟳  Compute", command=self._compute_rolling_hrv,
            fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
            font=FONT_BTN_SEC, height=28, corner_radius=5)
        self.btn_roll_compute.pack(side="left", padx=(SPACE_L, 0))

        self.lbl_roll_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_roll_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        ctk.CTkButton(bar, text="📋 Copy TSV", command=self._copy_rolling_tsv,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=28, corner_radius=5
                      ).pack(side="right")

        plot_frame = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=0)
        plot_frame.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_M))
        plot_frame.grid_rowconfigure(0, weight=1)
        plot_frame.grid_columnconfigure(0, weight=1)
        self._slots["rolling_hrv"] = CanvasSlot(plot_frame, 14, 6.0, toolbar=False)

    def _build_tab_arrhythmias(self) -> None:
        """Build the Arrhythmias tab: event cards (left) + ECG viewer (right)."""
        t = self.tabs.tab("Arrhythmias")
        t.grid_rowconfigure(1, weight=1)
        t.grid_columnconfigure(0, weight=1)

        # ── Row 0 : Action bar ────────────────────────────────
        bar = ctk.CTkFrame(t, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))

        self.btn_run_arrhythmia = ctk.CTkButton(
            bar, text="⚡  Classify",
            command=self._run_arrhythmia_analysis,
            fg_color=RED, hover_color=RED_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=8,
        )
        self.btn_run_arrhythmia.pack(side="left")  # type: ignore[union-attr]

        # ── Baseline window ──
        ctk.CTkLabel(bar, text="Baseline (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(SPACE_L, SPACE_XS))
        self.ent_arr_baseline = ctk.CTkEntry(
            bar, width=48, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_baseline.insert(0, "30")
        self.ent_arr_baseline.pack(side="left", padx=(0, SPACE_XS))
        _tip = ctk.CTkLabel(bar, text="(pré-stimulus)", font=FONT_KPI_LABEL,
                            text_color=MUTED)
        _tip.pack(side="left", padx=(0, SPACE_M))

        # ── Brady/tachy threshold ──
        ctk.CTkLabel(bar, text="ΔHR threshold (%):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(0, SPACE_XS))
        self.ent_arr_brady_pct = ctk.CTkEntry(
            bar, width=44, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_brady_pct.insert(0, "20")
        self.ent_arr_brady_pct.pack(side="left", padx=(0, SPACE_M))

        # ── Minimum beats ──
        ctk.CTkLabel(bar, text="Min duration (beats):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(0, SPACE_XS))
        self.ent_arr_min_beats = ctk.CTkEntry(
            bar, width=44, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_min_beats.insert(0, "10")
        self.ent_arr_min_beats.pack(side="left", padx=(0, SPACE_M))

        self.lbl_arrhythmia_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_arrhythmia_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        ctk.CTkButton(bar, text="📋 Copy TSV",
                      command=self._copy_arrhythmia_tsv,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=28, corner_radius=5,
                      ).pack(side="right")

        # ── Row 1 : Split body ────────────────────────────────
        body = tk.Frame(t, bg=BG, bd=0, highlightthickness=0)
        body.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(0, SPACE_S))
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(0, weight=0)   # event list — fixed
        body.grid_columnconfigure(1, weight=1)   # ECG viewer — elastic

        # ── Left : scrollable event list ─────────────────────
        list_outer = ctk.CTkFrame(body, fg_color=PANEL, corner_radius=6, width=280)
        list_outer.grid(row=0, column=0, sticky="ns", padx=(0, SPACE_S))
        list_outer.grid_propagate(False)
        list_outer.grid_rowconfigure(1, weight=1)
        list_outer.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(list_outer, text="DETECTED EPISODES",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))

        self._arr_event_scroll = ctk.CTkScrollableFrame(
            list_outer, fg_color="transparent",
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=BORDER2,
        )
        self._arr_event_scroll.grid(row=1, column=0, sticky="nsew", padx=SPACE_XS, pady=(0, SPACE_S))
        self._arr_event_scroll.grid_columnconfigure(0, weight=1)
        self._arr_card_widgets = []

        # ── Right : ECG detail panel ──────────────────────────
        right = tk.Frame(body, bg=BG, bd=0, highlightthickness=0)
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)

        # Edit / nav controls bar
        ebar = ctk.CTkFrame(right, fg_color=PANEL, corner_radius=0)
        ebar.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, SPACE_XS))

        self.lbl_arr_event_title = ctk.CTkLabel(
            ebar, text="← Click on an episode",
            font=FONT_SIDEBAR_HDR, text_color=MUTED, anchor="w")
        self.lbl_arr_event_title.pack(side="left", padx=SPACE_M, pady=SPACE_S)  # type: ignore[union-attr]

        # Separator
        ctk.CTkFrame(ebar, width=1, fg_color=BORDER).pack(
            side="left", fill="y", padx=(SPACE_S, SPACE_S), pady=SPACE_S)

        self.btn_arr_edit = ctk.CTkButton(
            ebar, text="Edit Peaks", width=96, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._toggle_arr_edit_mode,
        )
        self.btn_arr_edit.pack(side="left", padx=SPACE_XS)

        self.btn_arr_undo = ctk.CTkButton(
            ebar, text="↩ Undo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._undo_edit,
        )
        self.btn_arr_undo.pack(side="left", padx=SPACE_XS)
        self.btn_arr_redo = ctk.CTkButton(
            ebar, text="↪ Redo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._redo_edit,
        )
        self.btn_arr_redo.pack(side="left", padx=SPACE_XS)

        self.lbl_arr_edit_hint = ctk.CTkLabel(
            ebar, text="L-click: exclude/restore   R-click: add",
            font=FONT_HINT, text_color=ORANGE,
        )
        # Nav: ← → buttons + window size
        self.btn_arr_prev = ctk.CTkButton(
            ebar, text="◀", width=32, height=26, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            command=lambda: self._arr_navigate(-1),
        )
        self.btn_arr_prev.pack(side="right", padx=SPACE_XS)
        self.btn_arr_next = ctk.CTkButton(
            ebar, text="▶", width=32, height=26, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            command=lambda: self._arr_navigate(+1),
        )
        self.btn_arr_next.pack(side="right", padx=SPACE_XS)
        ctk.CTkLabel(ebar, text="Window (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="right", padx=(0, SPACE_XS))
        self.ent_arr_win = ctk.CTkEntry(
            ebar, width=48, height=26, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_win.insert(0, "3")
        self.ent_arr_win.pack(side="right", padx=(0, SPACE_S))

        # ECG canvas
        ecg_frame = tk.Frame(right, bg=PLOT["bg"], bd=0, highlightthickness=0)
        ecg_frame.grid(row=1, column=0, sticky="nsew")
        self._slots["arr_detail"] = CanvasSlot(ecg_frame, 14, 5.0, toolbar=False)

        # Wire up click handler on the canvas
        def _wire_arr_click(slot=self._slots["arr_detail"]):
            slot.canvas.mpl_connect("button_press_event", self._on_arr_detail_click)
            slot.canvas.mpl_connect("scroll_event",       self._on_arr_scroll)
        self.after(200, _wire_arr_click)

    def _run_arrhythmia_analysis(self) -> None:
        if self._rpeaks_ok is None or self._fs is None:
            self._set_status("Run Core Analysis first.", RED)
            return

        rpeaks = self._windowed_peaks()
        if rpeaks is None or len(rpeaks) < 5:
            self._set_status("Not enough peaks in the analysis window.", ORANGE)
            return
        fs      = float(self._fs)
        ctx_key = self._exp_context

        try:
            baseline_s = max(5.0, float(self.ent_arr_baseline.get()))
        except (ValueError, AttributeError):
            baseline_s = 30.0
        try:
            brady_pct = max(5.0, min(60.0, float(self.ent_arr_brady_pct.get())))
        except (ValueError, AttributeError):
            brady_pct = 20.0
        try:
            min_beats = max(3, int(self.ent_arr_min_beats.get()))
        except (ValueError, AttributeError):
            min_beats = 10

        def _worker():
            return classify_arrhythmias(
                rpeaks, fs, ctx_key,
                baseline_s=baseline_s,
                brady_pct=brady_pct,
                min_brady_beats=min_beats,
            )

        def _done(events: "list[ArrhythmiaEvent]"):
            self._arrhythmia_events = events
            self._arr_selected_idx  = -1

            sev_colors = {"alert": RED_MID, "warning": AMBER, "info": BLUE_MID}
            kind_icons = {
                "bradycardia": "🔵", "tachycardia": "🔴", "pause": "⏸",
                "esv_run": "⚡", "irregular_run": "〰", "block_av": "🔶",
            }

            # ── RR timeline (right panel, initial state) ──────
            t_peaks = rpeaks / fs
            rr_ms   = np.diff(rpeaks).astype(float) / fs * 1000
            t_rr    = (t_peaks[:-1] + t_peaks[1:]) / 2

            def draw_rr_timeline(fig):
                ax = fig.subplots(1, 1)
                style_axes(ax)
                ax.plot(t_rr, rr_ms, color=PLOT.get("ecg",CYAN_BRIGHT), lw=0.8, zorder=2)
                ax.set_ylabel("RR (ms)"); ax.set_xlabel("Time (s)")

                # Baseline and threshold lines (only if events contain brady/tachy)
                brady_events = [e for e in events if e.kind in ("bradycardia","tachycardia")
                                and e.baseline_hr > 0]
                if brady_events:
                    bl_hr  = brady_events[0].baseline_hr
                    bl_rr  = 60_000.0 / bl_hr
                    try:
                        bpct = max(5.0, float(self.ent_arr_brady_pct.get()))
                    except Exception:
                        bpct = 20.0
                    brady_rr = 60_000.0 / (bl_hr * (1 - bpct / 100))
                    tachy_rr = 60_000.0 / (bl_hr * (1 + bpct / 100))
                    try:
                        bl_end = min(float(self.ent_arr_baseline.get()),
                                     float(t_rr[-1]) if len(t_rr) else 30.0)
                    except Exception:
                        bl_end = 30.0
                    # Shade baseline window
                    ax.axvspan(0, bl_end, alpha=0.08, color=GREEN_MID,
                               zorder=0, label="Baseline window")
                    ax.axvline(bl_end, color=GREEN_MID, lw=1.0,
                               ls=":", alpha=0.6, zorder=3)
                    # Baseline RR line
                    ax.axhline(bl_rr, color=GREEN_MID, lw=1.2,
                               ls="--", alpha=0.8, zorder=4,
                               label=f"Baseline {bl_hr:.0f} bpm")
                    # Brady / tachy threshold lines
                    ax.axhline(brady_rr, color=AMBER, lw=0.9,
                               ls="--", alpha=0.65, zorder=4,
                               label=f"Seuil brady −{bpct:.0f}%")
                    ax.axhline(tachy_rr, color=BLUE_MID, lw=0.9,
                               ls="--", alpha=0.65, zorder=4,
                               label=f"Seuil tachy +{bpct:.0f}%")

                ax.set_title(
                    "RR series — click an episode to zoom",
                    loc="left", fontsize=9)
                for ev in events:
                    c = sev_colors.get(ev.severity, "#888")
                    ax.axvspan(ev.t_start, max(ev.t_end, ev.t_start + 0.05),
                               alpha=0.20, color=c, zorder=1)
                    ax.axvline(ev.t_start, color=c, lw=0.7, ls="--",
                               alpha=0.55, zorder=3)
                if brady_events:
                    ax.legend(loc="upper right", fontsize=7,
                              facecolor=PLOT.get("bg",NAVY),
                              labelcolor=PLOT.get("text",GRAY_LIGHT),
                              edgecolor=PLOT.get("border","#333"))

            self._slots["arr_detail"].update(draw_rr_timeline)

            # ── Build clickable event cards ───────────────────
            for w in self._arr_card_widgets:
                try: w.destroy()
                except Exception: pass
            self._arr_card_widgets.clear()

            if not events:
                lbl = ctk.CTkLabel(
                    self._arr_event_scroll,
                    text="No arrhythmia detected\nfor the active context.",
                    font=FONT_SMALL, text_color=MUTED, justify="center",
                )
                lbl.grid(row=0, column=0, pady=SPACE_L)
                self._arr_card_widgets.append(lbl)
            else:
                for idx, ev in enumerate(events):
                    self._build_arrhythmia_card(idx, ev, sev_colors, kind_icons)

            # ── TSV store ─────────────────────────────────────
            tsv_rows = ["Type\tStart_s\tEnd_s\tDuration_s\tHR_bpm\tBaseline_bpm\tDelta_pct\tRR_ms\tSeverity\tDescription"]
            for ev in events:
                tsv_rows.append(
                    f"{ev.kind}\t{ev.t_start:.2f}\t{ev.t_end:.2f}\t"
                    f"{ev.duration_s:.2f}\t{ev.hr_mean:.1f}\t"
                    f"{ev.baseline_hr:.1f}\t{ev.delta_pct:.1f}\t"
                    f"{ev.rr_mean:.1f}\t{ev.severity}\t{ev.label}"
                )
            self._arrhythmia_tsv = "\n".join(tsv_rows)

            n = len(events)
            self.lbl_arrhythmia_status.configure(  # type: ignore[union-attr]
                text=f"  {n} episode{'s' if n != 1 else ''} — click to explore",
                text_color=RED if any(e.severity=="alert" for e in events)
                           else (ORANGE if n else GREEN),
            )
            self._set_status(f"Arrhythmia classification — {n} episode(s)", GREEN)
            self.tabs.set("Arrhythmias")

        if self.btn_run_arrhythmia is None:
            return
        self._start_async_result(
            self.btn_run_arrhythmia, "Classifying…", _worker, _done)

    # ── Event card builder ────────────────────────────────────

    def _build_arrhythmia_card(
        self, idx: int, ev: "ArrhythmiaEvent",
        sev_colors: dict, kind_icons: dict,
    ) -> None:
        c_sev = sev_colors.get(ev.severity, MUTED)
        icon  = kind_icons.get(ev.kind, "·")

        card = ctk.CTkFrame(
            self._arr_event_scroll,
            fg_color=CARD, corner_radius=6,
            border_width=2, border_color=BORDER,
        )
        card.grid(row=idx, column=0, sticky="ew", padx=SPACE_S, pady=(0, SPACE_S))
        card.grid_columnconfigure(1, weight=1)

        ctk.CTkFrame(card, width=4, fg_color=c_sev,
                     corner_radius=0).grid(row=0, column=0, rowspan=3, sticky="ns")

        hdr = ctk.CTkFrame(card, fg_color="transparent")
        hdr.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, SPACE_S), pady=(SPACE_S, SPACE_XS))
        hdr.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(hdr, text=f"{icon}  {ev.kind.replace('_', ' ').title()}",
                     font=FONT_SIDEBAR_HDR, text_color=TEXT,
                     anchor="w").grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(hdr, text=ev.severity.upper(),
                     font=FONT_BADGE, text_color=c_sev,
                     anchor="e").grid(row=0, column=1, sticky="e")

        ctk.CTkLabel(
            card,
            text=f"  {ev.t_start:.2f} s → {ev.t_end:.2f} s  "
                 f"({ev.duration_s:.1f} s)   {ev.hr_mean:.0f} bpm",
            font=FONT_HINT, text_color=LIGHT, anchor="w",
        ).grid(row=1, column=1, sticky="ew", padx=(SPACE_XS, SPACE_S))

        # Delta vs baseline line (only for brady/tachy)
        if ev.baseline_hr > 0 and ev.kind in ("bradycardia", "tachycardia"):
            arrow  = "↓" if ev.delta_pct < 0 else "↑"
            d_col  = RED_LIGHT if ev.delta_pct < 0 else CORAL
            d_text = (f"  {arrow}{abs(ev.delta_pct):.0f}% vs baseline "
                      f"({ev.baseline_hr:.0f} → {ev.hr_mean:.0f} bpm)")
            ctk.CTkLabel(
                card, text=d_text,
                font=FONT_SUBSECTION, text_color=d_col, anchor="w",
            ).grid(row=2, column=1, sticky="ew", padx=(SPACE_XS, SPACE_S))
            desc_row = 3
        else:
            desc_row = 2

        ctk.CTkLabel(
            card, text=f"  {ev.label}",
            font=FONT_KPI_LABEL, text_color=MUTED, anchor="w",
            wraplength=230, justify="left",
        ).grid(row=desc_row, column=1, sticky="ew", padx=(SPACE_XS, SPACE_S), pady=(0, SPACE_S))

        def _on_click(_event=None, _idx=idx):
            self._select_arrhythmia_event(_idx)

        for widget in [card] + list(card.winfo_children()):
            widget.bind("<Button-1>", _on_click)
            try:
                for sub in widget.winfo_children():
                    sub.bind("<Button-1>", _on_click)
            except Exception as e:
                log.debug("winfo_children bind failed: %s", e)

        self._arr_card_widgets.append(card)

    # ── Event selection & ECG viewer ─────────────────────────

    def _select_arrhythmia_event(self, idx: int) -> None:
        """Highlight selected card and load the ECG window for this event."""
        self.nav_ctrl.select_arrhythmia_event(idx)

    def _draw_arr_detail(self) -> None:
        """Draw ECG strip for the selected arrhythmia event, with editable R peaks."""
        if self._signal_flt is None or self._time is None:
            return

        sig_flt  = self._signal_flt
        time     = self._time
        fs       = self._fs

        try:
            win = float(self.ent_arr_win.get())
        except Exception:
            win = self._arr_win
        win = max(0.5, win)
        self._arr_win = win

        t_start = self._arr_nav_pos
        t_end   = min(float(time[-1]), t_start + win)
        # Index-slice instead of a full-length boolean mask: `time` is always
        # uniformly spaced (np.arange(n)/fs), so the window bounds map
        # directly to a sample-index range. A full `(time >= t0) & (time <= t1)`
        # mask costs O(n_samples_total) on every redraw — this view redraws on
        # every navigation step/click, which becomes noticeably slow on long
        # or high-fs recordings. Slicing is O(window size) instead.
        i0 = max(0, int(t_start * fs))
        i1 = min(len(time), int(t_end * fs) + 1)
        mask_t  = slice(i0, i1)

        # Peak arrays
        rp_ok    = self._rpeaks_ok    if self._rpeaks_ok    is not None else np.array([])
        rp_excl  = self._rpeaks_manual_excl  if self._rpeaks_manual_excl  is not None else np.array([])
        rp_added = self._rpeaks_manual_added if self._rpeaks_manual_added is not None else np.array([])

        def _in_win(idx: np.ndarray) -> np.ndarray:
            return (idx / fs >= t_start) & (idx / fs <= t_end)

        mask_ok    = _in_win(rp_ok)    if len(rp_ok)    else np.array([], bool)
        mask_excl  = _in_win(rp_excl)  if len(rp_excl)  else np.array([], bool)
        mask_added = _in_win(rp_added) if len(rp_added) else np.array([], bool)

        # Selected event span
        ev = (self._arrhythmia_events[self._arr_selected_idx]
              if 0 <= self._arr_selected_idx < len(self._arrhythmia_events)
              else None)
        ev_t_start = ev.t_start if ev else None
        ev_t_end   = ev.t_end   if ev else None
        sev_color  = {"alert": RED_MID, "warning": AMBER, "info": BLUE_MID
                      }.get(ev.severity, MUTED) if ev else MUTED
        ev_label   = ev.label if ev else ""
        edit_mode  = self._arr_edit_mode

        n_in_win   = int(mask_ok.sum())
        t_amp      = self._thresh_amp

        def draw(fig):
            ax = fig.add_subplot(111)
            style_axes(ax)

            # ECG trace
            ax.plot(time[mask_t], sig_flt[mask_t],
                    color=PLOT.get("signal", CYAN), lw=0.9, zorder=2,
                    label="ECG filtré")

            # Event span highlight
            if ev_t_start is not None and ev_t_end is not None:
                _span_lo = max(ev_t_start, t_start)
                _span_hi = min(max(ev_t_end, ev_t_start + 0.05), t_end)
                if _span_lo < t_end and _span_hi > t_start:
                    ax.axvspan(_span_lo, _span_hi,
                               color=sev_color, alpha=0.14, zorder=1, linewidth=0)
                    ax.axvline(ev_t_start, color=sev_color, lw=1.0, ls="--",
                               alpha=0.7, zorder=3)
                    if ev_t_end > t_start:
                        ax.axvline(min(ev_t_end, t_end), color=sev_color,
                                   lw=1.0, ls="--", alpha=0.7, zorder=3)
                    # Label inside the span
                    _lx = max(ev_t_start, t_start) + 0.02
                    if _lx < t_end:
                        ylo, yhi = ax.get_ylim()
                        ax.text(_lx, yhi * 0.90, ev_label,
                                ha="left", va="top", fontsize=8, color=sev_color,
                                fontweight="bold", zorder=8,
                                bbox=dict(boxstyle="round,pad=0.2",
                                          fc=PLOT.get("bg",NAVY),
                                          ec=sev_color, alpha=0.85, lw=0.8))

            # R peaks
            if mask_excl.any():
                ax.scatter(rp_excl[mask_excl] / fs, sig_flt[rp_excl[mask_excl]],
                           color=RED, s=90, zorder=6, marker="x", linewidths=2,
                           label="Excluded")
            if mask_ok.any():
                ax.scatter(rp_ok[mask_ok] / fs, sig_flt[rp_ok[mask_ok]],
                           color=PLOT.get("rpeak_ok","#00E676"), s=55, zorder=5,
                           marker="o", label="Acceptés")
            if mask_added.any():
                ax.scatter(rp_added[mask_added] / fs, sig_flt[rp_added[mask_added]],
                           color=CYAN, s=140, zorder=7,
                           marker="*", linewidths=1.2, edgecolors="#006064",
                           label="Added")

            ax.axhline(t_amp, color=PLOT.get("threshold",AMBER_DARK),
                       lw=1.2, ls="--", alpha=0.6)
            ax.set_xlabel("Time (s)")
            ax.set_ylabel("Amplitude (norm.)")

            edit_tag = "  ·  ✏ EDIT" if edit_mode else ""
            ax.set_title(
                f"{t_start:.2f}–{t_end:.2f} s  ·  {n_in_win} peaks{edit_tag}",
                loc="left",
                color=ORANGE if edit_mode else PLOT.get("text",GRAY_LIGHT),
                fontsize=9,
            )
            ax.legend(framealpha=0, loc="upper right", fontsize=8)

        self._slots["arr_detail"].update(draw)

    # ── Edit mode toggle for arrhythmia tab ──────────────────

    def _toggle_arr_edit_mode(self) -> None:
        self._arr_edit_mode = not self._arr_edit_mode
        if self._arr_edit_mode:
            self.btn_arr_edit.configure(
                fg_color=ORANGE, hover_color=ORANGE_DEEP,
                text_color="white", text="Edit Mode ON",
            )
            self.lbl_arr_edit_hint.pack(side="left", padx=(SPACE_S, 0))
        else:
            self.btn_arr_edit.configure(
                fg_color=BORDER, hover_color=BORDER2,
                text_color=MUTED, text="Edit Peaks",
            )
            self.lbl_arr_edit_hint.pack_forget()
        self._draw_arr_detail()

    # ── Click handler (mirrors _on_detail_click) ─────────────

    def _on_arr_detail_click(self, event) -> None:
        """Left-click: toggle exclusion.  Right-click: add/remove peak."""
        if not self._arr_edit_mode:
            return
        if event.xdata is None or self._signal_flt is None:
            return
        if self._rpeaks_ok is None:
            return

        fs         = self._fs
        click_time = float(event.xdata)
        click_samp = int(np.clip(int(round(click_time * fs)),
                                 0, len(self._signal_flt) - 1))

        tol_s    = max(MouseECG.MIN_RR_MS / 1000 / 2, self._arr_win * 0.03)
        tol_samp = int(tol_s * fs)
        is_left  = (event.button == 1)
        is_right = (event.button == 3)

        if is_right:
            # Remove manually added peak near click
            if (self._rpeaks_manual_added is not None
                    and len(self._rpeaks_manual_added)):
                dists = np.abs(self._rpeaks_manual_added - click_samp)
                ni = int(np.argmin(dists))
                if dists[ni] <= tol_samp:
                    self._push_edit_undo()
                    self._manual_added.discard(int(self._rpeaks_manual_added[ni]))
                    self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
                    self._draw_arr_detail()
                    self._draw_detail(self._nav_pos)
                    self._set_status(f"Peak removed at {click_time:.3f} s", ORANGE)
                    self._update_undo_btns()
                    return
            # Snap to local max and add
            sig  = self._signal_flt
            lo   = max(0, click_samp - tol_samp)
            hi   = min(len(sig), click_samp + tol_samp + 1)
            new_samp = lo + int(np.argmax(sig[lo:hi]))
            if (self._rpeaks_ok is not None and len(self._rpeaks_ok)):
                if np.min(np.abs(self._rpeaks_ok - new_samp)) < int(MouseECG.MIN_RR_MS / 1000 * fs * 0.5):
                    self._set_status(
                        f"Too close to existing peak à {click_time:.3f} s", ORANGE)
                    return
            self._push_edit_undo()
            self._manual_added.add(new_samp)
            self._manual_excluded.discard(new_samp)
            self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
            self._draw_arr_detail()
            self._draw_detail(self._nav_pos)
            self._set_status(f"Peak added at {new_samp / fs:.3f} s", ORANGE)
            self._update_undo_btns()
            return

        if not is_left:
            return

        # Toggle exclusion of nearest peak
        added_set = self._manual_added
        base_ok   = np.array([p for p in self._rpeaks_ok if p not in added_set], int) \
                    if self._rpeaks_ok is not None else np.array([], int)
        excl_arr  = self._rpeaks_manual_excl if self._rpeaks_manual_excl is not None \
                    else np.array([], int)
        candidates = np.concatenate([base_ok, excl_arr])
        if len(candidates) == 0:
            return
        dists = np.abs(candidates / fs - click_time)
        ni    = int(np.argmin(dists))
        if dists[ni] > tol_s:
            return
        peak_idx = int(candidates[ni])
        self._push_edit_undo()
        if peak_idx in self._manual_excluded:
            self._manual_excluded.discard(peak_idx)
            msg = f"Peak restored at {peak_idx / fs:.3f} s"
        else:
            self._manual_excluded.add(peak_idx)
            msg = f"Peak excluded at {peak_idx / fs:.3f} s"
        self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
        self._draw_arr_detail()
        self._draw_detail(self._nav_pos)
        self._set_status(msg + "  — rerun Core Analysis pour mettre à jour HRV", ORANGE)
        self._update_undo_btns()

    # ── Scroll zoom on arrhythmia ECG ────────────────────────

    def _on_arr_scroll(self, event) -> None:
        if event.xdata is None or self._time is None:
            return
        factor   = 0.8 if event.button == "up" else 1.25
        new_win  = max(0.3, min(float(self._time[-1]), self._arr_win * factor))
        cursor_x = float(event.xdata)
        frac     = (cursor_x - self._arr_nav_pos) / max(self._arr_win, 1e-6)
        t_start  = max(0.0, cursor_x - frac * new_win)
        t_start  = min(t_start, max(0.0, float(self._time[-1]) - new_win))
        self._arr_win     = new_win
        self._arr_nav_pos = t_start
        try:
            self.ent_arr_win.delete(0, "end")
            self.ent_arr_win.insert(0, f"{new_win:.2f}")
        except Exception as e:
            log.debug("ent_arr_win update failed: %s", e)
        self._draw_arr_detail()

    # ── ◀ ▶ navigation ───────────────────────────────────────

    def _arr_navigate(self, direction: int) -> None:
        self.nav_ctrl.arr_navigate(direction)

    # ── sync undo/redo buttons in arrhythmia tab ─────────────

    def _update_undo_btns(self) -> None:
        """Update all undo/redo button instances (Detection + Arrhythmias tabs)."""
        n_u, n_r = len(self._edit_undo), len(self._edit_redo)
        for attr, n in [("btn_undo_edit", n_u), ("btn_redo_edit", n_r),
                        ("btn_arr_undo",  n_u), ("btn_arr_redo",  n_r)]:
            try:
                btn = getattr(self, attr)
                key = "undo" if "undo" in attr else "redo"
                sym = "↩" if key == "undo" else "↪"
                btn.configure(
                    state="normal" if n else "disabled",
                    text=f"{sym} {key.title()} ({n})" if n else f"{sym} {key.title()}",
                )
            except Exception as e:
                log.debug("undo/redo button configure failed: %s", e)

    def _copy_arrhythmia_tsv(self) -> None:
        tsv = getattr(self, "_arrhythmia_tsv", None)
        if not tsv:
            self._set_status("Run arrhythmia classification first.", RED)
            return
        self.clipboard_clear()
        self.clipboard_append(tsv)
        self._set_status("Arrhythmias copied to clipboard (Excel ready)", GREEN)

    def _build_tab_intervals(self) -> None:
        """Interval delineation tab: P/Q/R/S/T measurement per beat."""
        t = self.tabs.tab("Intervals")

        # ── Action bar ─────────────────────────────────────────────────────
        bar = ctk.CTkFrame(t, fg_color="transparent")
        bar.pack(side="top", fill="x", padx=SPACE_M, pady=(SPACE_M, SPACE_S))

        self.btn_run_ivl = ctk.CTkButton(
            bar, text="⚡  Delineate waves",
            command=self._run_intervals,
            fg_color=ORANGE, hover_color=ORANGE_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=8, state="disabled",
        )
        self.btn_run_ivl.pack(side="left")

        self.lbl_ivl_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_ivl_status.pack(side="left", padx=SPACE_M)

        # QTc formula selector
        ctk.CTkLabel(bar, text="QTc formula:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="right", padx=(0, SPACE_S))
        self.cb_qtc_formula = ctk.CTkComboBox(
            bar, width=140, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, button_color=BORDER2,
            text_color=TEXT, dropdown_fg_color=BG, dropdown_text_color=TEXT,
            values=["Mitchell (∛RR)", "Bazett (√RR)"],
            command=self._on_qtc_formula_change,
        )
        self.cb_qtc_formula.set("Mitchell (∛RR)")
        self.cb_qtc_formula.pack(side="right", padx=(0, SPACE_M))

        # Permissive bounds toggle
        self.sw_permissive = self._switch(bar, "Permissive bounds", dict(padx=0))
        self.sw_permissive.pack(side="right", padx=(0, SPACE_L))

        # ── Interval verifier nav bar (populated by _launch_interval_verifier) ──
        self.frm_ivl_nav = tk.Frame(t, bg=PANEL, bd=0, highlightthickness=0)
        self.frm_ivl_nav.pack(side="top", fill="x")

        # ── Body: violin plots (left) + annotated beat strip (right) ─────────
        body = tk.Frame(t, bg=BG, bd=0, highlightthickness=0)
        body.pack(side="top", fill="both", expand=True, padx=SPACE_S, pady=(0, SPACE_S))

        # Left: violin/box distributions  (30% width)
        left_card = tk.Frame(body, bg=CARD, bd=0, highlightthickness=0)
        left_card.pack(side="left", fill="both", expand=False)
        left_card.pack_propagate(False)
        left_card.configure(width=1)  # will be reset by pack weight

        # Use a PanedWindow for resizable split
        paned = tk.PanedWindow(body, orient=tk.HORIZONTAL,
                               bg=BORDER, sashwidth=4, sashrelief="flat",
                               handlesize=0)
        paned.pack(fill="both", expand=True)

        left_inner = tk.Frame(paned, bg=CARD, bd=0, highlightthickness=0)
        right_inner = tk.Frame(paned, bg=CARD, bd=0, highlightthickness=0)
        paned.add(left_inner, minsize=180, stretch="always")
        paned.add(right_inner, minsize=300, stretch="always")
        # Set initial split after widget is mapped
        def _set_sash(event=None):
            try:
                paned.sash_place(0, max(200, paned.winfo_width() // 3), 0)
            except Exception:
                pass
        paned.bind("<Map>", _set_sash)

        self._slots["intervals"] = CanvasSlot(left_inner, 6, 6, toolbar=False)
        self._slots["intervals_ecg"] = CanvasSlot(right_inner, 11, 6, toolbar=False)


    def _build_tab_beat_template(self) -> None:
        """Beat template tab: mean beat ± SD and morphology distributions."""
        t = self.tabs.tab("Beat Template")
        t.grid_rowconfigure(1, weight=3)
        t.grid_rowconfigure(2, weight=2)
        t.grid_columnconfigure(0, weight=1)

        # ── Header bar ──────────────────────────────────────────────────────
        bar = ctk.CTkFrame(t, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))
        ctk.CTkLabel(bar, text="BEAT TEMPLATE  —  mean ± SD  ·  morphology",
                     font=FONT_SIDEBAR_HDR, text_color=MUTED).pack(side="left")
        ctk.CTkButton(
            bar, text="⚙  Landmarks",
            command=self._open_wave_template_editor,
            fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
            font=FONT_BTN_SEC, height=28, corner_radius=6,
        ).pack(side="right")
        self.lbl_template_info = ctk.CTkLabel(
            bar, text="", font=FONT_HINT, text_color=MUTED, anchor="e")
        self.lbl_template_info.pack(side="right", padx=(0, SPACE_M))  # type: ignore[union-attr]

        # ── Mean beat plot ──────────────────────────────────────────────────
        beat_card = ctk.CTkFrame(t, fg_color=CARD, corner_radius=0)
        beat_card.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(0, SPACE_S))
        beat_card.grid_rowconfigure(0, weight=1)
        beat_card.grid_columnconfigure(0, weight=1)
        self._slots["beat"] = CanvasSlot(beat_card, 14, 5.0, toolbar=False)

        # ── Distributions: amplitude + correlation ──────────────────────────
        dist_card = ctk.CTkFrame(t, fg_color=CARD, corner_radius=0)
        dist_card.grid(row=2, column=0, sticky="nsew", padx=SPACE_S, pady=(0, SPACE_S))
        dist_card.grid_rowconfigure(0, weight=1)
        dist_card.grid_columnconfigure(0, weight=1)
        self._slots["beat_dist"] = CanvasSlot(dist_card, 14, 4.0, toolbar=False)


    # ════════════════════════════════════════════════════════
    #  QTC FORMULA
    # ════════════════════════════════════════════════════════

    def _qtc_formula(self) -> str:
        """Return 'mitchell' or 'bazett' from the Intervals tab combo selector."""
        try:
            val = self.cb_qtc_formula.get()  # type: ignore
            return "bazett" if "Bazett" in val else "mitchell"
        except Exception:
            return "mitchell"

    def _on_qtc_formula_change(self, _choice: str = "") -> None:
        """Re-compute QTc with the selected formula and refresh interval plots."""
        if self._results is None:
            return
        ivl = self._results.get("intervals")
        if ivl is None or ivl.empty or "QT_ms" not in ivl.columns:
            return
        formula = self._qtc_formula()
        qt_ms  = ivl["QT_ms"].values.astype(float)
        rr_arr = ivl["RR_ms"].values.astype(float)
        rr_s   = np.clip(rr_arr, MouseECG.RR_MIN_MS, None) / 1000.0
        if formula == "bazett":
            qtc = qt_ms / np.sqrt(rr_s)
        else:
            qtc = qt_ms / (rr_s ** (1.0 / 3.0))
        qtc = np.where((qtc < MouseECG.QTC_ABS_MIN) | (qtc > MouseECG.QTC_ABS_MAX),
                       np.nan, qtc)
        ivl = ivl.copy()
        ivl["QTc_ms"] = qtc
        self._results["intervals"] = ivl
        self._update_interval_plots()
        fname = "Mitchell (∛RR)" if formula == "mitchell" else "Bazett (√RR)"
        self._set_status(f"QTc recomputed — formula: {fname}  ✓", GREEN)

    def _update_interval_plots(self) -> None:
        """Redraw the interval violin plots after a QTc formula change."""
        if self._results is None:
            return
        ivl = self._results.get("intervals")
        if ivl is None or ivl.empty:
            return
        try:
            self._plot_intervals(self._results)
        except Exception as exc:
            log.debug("_update_interval_plots: %s", exc)

    # ════════════════════════════════════════════════════════
    #  ROLLING HRV
    # ════════════════════════════════════════════════════════

    def _compute_rolling_hrv(self) -> None:
        """Compute sliding-window HRV and render the timeline plot."""
        if self._rpeaks_ok is None or self._fs is None:
            self._set_status("Run Core Analysis first.", RED)
            return

        try:
            win_s  = max(5.0,  float(self.ent_roll_win.get()))
            step_s = max(1.0,  float(self.ent_roll_step.get()))
        except ValueError:
            self._set_status("Invalid window / step.", RED)
            return

        active = [m for m, cb in self._roll_metrics.items() if cb.get()]
        if not active:
            self._set_status("Sélectionner au moins une métrique.", RED)
            return

        rpeaks = self._windowed_peaks()
        if rpeaks is None or len(rpeaks) < 5:
            self._set_status("Not enough peaks in the analysis window.", ORANGE)
            return
        fs     = float(self._fs)

        def _worker():
            t_peaks  = rpeaks / fs
            t_start  = t_peaks[0]
            t_end    = t_peaks[-1]
            starts   = np.arange(t_start, t_end - win_s + step_s * 0.5, step_s)
            rows    = []
            n_wins  = len(starts)
            for i, t0 in enumerate(starts):
                t1   = t0 + win_s
                mask = (t_peaks >= t0) & (t_peaks < t1)
                ep   = rpeaks[mask]
                if len(ep) < 5:
                    continue
                rr = np.diff(ep).astype(float) / fs * 1000
                hr   = float(60_000.0 / rr.mean()) if len(rr) else np.nan
                sdnn = float(rr.std(ddof=1))        if len(rr) > 1 else np.nan
                rmssd = float(np.sqrt(np.mean(np.diff(rr)**2))) if len(rr) > 2 else np.nan
                diffs = np.abs(np.diff(rr))
                pnn6  = float(100.0 * np.sum(diffs > MouseECG.PNN_THRESHOLD) / len(diffs)) \
                        if len(diffs) else np.nan
                rows.append({
                    "t_mid": round(t0 + win_s / 2, 2),
                    "HR":    round(hr,    1),
                    "SDNN":  round(sdnn,  2),
                    "RMSSD": round(rmssd, 2),
                    "pNN6":  round(pnn6,  1),
                    "n_beats": len(ep),
                })
                pct = int((i + 1) / max(n_wins, 1) * 100)
                if i % max(1, n_wins // 20) == 0:
                    self.after(0, lambda p=pct, ii=i+1, tot=n_wins:
                               self._set_progress(p, f"Window {ii}/{tot}…"))
            return rows

        def _done(rows):
            if not rows:
                self.lbl_roll_status.configure(  # type: ignore[union-attr]
                    text="No valid windows — recording too short?",
                    text_color=RED)
                return

            df = pd.DataFrame(rows)
            self._rolling_hrv_df = df

            ctx    = EXPERIMENTAL_CONTEXTS.get(self._exp_context)
            colors = {"HR": ORANGE_DARK, "SDNN": BLUE_DARK,
                      "RMSSD": GREEN_DARK, "pNN6": PURPLE}
            ylabels = {"HR": "HR (bpm)", "SDNN": "SDNN (ms)",
                       "RMSSD": "RMSSD (ms)", "pNN6": "pNN6 (%)"}
            # reference bands per metric from active context
            ref_bands: "dict[str, tuple[float,float]]" = {}
            if ctx:
                ref_bands = {
                    "HR":    (ctx.hr_lo,    ctx.hr_hi),
                    "SDNN":  (ctx.sdnn_lo,  ctx.sdnn_hi),
                    "RMSSD": (ctx.rmssd_lo, ctx.rmssd_hi),
                    "pNN6":  (ctx.pnn6_lo,  ctx.pnn6_hi),
                }

            n_plots = len(active)

            def draw_rolling(fig):
                axes = fig.subplots(n_plots, 1, sharex=True)
                if n_plots == 1:
                    axes = [axes]
                fig.subplots_adjust(hspace=0.08)
                t = df["t_mid"].values

                for ax, metric in zip(axes, active):
                    style_axes(ax)
                    y     = df[metric].values
                    color = colors[metric]

                    # Shaded reference band from context
                    if metric in ref_bands:
                        lo, hi = ref_bands[metric]
                        ax.axhspan(lo, hi, alpha=0.10, color=color,
                                   linewidth=0, zorder=0)
                        ax.axhline(lo, color=color, lw=0.6,
                                   ls="--", alpha=0.45, zorder=1)
                        ax.axhline(hi, color=color, lw=0.6,
                                   ls="--", alpha=0.45, zorder=1)

                    ax.plot(t, y, color=color, lw=1.4, zorder=3)
                    ax.fill_between(t, y, alpha=0.08, color=color, zorder=2)
                    ax.set_ylabel(ylabels[metric], fontsize=8)
                    ax.set_title(
                        f"{metric}  —  window {win_s:.0f}s · pas {step_s:.0f}s",
                        loc="left", fontsize=8)
                    if ax is not axes[-1]:
                        ax.tick_params(labelbottom=False)

                axes[-1].set_xlabel("Time (s)")
                ctx_txt = ctx.label if ctx else ""
                if ctx_txt:
                    fig.text(0.99, 0.01, f"Contexte : {ctx_txt}",
                             ha="right", va="bottom", fontsize=7,
                             color=PLOT.get("muted", "#888"),
                             transform=fig.transFigure)

            self._slots["rolling_hrv"].update(draw_rolling)
            n = len(df)
            self.lbl_roll_status.configure(  # type: ignore[union-attr]
                text=f"  {n} windows · {win_s:.0f}s · pas {step_s:.0f}s  ✓",
                text_color=GREEN)
            self._set_status(f"Rolling HRV — {n} windows computed", GREEN)
            self.tabs.set("HRV"); self.after(50, lambda: self._on_hrv_view_change("Rolling"))

        self._start_async_result(
            self.btn_roll_compute, "Computing…", _worker, _done)

    def _copy_rolling_tsv(self) -> None:
        df = self._rolling_hrv_df
        if df is None:
            self._set_status("Compute Rolling HRV first.", RED)
            return
        self.clipboard_clear()
        self.clipboard_append(self._df_to_tsv(df))
        self._set_status("Rolling HRV copié dans le presse-papiers (Excel ready)", GREEN)

    def _build_tab_summary(self) -> None:
        t = self.tabs.tab("Summary")
        t.grid_rowconfigure(0, weight=0)   # KPI header — always visible
        t.grid_rowconfigure(1, weight=0)   # action bar
        t.grid_rowconfigure(2, weight=1)   # scrollable body
        t.grid_columnconfigure(0, weight=1)

        # ── KPI header strip ──────────────────────────────────────────────────
        kpi_frame = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0, height=76)
        kpi_frame.grid(row=0, column=0, sticky="ew")
        kpi_frame.pack_propagate(False)
        kpi_frame.grid_propagate(False)
        kpi_inner = ctk.CTkFrame(kpi_frame, fg_color="transparent")
        kpi_inner.pack(fill="both", expand=True, padx=SPACE_L, pady=SPACE_M)

        _KPI_DEFS = [
            ("hr_mean", "Mean HR",  "bpm", ORANGE_DARK),
            ("hr_min",  "Min HR",   "bpm", BLUE_DARK),
            ("hr_max",  "FC max",   "bpm", "#C62828"),
            ("sdnn",    "SDNN",     "ms",  "#1B5E20"),
            ("rmssd",   "RMSSD",    "ms",  PURPLE),
            ("pnn6",    "pNN6",     "%",   "#00695C"),
            ("porta",   "Porta",    "%",   "#37474F"),   # NEW — RR asymmetry
            ("lf_hf",   "LF/HF",   "",    BLUE_DARK),
            ("pr",      "PR",       "ms",  PINK),
            ("qrs",     "QRS",      "ms",  GREEN_DARK),
            ("qtc",     "QTc",      "ms",  ORANGE_DEEP),
            ("qt_disp", "QT disp.", "ms",  AMBER_DARK),   # NEW — QT dispersion
        ]
        self._sum_kpi_vals = {}
        for key, label, unit, color in _KPI_DEFS:
            card = ctk.CTkFrame(kpi_inner, fg_color=CARD, corner_radius=6)
            card.pack(side="left", fill="y", padx=(0, SPACE_S), expand=True)
            ctk.CTkLabel(card, text=label, font=FONT_BADGE,
                         text_color=color, anchor="center").pack(pady=(SPACE_S, 0), padx=SPACE_S)
            val_lbl = ctk.CTkLabel(card, text="—",
                                   font=FONT_TITLE,
                                   text_color=TEXT, anchor="center")
            val_lbl.pack(padx=SPACE_S)
            if unit:
                ctk.CTkLabel(card, text=unit, font=FONT_MICRO,
                             text_color=MUTED, anchor="center").pack(pady=(0, SPACE_S), padx=SPACE_S)
            else:
                ctk.CTkFrame(card, height=4, fg_color="transparent").pack()
            self._sum_kpi_vals[key] = val_lbl

        ctk.CTkFrame(t, height=1, fg_color=BORDER).grid(row=0, column=0, sticky="sew")

        # ── Action bar ────────────────────────────────────────────────────────
        btn_row = ctk.CTkFrame(t, fg_color="transparent", height=40)
        btn_row.grid(row=1, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, 0))
        btn_row.pack_propagate(False)
        for label, cmd in [
            ("📋  Copy Report",      self._copy_summary),
            ("💾  Enregistrer .txt",    self._save_summary_txt),
        ]:
            ctk.CTkButton(btn_row, text=label, command=cmd,
                          fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                          font=FONT_SMALL, height=28, corner_radius=5).pack(
                side="left", padx=(0, SPACE_S))
        self.lbl_epoch_info = ctk.CTkLabel(btn_row, text="", font=FONT_SMALL,
                                            text_color=MUTED)
        self.lbl_epoch_info.pack(side="right")

        # ── Scrollable body ───────────────────────────────────────────────────
        outer_scroll = ctk.CTkScrollableFrame(
            t, fg_color=BG,
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=BORDER2,
        )
        outer_scroll.grid(row=2, column=0, sticky="nsew", padx=0, pady=(SPACE_S, 0))
        outer_scroll.grid_columnconfigure(0, weight=1)
        self._sum_scroll = outer_scroll

        # ── Layout helpers ────────────────────────────────────────────────────
        def _section(parent, title: str, color: str = MUTED,
                     subtitle: str = "") -> ctk.CTkFrame:
            """Titled section with accent stripe, returns the content frame."""
            hdr = ctk.CTkFrame(parent, fg_color="transparent")
            hdr.pack(fill="x", padx=SPACE_M, pady=(SPACE_L, SPACE_S))
            stripe = ctk.CTkFrame(hdr, height=3, fg_color=color, corner_radius=2)
            stripe.pack(fill="x", pady=(0, SPACE_S))
            title_row = ctk.CTkFrame(hdr, fg_color="transparent")
            title_row.pack(fill="x")
            ctk.CTkLabel(title_row, text=title.upper(),
                         font=FONT_SIDEBAR_HDR,
                         text_color=color, anchor="w").pack(side="left")
            if subtitle:
                ctk.CTkLabel(title_row, text=f"  {subtitle}",
                             font=FONT_KPI_LABEL, text_color=MUTED,
                             anchor="w").pack(side="left", padx=(SPACE_S, 0))
            body = ctk.CTkFrame(parent, fg_color="transparent")
            body.pack(fill="x", padx=SPACE_S, pady=(SPACE_XS, 0))
            return body

        def _slot(parent, key: str, h: int, pad_right: bool = False) -> None:
            """Single CanvasSlot card, fixed height, fill available width."""
            card = ctk.CTkFrame(parent, fg_color=CARD, corner_radius=6, height=h)
            px = (0, 4) if pad_right else (0, 0)
            card.pack(side="left", fill="both", expand=True, padx=px)
            card.pack_propagate(False)
            inner = tk.Frame(card, bg=PLOT["bg"], bd=0, highlightthickness=0)
            inner.pack(fill="both", expand=True, padx=SPACE_XS, pady=SPACE_XS)
            self._slots[key] = CanvasSlot(inner, 10, h / 100, toolbar=False)

        def _row(parent, specs: "list[tuple[str,int,int]]") -> None:
            """specs = [(key, height_px, weight), ...]  — one horizontal row."""
            row_frame = ctk.CTkFrame(parent, fg_color="transparent")
            row_frame.pack(fill="x", pady=(0, SPACE_S))
            for i, (key, h, w) in enumerate(specs):
                pad_right = i < len(specs) - 1
                card = ctk.CTkFrame(row_frame, fg_color=CARD, corner_radius=6, height=h)
                px = (0, 4) if pad_right else (0, 0)
                card.pack(side="left", fill="both", expand=True, padx=px)
                card.pack_propagate(False)
                inner = tk.Frame(card, bg=PLOT["bg"], bd=0, highlightthickness=0)
                inner.pack(fill="both", expand=True, padx=SPACE_XS, pady=SPACE_XS)
                self._slots[key] = CanvasSlot(inner, 10 * w, h / 100, toolbar=False)

        # ━━━ SECTION 1 — Rythme cardiaque ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec1 = _section(outer_scroll, "❤  Heart Rate & Tachogram", ORANGE_DARK,
                        subtitle="RR tachogram + instantaneous HR")
        _row(sec1, [("sum_rr", 260, 1)])
        _row(sec1, [("sum_rr_hist", 220, 3), ("sum_rr_extra", 220, 2)])

        # ━━━ SECTION 2 — VFC domaine temps & fréquence ━━━━━━━━━━━━━━━━━━━━━━
        sec2 = _section(outer_scroll, "〰  HRV — Time & Frequency", BLUE_DARK,
                        subtitle="PSD · radar · Poincaré")
        _row(sec2, [("sum_psd", 280, 5), ("sum_radar", 280, 3)])
        _row(sec2, [("sum_poincare", 260, 3), ("sum_asymmetry", 260, 4)])   # NEW

        # ━━━ SECTION 3 — Morphologie des battements ━━━━━━━━━━━━━━━━━━━━━━━━━
        sec3 = _section(outer_scroll, "📊  Beat Morphology", PURPLE,
                        subtitle="mean template ± 1 SD · distributions")
        _row(sec3, [("sum_beat", 300, 1)])
        _row(sec3, [("sum_beat_dist", 220, 3), ("sum_quality_time", 220, 2)])  # NEW

        # ━━━ SECTION 4 — Intervalles ECG ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec4 = _section(outer_scroll, "📏  ECG Intervals", PINK,
                        subtitle="PR · QRS · QT · QTc  (delineation required)")
        _row(sec4, [("sum_intervals", 240, 3), ("sum_qt_disp", 240, 2)])     # NEW
        _row(sec4, [("sum_intervals_ecg", 300, 1)])

        # ━━━ SECTION 5 — VFC glissante ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec5 = _section(outer_scroll, "📈  Rolling HRV over time", "#00695C",
                        subtitle="SDNN · RMSSD · HR per rolling windows")
        _row(sec5, [("sum_rolling", 260, 1)])

        # ━━━ SECTION 6 — Rapport texte ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec6 = _section(outer_scroll, "📝  Detailed Report", MUTED)
        txt_card = ctk.CTkFrame(sec6, fg_color=CARD, corner_radius=6)
        txt_card.pack(fill="x", pady=(0, SPACE_L))
        self.txt_sum = ctk.CTkTextbox(txt_card, font=FONT_MONO, fg_color=CARD,
                                      text_color=TEXT, border_width=0, height=380)
        self.txt_sum.pack(fill="both", expand=True, padx=SPACE_M, pady=SPACE_M)

    # ════════════════════════════════════════════════════════
    #  SIGNAL PIPELINE  (detection engine)
    # ════════════════════════════════════════════════════════

    def _snapshot_params(self) -> dict:
        """Read every widget value on the main thread and return a plain dict snapshot.

        Delegates to ``FilterParams.from_widgets`` so that the field list is
        maintained in exactly one place (the FilterParams dataclass).  Any new
        parameter added to FilterParams is automatically picked up here and in
        _collect_session_state / _restore_session_worker.
        """
        params = FilterParams.from_widgets(self).to_dict()
        # Analysis window is independent from FilterParams (not a detection param)
        params["analysis_t_start"] = self._analysis_t_start
        params["analysis_t_end"]   = self._analysis_t_end
        return params

    @staticmethod
    def _compute_preview_bundle(
        sig_raw:     np.ndarray,
        fs:          int,
        params:      dict,
        progress_cb: "Optional[Callable[[int, str], None]]" = None,
    ) -> dict:
        """Pure computation — no access to *self*, no Tkinter calls.

        Filters, normalises, fixes polarity, and finds all R-peak candidates.
        Returns a plain dict that ``_on_preview_done`` writes to ``self.*``
        atomically on the main thread.

        This separation is the thread-safety contract:
        * Background workers call this method and return the bundle.
        * Only the main thread writes instance variables.
        """
        def _prog(pct: int, msg: str) -> None:
            if progress_cb:
                progress_cb(pct, msg)

        no_filter = params.get("no_filter", False)
        sig = sig_raw.copy()

        _prog(5, "Normalising raw signal…")
        signal_raw_norm = normalize(sig)

        if no_filter:
            log.info("_compute_preview_bundle: no_filter=True — skipping bandpass/notch/ecg_clean")
        else:
            _prog(10, "Bandpass filtering…")
            try:
                sig = bandpass(sig, fs, params["lp"], params["hp"])
            except Exception as exc:
                log.warning("bandpass skipped: %s", exc)

            if params["notch"]:
                _prog(25, "Notch filtering…")
                try:
                    sig = notch(sig, fs)
                except Exception as exc:
                    log.warning("notch skipped: %s — check fs vs notch frequency", exc)

            _prog(35, "NeuroKit2 clean…")
            try:
                assert nk is not None  # NK_AVAILABLE guard checked by caller
                sig = nk.ecg_clean(sig, sampling_rate=fs,
                                   method=params["clean_method"])
            except Exception as exc:
                log.warning("nk.ecg_clean skipped: %s", exc)

        _prog(45, "Normalising filtered signal…")
        sig = normalize(np.asarray(sig, dtype=float))

        # ── Manual polarity override ─────────────────────────────────────
        # If the user has explicitly toggled "Inverser signal", flip the
        # signal before auto-polarity detection.  fix_polarity will then
        # find the (now correct) positive R peaks and not re-flip.
        if params.get("invert_signal", False):
            sig = -sig
            log.info("_compute_preview_bundle: user-requested signal inversion applied")

        def _polarity_prog(pct: int, msg: str) -> None:
            _prog(45 + int(pct * 0.50), msg)

        # ── Detection method ─────────────────────────────────────────────
        det_method = params.get("detection_method", "auto")

        if "wavelet" in det_method.lower() or "cwt" in det_method.lower():
            # ── Wavelet (CWT) pipeline ────────────────────────────────────
            # fix_polarity AVANT la détection : le détecteur wavelet (comme SG)
            # travaille sur la dérivée positive → il faut que les R-peaks soient
            # positifs dans le signal, sinon seuls les artefacts positifs sont détectés.
            _prog(48, "Polarity correction…")
            sig, inverted, _, _ = fix_polarity(sig, fs, params["min_rr_ms"])
            _prog(50, "CWT — séparation bruit / QRS / J-wave…")
            try:
                peaks_wt, proms_wt, t_amp_wt = detect_peaks_wavelet(
                    sig,
                    fs=fs,
                    min_rr_ms=params["min_rr_ms"],
                    peak_distance_ms=params.get("peak_distance_ms", MouseECG.PEAK_DISTANCE_MS),
                )
            except ImportError:
                log.warning(
                    "PyWavelets non installé — "
                    "pip install PyWavelets  (fallback: auto)"
                )
                peaks_wt  = np.array([], dtype=int)
                proms_wt  = np.array([])
                t_amp_wt  = 0.0
            except Exception as exc:
                log.warning("Wavelet detection failed, falling back to auto: %s", exc)
                peaks_wt  = np.array([], dtype=int)
                proms_wt  = np.array([])
                t_amp_wt  = 0.0

            _prog(95, f"Wavelet candidates: {len(peaks_wt):,}")
            return {
                "signal_raw_norm": signal_raw_norm,
                "signal_flt":      sig,
                "no_filter_mode":  no_filter,
                "all_cands":       peaks_wt,
                "all_proms":       proms_wt if len(proms_wt) else np.ones(len(peaks_wt)),
                "inverted":        inverted,
            }

        if "sg" in det_method.lower() or "deriv" in det_method.lower():
            # ── SG + Derivative pipeline ──────────────────────────────────
            # fix_polarity automatique si l'utilisateur n'a pas déjà inversé
            # manuellement (évite la double inversion). Le détecteur SG
            # requiert des R positifs (upstroke = dérivée positive).
            if not params.get("invert_signal", False):
                _prog(48, "Polarity correction…")
                sig, inverted, _, _ = fix_polarity(sig, fs, params["min_rr_ms"])
            else:
                inverted = False  # utilisateur a géré la polarité manuellement
            _prog(50, "SG derivative detection…")
            try:
                sg_window_ms = float(params.get("sg_window_ms", 20.0))
            except (TypeError, ValueError):
                sg_window_ms = 20.0

            try:
                peaks_sg, proms_sg, t_amp_sg = detect_peaks_sg_derivative(
                    sig,
                    fs=fs,
                    sg_window_ms=sg_window_ms,
                    min_rr_ms=params["min_rr_ms"],
                    peak_distance_ms=params.get("peak_distance_ms", MouseECG.PEAK_DISTANCE_MS),
                    target_fs=float(params.get("sg_target_fs", 10000)),
                )
            except Exception as exc:
                log.warning("SG+derivative detection failed, falling back to auto: %s", exc)
                peaks_sg = np.array([], dtype=int)
                proms_sg = np.array([])
                t_amp_sg = 0.0

            _prog(95, f"SG+Deriv candidates: {len(peaks_sg):,}")
            return {
                "signal_raw_norm": signal_raw_norm,
                "signal_flt":      sig,
                "no_filter_mode":  no_filter,
                "all_cands":       peaks_sg,
                "all_proms":       proms_sg if len(proms_sg) else np.ones(len(peaks_sg)),
                "inverted":        False,
            }

        if "envelope" in det_method.lower() or "max" in det_method.lower():
            # ── Envelope Max pipeline ─────────────────────────────────────
            # Détection par maximum local d'amplitude — robuste aux signaux
            # saturés (ADC clipping) et aux morphologies atypiques où la
            # dérivée SG est peu discriminante.
            # fix_polarity requis : le détecteur sélectionne les maxima → les
            # R-peaks doivent être des extrema positifs dans le signal.
            if not params.get("invert_signal", False):
                _prog(48, "Polarity correction…")
                sig, inverted, _, _ = fix_polarity(sig, fs, params["min_rr_ms"])
            else:
                inverted = False
            _prog(55, "Envelope Max detection…")
            try:
                peaks_em, proms_em, t_amp_em = detect_peaks_envelope_max(
                    sig,
                    fs=fs,
                    min_rr_ms=params["min_rr_ms"],
                    peak_distance_ms=params.get("peak_distance_ms", MouseECG.PEAK_DISTANCE_MS),
                )
            except Exception as exc:
                log.warning("Envelope Max detection failed, falling back to auto: %s", exc)
                peaks_em = np.array([], dtype=int)
                proms_em = np.array([])
                t_amp_em = 0.0

            _prog(95, f"Envelope Max candidates: {len(peaks_em):,}")
            return {
                "signal_raw_norm": signal_raw_norm,
                "signal_flt":      sig,
                "no_filter_mode":  no_filter,
                "all_cands":       peaks_em,
                "all_proms":       proms_em if len(proms_em) else np.ones(len(peaks_em)),
                "inverted":        inverted,
            }
            
        # ── Auto (NeuroKit2) pipeline — original path ─────────────────────
        sig_out, inverted, cands, proms = fix_polarity(
            sig, fs, params["min_rr_ms"], progress_cb=_polarity_prog)

        _prog(95, f"Candidates found: {len(cands):,}")

        return {
            "signal_raw_norm": signal_raw_norm,
            "signal_flt":      sig_out,
            "no_filter_mode":  no_filter,
            "all_cands":       cands,
            "all_proms":       proms,
            "inverted":        inverted,
        }

    def _prepare_signal(
        self,
        params: dict,
        progress_cb: "Optional[Callable[[int, str], None]]" = None,
    ) -> None:
        """Filter, normalise, and fix polarity. Writes results to self.

        Thin wrapper around the pure ``_compute_preview_bundle`` for callers
        that already have signal data on self and are running on the main
        thread. Not currently called anywhere in the app — session restore
        now calls ``_compute_preview_bundle`` directly from
        ``_restore_session_worker`` (a background thread) and writes the
        result to self from ``_on_restore_session_done`` (main thread) —
        kept for any future main-thread caller that wants the write-to-self
        convenience.

        The background preview path uses ``_compute_preview_bundle`` directly
        and writes to self only from ``_on_preview_done`` (main thread).
        """
        if self._signal_raw is None:
            return
        bundle = self._compute_preview_bundle(
            self._signal_raw, self._fs, params, progress_cb)
        self._signal_raw_norm  = bundle["signal_raw_norm"]
        self._signal_flt       = bundle["signal_flt"]
        self._no_filter_mode   = bundle["no_filter_mode"]
        self._all_cands        = bundle["all_cands"]
        self._all_proms        = bundle["all_proms"]
        self._signal_inverted  = bundle.get("inverted", False)

    def _run_detection(self, thresh: float | None = None) -> int:
        """Apply current threshold to pre-computed candidates.

        Parameters
        ----------
        thresh : float | None
            If supplied, use this value directly (safe from a background
            thread).  If None, read ``self.sl_thr`` — only safe to call
            from the main thread.

        Returns the number of accepted peaks.  Fast — no signal processing.
        All Tkinter widget *writes* are marshalled through ``after(0, …)``
        so this method is safe to call from either thread.
        """
        if self._signal_flt is None or self._all_cands is None or self._all_proms is None:
            return 0

        if thresh is None:
            thresh = float(self.sl_thr.get())   # main-thread-only path  # type: ignore[union-attr]

        accepted, rejected, thresh_amp = apply_threshold(
            self._signal_flt, self._all_cands, self._all_proms, thresh,
            fs=self._fs)

        # ── Apply manual exclusions ──────────────────────────────
        if self._manual_excluded:
            manual_excl_mask = np.array([p in self._manual_excluded for p in accepted], dtype=bool)
            self._rpeaks_manual_excl = accepted[manual_excl_mask] if manual_excl_mask.any() else np.array([], int)
            accepted = accepted[~manual_excl_mask]
        else:
            self._rpeaks_manual_excl = np.array([], dtype=int)

        # ── Merge manually-added peaks ───────────────────────────
        # Added peaks are never in the candidate set; they bypass all thresholds.
        # They are removed from the exclusion set if present (can't be both).
        if self._manual_added:
            added_arr = np.array(sorted(self._manual_added), dtype=int)
            self._rpeaks_manual_added = added_arr
            # Remove from exclusion set if mistakenly present
            self._manual_excluded -= self._manual_added
            # Merge and sort
            accepted = np.unique(np.concatenate([accepted, added_arr]))
        else:
            self._rpeaks_manual_added = np.array([], dtype=int)

        self._rpeaks_ok  = accepted
        self._rpeaks_rej = rejected
        self._thresh_amp = thresh_amp
        n = len(accepted)

        # Widget writes: always on main thread via after(0, ...)
        color = GREEN if n > 10 else RED
        self.after(0, lambda _n=n, _c=color: self.lbl_npeaks.configure(  # type: ignore[union-attr]
            text=f"Peaks detected: {_n}", text_color=_c))
        # Enable the artifact review button as soon as peaks are available
        self.after(0, lambda: self.btn_review_art.configure(  # type: ignore[union-attr]
            state="normal" if n > 4 else "disabled"))
        self._update_signal_quality(accepted)
        return n

    def _update_signal_quality(self, accepted: np.ndarray) -> None:
        """Compute a 0–100 quality score and update the KPI label.

        Quality is based on:
        1. Beat morphology (primary): mean beat-to-template correlation from
           the last analysis run.  High correlation = clean, consistent QRS.
           Falls back to RR regularity if no template analysis has been run yet.
        2. Detection completeness (secondary): ratio of detected to expected
           beats, clipped to [0.5, 1.5] so it modulates but never dominates.

        The previous formula (1 - rr_cv) was unreliable because a healthy mouse
        at high HR during stress has a low rr_cv not from noise but from genuine
        sympathetic activation.
        """
        n = len(accepted)
        if n <= 5 or self._time is None:
            return
        dur         = self._time[-1]
        expected_n  = dur / 60 * MouseECG.HR_REST_BPM
        ratio       = float(np.clip(n / max(expected_n, 1), 0.5, 1.5))

        # Primary quality signal: mean beat-to-template correlation
        # beat_corr is computed in analyse_core and stored in _results.
        beat_corr = None
        if self._results is not None:
            beat_corr = self._results.get("beat_corr")

        if beat_corr is not None and len(beat_corr) > 0:
            morpho = float(np.nanmean(beat_corr))            # 0–1 (Pearson r)
            morpho = float(np.clip(morpho, 0.0, 1.0))
            quality = int(np.clip(100 * morpho * ratio, 0, 100))
        else:
            # Fallback before full analysis: RR regularity heuristic
            rr_tmp  = np.diff(accepted) / self._fs * 1000
            rr_cv   = rr_tmp.std() / (rr_tmp.mean() + 1e-6)
            quality = int(np.clip(100 * (1 - rr_cv) * ratio, 0, 100))

        self._sig_quality = quality
        color = GREEN if quality >= 70 else (ORANGE if quality >= 40 else RED)
        self.after(0, lambda q=quality, c=color:
                   self.lbl_quality.configure(text=f"Signal quality: {q}%", text_color=c))

    def _on_det_method_change(self, choice: str) -> None:
        """Show/hide SG options frame based on selected detection method."""
        if self._sg_frame is None:
            return
        if "SG" in choice or "Derivative" in choice:
            self._sg_frame.pack(fill="x")
        else:
            self._sg_frame.pack_forget()

    # ── No-filter master toggle ──────────────────────────────

    def _on_no_filter_toggle(self) -> None:
        """Enable/disable the no-filter mode and dim the filter controls.

        Utilise une liste plate pré-calculée (_filter_control_widgets) au lieu
        d'une traversée récursive winfo_children() — évite de configurer 50+
        widgets imbriqués à chaque toggle.
        """
        no_filter = bool(self.sw_no_filter.get())
        new_state = "disabled" if no_filter else "normal"

        # Chemin rapide : liste plate construite lors du démarrage
        if self._filter_control_widgets:
            for widget in self._filter_control_widgets:
                try:
                    widget.configure(state=new_state)
                except Exception as e:
                    log.debug("widget.configure(state) failed: %s", e)
        else:
            # Fallback : traversée récursive si la liste n'est pas encore prête
            def _set_state_recursive(frame) -> None:
                for widget in frame.winfo_children():
                    try:
                        widget.configure(state=new_state)
                    except Exception as e:
                        log.debug("widget.configure(state) failed: %s", e)
                    try:
                        _set_state_recursive(widget)
                    except Exception as e:
                        log.debug("_set_state_recursive failed: %s", e)
            _set_state_recursive(self._filter_widgets_frame)
            if self._adv_filters_frame is not None:
                _set_state_recursive(self._adv_filters_frame)

    # ── Raw / Filtered toggle ─────────────────────────────────

    def _on_show_raw_toggle(self) -> None:
        """Switch the overview and detail plots between raw and filtered signals.

        The raw signal is normalised (zero-mean, unit-variance) to match
        the amplitude scale of the filtered signal so that peak markers
        remain visually coherent regardless of which view is active.
        No re-processing is needed — both arrays are pre-computed.
        """
        self._show_raw = bool(self.sw_show_raw.get())
        if self._signal_flt is not None:
            # Invalidate the relevant downsampled cache so it is rebuilt next draw.
            # _ds_time is rebuilt alongside _ds_sig (via _downsample_pair),
            # so only reset it when the primary filtered signal cache is dropped.
            if self._show_raw:
                self._ds_raw_sig = None
            else:
                self._ds_sig  = None
                self._ds_time = None
            self._draw_detail(self._nav_pos)

    # ── Overview click-to-navigate ────────────────────────────

    def _on_overview_click(self, event) -> None:
        """Stub — overview removed."""
        self.nav_ctrl.on_overview_click(event)

    def _on_overview_scroll(self, event) -> None:
        """Stub — overview removed."""
        self.nav_ctrl.on_overview_scroll(event)

    # ── Detail scroll-wheel zoom ──────────────────────────────

    def _on_detail_scroll(self, event) -> None:
        """Zoom the detail view's x-axis in/out centred on the cursor position.

        Each scroll tick zooms by a factor of 1.25 (in) or 0.8 (out).
        After zooming, ``_nav_pos`` and the window-entry widget are updated
        to reflect the new visible range.

        The y-axis is intentionally unchanged — vertical zoom is handled by
        the matplotlib toolbar's Zoom-to-rectangle tool.
        """
        self.nav_ctrl.on_detail_scroll(event)

    # ── Undo / Redo for manual peak edits ────────────────────

    def _push_edit_undo(self) -> None:
        """Snapshot state before a destructive edit action."""
        snap = (frozenset(self._manual_excluded), frozenset(self._manual_added))
        self._edit_undo.append(snap)
        if len(self._edit_undo) > self._EDIT_UNDO_LIMIT:
            self._edit_undo.pop(0)
        self._edit_redo.clear()
        self._update_undo_btns()

    def _undo_edit(self, _event=None) -> None:
        """Ctrl+Z — restore previous peak-edit state."""
        if not self._edit_undo:
            self._set_status("Nothing to undo.", MUTED)
            return
        cur = (frozenset(self._manual_excluded), frozenset(self._manual_added))
        self._edit_redo.append(cur)
        excl, added = self._edit_undo.pop()
        self._manual_excluded = set(excl)
        self._manual_added    = set(added)
        self._apply_edit_state()
        n = len(self._edit_undo)
        self._set_status(f"Undone  ({str(n) + ' left' if n else 'none'})  — Ctrl+Y to redo", ORANGE)
        self._update_undo_btns()

    def _redo_edit(self, _event=None) -> None:
        """Ctrl+Y — rétablir après undo."""
        if not self._edit_redo:
            self._set_status("Nothing to redo.", MUTED)
            return
        cur = (frozenset(self._manual_excluded), frozenset(self._manual_added))
        self._edit_undo.append(cur)
        excl, added = self._edit_redo.pop()
        self._manual_excluded = set(excl)
        self._manual_added    = set(added)
        self._apply_edit_state()
        self._set_status(f"Redone  ({len(self._edit_redo)} remaining)", ORANGE)
        self._update_undo_btns()

    def _apply_edit_state(self) -> None:
        if self._signal_flt is not None and self._all_cands is not None:
            self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
            self._draw_detail(self._nav_pos)
            # Also refresh the arrhythmia ECG viewer if an event is selected
            if self._arr_selected_idx >= 0 and self._arrhythmia_events:
                self._draw_arr_detail()

    # ── Manual peak exclusion ─────────────────────────────────

    def _toggle_edit_mode(self) -> None:
        """Toggle the click-to-exclude edit mode on/off."""
        self._edit_mode = not self._edit_mode
        if self._edit_mode:
            self.btn_edit_mode.configure(
                fg_color=ORANGE, hover_color=ORANGE_DEEP,
                text_color="white", text="Edit Mode ON",
            )
            self.lbl_edit_hint.pack(side="left", padx=(SPACE_S, 0))
        else:
            self.btn_edit_mode.configure(
                fg_color=BORDER, hover_color=BORDER2,
                text_color=MUTED, text="Edit Peaks",
            )
            self.lbl_edit_hint.pack_forget()
            # Also turn off free placement when leaving edit mode
            if self._edit_free_placement:
                self._edit_free_placement = False
                try:
                    self.btn_free_placement.configure(
                        fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                        text="Free Placement",
                    )
                except Exception:
                    pass
        if self._signal_flt is not None:
            self._draw_detail(self._nav_pos)

    def _toggle_free_placement(self) -> None:
        """Toggle free-placement mode: bypass proximity constraint when adding peaks.

        When active, right-clicking adds a peak at the local max *regardless* of
        how close it is to an existing peak.  This is useful for very high-rate
        signals or for correcting closely-spaced double-peaks.

        Note: edit mode must be active for this to have any effect.
        """
        self._edit_free_placement = not self._edit_free_placement
        if self._edit_free_placement:
            self.btn_free_placement.configure(
                fg_color=BLUE, hover_color=BLUE_HOVER,
                text_color="white", text="Free Placement ON",
            )
            self._set_status(
                "Free Placement ON — R-click places a peak at the exact click position, "
                "no snapping, no proximity guard.", BLUE)
        else:
            self.btn_free_placement.configure(
                fg_color=BORDER, hover_color=BORDER2,
                text_color=MUTED, text="Free Placement",
            )
            self._set_status("Free Placement OFF — normal proximity guard restored.", MUTED)

    def _clear_manual_exclusions(self) -> None:
        """Re-include all manually excluded peaks, remove all manually added peaks."""
        if not self._manual_excluded and not self._manual_added:
            return
        self._push_edit_undo()
        n_excl  = len(self._manual_excluded)
        n_added = len(self._manual_added)
        self._manual_excluded.clear()
        self._manual_added.clear()
        # Invalidate any previous analysis — peaks have changed
        self._results  = None
        self._epoch_df = None
        self._rpeaks_manual_excl  = np.array([], dtype=int)
        self._rpeaks_manual_added = np.array([], dtype=int)
        if self._signal_flt is not None and self._all_cands is not None:
            self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
            self._draw_detail(self._nav_pos)
        self.lbl_npeaks.configure(  # type: ignore[union-attr]
            text=f"Peaks detected: {len(self._rpeaks_ok) if self._rpeaks_ok is not None else 0}",
            text_color=GREEN,
        )
        self._set_status(
            f"Manual edits cleared — {n_excl} exclusion(s), {n_added} addition(s) removed.",
            GREEN)

    def _on_detail_motion(self, event) -> None:
        """Track mouse position in edit mode and compute the preview peak position.

        In normal mode: snaps to the local maximum within ±tol_samp and shows
        an orange/red marker if it would land too close to an existing peak.
        In free placement mode: the preview follows the cursor exactly (no
        snapping) and is always shown in the "ok" colour.

        Redraws are throttled to 30 ms (≈33 fps) via after().
        """
        if not self._edit_mode:
            if self._hover_samp is not None:
                self._hover_samp = None
                self._hover_samp_near = False
                if self._hover_after_id is not None:
                    self.after_cancel(self._hover_after_id)
                    self._hover_after_id = None
                self._draw_detail(self._nav_pos)
            return

        if event.xdata is None or self._signal_flt is None or self._fs is None:
            if self._hover_samp is not None:
                self._hover_samp = None
                self._hover_samp_near = False
                self._draw_detail(self._nav_pos)
            return

        fs         = self._fs
        click_time = float(event.xdata)
        click_samp = int(np.clip(round(click_time * fs), 0, len(self._signal_flt) - 1))

        # Free placement: follow the cursor exactly, no snapping, always "ok"
        if self._edit_free_placement:
            new_samp = click_samp
            near     = False
        else:
            try:
                win = float(self.ent_window.get())
            except Exception:
                win = 2.0
            tol_samp = int(max(MouseECG.MIN_RR_MS / 1000 / 2, win * 0.03) * fs)

            # Snap to local maximum within ±tol_samp
            sig = self._signal_flt
            lo  = max(0, click_samp - tol_samp)
            hi  = min(len(sig), click_samp + tol_samp + 1)
            new_samp = lo + int(np.argmax(sig[lo:hi]))

            # Show warning colour if this would land too close to an existing peak
            near = False
            if self._rpeaks_ok is not None and len(self._rpeaks_ok):
                min_sep = int(MouseECG.MIN_RR_MS / 1000 * fs * 0.5)
                near    = int(np.min(np.abs(self._rpeaks_ok - new_samp))) < min_sep

        # Only schedule a redraw if the position actually changed
        if self._hover_samp == new_samp and self._hover_samp_near == near:
            return

        self._hover_samp      = new_samp
        self._hover_samp_near = near

        # Throttle: cancel any pending redraw, schedule a new one in 30 ms
        if self._hover_after_id is not None:
            self.after_cancel(self._hover_after_id)
        self._hover_after_id = self.after(30, self._flush_hover_redraw)

    def _flush_hover_redraw(self) -> None:
        """Execute the throttled hover redraw on the main thread."""
        self._hover_after_id = None
        self._draw_detail(self._nav_pos)

    def _on_detail_click(self, event) -> None:
        """Edit-mode click handler for the detail view.

        Left-click  (button 1) near an existing peak → toggle exclusion
        Right-click (button 3) anywhere              → add peak at local max,
                                                       or remove if clicking
                                                       a manually-added peak

        Only active when ``_edit_mode`` is True.
        """
        if not self._edit_mode:
            return
        if event.xdata is None or self._signal_flt is None:
            return
        if self._rpeaks_ok is None:
            return

        fs         = self._fs
        click_time = float(event.xdata)      # seconds
        click_samp = int(round(click_time * fs))
        click_samp = int(np.clip(click_samp, 0, len(self._signal_flt) - 1))

        # ── Tolerance ────────────────────────────────────────────────────────
        try:
            win = float(self.ent_window.get())
        except Exception:
            win = 2.0
        # Half the minimum RR interval (in seconds) is a natural click radius
        tol_s    = max(MouseECG.MIN_RR_MS / 1000 / 2, win * 0.03)
        tol_samp = int(tol_s * fs)

        is_left  = (event.button == 1)
        is_right = (event.button == 3)

        # ──────────────────────────────────────────────────────────────────────
        #  RIGHT-CLICK: add a new peak, or remove an existing manually-added one
        # ──────────────────────────────────────────────────────────────────────
        if is_right:
            # ── FREE PLACEMENT: always add at the exact clicked sample ────────
            # All guards (proximity, remove-nearby, local-max snapping) are
            # bypassed.  The peak lands precisely where the user clicked.
            if self._edit_free_placement:
                new_samp = click_samp
                self._push_edit_undo()
                self._manual_added.add(new_samp)
                self._manual_excluded.discard(new_samp)
                self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
                n_ok    = len(self._rpeaks_ok) if self._rpeaks_ok is not None else 0
                n_added = len(self._manual_added)
                self._set_status(
                    f"[Free] Added peak at {new_samp / fs:.3f} s  |  "
                    f"Total added: {n_added}  |  Accepted: {n_ok}  "
                    "— re-run Full Analysis to update HRV.",
                    BLUE,
                )
                self._draw_detail(self._nav_pos)
                return

            # ── NORMAL MODE ───────────────────────────────────────────────────
            # First check if click is near a manually-added peak → remove it
            if self._rpeaks_manual_added is not None and len(self._rpeaks_manual_added):
                dists = np.abs(self._rpeaks_manual_added - click_samp)
                nearest_i = int(np.argmin(dists))
                if dists[nearest_i] <= tol_samp:
                    peak_to_remove = int(self._rpeaks_manual_added[nearest_i])
                    self._push_edit_undo()
                    self._manual_added.discard(peak_to_remove)
                    self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
                    n_ok = len(self._rpeaks_ok) if self._rpeaks_ok is not None else 0
                    self._set_status(
                        f"Removed manually added peak at {click_time:.3f} s  |  "
                        f"Accepted: {n_ok}  — re-run Full Analysis to update HRV.",
                        ORANGE,
                    )
                    self._draw_detail(self._nav_pos)
                    return

            # Snap to local maximum within ±tol_samp of click
            sig      = self._signal_flt
            lo       = max(0, click_samp - tol_samp)
            hi       = min(len(sig), click_samp + tol_samp + 1)
            seg      = sig[lo:hi]
            local_max_offset = int(np.argmax(seg))
            new_samp = lo + local_max_offset

            # If too close to existing accepted peak → replace it
            # (exclude the old one, add the new one) instead of refusing.
            if self._rpeaks_ok is not None and len(self._rpeaks_ok):
                min_sep_samp = int(MouseECG.MIN_RR_MS / 1000 * fs * 0.5)
                dists_ok     = np.abs(self._rpeaks_ok - new_samp)
                nearest_idx  = int(np.argmin(dists_ok))
                nearest_dist = int(dists_ok[nearest_idx])
                if nearest_dist < min_sep_samp:
                    # Replace: exclude the nearby peak, add the new one
                    old_peak = int(self._rpeaks_ok[nearest_idx])
                    self._push_edit_undo()
                    self._manual_excluded.add(old_peak)
                    self._manual_added.discard(old_peak)
                    self._manual_added.add(new_samp)
                    self._manual_excluded.discard(new_samp)
                    self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
                    n_ok = len(self._rpeaks_ok) if self._rpeaks_ok is not None else 0
                    self._set_status(
                        f"Replaced peak {old_peak/fs:.3f} s → {new_samp/fs:.3f} s  "
                        f"({nearest_dist/fs*1000:.1f} ms apart)  |  Accepted: {n_ok}  "
                        "— re-run Full Analysis to update HRV.",
                        ORANGE,
                    )
                    self._draw_detail(self._nav_pos)
                    return

            self._push_edit_undo()
            self._manual_added.add(new_samp)
            # If this sample was previously excluded, unexclude it
            self._manual_excluded.discard(new_samp)
            self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
            n_ok    = len(self._rpeaks_ok) if self._rpeaks_ok is not None else 0
            n_added = len(self._manual_added)
            self._set_status(
                f"Added peak at {new_samp / fs:.3f} s (snapped to local max)  |  "
                f"Total added: {n_added}  |  Accepted: {n_ok}  "
                "— re-run Full Analysis to update HRV.",
                ORANGE,
            )
            self._draw_detail(self._nav_pos)
            return

        # ──────────────────────────────────────────────────────────────────────
        #  LEFT-CLICK: toggle exclusion of the nearest existing peak
        # ──────────────────────────────────────────────────────────────────────
        if not is_left:
            return

        # Pool: all accepted peaks + all currently excluded peaks
        # (manually added peaks are excluded from toggle — use right-click to remove)
        added_set  = self._manual_added
        base_ok    = np.array([p for p in self._rpeaks_ok if p not in added_set], int) \
                     if self._rpeaks_ok is not None else np.array([], int)
        excl_arr   = self._rpeaks_manual_excl if self._rpeaks_manual_excl is not None \
                     else np.array([], int)
        candidates = np.concatenate([base_ok, excl_arr])
        if len(candidates) == 0:
            return

        times_s   = candidates / fs
        distances = np.abs(times_s - click_time)
        nearest_i = int(np.argmin(distances))

        if distances[nearest_i] > tol_s:
            return   # click not close enough to any peak

        peak_idx = int(candidates[nearest_i])

        self._push_edit_undo()
        if peak_idx in self._manual_excluded:
            self._manual_excluded.discard(peak_idx)
        else:
            self._manual_excluded.add(peak_idx)

        self._run_detection(float(self.sl_thr.get()))  # type: ignore[union-attr]
        n_ok   = len(self._rpeaks_ok)  if self._rpeaks_ok   is not None else 0
        n_excl = len(self._manual_excluded)
        self._set_status(
            f"Manual exclusions: {n_excl}  |  Accepted peaks: {n_ok}  "
            "— re-run Full Analysis to update HRV.",
            ORANGE,
        )
        self._draw_detail(self._nav_pos)

    # ── Threshold slider / entry callbacks ────────────────────
    def _on_threshold_slide(self, value: float) -> None:
        """Called continuously while the slider is being dragged.

        Widget label and entry are updated immediately for visual feedback.
        Detection and redraws are debounced (80 ms) so rapid drag events do
        not flood the rendering pipeline — especially important for long
        recordings where apply_threshold() + two canvas draws take ~50 ms.
        """
        self.lbl_thr.configure(text=f"Sensitivity:  {value:.3f}")
        self.ent_thr.delete(0, "end")  # type: ignore[union-attr]
        self.ent_thr.insert(0, f"{value:.3f}")  # type: ignore[union-attr]

        if self._signal_flt is None or self._all_cands is None:
            return

        # Cancel any previously scheduled update and reschedule
        if self._thr_debounce_id is not None:
            self.after_cancel(self._thr_debounce_id)
        self._thr_debounce_id = self.after(80, lambda v=value: self._apply_threshold_ui(v))

    def _apply_threshold_ui(self, value: float) -> None:
        """Run detection and refresh plots — called after debounce delay.

        Always executes on the main thread (scheduled via after()), so it is
        safe to read the slider and write widgets directly.
        """
        self._thr_debounce_id = None
        # Peaks are about to change — invalidate stale results
        if self._results is not None:
            self._results  = None
            self._epoch_df = None
            self._reset_kpis()
        self._run_detection(value)
        self._draw_detail(self._nav_pos)

    def _on_threshold_entry(self, event=None) -> None:
        """Called when the user types a value in the exact-threshold entry.

        Applies immediately — no debounce — since this is a deliberate commit.
        """
        try:
            value = max(0.01, min(2.0, float(self.ent_thr.get())))  # type: ignore[union-attr]
            self.sl_thr.set(value)  # type: ignore[union-attr]
            self.lbl_thr.configure(text=f"Sensitivity:  {value:.3f}")
            if self._signal_flt is not None and self._all_cands is not None:
                self._apply_threshold_ui(value)
        except ValueError:
            pass

    # ════════════════════════════════════════════════════════
    #  ACTIONS  (file, preview, run)
    # ════════════════════════════════════════════════════════

    def _open_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Open Spike2 .mat",
            filetypes=[("MATLAB files", "*.mat"), ("All", "*.*")],
        )
        if path:
            self._load_path(path)

    def _show_channels(self) -> None:
        if not self._filepath:
            messagebox.showwarning("No file", "Open a .mat file first.")
            return
        try:
            messagebox.showinfo("Channels", list_channels(self._filepath))
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def _load_raw_only(self) -> None:
        """Load the file and display ONLY the raw signal.

        No bandpass/notch/NK-clean, no polarity correction, no detection —
        just the samples read from disk, windowed by the Analysis window
        fields if set, and z-score normalised purely for display scale.

        This is what runs automatically when a file is opened.  The user
        must click '1 ▶ Preview Detection' to run the actual DSP + detection
        pipeline (see _preview / _preview_worker).
        """
        if not self._filepath:
            return
        params = self._snapshot_params()
        self._start_async(
            self.btn_preview, "Loading…", "Loading raw signal…",
            lambda: self._load_raw_worker(params),
            self._on_raw_load_done,
            pass_result=True,
        )

    def _load_raw_worker(self, params: dict) -> dict:
        """Background worker — loads the file, returns the RAW signal only.

        Deliberately mirrors only the first few steps of _preview_worker
        (file read + time-window crop). It stops before any filtering,
        polarity correction, or detection call.
        """
        if self._filepath is None:
            raise ValueError("No file loaded.")

        def _prog(pct: int, msg: str) -> None:
            self.after(0, lambda p=pct, m=msg: self._set_progress(p, m))

        _prog(10, "Loading signal from file…")
        sig, detected_ch, _, detected_fs = load_mat_signal(
            self._filepath, params["channel"])

        fs = int(detected_fs) if detected_fs is not None else params["fs"]

        t0 = params["t_start"]
        t1 = params["t_end"]
        i0 = int(t0 * fs) if t0 > 0 else 0
        i1 = int(t1 * fs) if t1 > 0 else len(sig)
        sig = sig[i0:i1]

        if sig.std() < 1e-10:
            raise ValueError("Signal is flat — wrong channel.")

        n_samples = len(sig)
        dur_s     = n_samples / fs
        _prog(70, f"Signal loaded — {dur_s:.0f} s  ({n_samples:,} samples)")

        # ONLY processing step: z-score normalisation for display scale.
        # This is not a "filter" — it never changes morphology, polarity,
        # or which sample is the maximum; it's purely an axis convenience,
        # identical to what "Show raw signal" already displayed pre-preview.
        signal_raw_norm = normalize(sig)
        _prog(100, "Raw signal ready")

        return {
            "fs":           fs,
            "signal_raw":   sig,
            "signal_raw_norm": signal_raw_norm,
            "dur_s":        dur_s,
            "detected_ch":  detected_ch,
            "detected_fs":  detected_fs,
            "requested_ch": params["channel"],
            "fs_from_file": detected_fs is not None,
        }

    def _on_raw_load_done(self, bundle: dict) -> None:
        """Write raw-only state on the main thread and draw the raw trace.

        Mirrors the bookkeeping parts of _on_preview_done (channel/fs
        feedback, KPI reset, status message) but explicitly leaves every
        detection/filtering field at None — signal_flt, all_cands,
        rpeaks_ok, thresh_amp, etc. — so the rest of the app's existing
        "not previewed yet" guards (which already check for None) behave
        correctly until the user clicks Preview Detection.
        """
        fs        = bundle["fs"]
        sig_raw   = bundle["signal_raw"]
        n_samples = len(sig_raw)

        self._fs               = fs
        self._signal_raw       = sig_raw
        self._signal_raw_norm  = bundle["signal_raw_norm"]
        self._signal_flt       = None
        self._time             = np.arange(n_samples) / fs
        self._all_cands        = None
        self._all_proms        = None
        self._no_filter_mode   = True
        self._signal_inverted  = False
        self._raw_only_loaded  = True
        self._nav_pos          = 0.0
        self._ds_time          = None
        self._ds_sig           = None
        self._ds_sig_max       = None
        self._ds_sig_mid       = None
        self._ds_raw_sig       = None
        self._ds_raw_sig_max   = None
        self._ds_raw_sig_mid   = None
        self._ov_ylim          = None
        self._rpeaks_ok            = None
        self._rpeaks_rej           = None
        self._thresh_amp           = 0.0
        self._rpeaks_manual_excl   = np.array([], dtype=int)
        self._rpeaks_manual_added  = np.array([], dtype=int)
        self._manual_excluded.clear()
        self._manual_added.clear()
        self._results       = None
        self._epoch_df      = None
        self._annotations   = []
        self._wave_template = None
        self._session_dirty = False
        self._generation    = getattr(self, "_generation", 0) + 1
        self._reset_kpis()
        self._reset_result_plots()
        self._reset_tab_status_labels()

        if bundle["detected_ch"] != bundle["requested_ch"]:
            if self.ent_channel is not None:
                try:
                    self.ent_channel.delete(0, "end")
                    self.ent_channel.insert(0, bundle["detected_ch"])
                except Exception as _exc:
                    log.debug("Could not update channel entry: %s", _exc)
            self.lbl_file.configure(  # type: ignore[union-attr]
                text=f"Auto: {bundle['detected_ch']}", text_color=ORANGE)
        if bundle["fs_from_file"]:
            self._apply_detected_fs(bundle["detected_fs"])
        else:
            try:
                self.lbl_fs_source.configure(
                    text="Tip: fs not found in file — set manually above",
                    text_color=ORANGE)
            except Exception as e:
                log.debug("lbl_fs_source configure failed: %s", e)

        if self.lbl_npeaks is not None:
            self.lbl_npeaks.configure(  # type: ignore[union-attr]
                text="Peaks detected: — (click Preview Detection)", text_color=MUTED)
        if self.btn_review_art is not None:
            self.btn_review_art.configure(state="disabled")  # type: ignore[union-attr]

        dur = bundle["dur_s"]
        self._set_status(
            f"Raw signal loaded — {dur:.0f} s  |  {fs} Hz  "
            "→ click '1 ▶ Preview Detection' to filter and detect peaks.", BLUE)
        self.tabs.set("Detection")
        self._update_ann_count()
        self._nav_pos = 0.0
        self._sync_nav_pos_entry()
        if self.lbl_sig_duration is not None:
            self.lbl_sig_duration.configure(  # type: ignore[union-attr]
                text=f"durée totale : {dur:.1f} s", text_color=MUTED)
        self._draw_detail()

        self._analysis_t_start = 0.0
        self._analysis_t_end   = 0.0
        if self.lbl_analysis_window is not None:
            self.lbl_analysis_window.configure(  # type: ignore[union-attr]
                text=f"Raw signal loaded  ·  {dur:.1f} s  ·  not yet analysed",
                text_color=MUTED)

    def _preview(self) -> None:
        """Load, filter, and detect peaks — fast, no HRV."""
        if not self._filepath:
            messagebox.showwarning("No file", "Open a .mat file first.")
            return
        # Snapshot widget values on the main thread before spawning background work.
        params = self._snapshot_params()
        # pass_result=True: worker returns a bundle; _on_preview_done receives it.
        # This ensures ALL shared state writes happen on the main thread only.
        self._start_async(
            self.btn_preview, "Loading…", "Loading signal…",
            lambda: self._preview_worker(params),
            self._on_preview_done,
            pass_result=True,
        )

    def _preview_worker(self, params: dict) -> dict:
        """Background worker — MUST NOT write to self.

        Loads and processes the signal, then returns a plain data bundle.
        All instance-variable writes happen in ``_on_preview_done`` on the
        main thread, preventing data races with Tk resize/redraw callbacks.
        """
        if self._filepath is None:
            raise ValueError("No file loaded.")

        def _prog(pct: int, msg: str) -> None:
            """Thread-safe progress update — schedules on the main thread."""
            self.after(0, lambda p=pct, m=msg: self._set_progress(p, m))

        _prog(2, "Loading signal from file…")
        sig, detected_ch, _, detected_fs = load_mat_signal(
            self._filepath, params["channel"])

        # Determine effective fs
        if detected_fs is not None:
            fs = int(detected_fs)
        else:
            fs = params["fs"]

        # Estimation automatique min_rr supprimée.

        t0 = params["t_start"]
        t1 = params["t_end"]
        i0 = int(t0 * fs) if t0 > 0 else 0
        i1 = int(t1 * fs) if t1 > 0 else len(sig)
        sig = sig[i0:i1]

        if sig.std() < 1e-10:
            raise ValueError("Signal is flat — wrong channel.")

        n_samples = len(sig)
        dur_s     = n_samples / fs
        _prog(5, f"Signal loaded — {dur_s:.0f} s  ({n_samples:,} samples)")

        # _compute_preview_bundle does all DSP — pure, no self writes
        def _prep_prog(pct: int, msg: str) -> None:
            _prog(5 + int(pct * 0.85), msg)

        signal_bundle = self._compute_preview_bundle(sig, fs, params, _prep_prog)

        # Run threshold detection on the computed candidates (pure computation)
        thresh = params["thresh"]
        accepted, rejected, thresh_amp = apply_threshold(
            signal_bundle["signal_flt"],
            signal_bundle["all_cands"],
            signal_bundle["all_proms"],
            thresh,
            fs=fs,
        )

        _prog(100, "Done")

        return {
            # Signal identity
            "fs":              fs,
            "signal_raw":      sig,
            "dur_s":           dur_s,
            "detected_ch":     detected_ch,
            "detected_fs":     detected_fs,
            "requested_ch":    params["channel"],
            "fs_from_file":    detected_fs is not None,
            "thresh":          thresh,
            # Prepared signal bundle
            **signal_bundle,
            # Detection results
            "rpeaks_ok":       accepted,
            "rpeaks_rej":      rejected,
            "thresh_amp":      thresh_amp,
            "recommended_min_rr_ms": None,
        }

    def _apply_detected_fs(self, fs: float) -> None:
        """Update the fs entry and source label on the main thread."""
        try:
            self.ent_fs.delete(0, "end")
            self.ent_fs.insert(0, str(int(fs)))
        except Exception as _exc:
            log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 5605, _exc)
        self.lbl_fs_source.configure(
            text=f"✓ Auto-detected from file: {int(fs)} Hz",
            text_color=GREEN)

    def _on_preview_done(self, bundle: dict) -> None:
        """Atomically write all signal state on the main thread, then draw.

        This is the ONLY place that should assign signal/peak instance variables
        after a preview.  Because it runs via after(0, …) (scheduled by
        _start_async after the background worker finishes), it is guaranteed to
        execute on the Tk main thread with no concurrent background writes.
        """
        fs        = bundle["fs"]
        sig_raw   = bundle["signal_raw"]
        n_samples = len(sig_raw)

        # ── Atomic state update (main thread) ────────────────────────────────
        self._fs              = fs
        self._signal_raw      = sig_raw
        self._signal_raw_norm = bundle["signal_raw_norm"]
        self._signal_flt      = bundle["signal_flt"]
        self._time            = np.arange(n_samples) / fs
        self._all_cands       = bundle["all_cands"]
        self._all_proms       = bundle["all_proms"]
        self._no_filter_mode  = bundle["no_filter_mode"]
        self._signal_inverted = bundle.get("inverted", False)
        self._raw_only_loaded = False
        # recommended_min_rr_ms supprimé — ent_minrr non modifié automatiquement.
        self._nav_pos         = 0.0
        self._ds_time         = None
        self._ds_sig          = None
        self._ds_sig_max      = None
        self._ds_sig_mid      = None
        self._ds_raw_sig      = None
        self._ds_raw_sig_max  = None
        self._ds_raw_sig_mid  = None
        self._ov_ylim         = None   # clear y-zoom on new signal
        # Peak detection results (computed in worker from pure candidates)
        self._rpeaks_ok           = bundle["rpeaks_ok"]
        self._rpeaks_rej          = bundle["rpeaks_rej"]
        self._thresh_amp          = bundle["thresh_amp"]
        self._rpeaks_manual_excl  = np.array([], dtype=int)
        self._rpeaks_manual_added = np.array([], dtype=int)
        # Reset manual peak edits — new file, clean slate
        self._manual_excluded.clear()
        self._manual_added.clear()
        # Invalidate all previous analysis state — new file, clean slate
        self._results       = None
        self._epoch_df      = None
        self._annotations   = []    # annotations belong to a specific file
        self._wave_template = None  # template may not suit new signal
        self._session_dirty = False
        # Increment generation so any in-flight bg workers discard their results
        self._generation    = getattr(self, "_generation", 0) + 1
        # Reset UI to blank state
        self._reset_kpis()
        self._reset_result_plots()
        self._reset_tab_status_labels()

        # ── UI feedback for auto-detected channel / fs ────────────────────────
        def _subject_from_channel(name: str) -> Optional[str]:
            digits = "".join(ch for ch in name if ch.isdigit())
            return digits if digits else None

        if bundle["detected_ch"] != bundle["requested_ch"]:
            if self.ent_channel is not None:
                try:
                    self.ent_channel.delete(0, "end")
                    self.ent_channel.insert(0, bundle["detected_ch"])
                except Exception as _exc:
                    log.debug("Could not update channel entry: %s", _exc)

            if self.ent_subject is not None:
                subject_id = _subject_from_channel(bundle["detected_ch"])
                current_subject = self.ent_subject.get().strip()
                if subject_id and (not current_subject or current_subject.lower().startswith("subject")):
                    try:
                        self.ent_subject.delete(0, "end")
                        self.ent_subject.insert(0, subject_id)
                    except Exception as _exc:
                        log.debug("Could not update subject entry: %s", _exc)

            self.lbl_file.configure(  # type: ignore[union-attr]
                text=f"Auto: {bundle['detected_ch']}", text_color=ORANGE)
        if bundle["fs_from_file"]:
            self._apply_detected_fs(bundle["detected_fs"])
        else:
            try:
                self.lbl_fs_source.configure(
                    text="Tip: fs not found in file — set manually above",
                    text_color=ORANGE)
            except Exception as e:
                log.debug("lbl_fs_source configure failed: %s", e)

        # Update peak count label and quality score
        n = len(self._rpeaks_ok)  # type: ignore[union-attr]
        color = GREEN if n > 10 else RED
        if self.lbl_npeaks is not None:
            self.lbl_npeaks.configure(text=f"Peaks detected: {n}", text_color=color)  # type: ignore[union-attr]
        if self.btn_review_art is not None:
            self.btn_review_art.configure(state="normal" if n > 4 else "disabled")  # type: ignore[union-attr]
        self._update_signal_quality(self._rpeaks_ok)  # type: ignore[union-attr]

        dur = bundle["dur_s"]
        self._set_status(
            f"Signal ready — {n} peaks  |  {dur:.0f} s  |  {fs} Hz  "
            "→ adjust threshold then Run Full Analysis.", GREEN)
        self.tabs.set("Detection")
        self._update_ann_count()   # reflect cleared annotations immediately
        # Sync nav bar
        self._nav_pos = 0.0
        self._sync_nav_pos_entry()
        if self.lbl_sig_duration is not None:
            self.lbl_sig_duration.configure(  # type: ignore[union-attr]
                text=f"durée totale : {dur:.1f} s", text_color=MUTED)
        self._draw_detail()

        # Reset analysis window on new signal load and update feedback label
        self._analysis_t_start = 0.0
        self._analysis_t_end   = 0.0
        if self.lbl_analysis_window is not None:
            self.lbl_analysis_window.configure(  # type: ignore[union-attr]
                text=f"Full signal  ·  {n} peaks  ·  {dur:.1f} s",
                text_color=MUTED)

    def _windowed_peaks(self) -> "Optional[np.ndarray]":
        """Return a copy of _rpeaks_ok filtered to the current analysis window.

        If no window is set (both bounds = 0), returns the full array.
        Returns None if _rpeaks_ok is None.

        This is the single source of truth for all analysis methods
        (_run_freq, _run_nonlinear, _run_intervals, _run_arrhythmia_analysis,
        _compute_epochs, _compute_rolling_hrv) — they all call this instead of
        doing ``self._rpeaks_ok.copy()`` directly.
        """
        if self._rpeaks_ok is None or self._fs is None:
            return None
        rp = self._rpeaks_ok.copy()
        t0 = self._analysis_t_start
        t1 = self._analysis_t_end
        if t0 <= 0 and t1 <= 0:
            return rp          # no window — full signal
        fs   = self._fs
        mask = rp / fs >= t0
        if t1 > 0:
            mask &= rp / fs <= t1
        return rp[mask]

    def _apply_analysis_window(self) -> None:
        """Read the analysis window entries and store in _analysis_t_start/_end.

        Updates the feedback label with the peak count inside the window.
        Does NOT re-run detection or analysis — the window is applied on
        the next Core Analysis run.
        """
        if self.ent_analysis_t0 is None or self.ent_analysis_t1 is None:
            return

        try:
            t0_raw = self.ent_analysis_t0.get().strip()  # type: ignore[union-attr]
            t1_raw = self.ent_analysis_t1.get().strip()  # type: ignore[union-attr]
            t0 = float(t0_raw) if t0_raw else 0.0
            t1 = float(t1_raw) if t1_raw else 0.0
        except ValueError:
            self._set_status("Invalid window — enter numeric values.", RED)
            return

        # Validate
        if t1 > 0 and t0 >= t1:
            self._set_status("La borne de début doit être inférieure à la borne de fin.", RED)
            return
        if self._time is not None and t1 > float(self._time[-1]) + 0.1:
            self._set_status(
                f"La borne de fin dépasse la durée du signal ({self._time[-1]:.1f} s).", ORANGE)

        self._analysis_t_start = t0
        self._analysis_t_end   = t1

        # Feedback: count peaks in window
        if self._rpeaks_ok is not None and self._fs is not None:
            fs = self._fs
            t_end_eff = float(self._time[-1]) if (self._time is not None and t1 == 0) else t1
            mask = (self._rpeaks_ok / fs >= t0)
            if t1 > 0:
                mask &= (self._rpeaks_ok / fs <= t1)
            n = int(mask.sum())
            dur = (t_end_eff - t0) if t1 > 0 else (float(self._time[-1]) - t0 if self._time is not None else 0)
            label_txt = (f"✓  {n} peaks  ·  {t0:.1f} s → {t_end_eff:.1f} s  ({dur:.1f} s)"
                         if t0 > 0 or t1 > 0
                         else f"✓  {n} peaks  ·  full signal")
            color = GREEN if n >= 5 else ORANGE
        else:
            label_txt = "✓  Window applied — run analysis"
            color = MUTED

        if self.lbl_analysis_window is not None:
            self.lbl_analysis_window.configure(  # type: ignore[union-attr]
                text=label_txt, text_color=color)

        self._set_status(
            "Analysis window updated — re-run Core Analysis.", BLUE)

    def _reset_analysis_window(self) -> None:
        """Reset analysis window to full signal."""
        self._analysis_t_start = 0.0
        self._analysis_t_end   = 0.0
        if self.ent_analysis_t0 is not None:
            self.ent_analysis_t0.delete(0, "end")  # type: ignore[union-attr]
        if self.ent_analysis_t1 is not None:
            self.ent_analysis_t1.delete(0, "end")  # type: ignore[union-attr]
        if self.lbl_analysis_window is not None:
            # Recompute peak count for full signal
            if self._rpeaks_ok is not None:
                n = len(self._rpeaks_ok)
                dur = float(self._time[-1]) if self._time is not None else 0
                self.lbl_analysis_window.configure(  # type: ignore[union-attr]
                    text=f"✓  Full signal  ·  {n} peaks  ·  {dur:.1f} s",
                    text_color=GREEN)
            else:
                self.lbl_analysis_window.configure(  # type: ignore[union-attr]
                    text="", text_color=MUTED)
        self._set_status("Analysis window reset — full signal.", MUTED)

    def _run_analysis(self) -> None:
        if not NK_AVAILABLE:
            messagebox.showerror("Missing", "pip install neurokit2")
            return
        if self._signal_flt is None or self._rpeaks_ok is None:
            messagebox.showwarning("Not ready", "Click '\u25b6 Preview Detection' first.")
            return
        if len(self._rpeaks_ok) < 5:
            messagebox.showwarning(
                "Too few peaks",
                f"Only {len(self._rpeaks_ok)} peaks detected.\n"
                "Adjust threshold / detection settings.")
            return
        # Snapshot ALL widget values on the main thread before spawning the worker.
        # The background thread must never call .get() on any Tkinter widget --
        # doing so races with the event loop and causes freezes / crashes.
        params = self._snapshot_params()
        self._start_async(
            self.btn_run, "Analysing\u2026", "Running HRV analysis\u2026",
            lambda: self._analysis_worker(params),
            self._on_analysis_done,
            pass_result=True,
        )

    def _analysis_worker(self, params: dict) -> dict:
        """Background worker — MUST NOT write to self.

        Runs core analysis and optional artifact correction, then returns a
        plain bundle.  ``_on_analysis_done`` writes all results to self on
        the main thread, preventing races with Tk draw callbacks.
        """
        if self._rpeaks_ok is None:
            raise RuntimeError("No peaks available — run Preview Detection first.")
        # Take a snapshot of the peaks at worker-start time.  After this point
        # the worker operates only on local variables — no self writes.
        rp = self._rpeaks_ok.copy()

        def _prog(pct: int, msg: str) -> None:
            self.after(0, lambda p=pct, m=msg: self._set_progress(p, m))

        artifact_report = None
        if params["artifact_correction"]:
            _prog(2, "Artifact correction (auto)…")
            try:
                rp_corrected, artifact_report = correct_rr_artifacts(
                    rp, self._fs,
                    rr_min_ms=params.get("min_rr_ms", MouseECG.RR_MIN_MS),
                    rr_max_ms=MouseECG.RR_MAX_MS,
                    window_beats=11, dev_threshold=0.20,
                    signal=self._signal_flt,
                )
                rp = rp_corrected
                removed = artifact_report["n_in"] - artifact_report["n_out"]
                log.info("Artifact correction: −%d peaks (non-physio=%d ectopic=%d dup=%d)",
                         removed, artifact_report["n_nonphysio"],
                         artifact_report["n_ectopic"], artifact_report["n_duplicate"])
            except Exception as exc:
                log.warning("Artifact correction failed: %s", exc)
                artifact_report = None

        # ── Save full (artifact-corrected) peak set BEFORE windowing ─────────
        # This is what _on_analysis_done will write back to _rpeaks_ok so that
        # changing the analysis window on the next run always starts from the
        # complete detection result — not a previously-windowed subset.
        rp_full = rp.copy()

        # ── Apply analysis window AFTER artifact correction ───────────────────
        # Window is applied to the analysis only — rp_full is always preserved.
        ana_t0 = params.get("analysis_t_start", 0.0)
        ana_t1 = params.get("analysis_t_end",   0.0)
        if (ana_t0 > 0 or ana_t1 > 0) and self._fs is not None:
            fs_snap = self._fs
            mask = rp / fs_snap >= ana_t0
            if ana_t1 > 0:
                mask &= rp / fs_snap <= ana_t1
            rp_windowed = rp[mask]
            if len(rp_windowed) < 5:
                raise ValueError(
                    f"Analysis window too short: only {len(rp_windowed)} peaks "
                    f"entre {ana_t0:.1f} s et "
                    f"{'fin' if ana_t1 == 0 else f'{ana_t1:.1f} s'}.\n"
                    "Élargir la window ou la réinitialiser (bouton 'Tout').")
            rp = rp_windowed
            log.info("Analysis window applied: %.1f s → %s  (%d / %d peaks)",
                     ana_t0, f"{ana_t1:.1f} s" if ana_t1 > 0 else "end",
                     len(rp), len(rp_full))

        _prog(10, "Core analysis (RR, HR, time-domain HRV, beat template)…")
        if self._signal_flt is None:
            raise RuntimeError("Signal not loaded — run Preview Detection first.")
        results = analyse_core(self._signal_flt, rp, self._fs,
                               progress_cb=lambda p, m: _prog(10 + int(p * 0.9), m))

        return {
            "results":         results,
            "rpeaks_ok":       rp_full,   # ← always full set: never overwrite with windowed
            "rpeaks_analysed": rp,         # ← windowed subset used for this analysis
            "artifact_report": artifact_report,
            "auto_epochs":     params.get("auto_epochs", False),
        }

    def _on_analysis_done(self, bundle: dict) -> None:
        # ── Atomic state update (main thread) ────────────────────────────────
        self._results         = bundle["results"]
        # rpeaks_ok in the bundle is always the FULL artifact-corrected set,
        # never the windowed subset — so the signal view and next analysis
        # always start from the complete detection result.
        self._rpeaks_ok       = bundle["rpeaks_ok"]
        self._artifact_report = bundle["artifact_report"]

        if self._results is None:
            return
        # Sync the peak label: show windowed count if a window was active,
        # otherwise show total artifact-corrected count.
        _rp_ok   = self._rpeaks_ok
        _rp_used = bundle.get("rpeaks_analysed", _rp_ok)  # windowed subset
        n_total  = len(_rp_ok)  if _rp_ok   is not None else 0
        n_used   = len(_rp_used) if _rp_used is not None else 0
        n_peaks  = n_used  # use analysed count for status messages

        windowed = (n_used < n_total)
        if self.lbl_npeaks is not None:
            _c = GREEN if n_total > 10 else RED
            arep = bundle.get("artifact_report")
            _suffix = "  (after correction)" if (arep and arep["n_in"] != arep["n_out"]) else ""
            if windowed:
                _suffix += f"  [{n_used}/{n_total} in window]"
            self.lbl_npeaks.configure(  # type: ignore[union-attr]
                text=f"Peaks detected: {n_total}{_suffix}", text_color=_c)
        arep = self._artifact_report
        if arep and self._snapshot_params().get("artifact_correction"):
            removed = arep["n_in"] - arep["n_out"]
            art_str = (f"  |  −{removed} artifacts" if removed > 0
                       else "  |  no artifacts")
        else:
            art_str = ""
        n_valid = self._results["hr"].get("n_valid", n_peaks)
        win_str = f" (window: {n_used} peaks)" if windowed else ""
        self._set_status(
            f"Core analysis done — {n_used} peaks analysed{win_str} / {n_valid} valid{art_str}  |  rendering…", GREEN)
        self._update_kpis()
        self._draw_detail()
        # Enable the per-tab buttons now that core results are available
        for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl", "btn_run_arrhythmia"):
            if getattr(self, btn_attr, None) is not None:
                getattr(self, btn_attr).configure(state="normal")
        # Update per-tab status labels
        if self.lbl_freq_status is not None:
            self.lbl_freq_status.configure(  # type: ignore[union-attr]
                text="  Core done — click to compute LF / HF", text_color=BLUE)
        if self.lbl_nonlin_status is not None:
            self.lbl_nonlin_status.configure(  # type: ignore[union-attr]
                text="  Core done — click to compute SampEn / DFA (slow!)", text_color="#9C27B0")
        if self.lbl_ivl_status is not None:
            self.lbl_ivl_status.configure(  # type: ignore[union-attr]
                text="  Core done — click to delineate P/Q/S/T waves", text_color=ORANGE)
        try:
            self._draw_core_results(
                on_complete=lambda: self._set_status(
                    f"Core analysis done — {n_used} peaks / {n_valid} valid{art_str}{win_str}  "
                    "| Use per-tab buttons for Freq / Non-linear / Intervals", GREEN),
                auto_epochs=bool(bundle.get("auto_epochs", False)),
            )
        except Exception:
            log.exception("_draw_core_results failed")
        # Populate interpretation tab with core values (freq/nonlinear added later)
        pass  # interpretation removed
        # Update analysis window label to reflect the window that was actually used
        if self.lbl_analysis_window is not None:
            t0 = self._analysis_t_start
            t1 = self._analysis_t_end
            if t0 > 0 or t1 > 0:
                dur_str = f"{t0:.1f} s → {t1:.1f} s" if t1 > 0 else f"{t0:.1f} s → fin"
                self.lbl_analysis_window.configure(  # type: ignore[union-attr]
                    text=f"✓  Analysed  ·  {n_used}/{n_total} peaks  ·  {dur_str}",
                    text_color=GREEN)
            else:
                self.lbl_analysis_window.configure(  # type: ignore[union-attr]
                    text=f"✓  Analysed  ·  {n_total} peaks  ·  full signal",
                    text_color=GREEN)
        # Enable Save Session now that we have results
        self._update_quality_badge()
        if self.btn_save_session is not None:
            self.btn_save_session.configure(state="normal")  # type: ignore[union-attr]
        self._session_dirty = True
        # Update session/template info labels
        self._update_session_ui(
            has_session=bool(self._filepath and load_session(self._filepath) is not None))


    # ── Per-module on-demand analysis ────────────────────────

    def _open_artifact_review(self) -> None:
        """Detect artifact candidates and open the interactive review dialog."""
        if self._signal_flt is None or self._rpeaks_ok is None or len(self._rpeaks_ok) < 4:
            messagebox.showwarning("Not ready",
                                   "Run Preview Detection first to load peaks.")
            return

        rp  = self._rpeaks_ok.copy()
        fs  = self._fs
        sig = self._signal_flt
        rr_min = float(self._safe_float(self.ent_minrr, MouseECG.RR_MIN_MS))

        self._set_status("Detecting artifacts…", MUTED)
        self.btn_review_art.configure(state="disabled", text="Detecting…")  # type: ignore[union-attr]
        self.update_idletasks()

        try:
            candidates = detect_rr_artifacts(
                rp, fs,
                rr_min_ms    = rr_min,
                rr_max_ms    = MouseECG.RR_MAX_MS,
                window_beats = 11,
                dev_threshold= 0.20,
                signal       = sig,
            )
        except Exception as exc:
            messagebox.showerror("Detection error", str(exc))
            self.btn_review_art.configure(state="normal", text="🔍  Review Artifacts")  # type: ignore[union-attr]
            return

        self.btn_review_art.configure(state="normal", text="🔍  Review Artifacts")  # type: ignore[union-attr]

        n = len(candidates)
        if n == 0:
            self._set_status("No artifacts detected — signal looks clean ✓", GREEN)
            messagebox.showinfo("No artifacts",
                                "No artifact candidates found with the current settings.\n\n"
                                "If you suspect issues, try lowering the Min R-R distance "
                                "or adjusting the sensitivity threshold.")
            return

        self._set_status(f"{n} artifact candidates found — opening review…", ORANGE)

        # Count by type for display
        counts = {}
        for c in candidates:
            counts[c["type"]] = counts.get(c["type"], 0) + 1
        detail = "  ·  ".join(f"{v} {k}" for k, v in counts.items())
        self._set_status(f"Reviewing {n} candidates ({detail})", ORANGE)

        dlg = ArtifactReviewDialog(self, sig, rp, fs, candidates, rr_min_ms=rr_min)
        self.wait_window(dlg)   # blocks until dialog closes

        result = dlg.get_result()
        if result is None:
            self._set_status("Artifact review cancelled", MUTED)
            return

        corrected, report = apply_artifact_decisions(rp, result)
        removed = report["n_in"] - report["n_out"]

        self._rpeaks_ok       = corrected
        self._artifact_report = report

        # Clear any manual exclusions that overlapped removed peaks
        if removed > 0:
            removed_samples = {c["sample"] for c in result if c["decision"] == "remove"}
            self._manual_excluded -= removed_samples

        art_str = (f"Artifact review: −{removed} beats "
                   f"(non-physio={report['n_nonphysio']}  "
                   f"ectopic={report['n_ectopic']}  "
                   f"dup={report['n_duplicate']}  "
                   f"kept={report['n_kept']})")

        if removed > 0:
            # Peaks changed → previous HRV metrics are stale.  Discard them so
            # the user cannot export results that don't match the corrected peaks.
            self._results = None
            self._epoch_df = None
            stale_note = "  ⚠ Re-run Core Analysis to update HRV metrics."
            self._set_status(art_str + stale_note, ORANGE)
            # Visual warning on every analysis tab
            warn_text = "⚠  Peaks changed after artifact review — re-run Core Analysis"
            for attr in ("lbl_freq_status", "lbl_nonlin_status", "lbl_ivl_status"):
                lbl = getattr(self, attr, None)
                if lbl is not None:
                    lbl.configure(text=warn_text, text_color=ORANGE)
            for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl"):
                btn = getattr(self, btn_attr, None)
                if btn is not None:
                    btn.configure(state="disabled")
        else:
            self._set_status(art_str, GREEN)

        # Refresh the overview / detail with the cleaned peaks
        self._draw_detail()
        color = GREEN if len(corrected) > 10 else RED
        self.lbl_npeaks.configure(  # type: ignore[union-attr]
            text=f"Peaks detected: {len(corrected)}  (after review)",
            text_color=color)

    def _run_freq(self) -> None:
        """Compute frequency-domain HRV in background, then render."""
        if self._results is None or self._rpeaks_ok is None:
            messagebox.showwarning("Not ready", "Run Core Analysis first.")
            return
        rp = self._windowed_peaks()
        if rp is None or len(rp) < 5:
            messagebox.showwarning("Not ready", "Not enough peaks in the analysis window.")
            return
        fs  = self._fs
        _gen = getattr(self, "_generation", 0)  # snapshot — detect file change
        if self.lbl_freq_status is not None:
            self.lbl_freq_status.configure(text="  Computing…", text_color=ORANGE)  # type: ignore[union-attr]

        def _worker():
            def _prog(p, m):
                self.after(0, lambda pp=p, mm=m: self._set_progress(pp, mm))
            return analyse_hrv_freq(rp, fs, progress_cb=_prog)

        def _done(result):
            if self._results is None or getattr(self, "_generation", 0) != _gen:
                log.info("_run_freq: stale result discarded (file changed)")
                return
            self._results["hrv_freq"] = result
            results: dict = self._results  # narrow for type checkers
            tasks = [
                ("PSD",       lambda: self._plot_psd(results)),
                ("HRV radar", lambda: self._plot_radar(results)),
                ("HRV tables (freq)", lambda: self._plot_hrv_tables(results)),
                ("Summary",   lambda: self._plot_summary(results)),
            ]
            n_lf = n_hf = "—"
            try:
                n_lf = f"{float(result['HRV_LF'].values[0])*100:.1f}%"
                n_hf = f"{float(result['HRV_HF'].values[0])*100:.1f}%"
            except Exception as _exc:
                log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 5940, _exc)
            if self.lbl_freq_status is not None:
                self.lbl_freq_status.configure(  # type: ignore[union-attr]
                    text=f"  Done  LF={n_lf}  HF={n_hf}", text_color=GREEN)
            self._run_plot_chain(
                tasks,
                on_complete=lambda: self._set_status("Frequency HRV done", GREEN))

        self._start_async_result(self.btn_run_freq, "Computing…", _worker, _done)  # type: ignore[arg-type]

    def _run_nonlinear(self) -> None:
        """Compute non-linear HRV in background, then render."""
        if self._results is None or self._rpeaks_ok is None:
            messagebox.showwarning("Not ready", "Run Core Analysis first.")
            return
        rp = self._windowed_peaks()
        if rp is None or len(rp) < 5:
            messagebox.showwarning("Not ready", "Not enough peaks in the analysis window.")
            return
        sig = self._signal_flt
        if sig is None:
            messagebox.showwarning("Not ready", "Signal not loaded.")
            return
        fs  = self._fs
        _gen = getattr(self, "_generation", 0)
        if self.lbl_nonlin_status is not None:
            self.lbl_nonlin_status.configure(  # type: ignore[union-attr]
                text="  Computing SampEn / DFA… (may take 30 s+)", text_color=ORANGE)

        def _worker():
            def _prog(p, m):
                self.after(0, lambda pp=p, mm=m: self._set_progress(pp, mm))
            return analyse_hrv_nonlinear(sig, rp, fs, progress_cb=_prog)

        def _done(result):
            if self._results is None or getattr(self, "_generation", 0) != _gen:
                log.info("_run_nonlinear: stale result discarded (file changed)")
                return
            self._results["hrv_nonlin"] = result
            results: dict = self._results  # narrow for type checkers
            tasks = [
                ("Non-linear metrics", lambda: self._plot_nonlinear(results)),
                ("Summary",            lambda: self._plot_summary(results)),
            ]
            sampen = "—"
            try:
                sampen = f"{float(result['HRV_SampEn'].values[0]):.3f}"
            except Exception as _exc:
                log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 5979, _exc)
            if self.lbl_nonlin_status is not None:
                self.lbl_nonlin_status.configure(  # type: ignore[union-attr]
                    text=f"  Done  SampEn={sampen}", text_color=GREEN)
            self._run_plot_chain(
                tasks,
                on_complete=lambda: self._set_status("Non-linear HRV done", GREEN))

        self._start_async_result(self.btn_run_nonlin, "Computing…", _worker, _done)  # type: ignore[arg-type]

    def _open_annotations(self) -> None:
        """Open the annotation manager dialog."""
        AnnotationManagerDialog(self)

    def _update_ann_count(self) -> None:
        """Refresh the annotation count badge in the toolbar."""
        if self.lbl_ann_count is not None:
            n = len(self._annotations)
            self.lbl_ann_count.configure(  # type: ignore[union-attr]
                text=f"{n}" if n else "",
                text_color=ORANGE if n else MUTED)

    def _run_intervals(self) -> None:
        """Compute interval delineation in background, then launch verifier."""
        if self._results is None or self._signal_flt is None or self._rpeaks_ok is None:
            messagebox.showwarning("Not ready", "Run Core Analysis first.")
            return
        rp = self._windowed_peaks()
        if rp is None or len(rp) < 5:
            messagebox.showwarning("Not ready", "Not enough peaks in the analysis window.")
            return
        fs  = self._fs
        sig = self._signal_flt
        # Recompute rr_ms from the windowed rp — do NOT use self._results["rr_ms"]
        # which came from a potentially different window in Core Analysis.
        rr  = np.diff(rp).astype(float) / fs * 1000
        _gen = getattr(self, "_generation", 0)

        if self.lbl_ivl_status is not None:
            self.lbl_ivl_status.configure(  # type: ignore[union-attr]
                text="  Delineating all beats…", text_color=ORANGE)

        # Load or create template — always pass it so auto-update works
        _wt = self._wave_template
        if _wt is None:
            _wt = WaveTemplate.load()
            self._wave_template = _wt
        wt_for_worker = _wt   # always pass (confirmed or not)
        permissive_for_worker = (bool(self.sw_permissive.get())  # type: ignore[union-attr]
                                  if self.sw_permissive is not None else False)

        def _worker() -> "tuple[pd.DataFrame, np.ndarray, np.ndarray]":
            def _prog(p: int, m: str) -> None:
                self.after(0, lambda pp=p, mm=m: self._set_progress(pp, mm))

            # Build beat matrix here so it can be passed to the verifier
            fixed_hw  = int(MouseECG.BEAT_HALF_WIN_S * fs)
            rr_samp   = np.diff(rp) if len(rp) > 1 else np.array([fixed_hw * 2])
            rr_min_s  = int(rr_samp.min()) if len(rr_samp) else fixed_hw * 2
            half_win  = max(20, min(fixed_hw, int(rr_min_s * 0.45)))
            bt_ms     = np.arange(-half_win, half_win) / fs * 1000
            mask_v    = (rp - half_win >= 0) & (rp + half_win < len(sig))
            valid_rp  = rp[mask_v]
            if len(valid_rp) >= 2:
                idx_mat  = valid_rp[:, None] + np.arange(-half_win, half_win)
                beat_mat = sig[idx_mat].astype(float)
            else:
                beat_mat = np.zeros((0, half_win * 2))

            df = analyse_intervals(sig, rp, fs, rr,
                                   progress_cb=_prog,
                                   wave_template=wt_for_worker,
                                   permissive_bounds=permissive_for_worker)
            return df, beat_mat, bt_ms

        def _done(result: "tuple[pd.DataFrame, np.ndarray, np.ndarray]") -> None:
            if self._results is None or getattr(self, "_generation", 0) != _gen:
                log.info("_run_intervals: stale result discarded (file changed)")
                return
            df, beat_mat, bt_ms = result
            self._results["intervals"] = df

            interval_cols = [c for c in ["PR_ms", "QRS_ms", "QT_ms"] if c in df.columns]
            n_ok    = int((~df[interval_cols].isna().any(axis=1)).sum()) if interval_cols else 0  # type: ignore[arg-type]
            n_total = len(df)

            wt        = self._wave_template
            tmpl_note = f"  template:{wt.source}" if wt else ""
            note      = f"  {n_ok}/{n_total} complete — verify in panel below{tmpl_note}"
            note_color = GREEN if n_ok > 0 else ORANGE
            if n_ok == 0 and n_total > 0:
                note += "  ⚠ check template / filters"
            if self.lbl_ivl_status is not None:
                self.lbl_ivl_status.configure(text=note, text_color=note_color)  # type: ignore[union-attr]

            # Launch interactive verifier (replaces static beat strip)
            self._launch_interval_verifier(df, beat_mat, bt_ms)

            # Plot distributions immediately with all beats
            results: dict = self._results
            self._run_plot_chain(
                [("ECG intervals", lambda: self._plot_intervals(results)),
                 ("Summary",       lambda: self._plot_summary(results))],
                on_complete=lambda: self._set_status(
                    f"Interval delineation done — {n_ok}/{n_total} beats  "
                    "| verify in panel below then click Finalise", note_color))

        self._start_async_result(self.btn_run_ivl, "Delineating…", _worker, _done)  # type: ignore[arg-type]

    def _launch_interval_verifier(
        self,
        df:       "pd.DataFrame",
        beat_mat: "np.ndarray",
        beat_time:"np.ndarray",
    ) -> None:
        """Instantiate IntervalVerifierPanel in the intervals tab.

        Called from _run_intervals _done callback (main thread only).
        The verifier renders into the intervals_ecg CanvasSlot and places
        its navigation bar into frm_ivl_nav.
        """
        slot = self._slots.get("intervals_ecg")
        nav  = self.frm_ivl_nav
        if slot is None or nav is None:
            return

        def _on_finalise(verified_df: "pd.DataFrame") -> None:
            """Replace stored intervals with verified subset; re-plot."""
            if self._results is None:
                return
            self._results["intervals"] = verified_df
            results: dict = self._results
            n_ok    = int((~verified_df[["PR_ms","QRS_ms","QT_ms"]]
                           .isna().any(axis=1)).sum()) if all(  # type: ignore[arg-type]
                               c in verified_df.columns for c in
                               ["PR_ms","QRS_ms","QT_ms"]) else 0
            if self.lbl_ivl_status is not None:
                self.lbl_ivl_status.configure(  # type: ignore[union-attr]
                    text=f"  ✓ Finalised — {n_ok}/{len(verified_df)} beats accepted",
                    text_color=GREEN)
            self._run_plot_chain(
                [("ECG intervals", lambda: self._plot_intervals(results)),
                 ("Summary",       lambda: self._plot_summary(results))],
                on_complete=lambda: self._set_status(
                    f"Intervals finalised — {n_ok}/{len(verified_df)} beats", GREEN))

        n_verifier = min(len(df), len(beat_mat))
        self._ivl_verifier = IntervalVerifierPanel(
            df        = df.iloc[:n_verifier],
            beat_mat  = beat_mat[:n_verifier],
            beat_time = beat_time,
            fs        = self._fs,
            slot      = slot,
            nav_frame = nav,
            on_finalise = _on_finalise,
        )

    def _start_async_result(
        self,
        button: ctk.CTkButton,
        busy_label: str,
        worker: Callable[[], Any],
        on_done: Callable[[Any], None],
        ) -> None:
        """Convenience wrapper: like _start_async but forwards the worker return value.

        Calls _start_async with pass_result=True so that on_done receives the
        value returned by worker().  All threading, progress, and error handling
        are identical to _start_async.
        """
        self._start_async(
            button,
            busy_label,
            "",          # no separate status message — caller sets its own
            worker,
            on_done,
            pass_result=True,
        )

    # ── Async helper ──────────────────────────────────────────

    def _set_progress(self, pct: int, msg: str) -> None:
        """Update the deterministic progress bar + stage label (main thread only).

        Does NOT call update_idletasks() — that would block the event loop and
        cause the UI to freeze mid-analysis.  The progress bar is updated
        asynchronously by the normal Tk event loop.
        """
        frac = max(0.0, min(1.0, pct / 100.0))
        self.progress.set(frac)
        self.lbl_progress.configure(text=f"{pct}%  {msg}")

    def _start_async(
        self,
        button: ctk.CTkButton,
        btn_busy_label: str,
        status_msg: str,
        worker: Callable[[], Any],
        on_done: Callable[..., None],
        original_label: "str | None" = None,
        pass_result: bool = False,
    ) -> None:
        """Disable *button*, show the progress bar, and run *worker* in a thread.

        On success, *on_done* is scheduled on the main thread via ``after(0, …)``.
        On failure, an error dialog is shown and the button is re-enabled.
        The original button text is restored in both cases.

        Parameters
        ----------
        button          : Button to disable while busy.
        btn_busy_label  : Text shown on the button while running.
        status_msg      : Status bar text shown while running (pass "" to skip).
        worker          : Callable executed in the background thread.
                          Must not access Tkinter widgets directly.
        on_done         : Callable executed on the main thread on success.
                          If pass_result=True it receives worker()'s return value.
        original_label  : Button text to restore; auto-detected if None.
        pass_result     : If True, worker() return value is passed to on_done(result).
                          If False (default), on_done is called with no arguments.
        """
        original_text = original_label or button.cget("text")
        button.configure(state="disabled", text=btn_busy_label)
        if status_msg:
            self._set_status(status_msg, ORANGE)
        self.progress.set(0)
        self._prog_row.pack(side="bottom", fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_S))

        # ── Heartbeat animation ───────────────────────────────────────────────
        # Pulses the progress bar between its current value and (value + 3%)
        # every 400 ms so the user can always tell the app is alive, even during
        # a long numpy/scipy call that produces no intermediate callbacks.
        _pulse_after_id: str | None = None
        _pulse_direction: int = 1     # 1 = growing, -1 = shrinking
        _pulse_base: float    = 0.0   # last "real" progress value

        def _pulse() -> None:
            nonlocal _pulse_after_id, _pulse_direction, _pulse_base
            if _pulse_after_id is None:
                return
            cur = self.progress.get()
            # If the real progress moved forward, update base and reset pulse
            if cur > _pulse_base + 0.04 or cur < _pulse_base:
                _pulse_base      = cur
                _pulse_direction = 1
            delta = 0.03 * _pulse_direction
            nxt   = cur + delta
            # Clamp so pulse never exceeds base+4% and never goes below base
            if nxt > _pulse_base + 0.04:
                nxt              = _pulse_base + 0.04
                _pulse_direction = -1
            elif nxt < _pulse_base:
                nxt              = _pulse_base
                _pulse_direction = 1
            self.progress.set(max(0.0, min(0.99, nxt)))
            _pulse_after_id = self.after(400, _pulse)

        _pulse_after_id = self.after(400, _pulse)

        def _thread_target() -> None:
            try:
                result = worker()
                if pass_result:
                    self.after(0, lambda r=result: _finish_success(r))
                else:
                    self.after(0, lambda: _finish_success(None))
            except Exception as exc:
                tb = traceback.format_exc()
                self.after(0, lambda e=exc, t=tb: _finish_error(e, t))

        def _finish_success(result: Any) -> None:
            nonlocal _pulse_after_id
            _pulse_after_id = None          # stop heartbeat
            self._stop_progress(button, original_text)
            if pass_result:
                on_done(result)
            else:
                on_done()  # noqa: pass_result=False branch — on_done takes no args

        def _finish_error(exc: Exception, tb: str) -> None:
            nonlocal _pulse_after_id
            _pulse_after_id = None          # stop heartbeat
            self._stop_progress(button, original_text)
            self._set_status(f"Error: {exc}", RED)
            messagebox.showerror("Error", f"{exc}\n\n{tb}")

        threading.Thread(target=_thread_target, daemon=True).start()

    def _stop_progress(self, button: ctk.CTkButton, original_label: str) -> None:
        self.progress.set(1.0)
        self._prog_row.pack_forget()
        self.lbl_progress.configure(text="")
        button.configure(state="normal", text=original_label)

    # ════════════════════════════════════════════════════════
    #  DRAWING — overview & detail
    # ════════════════════════════════════════════════════════

    def _draw_overview(self) -> None:
        """Stub — overview removed; detail view is the sole signal display."""
        pass  # all callers now use _draw_detail() directly

    def _compute_filter_preview_segment(
        self, t_start: float, t_end: float,
    ) -> "Optional[tuple[np.ndarray, np.ndarray, np.ndarray]]":
        """Compute a live-filtered preview of the signal over [t_start, t_end].

        Operates ONLY on the visible window (+ a short margin to absorb
        filtfilt edge transients) using the CURRENT filter widget values
        (HP/LP cutoffs, notch, clean method) — never on the full recording.
        This is intentionally cheap: it never touches ``self._signal_flt``
        or any detection state, so it's safe to recompute on every redraw
        without affecting Preview Detection / Run Full Analysis results.

        Returns (t_slice, raw_slice_norm, filt_slice_norm), or None if the
        raw signal isn't loaded yet or the window is too short to filter.
        """
        if self._signal_raw is None or self._fs is None or self._time is None:
            return None
        fs = self._fs

        margin = 1.0  # seconds — absorbs filtfilt edge transients
        lo_t = max(0.0, t_start - margin)
        hi_t = min(float(self._time[-1]), t_end + margin)
        lo_i = int(lo_t * fs)
        hi_i = int(hi_t * fs)
        if hi_i - lo_i < int(0.5 * fs):
            return None  # too short to filter meaningfully

        seg_raw = np.asarray(self._signal_raw[lo_i:hi_i], dtype=float)

        lp_v         = self._safe_float(self.ent_lp, MouseECG.BP_LO_HZ)
        hp_v         = self._safe_float(self.ent_hp, MouseECG.BP_HI_HZ)
        notch_on     = bool(self.sw_notch.get()) if self.sw_notch is not None else False
        clean_method = self.cb_clean.get() if self.cb_clean is not None else "neurokit"

        seg = seg_raw.copy()
        try:
            seg = bandpass(seg, fs, lp_v, hp_v)
        except Exception as exc:
            log.debug("filter preview: bandpass skipped — %s", exc)
        if notch_on:
            try:
                seg = notch(seg, fs)
            except Exception as exc:
                log.debug("filter preview: notch skipped — %s", exc)
        try:
            if nk is not None:
                seg = nk.ecg_clean(seg, sampling_rate=fs, method=clean_method)
        except Exception as exc:
            log.debug("filter preview: ecg_clean skipped — %s", exc)

        seg      = normalize(np.asarray(seg, dtype=float))
        raw_norm = normalize(seg_raw)

        # Trim the margin back off — only the visible window is returned
        off0 = int(round((t_start - lo_t) * fs))
        off1 = off0 + int(round((t_end - t_start) * fs))
        off1 = min(off1, len(seg), len(raw_norm))
        off0 = min(off0, off1)
        t_slice = self._time[lo_i:hi_i][off0:off1]
        return t_slice, raw_norm[off0:off1], seg[off0:off1]

    def _on_filter_preview_toggle(self) -> None:
        """Toggle the before/after filter overlay and redraw."""
        self._filter_preview_on = bool(self.sw_filter_preview.get()) if self.sw_filter_preview is not None else False
        self._draw_detail()

    def _refresh_filter_preview(self) -> None:
        """Recompute the filter-preview overlay with current widget values.

        Bound to HP/LP entry <Return>/<FocusOut> and notch/clean-method
        changes — the preview segment isn't auto-reactive to keystrokes,
        only re-evaluated on these discrete commit events, matching how
        the rest of the sidebar (Preview Detection button) already works.
        """
        if self._filter_preview_on:
            self._draw_detail()

    def _draw_detail(self, t_start: float | None = None) -> None:
        """Draw the time-windowed detail view with peak markers.

        The active signal (raw or filtered) is drawn at full opacity; the
        other signal is drawn as a ghost at 20 % opacity so the filtering
        effect is always visible.  Peak markers are computed from the filtered
        signal regardless of mode — they are not re-detected on toggle.

        Also handles two extra states, both display-only (no detection state
        is touched):
        • Raw-only (just opened, Preview Detection not yet run): signal_flt
          is None, so only the raw trace is drawn, with no peaks/threshold.
        • Filter preview overlay (self._filter_preview_on): an on-the-fly
          filtered version of the visible window, computed from the current
          filter widget values, overlaid on the raw trace so the user can
          judge filter settings before committing to Preview Detection.
        """
        if self._time is None:
            return

        sig_flt = self._signal_flt
        sig_raw = self._signal_raw_norm
        time    = self._time
        fs      = self._fs
        raw_only = sig_flt is None
        # Force "show raw" while no filtered signal exists yet — there is
        # nothing else to display, and the toggle would otherwise blank the plot.
        show_raw = self._show_raw or raw_only

        try:
            win = float(self.ent_window.get())
            if not (0 < win < 1e6):
                win = 10.0
        except Exception:
            win = 10.0

        if t_start is None:
            t_start = self._nav_pos
        t_end  = min(time[-1], t_start + win)
        # Index-slice instead of a full-length boolean mask: `time` is always
        # uniformly spaced (np.arange(n)/fs), so the window bounds map
        # directly to a sample-index range. A full `(time >= t0) & (time <= t1)`
        # mask costs O(n_samples_total) on every redraw; in edit mode this
        # runs up to ~33x/sec (hover throttle), so on long/high-fs recordings
        # that cost is easily noticeable. Slicing is O(window size) instead.
        _i0 = max(0, int(t_start * fs))
        _i1 = min(len(time), int(t_end * fs) + 1)
        mask_t = slice(_i0, _i1)

        # ── Filter preview overlay (before/after) ──────────────────────────
        # Computed for the visible window only — cheap, display-only, never
        # touches signal_flt or detection state. Works both pre- and
        # post-Preview Detection so the user can audition new filter values.
        filt_preview = None
        if self._filter_preview_on:
            filt_preview = self._compute_filter_preview_segment(t_start, t_end)

        def _time_for(sig: np.ndarray | None) -> np.ndarray:
            if sig is None or len(sig) == len(time):
                return time
            log.warning(
                "Time axis length %d does not match signal length %d; truncating for plotting",
                len(time), len(sig)
            )
            return time[:len(sig)]

        def _window_mask(sig: np.ndarray | None) -> np.ndarray:
            if sig is None or len(sig) == len(time):
                return np.arange(len(time))[mask_t]
            t = time[:len(sig)]
            return (t >= t_start) & (t <= t_end)

        rp_ok          = self._rpeaks_ok  if self._rpeaks_ok  is not None else np.array([])
        rp_rej         = self._rpeaks_rej if self._rpeaks_rej is not None else np.array([])
        rp_excl        = self._rpeaks_manual_excl  if self._rpeaks_manual_excl  is not None else np.array([])
        rp_added       = self._rpeaks_manual_added if self._rpeaks_manual_added is not None else np.array([])
        t_amp          = self._thresh_amp
        edit_mode      = self._edit_mode
        no_filter_mode  = self._no_filter_mode
        signal_inverted = self._signal_inverted
        _ann_snap       = list(self._annotations)
        hover_samp      = self._hover_samp
        hover_near      = self._hover_samp_near

        def _in_view(idx: np.ndarray) -> np.ndarray:
            return (idx / fs >= t_start) & (idx / fs <= t_end)

        mask_ok    = _in_view(rp_ok)    if len(rp_ok)    else np.array([], bool)
        mask_rej   = _in_view(rp_rej)   if len(rp_rej)   else np.array([], bool)
        mask_excl  = _in_view(rp_excl)  if len(rp_excl)  else np.array([], bool)
        mask_added = _in_view(rp_added) if len(rp_added) else np.array([], bool)
        n_in_view  = int(mask_ok.sum())
        n_excl     = int(mask_excl.sum())
        n_added_view = int(mask_added.sum())

        primary_sig   = sig_raw   if show_raw else sig_flt
        primary_color = PLOT["raw"]    if show_raw else PLOT["signal"]
        ghost_sig     = sig_flt   if show_raw else sig_raw
        ghost_color   = PLOT["signal"] if show_raw else PLOT["raw"]
        if raw_only:
            label_mode = "Raw (not yet analysed)"
        elif no_filter_mode:
            label_mode = "Unfiltered" if not show_raw else "Pre-norm baseline"
        else:
            label_mode = "Raw" if show_raw else "Filtered"

        def draw(fig):
            ax = fig.add_subplot(111)
            style_axes(ax)

            # Ghost trace — suppressed in no-filter mode (signals identical)
            # and in raw-only mode (no filtered signal exists yet).
            if ghost_sig is not None and not no_filter_mode and not raw_only:
                t_ghost = _time_for(ghost_sig)
                m_ghost = _window_mask(ghost_sig)
                ax.plot(t_ghost[m_ghost], ghost_sig[m_ghost],
                        color=ghost_color, lw=0.5, alpha=0.22, zorder=1,
                        label="Filtered" if show_raw else "Raw")

            # Primary trace
            if primary_sig is not None:
                t_primary = _time_for(primary_sig)
                m_primary = _window_mask(primary_sig)
                ax.plot(t_primary[m_primary], primary_sig[m_primary],
                        color=primary_color, lw=0.9, zorder=2, label=label_mode)

            # ── Filter preview overlay (before/after, current widget values) ───
            # Independent of raw/filtered toggle and of raw_only state — shows
            # what Preview Detection WOULD produce with the current filter
            # settings, computed live on just the visible window.
            if filt_preview is not None:
                t_fp, raw_fp, filt_fp = filt_preview
                ax.plot(t_fp, filt_fp, color=PLOT["signal"], lw=1.1,
                        zorder=3, alpha=0.9, label="Filtered (preview)")

            # Rejected candidates (light grey circles)
            if mask_rej.any() and sig_flt is not None:
                ax.scatter(rp_rej[mask_rej] / fs, sig_flt[rp_rej[mask_rej]],
                           color=PLOT["rpeak_bad"], s=30, zorder=4,
                           marker="o", label="Rejected", alpha=0.5)
            # Manually excluded peaks (red X markers)
            if mask_excl.any() and sig_flt is not None:
                ax.scatter(rp_excl[mask_excl] / fs, sig_flt[rp_excl[mask_excl]],
                           color=RED, s=90, zorder=6,
                           marker="x", linewidths=2,
                           label=f"Excluded ({n_excl})")
            # Accepted peaks (green dots)
            if mask_ok.any() and sig_flt is not None:
                ax.scatter(rp_ok[mask_ok] / fs, sig_flt[rp_ok[mask_ok]],
                           color=PLOT["rpeak_ok"], s=55, zorder=5,
                           marker="o", label="Accepted")
            # Manually added peaks (cyan star — rendered on top of everything)
            if mask_added.any() and sig_flt is not None:
                ax.scatter(rp_added[mask_added] / fs, sig_flt[rp_added[mask_added]],
                           color=CYAN, s=140, zorder=7,
                           marker="*", linewidths=1.2, edgecolors="#006064",
                           label=f"Added ({n_added_view})")

            # ── Hover preview (edit mode — shows snapped R-peak position) ───
            if edit_mode and hover_samp is not None and sig_flt is not None:
                h_t = hover_samp / fs
                if t_start <= h_t <= t_end:
                    h_amp = float(sig_flt[hover_samp])
                    # Color: orange = replaces nearby peak, cyan = free placement
                    h_color = ORANGE if hover_near else CYAN
                    h_label = "→ replaces nearby peak" if hover_near else "→ add here"
                    # Dashed vertical guide line
                    ax.axvline(h_t, color=h_color, lw=1.0, ls="--",
                               alpha=0.65, zorder=8)
                    # Diamond marker at snapped amplitude
                    ax.scatter([h_t], [h_amp],
                               color=h_color, s=180, marker="D",
                               alpha=0.80, zorder=10, linewidths=1.4,
                               edgecolors="white", label=h_label)
                    # Small text annotation above the marker
                    ax.annotate(
                        f"{h_t:.3f} s",
                        xy=(h_t, h_amp),
                        xytext=(0, 14), textcoords="offset points",
                        ha="center", va="bottom", fontsize=8,
                        color=h_color, fontweight="bold",
                        bbox=dict(boxstyle="round,pad=0.25",
                                  fc="white", ec=h_color, alpha=0.85, lw=0.8),
                        zorder=11,
                    )

            # ── Annotation spans ─────────────────────────────────────────
            for ann in _ann_snap:
                t0  = float(ann["t_start"])
                t1  = float(ann["t_end"])
                col = ann.get("color", ORANGE_DARK)
                lbl = ann.get("label", "")
                if t1 < t_start or t0 > t_end:
                    continue
                _ylo, _yhi = ax.get_ylim()
                ax.axvspan(max(t0, t_start), min(t1, t_end),
                           color=col, alpha=0.12, zorder=6, linewidth=0)
                for _tx in (t0, t1):
                    if t_start <= _tx <= t_end:
                        ax.axvline(_tx, color=col, lw=1.2, ls="-",
                                   alpha=0.8, zorder=7)
                # Label: prefer at left edge, else at midpoint
                _lx = t0 if t_start <= t0 <= t_end else (t0 + t1) / 2
                if lbl and t_start <= _lx <= t_end:
                    ax.text(_lx + 0.01, _yhi * 0.96, lbl,
                            ha="left", va="top", fontsize=8,
                            color=col, fontweight="bold", zorder=8,
                            bbox=dict(boxstyle="round,pad=0.2",
                                      fc="white", ec=col, alpha=0.88, lw=0.8))

            if not raw_only:
                ax.axhline(t_amp, color=PLOT["threshold"], lw=1.4, ls="--",
                           label=f"Threshold ({t_amp:.3f})")
            ax.set_xlabel("Time (s)")
            ax.set_ylabel("Amplitude (norm.)")
            filter_tag    = ""   # no_filter is the default — no need for a warning tag
            inverted_tag  = "  ·  ↕ auto-inverted" if signal_inverted else ""
            if edit_mode:
                title_suffix = "  ·  ✏ EDIT — L-click: exclude/restore   R-click: add/replace"
            else:
                title_suffix = ""
            title_color  = ORANGE if edit_mode else PLOT["text"]
            if raw_only:
                ax.set_title(
                    f"Detail  {t_start:.1f}–{t_end:.1f} s  ·  {label_mode}"
                    "  ·  click '1 ▶ Preview Detection' to filter & detect",
                    loc="left", color=title_color)
            else:
                ax.set_title(
                    f"Detail  {t_start:.1f}–{t_end:.1f} s  ·  {n_in_view} peaks"
                    f"  ·  {label_mode}{filter_tag}{inverted_tag}{title_suffix}",
                    loc="left", color=title_color)
            ax.legend(framealpha=0, loc="upper right")

        self._slots["detail"].update(draw)

    def _kb_navigate(self, direction: int) -> None:
        """Keyboard left/right arrow navigation — only active on Detection tab."""
        self.nav_ctrl.kb_navigate(direction)

    def _navigate(self, direction: int) -> None:
        """Shift the detail view left/right by 80 % of the window width."""
        self.nav_ctrl.navigate(direction)

    def _navigate_big(self, direction: int) -> None:
        """Jump by 10× the current window width."""
        self.nav_ctrl.navigate_big(direction)

    def _nav_reset(self) -> None:
        self.nav_ctrl.nav_reset()

    def _nav_end(self) -> None:
        """Jump to the end of the signal."""
        self.nav_ctrl.nav_end()

    def _nav_goto(self) -> None:
        """Jump to the time entered in the position field."""
        self.nav_ctrl.nav_goto()

    def _sync_nav_pos_entry(self) -> None:
        """Update the position entry widget to reflect _nav_pos."""
        self.nav_ctrl.sync_nav_pos_entry()

    # ════════════════════════════════════════════════════════
    #  RESULT PLOTS
    # ════════════════════════════════════════════════════════

    def _run_plot_chain(
        self,
        tasks: list,
        on_complete: "Optional[Callable[[], None]]" = None,
        auto_epochs: bool = False,
    ) -> None:
        """Run a list of (label, fn) plot tasks sequentially via after() chain."""
        total = len(tasks)

        def _run_next(idx: int) -> None:
            if idx >= total:
                self._set_progress(100, "Done")
                if on_complete:
                    on_complete()
                if auto_epochs:
                    self.after(200, self._compute_epochs)
                return
            label, fn = tasks[idx]
            pct = int(100 * (idx + 1) / total)
            self._set_progress(pct, f"Rendering {label}…")
            try:
                fn()
            except Exception:
                log.exception("Plot task '%s' failed", label)
            self.after(25, lambda i=idx + 1: _run_next(i))

        _run_next(0)

    def _draw_core_results(
        self,
        on_complete: "Optional[Callable[[], None]]" = None,
        auto_epochs: bool = False,
    ) -> None:
        """Render only the fast core plots (RR, Beat, Summary, Poincaré).

        Called immediately after core analysis.  Freq / non-linear / intervals
        are rendered separately when their per-tab buttons are clicked.
        """
        r = self._results
        if r is None:
            return
        results: dict = r  # narrow for type checkers
        tasks = [
            ("RR / HR tachogram", lambda: self._plot_rr(results)),
            ("HRV time-domain",   lambda: self._plot_hrv_tables(results)),
            ("Poincaré",          lambda: self._plot_nonlinear(results)),
            ("Beat template",     lambda: self._plot_beat_template(results)),
            ("Summary",           lambda: self._plot_summary(results)),
        ]
        self._run_plot_chain(tasks, on_complete=on_complete, auto_epochs=auto_epochs)

    def _draw_all_results(
        self,
        on_complete: "Optional[Callable[[], None]]" = None,
        auto_epochs: bool = False,
    ) -> None:
        """Render ALL result plots (used by export and legacy callers)."""
        r = self._results
        if r is None:
            log.warning("_draw_all_results called with no results")
            return
        results: dict = r  # narrow for type checkers
        tasks = [
            ("RR / HR tachogram",     lambda: self._plot_rr(results)),
            ("HRV tables",            lambda: self._plot_hrv_tables(results)),
            ("Poincaré / non-linear", lambda: self._plot_nonlinear(results)),
            ("PSD",                   lambda: self._plot_psd(results)),
            ("HRV radar",             lambda: self._plot_radar(results)),
            ("ECG intervals ECG preview", lambda: self._plot_intervals_ecg(results)),
            ("ECG intervals",         lambda: self._plot_intervals(results)),
            ("Beat template",         lambda: self._plot_beat_template(results)),
            ("Summary",               lambda: self._plot_summary(results)),
        ]
        self._run_plot_chain(tasks, on_complete=on_complete, auto_epochs=auto_epochs)

    def _plot_rr(self, r: dict) -> None:
        """Plot RR tachogram, HR trace, and RR distribution histogram.

        Drastic RR changes are detected and shown as orange/red markers on
        the tachogram.  Clicking any point navigates to that beat in Detection.
        Right-clicking jumps specifically to the nearest spike.
        """
        import datetime as _dt
        rdf       = r["rr_df"]
        rr_ms_raw = r.get("rr_ms", np.array([]))

        # Fall back to raw RR data if filtered dataframe is empty
        if rdf.empty and len(rr_ms_raw) > 1 and self._rpeaks_ok is not None:
            _wp_rr = self._windowed_peaks()
            rdf = pd.DataFrame({
                "Time_s": (_wp_rr if _wp_rr is not None else self._rpeaks_ok)[1:len(rr_ms_raw) + 1] / self._fs,
                "RR_ms":  rr_ms_raw,
                "HR_bpm": 60_000.0 / np.clip(rr_ms_raw, 1, None),
            })
        if rdf.empty:
            log.warning("_plot_rr: empty rdf — skipping")
            return

        t_all  = np.asarray(rdf["Time_s"].values, dtype=float)
        rr_all = np.asarray(rdf["RR_ms"].values,  dtype=float)
        hr_all = np.asarray(rdf["HR_bpm"].values,  dtype=float)

        # ── Detect drastic RR changes ──────────────────────────────────────
        # A beat is a "spike" if its RR deviates more than spike_thr standard
        # deviations from the local rolling median (window = 15 beats).
        spike_thr  = 2.5   # SD threshold
        spike_idx  = np.array([], dtype=int)
        spike_t    = np.array([], dtype=float)
        spike_rr   = np.array([], dtype=float)
        spike_mag  = np.array([], dtype=float)  # delta-RR in ms

        if len(rr_all) >= 10:
            # Rolling median via scipy — O(n·log(w)) au lieu de O(n·w)
            from scipy.ndimage import median_filter as _median_filter
            win = min(15, len(rr_all) // 2)
            roll_med = _median_filter(rr_all.astype(float), size=win, mode="nearest")
            delta = rr_all - roll_med
            # Also flag beats where consecutive delta-RR is extreme
            drr = np.diff(rr_all)
            drr_padded = np.concatenate([[0], drr])
            rr_sd = max(float(rr_all.std()), 1.0)
            # Spike = large deviation from local median OR large consecutive jump
            spike_mask = (np.abs(delta) > spike_thr * rr_sd) | \
                         (np.abs(drr_padded) > spike_thr * rr_sd * 1.2)
            spike_idx = np.where(spike_mask)[0]
            if len(spike_idx):
                spike_t   = t_all[spike_idx]
                spike_rr  = rr_all[spike_idx]
                spike_mag = delta[spike_idx]

        # Downsample for display
        t_ds   = downsample_for_display(t_all)
        rr_ds  = downsample_for_display(rr_all)
        hr_ds  = downsample_for_display(hr_all)
        rr_mean = float(rr_all.mean())
        hr_mean = float(hr_all.mean())
        rr_sd_v = float(rr_all.std())
        c_rr, c_hr = "#388E3C", ORANGE_DARK
        n_spikes   = len(spike_idx)
        updated_at = _dt.datetime.now().strftime("%H:%M:%S")

        # Spike colours: orange = moderate, red = severe
        def _spike_color(mag: float) -> str:
            return RED if abs(mag) > 3.5 * rr_sd_v else AMBER

        def draw_tachogram(fig):
            axes = fig.subplots(2, 1, sharex=True)
            for ax in axes:
                style_axes(ax)

            # ── RR tachogram ────────────────────────────────────
            axes[0].plot(t_ds, rr_ds, color=c_rr, lw=0.8, zorder=2)
            axes[0].axhline(rr_mean, color=c_rr, ls="--", lw=0.9, alpha=0.5, zorder=1)

            # ±1 SD reference band
            axes[0].axhspan(rr_mean - rr_sd_v, rr_mean + rr_sd_v,
                            alpha=0.06, color=c_rr, zorder=0)
            axes[0].axhline(rr_mean + spike_thr * rr_sd_v,
                            color=AMBER, lw=0.6, ls=":", alpha=0.5, zorder=1)
            axes[0].axhline(rr_mean - spike_thr * rr_sd_v,
                            color=AMBER, lw=0.6, ls=":", alpha=0.5, zorder=1)

            # Spike markers
            if len(spike_t):
                for st, sr, sm in zip(spike_t, spike_rr, spike_mag):
                    col = _spike_color(sm)
                    axes[0].scatter([st], [sr], s=55, color=col,
                                    marker="v" if sm < 0 else "^",
                                    zorder=5, edgecolors="white",
                                    linewidths=0.6, alpha=0.9)

            spike_note = (f"  ·  {n_spikes} spike{'s' if n_spikes != 1 else ''} détecté{'s' if n_spikes != 1 else ''}"
                          if n_spikes else "")
            axes[0].set_ylabel("RR (ms)")
            axes[0].set_title(
                f"RR Intervals  ·  moy. {rr_mean:.1f} ms  ·  SD {rr_sd_v:.1f} ms{spike_note}",
                loc="left", fontsize=9)
            axes[0].set_title(f"click=navigate  r-click=next spike  ·  {updated_at}",
                              loc="right", fontsize=7,
                              color=PLOT.get("muted", "#666"))
            axes[0].tick_params(labelbottom=False)

            # ── HR trace ────────────────────────────────────────
            axes[1].plot(t_ds, hr_ds, color=c_hr, lw=0.8, zorder=2)
            axes[1].axhline(hr_mean, color=c_hr, ls="--", lw=0.9, alpha=0.5, zorder=1)

            # Mirror spikes on HR axis
            if len(spike_t):
                spike_hr = 60_000.0 / np.clip(spike_rr, 1, None)
                for st, shr, sm in zip(spike_t, spike_hr, spike_mag):
                    col = _spike_color(sm)
                    axes[1].scatter([st], [shr], s=45, color=col,
                                    marker="v" if sm < 0 else "^",
                                    zorder=5, edgecolors="white",
                                    linewidths=0.6, alpha=0.9)

            axes[1].set_ylabel("HR (bpm)")
            axes[1].set_xlabel(
                "Time (s)  —  ▲ accélération soudaine  ▼ décélération soudaine  (seuil ±2.5 SD)")
            axes[1].set_title(f"Instantaneous HR  ·  moy. {hr_mean:.0f} bpm",
                              loc="left", fontsize=9)

        # Capture annotations for closure
        _anns = list(self._annotations)
        _draw_fn_orig = draw_tachogram

        def draw_tachogram_annotated(fig):
            _draw_fn_orig(fig)
            axs = fig.axes
            if not axs or not _anns:
                return
            ax = axs[0]          # RR axis
            ax2 = axs[1] if len(axs) > 1 else None
            for ann in _anns:
                col = ann.get("color", ORANGE_DARK)
                lbl = ann.get("label", "")
                ts, te = ann["t_start"], ann["t_end"]
                for _ax in ([ax] if ax2 is None else [ax, ax2]):
                    _ax.axvspan(ts, te, alpha=0.10, color=col, zorder=0)
                    _ax.axvline(ts, color=col, lw=1.2, ls="-", alpha=0.8, zorder=6)
                    _ax.axvline(te, color=col, lw=0.7, ls="--", alpha=0.5, zorder=6)
                if lbl:
                    ylo, yhi = ax.get_ylim()
                    ax.text((ts + te) / 2, yhi, lbl,
                            ha="center", va="top", fontsize=7, color=col,
                            fontweight="bold", zorder=8,
                            bbox=dict(boxstyle="round,pad=0.15",
                                      fc=PLOT.get("bg","#1A1A2E"), ec=col,
                                      alpha=0.8, lw=0.6))

        self._slots["rr"].update(draw_tachogram_annotated)

        # ── Wire click-to-navigate ─────────────────────────────────────────
        if self._rr_click_cid is not None:
            try:
                self._slots["rr"].canvas.mpl_disconnect(self._rr_click_cid)
            except Exception as e:
                log.debug("mpl_disconnect (rr click) failed: %s", e)

        # Store spike times for right-click navigation
        _spike_times = spike_t.copy() if len(spike_t) else np.array([], dtype=float)

        def _on_rr_click(event):
            if event.xdata is None or event.inaxes is None:
                return
            t_clicked = float(event.xdata)

            if event.button == 3 and len(_spike_times):
                # Right-click: jump to the nearest spike
                dists = np.abs(_spike_times - t_clicked)
                nearest_spike_t = float(_spike_times[int(np.argmin(dists))])
                t_nav = nearest_spike_t
                spike_info = f"spike à {nearest_spike_t:.3f} s"
            elif event.button == 1:
                # Left-click: navigate to clicked time
                t_nav = t_clicked
                spike_info = None
            else:
                return

            if self._time is not None:
                sig_dur = float(self._time[-1])
                try:
                    win = float(self.ent_window.get())  # type: ignore[union-attr]
                except Exception:
                    win = 2.0
                self._nav_pos = max(0.0, min(t_nav - win / 2, sig_dur - win))
            self._sync_nav_pos_entry()
            try:
                self.tabs.set("Detection")
            except Exception as e:
                log.debug("tabs.set Detection failed: %s", e)
            self._draw_detail()
            if spike_info:
                self._set_status(f"Navigation → {spike_info}", AMBER)

        self._rr_click_cid = self._slots["rr"].canvas.mpl_connect(
            "button_press_event", _on_rr_click)
        _rr_desc = rdf["RR_ms"].describe()
        _rr_tsv  = "Metric\tRR_ms\n" + "\n".join(
            f"{k}\t{v:.5g}" for k, v in _rr_desc.items())
        self._set_textbox(self.txt_rr,
            "\n".join(f"  {k:<14} {v:>10.2f}" for k, v in _rr_desc.items()),
            tsv=_rr_tsv)

        rr_clipped = rdf["RR_ms"].clip(MouseECG.RR_MIN_MS, MouseECG.RR_MAX_MS).values

        def draw_histogram(fig):
            ax = fig.add_subplot(111)
            style_axes(ax)
            ax.hist(rr_clipped, bins=50, color=c_rr, alpha=0.7,
                    edgecolor="white", lw=0.3)
            ax.set_xlabel("RR (ms)")
            ax.set_ylabel("Count")
            ax.set_title("Distribution RR", loc="left")
            ax.xaxis.set_major_locator(matplotlib.ticker.MultipleLocator(1))

        self._slots["rr_hist"].update(draw_histogram)

    def _plot_hrv_tables(self, r: dict) -> None:
        """Populate time-domain and frequency-domain HRV text boxes."""
        self._set_textbox(self.txt_td, self._df_to_text(r["hrv_time"]),
                          tsv=self._df_to_tsv(r["hrv_time"]))

        fd_df = r["hrv_freq"]
        if fd_df is None or fd_df.empty:
            self._set_textbox(self.txt_fd, "  (not computed)")
            return

        lines: list[str] = []
        for col in fd_df.columns:
            try:
                v    = float(fd_df[col].values[0])
                name = col.replace("HRV_", "")
                if not np.isfinite(v):
                    continue
                if name in ("LF", "HF", "VLF"):
                    lines.append(f"  {name:<26} {v * 100:>10.1f} %")
                elif name == "LFHF":
                    lines.append(f"  {'LF/HF ratio':<26} {v:>10.3f}")
                elif name in ("LF_peak", "HF_peak", "VLF_peak"):
                    lines.append(f"  {name + ' (Hz)':<26} {v:>10.4f}")
                else:
                    lines.append(f"  {name:<26} {v:>10.4f}")
            except Exception as exc:
                log.debug("_plot_hrv_tables fd skip '%s': %s", col, exc)

        # Build TSV alongside the display text
        _fd_tsv_rows = ["Metric\tValue"]
        for col in fd_df.columns:
            try:
                v = float(fd_df[col].values[0])
                if np.isfinite(v):
                    _fd_tsv_rows.append(f"{col.replace('HRV_', '')}\t{v:.6g}")
            except (TypeError, ValueError) as _tsv_exc:
                log.debug("_plot_hrv_tables TSV: skip col %s: %s", col, _tsv_exc)
        _fd_tsv = "\n".join(_fd_tsv_rows)
        self._set_textbox(self.txt_fd, "\n".join(lines) if lines else "  (not computed)",
                          tsv=_fd_tsv if len(_fd_tsv_rows) > 1 else None)

    def _plot_psd(self, r: dict) -> None:
        """Welch PSD with mouse-specific VLF / LF / HF band shading.

        RR intervals are resampled to a uniform time grid using a cubic spline
        before computing the Welch periodogram.  Cubic (vs linear) resampling
        preserves spectral shape and avoids the artificial high-frequency power
        that linear interpolation introduces.

        Mouse-specific design choices
        ─────────────────────────────
        • Interpolation rate: 20 Hz  (Nyquist = 10 Hz >> HF ceiling of 5 Hz)
        • nperseg: aims for ≥ 0.02 Hz resolution — enough to separate
          VLF (0–0.4), LF (0.4–1.5) and HF (1.5–5.0) bands cleanly.
          Formula: nperseg = max(256, min(fs_interp / 0.02, N // 2))
          e.g. 20 / 0.02 = 1000, so for long recordings nperseg = 1000.
        • noverlap: 75 % of nperseg (Welch variance reduction)
        • Window: Hann (default scipy) — good sidelobe suppression
        """
        rr_ms = r["rr_ms"]
        MIN_BEATS = 60
        if len(rr_ms) < MIN_BEATS:
            log.warning("_plot_psd: too few RR intervals (%d, need ≥ %d)", len(rr_ms), MIN_BEATS)
            def draw_warn(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                ax.text(0.5, 0.5,
                        f"Spectral HRV requires ≥ {MIN_BEATS} beats\n"
                        f"(recording has {len(rr_ms)})",
                        ha="center", va="center", color=PLOT["muted"],
                        transform=ax.transAxes, fontsize=11)
                ax.set_axis_off()
            self._slots["psd"].update(draw_warn)
            return
        try:
            # Build a uniformly sampled RR series via cubic spline
            ts    = np.cumsum(rr_ms) / 1000.0          # cumulative time in seconds
            dt    = 1.0 / MouseECG.PSD_INTERP_FS
            t_uni = np.arange(ts[0], ts[-1], dt)
            if len(t_uni) < 32:
                log.warning("_plot_psd: interpolated series too short (%d pts)", len(t_uni))
                return

            cs        = CubicSpline(ts, rr_ms)
            rr_interp = cs(t_uni)
            N         = len(rr_interp)

            # nperseg chosen for ≥ 0.02 Hz resolution (resolves LF band floor at 0.4 Hz cleanly)
            # Minimum 256, maximum N//2
            target_res_hz = 0.020
            nperseg = int(np.clip(
                MouseECG.PSD_INTERP_FS / target_res_hz,
                256, N // 2,
            ))
            noverlap = int(nperseg * 0.75)

            freqs, psd = _scipy_welch(
                rr_interp - rr_interp.mean(),
                fs=MouseECG.PSD_INTERP_FS,
                nperseg=nperseg,
                noverlap=noverlap,
                window="hann",
                scaling="density",
            )

            # Mouse-specific band definitions (Thireau 2008)
            bands = [
                (MouseECG.VLF[0], MouseECG.VLF[1], "VLF",
                 f"VLF  {MouseECG.VLF[0]}–{MouseECG.VLF[1]} Hz", BLUE_DARK),
                (MouseECG.LF[0],  MouseECG.LF[1],  "LF",
                 f"LF   {MouseECG.LF[0]}–{MouseECG.LF[1]} Hz  (baroreflex)", PURPLE),
                (MouseECG.HF[0],  MouseECG.HF[1],  "HF",
                 f"HF   {MouseECG.HF[0]}–{MouseECG.HF[1]} Hz  (respiratory)", "#1B5E20"),
            ]

            # Compute per-band power for annotation
            def _band_power(lo: float, hi: float) -> float:
                m = (freqs >= lo) & (freqs <= hi)
                # np.trapz was renamed np.trapezoid in NumPy 2.0; the shim at the top
                # of the module ensures np.trapz always exists.
                return float(np.trapz(psd[m], freqs[m])) if m.any() else 0.0  # type: ignore[attr-defined]

            band_powers = {name: _band_power(lo, hi) for lo, hi, name, _, _ in bands}
            total_power = sum(band_powers.values()) + 1e-12

            # Frequency resolution actually achieved
            df = freqs[1] - freqs[0] if len(freqs) > 1 else float("nan")

            def draw_psd(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                ax.semilogy(freqs, psd, color="#546E7A", lw=1.0, zorder=3)
                for lo, hi, name, legend_label, color in bands:
                    m = (freqs >= lo) & (freqs <= hi)
                    pct = band_powers[name] / total_power * 100
                    ax.fill_between(freqs, psd, where=m, alpha=0.35, color=color,
                                    label=f"{legend_label}  ({pct:.1f}%)", zorder=2)
                    # Vertical band boundary lines
                    for boundary in (lo, hi):
                        if 0 < boundary < MouseECG.PSD_XLIM:
                            ax.axvline(boundary, color=color, lw=0.8, ls=":",
                                       alpha=0.6, zorder=1)
                ax.set_xlabel("Frequency (Hz)")
                ax.set_ylabel("PSD (ms²/Hz)")
                ax.set_xlim(0, MouseECG.PSD_XLIM)
                ax.legend(framealpha=0, loc="upper right", fontsize=9)
                ax.set_title(
                    f"Power spectral density  (Welch · Δf={df:.3f} Hz · "
                    f"n={N} pts)",
                    loc="left",
                )

            self._slots["psd"].update(draw_psd)
        except Exception as exc:
            log.warning("_plot_psd failed: %s", exc)

    def _plot_radar(self, r: dict) -> None:
        """Normalised HRV spider / radar chart."""
        try:
            metrics: dict[str, float] = {}
            for df, keys in [
                (r["hrv_time"],   ["HRV_SDNN", "HRV_RMSSD", "HRV_pNN6"]),
                (r["hrv_freq"],   ["HRV_LF",   "HRV_HF",    "HRV_LFHF"]),
                (r["hrv_nonlin"], ["HRV_SD1",  "HRV_SD2",   "HRV_SampEn"]),
            ]:
                if df is None or df.empty:
                    continue
                for k in keys:
                    if k not in df.columns:
                        continue
                    try:
                        v = float(df[k].values[0])
                        if np.isfinite(v):
                            metrics[k.replace("HRV_", "")] = v
                    except Exception as exc:
                        log.debug("_plot_radar skip '%s': %s", k, exc)

            if len(metrics) < 3:
                return

            labels   = list(metrics.keys())
            values   = np.array(list(metrics.values()))
            v_range  = values.max() - values.min()
            v_norm   = (values - values.min()) / (v_range + 1e-9)
            n        = len(labels)
            angles   = np.linspace(0, 2 * np.pi, n, endpoint=False).tolist()
            v_closed = v_norm.tolist() + [v_norm[0]]
            a_closed = angles + angles[:1]

            # Rich labels: name + value — placed by thetagrids (stays inside bounds)
            rich_labels = [f"{lbl}\n{val:.3g}" for lbl, val in zip(labels, values)]

            def draw_radar(fig):
                ax = fig.add_subplot(111, polar=True)
                ax.set_facecolor(PLOT["axes"])
                ax.plot(a_closed, v_closed, color=RED, lw=2)
                ax.fill(a_closed, v_closed, color=RED, alpha=0.15)
                ax.set_thetagrids(np.degrees(angles), rich_labels, color=PLOT["text"])
                ax.set_ylim(0, 1)
                ax.set_yticks([0.25, 0.5, 0.75])
                ax.set_yticklabels(["25%", "50%", "75%"], color=PLOT["muted"])
                ax.grid(color=PLOT["grid"], alpha=0.5)
                ax.spines["polar"].set_color(PLOT["border"])
                ax.set_title("HRV Profile (normalised)", pad=8, color=PLOT["text"])

            self._slots["radar"].update(draw_radar)
        except Exception as exc:
            log.warning("_plot_radar failed: %s", exc)

    def _plot_nonlinear(self, r: dict) -> None:
        """Poincaré plot and non-linear HRV metric table."""
        self._set_textbox(self.txt_nl, self._df_to_text(r["hrv_nonlin"]),
                          tsv=self._df_to_tsv(r["hrv_nonlin"]))

        rr_ms = r["rr_ms"]
        if len(rr_ms) < 2:
            return

        nl   = r["hrv_nonlin"]
        sd1  = self._safe_df_val(nl, "HRV_SD1", 1)
        sd2  = self._safe_df_val(nl, "HRV_SD2", 1)
        rr_a = rr_ms[:-1]
        rr_b = rr_ms[1:]
        lim  = [float(rr_ms.min()) - 20, float(rr_ms.max()) + 20]

        def draw_poincare(fig):
            ax = fig.add_subplot(111)
            style_axes(ax)
            ax.scatter(rr_a, rr_b, s=3, alpha=0.25, color=BLUE, rasterized=True)
            ax.plot(lim, lim, color=BORDER2, lw=1, ls="--", alpha=0.7)
            ax.set_xlim(lim)
            ax.set_ylim(lim)
            # Don't use set_aspect("equal") — it creates dead whitespace when
            # the container isn't square. Force equal axes via xlim/ylim instead.
            ax.set_xlabel("RR_n (ms)")
            ax.set_ylabel("RR_n+1 (ms)")
            ax.set_title(f"Poincaré diagram  SD1={sd1}  SD2={sd2}", loc="left")

        self._slots["poincare"].update(draw_poincare)


    def _plot_intervals_ecg(self, r: dict) -> None:
        """ECG beat strip annotated with P / Q / R / S / T landmarks.

        Design
        ------
        • X-axis is relative time from R peak (ms) — always centred at 0.
        • 3 beats are selected with the most complete wave annotation.
        • Plus a 4th "anatomy" reference panel on the right.
        • R_peak_s is read directly from the DataFrame (no index-mapping guesses).
        """
        ivl = r.get("intervals")
        sig = self._signal_flt
        fs  = self._fs

        slot = self._slots.get("intervals_ecg")
        if slot is None:
            return

        # Guard: need at least a signal and some delineated beats with R_peak_s
        has_data = (
            ivl is not None
            and not ivl.empty
            and sig is not None
            and fs is not None
            and "R_peak_s" in ivl.columns
        )

        if not has_data:
            def draw_hint(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                ax.text(0.5, 0.5, "Run Interval Delineation to see annotated beats",
                        ha="center", va="center", color=PLOT["muted"],
                        transform=ax.transAxes, fontsize=11)
                ax.set_axis_off()
            slot.update(draw_hint)
            return

        # has_data is True → narrow types for static checkers
        assert ivl is not None
        assert sig is not None
        assert fs is not None

        # ── Select up to 3 beats with the most complete wave annotation ──
        WAVE_POS_COLS = ["P_peak_s", "Q_peak_s", "S_peak_s", "T_peak_s"]
        available = [c for c in WAVE_POS_COLS if c in ivl.columns]

        if available:
            completeness = ivl[available].notna().sum(axis=1)
            # Sort by completeness desc, then by proximity to median RR
            med_rr = float(ivl["RR_ms"].median()) if "RR_ms" in ivl.columns else 100.0
            rr_dist = (ivl["RR_ms"] - med_rr).abs() if "RR_ms" in ivl.columns else pd.Series(0, index=ivl.index)
            sort_key = -completeness * 10 + rr_dist.fillna(999)
            best_order = sort_key.argsort()
            picks = list(best_order[:3])
        else:
            # No wave positions at all — just pick 3 beats near median RR
            picks = list(range(min(3, len(ivl))))

        if not picks:
            slot.update(lambda fig: None)
            return

        # ── Beat window: -PRE_MS to +POST_MS relative to R peak ────────
        PRE_MS  = 65    # ms before R  (covers P onset at ~-55ms)
        POST_MS = 110   # ms after  R  (covers T offset at ~+90ms in mice)

        row_data = ivl.iloc[picks]

        def draw_ecg_preview(fig):
            # Capture colors in local scope to ensure availability in draw function
            purple_col = PURPLE
            teal_col = TEAL
            orange_deep_col = ORANGE_DEEP
            
            # Colour / marker per wave landmark
            landmark_style = {
                "P_peak_s":   dict(color="#1A56DB", marker="^",  ms=8,  mew=1.0, label="P"),
                "Q_peak_s":   dict(color=purple_col, marker="v",  ms=8,  mew=1.0, label="Q"),
                "S_peak_s":   dict(color=purple_col, marker="v",  ms=8,  mew=1.0, label="S"),
                "J_peak_s":   dict(color=teal_col, marker="^",  ms=7,  mew=1.0, label="J"),
                "T_peak_s":   dict(color="#D84315", marker="^",  ms=8,  mew=1.0, label="T"),
            }

            # Interval spans: (start_col, end_col, colour, label)
            # None means use R=0 ms as boundary
            span_defs = [
                ("P_peak_s",  "Q_peak_s",  "#1A56DB", "P"),    # P peak → Q
                ("Q_peak_s",  "S_peak_s",  purple_col, "QRS"),  # QRS complex
                (None,        "T_peak_s",  orange_deep_col, "QT"),   # R → T peak
            ]
            # 3 beat columns + 1 anatomy column
            n_beats = len(picks)
            from matplotlib.gridspec import GridSpec
            ratios = [1] * n_beats + [0.55]
            gs = GridSpec(1, n_beats + 1, figure=fig,
                          width_ratios=ratios,
                          left=0.06, right=0.98, top=0.88, bottom=0.12,
                          wspace=0.30)
            axes = [fig.add_subplot(gs[0, i]) for i in range(n_beats + 1)]

            for ax_idx, (_, row) in enumerate(row_data.iterrows()):
                ax = axes[ax_idx]
                r_t = float(row.get("R_peak_s", float("nan")))
                if not np.isfinite(r_t):
                    ax.set_axis_off()
                    continue

                r_samp = int(round(r_t * fs))
                s0 = max(0, int((r_t - PRE_MS / 1000) * fs))
                s1 = min(len(sig), int((r_t + POST_MS / 1000) * fs))
                if s1 - s0 < 4:
                    ax.set_axis_off()
                    continue

                # Relative time axis (ms, 0 = R peak)
                t_rel = (np.arange(s0, s1) - r_samp) / fs * 1000

                style_axes(ax)
                ax.plot(t_rel, sig[s0:s1], color=PLOT["signal"], lw=1.3, zorder=3)
                ax.axvline(0, color=PLOT["rpeak_ok"], lw=1.0, ls="--", alpha=0.6)

                # R peak marker
                r_samp_clipped = min(max(r_samp, s0), s1 - 1)
                ax.plot(0, sig[r_samp_clipped],
                        marker="o", color=PLOT["rpeak_ok"], ms=8, zorder=7)

                # ── Wave landmark markers ────────────────────────────────
                for col, style in landmark_style.items():
                    if col not in row.index:
                        continue
                    wt_s = float(row[col]) if pd.notna(row[col]) else float("nan")
                    if not np.isfinite(wt_s):
                        continue
                    wt_rel = (wt_s - r_t) * 1000       # ms relative to R
                    if not (t_rel[0] - 1 <= wt_rel <= t_rel[-1] + 1):
                        continue
                    wt_smp = int(round(wt_s * fs))
                    wt_smp = min(max(wt_smp, s0), s1 - 1)
                    w_amp  = sig[wt_smp]
                    ax.plot(wt_rel, w_amp,
                            linestyle="none", color=style["color"],
                            marker=style["marker"], ms=style["ms"],
                            mew=style["mew"], zorder=6)
                    # Wave letter label close to the marker
                    va = "bottom" if style["marker"] == "^" else "top"
                    offset = 0.04 * (ax.get_ylim()[1] - ax.get_ylim()[0])
                    y_lbl  = w_amp + offset if va == "bottom" else w_amp - offset
                    letter = str(style["label"]).split("\n")[0]   # 'P', 'Q', 'S', 'T'
                    ax.text(wt_rel, y_lbl, letter,
                            ha="center", va=va, fontsize=8,
                            color=style["color"], fontweight="bold", zorder=8)

                # ── Interval spans (drawn in axes-fraction y to avoid ylim issues) ─
                ylo, yhi = ax.get_ylim()
                yspan    = yhi - ylo if yhi != ylo else 1.0

                def _span(c0, c1, fc, lab):
                    # x0 / x1 in relative ms
                    try:
                        x0 = (float(row[c0]) - r_t) * 1000 if (c0 and c0 in row.index and pd.notna(row[c0])) else 0.0
                        x1 = (float(row[c1]) - r_t) * 1000 if (c1 and c1 in row.index and pd.notna(row[c1])) else 0.0
                    except Exception as e:
                        log.warning("Interval float conversion failed at beat row: %s", e)
                        return
                    if not (np.isfinite(x0) and np.isfinite(x1)):
                        return
                    # For PR span: goes P_peak → R (x1=0)
                    # For QT span: goes R (x0=0) → T_peak
                    x_lo, x_hi = min(x0, x1), max(x0, x1)
                    if x_hi - x_lo < 0.5:   # skip degenerate spans
                        return
                    ax.axvspan(x_lo, x_hi, ymin=0, ymax=1,
                               color=fc, alpha=0.10, zorder=1)
                    ax.text((x_lo + x_hi) / 2, yhi - 0.04 * yspan,
                            lab, ha="center", va="top", fontsize=8,
                            color=fc, fontweight="bold", zorder=9,
                            bbox=dict(fc="white", ec="none", alpha=0.7, pad=1))

                # PR interval: P_peak → R (0)
                if "P_peak_s" in row.index and pd.notna(row["P_peak_s"]):
                    _span("P_peak_s", None, "#1B5E20", "PR")
                # QRS: Q → S
                if "Q_peak_s" in row.index and "S_peak_s" in row.index:
                    _span("Q_peak_s", "S_peak_s", purple_col, "QRS")
                # QT: R (0) → T peak
                if "T_peak_s" in row.index and pd.notna(row.get("T_peak_s", float("nan"))):
                    _span(None, "T_peak_s", orange_deep_col, "QT")

                # ── Title: measured interval values ──────────────────────
                parts = []
                for col_name, label in [("PR_ms","PR"), ("QRS_ms","QRS"),
                                         ("QT_ms","QT"), ("RR_ms","RR")]:
                    v = row.get(col_name, float("nan"))
                    if pd.notna(v) and np.isfinite(float(v)):
                        parts.append(f"{label} {float(v):.0f}")
                ax.set_title("  ".join(parts) + " ms" if parts else "Pas d'intervalle",
                             fontsize=8, color=PLOT["text"])
                ax.set_xlabel("ms / R", fontsize=7)
                ax.set_xlim(-PRE_MS, POST_MS)
                # Auto-scale y to signal content (not matplotlib default)
                seg_vis = sig[s0:s1]
                if len(seg_vis) > 0:
                    sv_lo = float(np.percentile(seg_vis, 2))
                    sv_hi = float(np.percentile(seg_vis, 98))
                    sv_pad = max((sv_hi - sv_lo) * 0.25, 0.05)
                    ax.set_ylim(sv_lo - sv_pad, sv_hi + sv_pad)
                if ax_idx == 0:
                    ax.set_ylabel("Amplitude", fontsize=7)

            # ── Anatomy reference panel ──────────────────────────────────
            ax_ref = axes[-1]
            ax_ref.set_facecolor(PLOT["axes"])
            ax_ref.set_xlim(0, 10)
            ax_ref.set_ylim(0, 10)
            ax_ref.set_axis_off()
            ax_ref.set_title("Mouse ECG\ncomplex", fontsize=9,
                             color=PLOT["text"], fontweight="bold")

            # Draw a schematic mouse ECG with J wave (early repolarisation hump)
            ecg_x = [0,  1.0, 1.5, 2.0,        # isoelectric
                     2.5, 2.8,                   # P wave
                     3.1, 3.4,                   # back to iso
                     3.7, 3.9,  4.0, 4.2, 4.4,  # Q / R / S
                     4.6, 4.9,                   # J hump (early repol.)
                     5.3, 5.8, 6.5, 7.0, 7.8, 8.2, 9.5, 10.0]  # T, return
            ecg_y = [5,  5.0, 5.2, 5.0,
                     5.5, 5.8,
                     5.5, 5.0,
                     4.7, 2.0, 9.0, 2.5, 4.7,   # Q-dip, R-peak, S-dip
                     5.1, 5.4,                   # J hump
                     5.1, 5.0, 5.7, 6.1, 5.7, 5.1, 5.0, 5.0]   # T wave

            ax_ref.plot(ecg_x, ecg_y, color=PLOT["signal"], lw=2.0, zorder=3)

            # Landmark annotations
            annotations = [
                (2.8, 6.1, "P",   "#1A56DB",  8),   # P peak
                (4.0, 9.2, "R",   "#1B5E20",  9),   # R peak
                (3.9, 1.6, "Q",   purple_col,  8),   # Q dip
                (4.4, 4.2, "S",   purple_col,  8),   # S dip
                (4.9, 5.7, "J",   teal_col,  8),   # J hump (new)
                (6.5, 6.4, "T",   "#D84315",  8),   # T peak
            ]
            for (x, y, lbl, col, fs_) in annotations:
                ax_ref.text(x, y, lbl, ha="center", va="center",
                            fontsize=fs_, color=col, fontweight="bold", zorder=6)

            # Bracket spans for PR / QRS / QT
            def _bracket(x0, x1, y, label, col):
                ax_ref.annotate("", xy=(x1, y), xytext=(x0, y),
                                arrowprops=dict(arrowstyle="<->", color=col, lw=1.2))
                ax_ref.text((x0 + x1) / 2, y - 0.5, label,
                            ha="center", va="top", fontsize=7,
                            color=col, fontweight="bold")

            _bracket(2.0, 4.0, 1.2, "PR",  "#1B5E20")
            _bracket(3.7, 4.4, 0.0, "QRS", purple_col)
            _bracket(4.0, 7.0, 2.1, "QT",  orange_deep_col)

            # Reference values text
            ref_lines = [
                ("PR",  "30–55 ms",  "#1B5E20"),
                ("QRS", " 8–25 ms",  purple_col),
                ("J",   "15–25 ms",  teal_col),
                ("QT",  "30–90 ms",  "#D84315"),
            ]
            for i, (lbl, val, col) in enumerate(ref_lines):
                ax_ref.text(0.1, 9.6 - i * 1.1, f"{lbl:4s} {val}",
                            ha="left", va="top", fontsize=7.5,
                            color=col, fontfamily="monospace")

            fig.suptitle("ECG beat annotation preview  —  best delineated beats",
                         fontsize=9, color=PLOT["muted"])

        slot.update(draw_ecg_preview)

    def _plot_intervals(self, r: dict) -> None:
        """Violin + box plot for PR / QRS / QT / QTc intervals."""
        ivl   = r["intervals"]
        if ivl is None or ivl.empty:
            def draw_unavailable_early(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                ax.text(0.5, 0.5, "Interval delineation not computed yet",
                        ha="center", va="center",
                        color=PLOT["muted"], transform=ax.transAxes)
                ax.set_axis_off()
            self._slots["intervals"].update(draw_unavailable_early)
            return
        cols  = [c for c in ["PR_ms", "QRS_ms", "QT_ms", "QTc_ms"]
                 if c in ivl.columns and ivl[c].notna().sum() > 3]

        if not cols:
            def draw_unavailable(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                ax.text(0.5, 0.5,
                        "Interval delineation not available\n"
                        "(requires clear P/Q/S/T waves at high SNR)",
                        ha="center", va="center",
                        color=PLOT["muted"], transform=ax.transAxes, linespacing=1.8)
                ax.set_xticks([])
                ax.set_yticks([])
                for sp in ax.spines.values():
                    sp.set_visible(False)
            self._slots["intervals"].update(draw_unavailable)
            return

        palette  = [BLUE_DARK, GREEN_DARK, PINK, ORANGE_DARK]
        col_data = [(col, ivl[col].dropna().values, color)
                    for col, color in zip(cols, palette)]

        # Reference ranges for drawing expected-value bands
        _ref = {
            "PR_ms":   MouseECG.PR_NORMAL,
            "QRS_ms":  MouseECG.QRS_NORMAL,
            "QT_ms":   MouseECG.QT_NORMAL,
            "QTc_ms":  MouseECG.QTC_NORMAL,
        }

        def draw_intervals(fig):
            from matplotlib.gridspec import GridSpec
            n_cols = len(col_data)
            # Use explicit GridSpec to avoid constrained_layout conflicts
            gs = GridSpec(1, n_cols, figure=fig,
                          left=0.10, right=0.97, top=0.88, bottom=0.08,
                          wspace=0.40)
            for ci, (col, data, color) in enumerate(col_data):
                ax = fig.add_subplot(gs[0, ci])
                style_axes(ax)
                finite = data[np.isfinite(data)] if len(data) >= 2 else np.array([])
                if len(finite) < 2:
                    ax.text(0.5, 0.5, f"{col.replace('_ms','')}\nn<2",
                            ha="center", va="center", color=PLOT["muted"],
                            transform=ax.transAxes, fontsize=9)
                    ax.set_axis_off()
                    continue
                # Violin
                try:
                    vp = ax.violinplot(finite, positions=[0], widths=0.7,
                                       showmedians=False, showextrema=False)
                    for body in vp["bodies"]:
                        body.set_facecolor(color)
                        body.set_alpha(0.25)
                        body.set_edgecolor(color)
                        body.set_linewidth(0.8)
                except Exception:
                    pass
                # Box on top of violin
                ax.boxplot(finite, positions=[0], widths=0.18, patch_artist=True,
                           boxprops=dict(facecolor=color, alpha=0.35, linewidth=0.8),
                           medianprops=dict(color="white", lw=2.0),
                           whiskerprops=dict(color=color, lw=1.0, alpha=0.7),
                           capprops=dict(color=color, lw=1.0),
                           flierprops=dict(marker=".", color=MUTED, ms=2, alpha=0.4))
                # Reference range band
                lo_ref, hi_ref = _ref.get(col, (None, None))
                if lo_ref is not None:
                    ax.axhspan(lo_ref, hi_ref, color=color, alpha=0.08, zorder=0)
                    ax.axhline(lo_ref, color=color, lw=0.7, ls=":", alpha=0.5)
                    ax.axhline(hi_ref, color=color, lw=0.7, ls=":", alpha=0.5)
                # Stats annotation (compact)
                med = float(np.median(finite))
                ax.text(0.5, 0.97,
                        f"med {med:.1f}  ±{finite.std():.1f}",
                        ha="center", va="top", color=PLOT["muted"], fontsize=8,
                        transform=ax.transAxes,
                        bbox=dict(boxstyle="round,pad=0.2",
                                  fc=PLOT["axes"], ec="none", alpha=0.8))
                label = col.replace("_ms", "")
                ax.set_title(label, color=color, fontsize=10, pad=4)
                if ci == 0:
                    ax.set_ylabel("ms", fontsize=8)
                ax.set_xticks([])
                # Auto-range with padding
                p2, p98 = np.percentile(finite, [2, 98])
                pad = max((p98 - p2) * 0.25, 5)
                ax.set_ylim(p2 - pad, p98 + pad)

        self._slots["intervals"].update(draw_intervals)
        # Only describe the interval measurement columns, not wave-position columns
        ivl_stats = ivl[["RR_ms", "PR_ms", "QRS_ms", "QT_ms", "QTc_ms"]].copy()
        ivl_stats = ivl_stats[[c for c in ivl_stats.columns if c in ivl.columns]]
        _ivl_desc = ivl_stats.describe().round(2)
        # self._set_textbox(self.txt_ivl, _ivl_desc.to_string(),  # Attribut supprimé
        #                   tsv=self._describe_to_tsv(_ivl_desc))

    def _plot_beat_template(self, r: dict) -> None:
        """Average beat template, ±1 SD band, and amplitude / morphology distributions.

        All heavy numpy work (beat matrix, SD, per-beat correlations) was pre-computed
        in run_full_analysis() on the background thread.  This function only renders.
        """
        beat_time   = r.get("beat_time")
        mean_beat   = r.get("beat_template")
        beat_matrix = r.get("beat_matrix")
        beat_sd     = r.get("beat_sd")
        beat_corr   = r.get("beat_corr")
        peak_amps   = r.get("peak_amps")

        if beat_time is None or mean_beat is None or beat_matrix is None:
            return

        n_beats = len(beat_matrix)
        if n_beats < 4:
            log.warning("_plot_beat_template: only %d valid beats — skipping", n_beats)
            return

        stride = max(1, n_beats // 60)  # show at most ~60 individual ghost traces

        wt = self._wave_template

        def draw_template(fig):
            ax = fig.add_subplot(111)
            style_axes(ax)
            # Ghost traces (subsampled for performance)
            for beat in beat_matrix[::stride]:
                ax.plot(beat_time, beat, color=PLOT["grid"], lw=0.3, alpha=0.3)
            if beat_sd is not None:
                ax.fill_between(beat_time, mean_beat - beat_sd, mean_beat + beat_sd,
                                color=BLUE, alpha=0.2, label="±1 SD")
            ax.plot(beat_time, mean_beat, color=BLUE, lw=2.0, label="Mean beat")
            ax.axvline(0, color=RED, lw=1.2, ls="--", alpha=0.8, label="R peak")

            # ── Overlay wave template landmarks if confirmed ───────────────
            if wt is not None:
                for wkey, (center_ms, half_ms) in wt.landmarks.items():
                    col  = WaveTemplateEditor.WAVE_COLORS.get(wkey, GRAY)
                    name = WaveTemplateEditor.WAVE_SHORT_LABELS.get(wkey, wkey)
                    ax.axvline(center_ms, color=col, lw=1.4, ls=":",
                               alpha=0.85, zorder=6)
                    ax.axvspan(center_ms - half_ms, center_ms + half_ms,
                               alpha=0.08, color=col, zorder=0)
                    # Label at the bottom of the axis
                    ax.text(center_ms, 0.02, name,
                            transform=ax.get_xaxis_transform(),
                            ha="center", va="bottom",
                            color=col, fontsize=9, fontweight="bold", zorder=7)
                src_note = f"  ·  template: {wt.source}" if wt is not None else ""
            else:
                src_note = ""

            ax.set_xlabel("Time relative to R peak (ms)")
            ax.set_ylabel("Amplitude (norm.)")
            ax.set_title(f"Mean template  (n={n_beats} beats){src_note}", loc="left")
            ax.legend(framealpha=0, loc="upper right")

        self._slots["beat"].update(draw_template)

        if beat_corr is None or peak_amps is None:
            return
        mean_corr = float(np.nanmean(beat_corr))

        n_bad = int(np.sum(beat_corr < 0.90)) if beat_corr is not None else 0

        def draw_distributions(fig):
            ax_amp, ax_corr = fig.subplots(1, 2)
            for ax in (ax_amp, ax_corr):
                style_axes(ax)
            ax_amp.hist(peak_amps, bins=min(50, max(10, n_beats // 4)),
                        color=BLUE, alpha=0.75, edgecolor="none")
            ax_amp.set_xlabel("R-peak amplitude (norm.)")
            ax_amp.set_ylabel("Count")
            ax_amp.set_title("Amplitude des peaks R", loc="left")

            bins_corr = min(40, max(10, n_beats // 4))
            ax_corr.hist(beat_corr, bins=bins_corr,
                         color=GREEN, alpha=0.75, edgecolor="none")
            ax_corr.axvline(mean_corr, color=RED, lw=1.5, ls="--",
                            label=f"mean={mean_corr:.3f}")
            ax_corr.axvline(0.90, color=ORANGE, lw=1.0, ls=":",
                            label="0.90 threshold")
            ax_corr.set_xlabel("Correlation with template")
            ax_corr.set_ylabel("Count")
            title = f"Beat Morphology  ({n_bad} beats < 0.90)"
            ax_corr.set_title(title, loc="left",
                              color=(ORANGE if n_bad > n_beats * 0.1 else PLOT["text"]))
            ax_corr.legend(framealpha=0)

        self._slots["beat_dist"].update(draw_distributions)

    def _plot_summary(self, r: dict) -> None:
        """Populate the Summary tab: KPI cards, all plots, and text report."""
        hr  = r["hr"]
        td  = r["hrv_time"]
        fd  = r["hrv_freq"]
        nl  = r["hrv_nonlin"]
        ivl = r["intervals"]
        val = self._safe_df_val   # shorthand

        # Default values for metrics computed later (referenced in report text)
        porta_pct: float = float("nan")
        guzik_pct: float = float("nan")
        n_dec: int = 0; n_acc: int = 0; n_tot: int = 0
        qt_disp: float = float("nan")
        _qt: "np.ndarray" = np.array([])

        # ── KPI cards ─────────────────────────────────────────────────────────
        def _kpi(key: str, text: str) -> None:
            lbl = self._sum_kpi_vals.get(key)
            if lbl is not None:
                lbl.configure(text=text)

        rdf = r.get("rr_df")
        if rdf is not None and not rdf.empty:
            _kpi("hr_mean", f"{rdf['HR_bpm'].mean():.0f}")
            _kpi("hr_min",  f"{rdf['HR_bpm'].quantile(0.02):.0f}")
            _kpi("hr_max",  f"{rdf['HR_bpm'].quantile(0.98):.0f}")

        _kpi("sdnn",  val(td, "HRV_SDNN",  1))
        _kpi("rmssd", val(td, "HRV_RMSSD", 1))
        _kpi("pnn6",  val(td, "HRV_pNN6",  1))

        try:
            lfhf = float(fd["HRV_LFHF"].values[0]) if (fd is not None and "HRV_LFHF" in fd.columns) else float("nan")
        except Exception:
            lfhf = float("nan")
        _kpi("lf_hf", f"{lfhf:.2f}" if np.isfinite(lfhf) else "—")

        if ivl is not None and not ivl.empty:
            for col, key in [("PR_ms", "pr"), ("QRS_ms", "qrs"), ("QTc_ms", "qtc")]:
                if col in ivl.columns:
                    d = ivl[col].dropna()
                    if len(d):
                        _kpi(key, f"{d.median():.0f}")

        # ── Mirror existing draw_fn into summary slots ─────────────────────────
        # Each primary slot holds a fully-bound draw closure from the analysis
        # pipeline. Replaying it into the summary slot renders the same plot.
        _MIRRORS = [
            ("rr",            "sum_rr"),
            ("rr_hist",       "sum_rr_hist"),
            ("psd",           "sum_psd"),
            ("radar",         "sum_radar"),
            ("poincare",      "sum_poincare"),
            ("beat",          "sum_beat"),
            ("beat_dist",     "sum_beat_dist"),
            ("intervals",     "sum_intervals"),
            ("intervals_ecg", "sum_intervals_ecg"),
            ("rolling_hrv",   "sum_rolling"),
        ]
        for src_key, dst_key in _MIRRORS:
            src = self._slots.get(src_key)
            dst = self._slots.get(dst_key)
            if src is None or dst is None:
                continue
            fn = getattr(src, "_draw_fn", None)
            if fn is not None:
                try:
                    dst.update(fn)
                except Exception as exc:
                    log.debug("sum mirror %s→%s: %s", src_key, dst_key, exc)

        # ── sum_rr_extra: FC distribution + qualité morphologique ─────────────
        rr_ms     = r.get("rr_ms", np.array([]))
        beat_corr = r.get("beat_corr")
        if len(rr_ms) > 4:
            hr_bpm = 60_000.0 / np.clip(rr_ms, 1, None)

            def draw_rr_extra(fig):
                axes = fig.subplots(1, 2)
                style_axes(axes[0]); style_axes(axes[1])
                bins = min(60, max(10, len(hr_bpm) // 40))
                axes[0].hist(hr_bpm, bins=bins, color=ORANGE_DARK, alpha=0.72, edgecolor="none")
                med_hr = float(np.median(hr_bpm))
                axes[0].axvline(med_hr, color=ORANGE_DARK, lw=1.6, ls="--",
                                label=f"méd. {med_hr:.0f}")
                axes[0].set_xlabel("HR (bpm)")
                axes[0].set_ylabel("Beats")
                axes[0].set_title("HR Distribution", loc="left")
                axes[0].legend(framealpha=0, fontsize=8)
                if beat_corr is not None and len(beat_corr) > 4:
                    axes[1].hist(beat_corr,
                                 bins=min(40, max(8, len(beat_corr) // 40)),
                                 color=RED, alpha=0.70, edgecolor="none")
                    axes[1].axvline(0.90, color=ORANGE, lw=1.2, ls=":",
                                    label="threshold 0.90")
                    mc = float(np.nanmean(beat_corr))
                    axes[1].axvline(mc, color=RED, lw=1.6, ls="--",
                                    label=f"moy. {mc:.3f}")
                    axes[1].set_xlabel("Correlation to template")
                    axes[1].set_ylabel("Beats")
                    axes[1].set_title("Morphological quality", loc="left")
                    axes[1].legend(framealpha=0, fontsize=8)
                else:
                    axes[1].text(0.5, 0.5, "Template not computed",
                                 ha="center", va="center",
                                 color=PLOT["muted"], transform=axes[1].transAxes)
                    axes[1].set_axis_off()

            dst = self._slots.get("sum_rr_extra")
            if dst is not None:
                dst.update(draw_rr_extra)

        # ── sum_asymmetry: Asymétrie RR (Porta / Guzik index) ─────────────────
        # Porta index P0: fraction of beats with RR_n+1 < RR_n  (sym → 50 %)
        # Guzik index GI: contribution of decelerations to total variation
        # Both are markers of autonomic nervous system balance asymmetry.
        if len(rr_ms) > 8:
            rr_a = rr_ms[:-1].astype(float)
            rr_b = rr_ms[1:].astype(float)
            diff = rr_b - rr_a
            n_dec = int(np.sum(diff > 0))    # decelerations (RR lengthens)
            n_acc = int(np.sum(diff < 0))    # accelerations (RR shortens)
            n_tot = len(diff)
            porta_raw  = n_acc / n_tot if n_tot else 0.5
            guzik_num  = float(np.sum(diff[diff > 0] ** 2))
            guzik_den  = float(np.sum(diff ** 2))
            guzik_raw  = (1.0 - guzik_num / guzik_den) if guzik_den > 0 else 0.5
            porta_pct  = porta_raw  * 100
            guzik_pct  = guzik_raw  * 100
            # ΔRR histogram (signed)
            drr_clip = np.clip(diff, -60, 60)

            _porta_pct  = porta_pct
            _guzik_pct  = guzik_pct
            _n_dec = n_dec; _n_acc = n_acc; _n_tot = n_tot
            _drr_clip = drr_clip

            def draw_asymmetry(fig):
                ax_bar, ax_hist = fig.subplots(1, 2)
                style_axes(ax_bar); style_axes(ax_hist)

                # Bar chart: proportions
                cats  = ["Decelerations\n(RR↑)", "Neutral", "Accelerations\n(RR↓)"]
                n_neu = _n_tot - _n_dec - _n_acc
                vals2 = [_n_dec / _n_tot * 100,
                         n_neu  / _n_tot * 100,
                         _n_acc / _n_tot * 100]
                colors2 = ["#C62828", "#455A64", BLUE_DARK]
                bars = ax_bar.bar(cats, vals2, color=colors2, alpha=0.80, width=0.5)
                ax_bar.axhline(50, color=BORDER2, lw=1.2, ls="--", alpha=0.7)
                ax_bar.set_ylabel("% beats")
                ax_bar.set_ylim(0, 105)
                for bar, v in zip(bars, vals2):
                    ax_bar.text(bar.get_x() + bar.get_width() / 2,
                                v + 1.5, f"{v:.1f}%",
                                ha="center", va="bottom", fontsize=8, color=PLOT["text"])
                porta_str = f"Porta={_porta_pct:.1f}%  Guzik={_guzik_pct:.1f}%"
                ax_bar.set_title(f"RR Asymmetry  ({porta_str})", loc="left", fontsize=8)
                ax_bar.tick_params(axis="x", labelsize=8)

                # ΔRR histogram
                ax_hist.hist(_drr_clip[_drr_clip < 0], bins=30,
                             color=BLUE_DARK, alpha=0.70, label="accelerations")
                ax_hist.hist(_drr_clip[_drr_clip > 0], bins=30,
                             color="#C62828", alpha=0.70, label="decelerations")
                ax_hist.axvline(0, color=BORDER2, lw=1.2, ls="--")
                ax_hist.set_xlabel("ΔRR (ms)")
                ax_hist.set_ylabel("Beats")
                ax_hist.set_title("ΔRR Distribution", loc="left", fontsize=8)
                ax_hist.legend(framealpha=0, fontsize=8)

            dst = self._slots.get("sum_asymmetry")
            if dst is not None:
                dst.update(draw_asymmetry)
            # KPI
            _kpi("porta", f"{porta_pct:.0f}")

        # ── sum_quality_time: Qualité morphologique dans le temps ──────────────
        if beat_corr is not None and len(beat_corr) > 4 and self._rpeaks_ok is not None:
            _bc   = np.asarray(beat_corr, dtype=float)
            _wp = self._windowed_peaks()
            _rp   = (_wp if _wp is not None else self._rpeaks_ok).astype(float) / self._fs   # peak times (windowed)
            # Align: beat_corr has 1 value per accepted beat, first peak has no interval
            _t_bc = _rp[:len(_bc)] if len(_rp) >= len(_bc) else _rp

            # Rolling 50-beat mean
            _win  = min(50, max(10, len(_bc) // 20))
            _kern = np.ones(_win) / _win
            _roll = np.convolve(_bc, _kern, mode="valid")
            _roll_t = _t_bc[_win - 1: _win - 1 + len(_roll)]

            n_pts = min(len(_t_bc), len(_bc))
            _bc_t = _t_bc[:n_pts]
            _bc_v = _bc[:n_pts]

            def draw_quality_time(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                # Scatter individual beats, coloured by quality
                if len(_bc_t) != len(_bc_v):
                    log.debug("quality plot length mismatch: %d vs %d",
                              len(_bc_t), len(_bc_v))
                sc = ax.scatter(_bc_t, _bc_v, s=2, c=_bc_v, cmap="RdYlGn",
                                vmin=0.7, vmax=1.0, alpha=0.35, rasterized=True, zorder=2)
                ax.plot(_roll_t, _roll, color=BLUE, lw=1.8, zorder=3,
                        label=f"rolling mean (n={_win})")
                ax.axhline(0.90, color=ORANGE, lw=1.0, ls=":", alpha=0.8,
                           label="threshold 0.90")
                ax.set_xlabel("Time (s)")
                ax.set_ylabel("Correlation to template")
                ax.set_title("Morphological quality over time", loc="left")
                ax.set_ylim(max(0, float(np.nanmin(_bc_v)) - 0.05), 1.02)
                ax.legend(framealpha=0, fontsize=8, loc="lower right")
                try:
                    fig.colorbar(sc, ax=ax, fraction=0.025, pad=0.02,
                                 label="corrélation")
                except Exception as e:
                    log.debug("colorbar render failed: %s", e)

            dst = self._slots.get("sum_quality_time")
            if dst is not None:
                dst.update(draw_quality_time)

        # ── sum_qt_disp: Dispersion QT + relation QT/RR ────────────────────────
        if ivl is not None and not ivl.empty:
            _qt   = ivl["QT_ms"].dropna().values.astype(float)  if "QT_ms"  in ivl.columns else np.array([])
            _qtc  = ivl["QTc_ms"].dropna().values.astype(float) if "QTc_ms" in ivl.columns else np.array([])
            _rr_ivl = ivl["RR_ms"].dropna().values.astype(float) if "RR_ms" in ivl.columns else np.array([])

            qt_disp = float(np.nanmax(_qt) - np.nanmin(_qt)) if len(_qt) > 3 else float("nan")
            if np.isfinite(qt_disp):
                _kpi("qt_disp", f"{qt_disp:.0f}")

            if len(_qt) > 10 and len(_rr_ivl) > 10:
                _n_min = min(len(_qt), len(_rr_ivl))
                _qt_al = _qt[:_n_min]
                _rr_al = _rr_ivl[:_n_min]
                # Simple linear regression QT ~ RR
                try:
                    _mask = np.isfinite(_qt_al) & np.isfinite(_rr_al)
                    _rr_f = _rr_al[_mask]; _qt_f = _qt_al[_mask]
                    _coef = np.polyfit(_rr_f, _qt_f, 1)
                    _fit_rr = np.linspace(_rr_f.min(), _rr_f.max(), 100)
                    _fit_qt = np.polyval(_coef, _fit_rr)
                    _slope  = _coef[0]
                    _has_fit = True
                except Exception:
                    _has_fit = False
                    _rr_f = _qt_f = np.array([])
                    _fit_rr = _fit_qt = _slope = None

                _qt_disp_val = qt_disp
                _qtc_al = _qtc[:_n_min] if len(_qtc) >= _n_min else _qtc

                def draw_qt_disp(fig):
                    ax_disp, ax_rel = fig.subplots(1, 2)
                    style_axes(ax_disp); style_axes(ax_rel)

                    # Left: QT & QTc distribution with dispersion band
                    if len(_qt_al) > 3:
                        p2_qt,  p98_qt  = np.percentile(_qt_al,  [2, 98])
                        ax_disp.violinplot(_qt_al,  positions=[0], widths=0.6,
                                           showmedians=False, showextrema=False)
                        ax_disp.boxplot(_qt_al, positions=[0], widths=0.18,
                                        patch_artist=True,
                                        boxprops=dict(facecolor=PINK, alpha=0.3, lw=0.7),
                                        medianprops=dict(color="white", lw=2),
                                        whiskerprops=dict(color=PINK, lw=0.8),
                                        capprops=dict(color=PINK, lw=0.8),
                                        flierprops=dict(marker=".", ms=2, color=MUTED, alpha=0.4))
                    if len(_qtc_al) > 3:
                        ax_disp.violinplot(_qtc_al, positions=[1], widths=0.6,
                                           showmedians=False, showextrema=False)
                        ax_disp.boxplot(_qtc_al, positions=[1], widths=0.18,
                                        patch_artist=True,
                                        boxprops=dict(facecolor=AMBER_DARK, alpha=0.3, lw=0.7),
                                        medianprops=dict(color="white", lw=2),
                                        whiskerprops=dict(color=AMBER_DARK, lw=0.8),
                                        capprops=dict(color=AMBER_DARK, lw=0.8),
                                        flierprops=dict(marker=".", ms=2, color=MUTED, alpha=0.4))
                    ax_disp.set_xticks([0, 1])
                    ax_disp.set_xticklabels(["QT (ms)", "QTc (ms)"], fontsize=9)
                    ax_disp.set_ylabel("ms")
                    disp_str = f"disp. {_qt_disp_val:.0f} ms" if np.isfinite(_qt_disp_val) else ""
                    ax_disp.set_title(f"QT Variability  {disp_str}", loc="left", fontsize=8)

                    # Right: QT vs RR scatter + regression line
                    ax_rel.scatter(_rr_f, _qt_f, s=2, alpha=0.25,
                                   color=PINK, rasterized=True, zorder=2)
                    if _has_fit and _fit_rr is not None:
                        ax_rel.plot(_fit_rr, _fit_qt, color=ORANGE, lw=1.8, zorder=3,
                                    label=f"pente {_slope:.3f}")
                        ax_rel.legend(framealpha=0, fontsize=8)
                    ax_rel.set_xlabel("RR (ms)")
                    ax_rel.set_ylabel("QT (ms)")
                    ax_rel.set_title("QT/RR relationship (dynamic correction)", loc="left", fontsize=8)

                dst = self._slots.get("sum_qt_disp")
                if dst is not None:
                    dst.update(draw_qt_disp)

        # ── Texte du rapport ─────────────────────────────────────────────────
        filter_note = "  ⚠ Signal brut (sans filtres)" if self._no_filter_mode else "  Bandpass + notch + NK clean"
        arep = self._artifact_report
        if arep:
            removed  = arep["n_in"] - arep["n_out"]
            art_lines = [
                "", "  ARTIFACT CORRECTION",
                f"    Before        {arep['n_in']}  beats",
                f"    After         {arep['n_out']}  beats",
                f"    Removed       {removed}",
                f"      Non-physio  {arep['n_nonphysio']}",
                f"      Ectopic     {arep['n_ectopic']}",
                f"      Duplicates  {arep['n_duplicate']}",
            ]
        else:
            art_lines = ["", "  ARTIFACT CORRECTION", "    Not applied"]

        # ── asymmetry metrics for report ──────────────────────────────────────
        asym_lines: list[str] = []
        if len(rr_ms) > 8:
            asym_lines = [
                "", "  RR ASYMMETRY (autonomic system)",
                f"    Porta index (acc.)   {porta_pct:.1f} %  (symétrie → 50 %)",
                f"    Guzik index  (acc.)  {guzik_pct:.1f} %",
                f"    Décélérations        {n_dec} / {n_tot}",
                f"    Accélérations        {n_acc} / {n_tot}",
            ]

        lines = [
            "═" * 62,
            "  ECG ANALYSIS  —  Summary Report",
            f"  Subject  :  {self.ent_subject.get()}",
            f"  Date     :  {datetime.now():%Y-%m-%d  %H:%M}",
            f"  File     :  {os.path.basename(self._filepath or '')}",
            f"  Filters  :{filter_note}",
            "═" * 62, "",
            "  HEART RATE",
            f"    Moyenne          {hr['mean']:.1f} bpm",
            f"    Min  (2e %ile)   {hr['min']:.1f} bpm",
            f"    Max  (98e %ile)  {hr['max']:.1f} bpm",
            f"    SD               {hr['std']:.2f} bpm",
            f"    N battements     {hr['n']}", "",
            "  HRV — TIME DOMAIN",
            f"    MeanNN   {val(td, 'HRV_MeanNN')} ms",
            f"    SDNN     {val(td, 'HRV_SDNN')} ms",
            f"    RMSSD    {val(td, 'HRV_RMSSD')} ms",
            f"    pNN6     {val(td, 'HRV_pNN6')} %  (>{MouseECG.PNN_THRESHOLD} ms)",
            f"    pNN20    {val(td, 'HRV_pNN20')} %", "",
            "  HRV — FREQUENCY DOMAIN",
            f"    VLF      {val(fd, 'HRV_VLF')} n.u.",
            f"    LF       {val(fd, 'HRV_LF')} n.u.",
            f"    HF       {val(fd, 'HRV_HF')} n.u.",
            f"    LF/HF    {val(fd, 'HRV_LFHF')}", "",
            "  HRV — NON-LINEAR",
            f"    SD1      {val(nl, 'HRV_SD1')} ms",
            f"    SD2      {val(nl, 'HRV_SD2')} ms",
            f"    SampEn   {val(nl, 'HRV_SampEn')}",
            f"    ApEn     {val(nl, 'HRV_ApEn')}",
            f"    DFA α1   {val(nl, 'HRV_DFA_alpha1')}",
            f"    DFA α2   {val(nl, 'HRV_DFA_alpha2')}",
        ]
        if ivl is not None and not ivl.empty and "QT_ms" in ivl.columns:
            lines += ["", "  ECG INTERVALS  (median ± SD)"]
            for col in ["PR_ms", "QRS_ms", "QT_ms", "QTc_ms"]:
                if col in ivl.columns:
                    data = ivl[col].dropna()
                    if len(data):
                        lines.append(
                            f"    {col:<16} {data.median():.1f} ± {data.std():.1f} ms")
            if len(_qt) > 3:
                lines.append(f"    QT dispersion    {qt_disp:.1f} ms")
        lines += asym_lines
        lines += art_lines
        lines += ["", "═" * 62]
        self._set_textbox(self.txt_sum, "\n".join(lines))

    # ── KPI bar update ────────────────────────────────────────

    def _reset_result_plots(self) -> None:
        """Clear stored draw_fn on every result-plot slot.

        Prevents stale draw functions from a previous file replaying
        on window resize after a new file is loaded.
        """
        result_slots = (
            "rr", "rr_hist",
            "poincare", "psd", "radar",
            "intervals", "intervals_ecg",
            "beat", "beat_dist",
            "epochs", "rolling_hrv",
            "arr_detail",
            # Summary mirrors
            "sum_rr", "sum_rr_hist", "sum_rr_extra",
            "sum_psd", "sum_radar", "sum_poincare",
            "sum_beat", "sum_beat_dist",
            "sum_intervals", "sum_intervals_ecg",
            "sum_rolling",
            # New summary panels
            "sum_asymmetry", "sum_quality_time", "sum_qt_disp",
        )
        for name in result_slots:
            slot = self._slots.get(name)
            if slot is not None:
                slot._draw_fn = None
                slot._show_placeholder()

    def _reset_tab_status_labels(self) -> None:
        """Reset per-tab status labels and disable action buttons.

        Called on new file load so labels from the previous analysis
        (e.g. "Done LF=42%") don't persist after loading a new file.
        """
        _neutral = "  Run Core Analysis first"
        if self.lbl_freq_status is not None:
            self.lbl_freq_status.configure(text=_neutral, text_color=PLOT["muted"])  # type: ignore[union-attr]
        if self.lbl_nonlin_status is not None:
            self.lbl_nonlin_status.configure(text=_neutral, text_color=PLOT["muted"])  # type: ignore[union-attr]
        lbl_ivl = getattr(self, "lbl_ivl_status", None)
        if lbl_ivl is not None:
            lbl_ivl.configure(text=_neutral, text_color=PLOT["muted"])
        lbl_arr = getattr(self, "lbl_arrhythmia_status", None)
        if lbl_arr is not None:
            lbl_arr.configure(text="  Run Core Analysis first", text_color=PLOT["muted"])
        lbl_roll = getattr(self, "lbl_roll_status", None)
        if lbl_roll is not None:
            lbl_roll.configure(text="  Run Core Analysis first", text_color=PLOT["muted"])
        for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl", "btn_run_arrhythmia"):
            btn = getattr(self, btn_attr, None)
            if btn is not None:
                btn.configure(state="disabled")

    def _reset_kpis(self) -> None:
        """Reset all KPI labels to dash when results are invalidated."""
        for key in ("hr_mean", "hr_range", "rr_mean", "n_beats",
                    "sdnn", "rmssd", "pnn50", "dur"):
            widget = self._kpi.get(key)
            if widget is not None:
                widget.configure(text="--")

    def _update_kpis(self) -> None:
        if self._results is None:
            return
        r   = self._results
        hr  = r["hr"]
        rdf = r["rr_df"]
        td  = r["hrv_time"]

        def hrv_val(key: str) -> str:
            try:
                return f"{float(td[key].values[0]):.1f}"
            except Exception:
                return "—"

        self._kpi["hr_mean"].configure(text=f"{hr['mean']:.0f} bpm")
        self._kpi["hr_range"].configure(text=f"{hr['min']:.0f}–{hr['max']:.0f}")
        try:
            rr_mean = float(np.nanmean(r["rr_ms"]))
            self._kpi["rr_mean"].configure(text=f"{rr_mean:.0f} ms")
        except Exception:
            self._kpi["rr_mean"].configure(text="—")
        n_valid = hr.get("n_valid", hr["n"])
        self._kpi["n_beats"].configure(text=str(n_valid))
        self._kpi["sdnn"].configure(text=hrv_val("HRV_SDNN"))
        self._kpi["rmssd"].configure(text=hrv_val("HRV_RMSSD"))
        self._kpi["pnn50"].configure(text=hrv_val("HRV_pNN6"))
        try:
            self._kpi["dur"].configure(text=f"{rdf['Time_s'].iloc[-1]:.0f} s")
        except Exception:
            self._kpi["dur"].configure(text="—")

    # ════════════════════════════════════════════════════════
    #  EPOCH ANALYSIS
    # ════════════════════════════════════════════════════════

    def _compute_epochs(self) -> None:
        """Compute epoch-level HRV in a background thread to keep the UI responsive.

        nk.hrv_time() is called once per epoch.  On long recordings with many
        short epochs this adds up to several seconds — enough to freeze the UI
        noticeably.  The calculation is therefore moved to a daemon thread via
        _start_async_result, exactly like _run_freq / _run_nonlinear.
        """
        if self._rpeaks_ok is None or len(self._rpeaks_ok) < 10:
            messagebox.showwarning("No data", "Run Preview Detection first.")
            return
        if self._time is None:
            return

        epoch_s   = max(MouseECG.EPOCH_MIN_S,
                        self._safe_float(self.ent_epoch, MouseECG.EPOCH_DEFAULT_S))
        overlap_s = max(0.0, self._safe_float(self.ent_overlap, 0.0))
        fs        = self._fs
        rp        = self._windowed_peaks()   # respects analysis window
        if rp is None or len(rp) < 10:
            messagebox.showwarning("No data", "Not enough peaks in the analysis window.")
            return
        # Duration from windowed peaks rather than full signal
        t_peaks = rp / fs
        dur     = float(t_peaks[-1] - t_peaks[0])

        if overlap_s >= epoch_s:
            messagebox.showwarning("Bad overlap",
                f"Overlap ({overlap_s:.0f}s) must be less than epoch ({epoch_s:.0f}s).")
            return
        step      = max(1.0, epoch_s - overlap_s)
        t_win_start = float(t_peaks[0])    # absolute start of the windowed range
        starts    = np.arange(t_win_start, t_win_start + dur - epoch_s + step * 0.5, step)
        if len(starts) < 2:
            messagebox.showwarning(
                "Too few epochs",
                f"Recording too short for {epoch_s:.0f}s epochs. "
                f"Try a shorter epoch (e.g. {int(dur // 3)}s).")
            return

        def _worker():
            """Runs on background thread — no Tkinter access allowed."""
            rows = []
            n_ep = len(starts)
            for idx_ep, t0 in enumerate(starts):
                t1    = t0 + epoch_s
                ep_rp = rp[(rp / fs >= t0) & (rp / fs < t1)]
                if len(ep_rp) < 5:
                    continue
                rr = np.diff(ep_rp).astype(float) / fs * 1000
                try:
                    assert nk is not None  # NK_AVAILABLE checked by caller
                    hrv_ep = nk.hrv_time(ep_rp, sampling_rate=int(fs), show=False)
                    sdnn   = float(hrv_ep["HRV_SDNN"].values[0])
                    rmssd  = float(hrv_ep["HRV_RMSSD"].values[0])
                except Exception as exc:
                    log.warning("Epoch hrv_time failed, manual calc: %s", exc)
                    sdnn  = float(rr.std())
                    rmssd = float(np.sqrt(np.mean(np.diff(rr) ** 2)))
                rows.append({
                    "Epoch_start_s": round(t0, 1),
                    "Epoch_end_s":   round(t1, 1),
                    "N_beats":       len(ep_rp),
                    "HR_mean":       round(float(60_000 / rr.mean()), 1),
                    "MeanNN":        round(float(rr.mean()), 1),
                    "SDNN":          round(sdnn, 2),
                    "RMSSD":         round(rmssd, 2),
                })
                # Progress via after() — safe cross-thread call, throttled to every 5%
                pct = int((idx_ep + 1) / max(n_ep, 1) * 100)
                if idx_ep % max(1, n_ep // 20) == 0:
                    self.after(0, lambda p=pct, e=idx_ep+1, tot=n_ep:
                               self._set_progress(p, f"Epoch {e}/{tot}…"))
            return rows

        def _done(rows):
            if not rows:
                messagebox.showwarning("No epochs", "No valid epochs found.")
                return
            df    = pd.DataFrame(rows)
            self._epoch_df = df
            t_mid = (df["Epoch_start_s"] + df["Epoch_end_s"]) / 2

            plot_specs = [
                ("HR_mean", "HR (bpm)",   ORANGE_DARK, "Heart Rate per Epoch"),
                ("SDNN",    "SDNN (ms)",  BLUE_DARK, "SDNN per Epoch"),
                ("RMSSD",   "RMSSD (ms)", GREEN_DARK, "RMSSD per Epoch"),
            ]

            def draw_epochs(fig):
                axes = fig.subplots(3, 1, sharex=True)
                for ax, (col, ylabel, color, title) in zip(axes, plot_specs):
                    style_axes(ax)
                    y = df[col].values
                    ax.plot(t_mid, y, color=color, lw=1.5, marker="o", ms=3.5)
                    ax.fill_between(t_mid, y, alpha=0.10, color=color)
                    ax.set_ylabel(ylabel)
                    ax.set_title(title, loc="left")
                    if ax is not axes[-1]:
                        ax.tick_params(labelbottom=False)
                axes[-1].set_xlabel("Time (s)")

            self._slots["epochs"].update(draw_epochs)
            self._set_textbox(self.txt_epochs, df.to_string(index=False),
                              tsv=self._df_to_tsv(df))
            n_ep = len(df)
            # Update the label in the Epochs tab header
            self.lbl_epoch_count.configure(
                text=f"{n_ep} epochs × {epoch_s:.0f}s", text_color=BLUE)
            # Update the label in the Summary tab (shows last-computed epoch info)
            self.lbl_epoch_info.configure(
                text=f"Last epoch run: {n_ep} × {epoch_s:.0f}s", text_color=MUTED)
            self.tabs.set("HRV"); self.after(50, lambda: self._on_hrv_view_change("Epochs"))
            self._set_status(f"Epoch analysis done — {n_ep} epochs", GREEN)

        self._start_async_result(self.btn_compute_epochs, "Computing…", _worker, _done)

    # ════════════════════════════════════════════════════════
    #  EXPORT
    # ════════════════════════════════════════════════════════

    def _build_excel_workbook(self) -> "Workbook":
        """Build a formatted openpyxl Workbook from the current results."""
        return self.export_ctrl.build_excel_workbook()

    def _write_excel(self, destination) -> None:
        """Write the formatted workbook to *destination* (path or BytesIO)."""
        self.export_ctrl.write_excel(destination)

    def _export_excel(self) -> None:
        self.export_ctrl.export_excel()

    def _export_zip(self) -> None:
        self.export_ctrl.export_zip()

    def _export_pdf_report(self) -> None:
        """Generate a one-page PDF summary: ECG strip + KPI table + interpretation."""
        self.export_ctrl.export_pdf_report()

    def _export_prism(self) -> None:
        """Export all analysis results to a GraphPad Prism .pzfx file."""
        self.export_ctrl.export_prism()


    def _add_recent(self, path: str) -> None:
        if path in self._recent:
            self._recent.remove(path)
        self._recent.insert(0, path)
        self._recent = self._recent[:8]

    def _open_recent(self) -> None:
        """Show recent recordings from SQLite registry with summary stats."""
        db_rows = recent_recordings(limit=20) if _DB_AVAILABLE else []
        db_paths = {r["filepath"] for r in db_rows}
        extra = [p for p in self._recent if p not in db_paths and os.path.exists(p)]

        if not db_rows and not extra:
            messagebox.showinfo("Recent files", "No recent recordings yet.")
            return

        win = ctk.CTkToplevel(self)
        win.title("Recent recordings")
        win.geometry("740x480")
        win.configure(fg_color=BG)
        win.grab_set(); win.lift()

        hdr = ctk.CTkFrame(win, fg_color=PANEL, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Recent recordings", font=FONT_CARD_TITLE,
                     text_color=TEXT, anchor="w").pack(side="left", padx=SPACE_L, pady=SPACE_M)

        scroll = ctk.CTkScrollableFrame(win, fg_color=BG)
        scroll.pack(fill="both", expand=True, padx=SPACE_M, pady=SPACE_M)

        def _entry(path: str, hr: str = "—", sdnn: str = "—",
                   dur: str = "—", notes: str = "") -> None:
            if not os.path.exists(path):
                return
            card = ctk.CTkFrame(scroll, fg_color=CARD, corner_radius=6)
            card.pack(fill="x", pady=(0, SPACE_S))
            top = ctk.CTkFrame(card, fg_color="transparent")
            top.pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
            ctk.CTkButton(top, text=os.path.basename(path), anchor="w",
                          fg_color="transparent", hover_color=BORDER,
                          text_color=BLUE, font=FONT_SMALL,
                          command=lambda p=path: (win.destroy(), self._load_path(p))
                          ).pack(side="left", fill="x", expand=True)
            ctk.CTkLabel(top, text=f"HR {hr}  SDNN {sdnn}  {dur}",
                         font=FONT_KPI_LABEL, text_color=MUTED).pack(side="right")
            if notes:
                ctk.CTkLabel(card, text=f"📝 {notes[:90]}",
                             font=FONT_KPI_LABEL, text_color=LIGHT,
                             anchor="w").pack(padx=SPACE_M, pady=(0, SPACE_S), fill="x")

        for r in db_rows:
            _entry(
                r["filepath"],
                hr=f"{r['hr_mean']:.0f} bpm" if r.get("hr_mean") else "—",
                sdnn=f"{r['sdnn']:.1f} ms"   if r.get("sdnn") else "—",
                dur=f"{r['duration_s']:.0f} s" if r.get("duration_s") else "",
                notes=r.get("notes", ""),
            )
        for p in extra:
            _entry(p)



    def _load_path(self, path: str) -> None:
        if not os.path.exists(path):
            messagebox.showerror("Not found", f"File not found:\n{path}")
            return
        # Full reset before loading a new file so the app looks exactly like
        # it did at startup — no stale results, plots, or arrhythmia cards
        # from the previous recording.
        self._reset_for_new_file()
        self._filepath = path
        self._recording_notes = get_notes(path) if _DB_AVAILABLE else ""
        self.lbl_file.configure(text=os.path.basename(path), text_color=GREEN)  # type: ignore[union-attr]
        self._add_recent(path)
        # ── Try to restore a previously saved session ───────────────
        if self._try_restore_session(path):
            return   # session restored — skip raw load
        self._load_raw_only()

    def _reset_for_new_file(self) -> None:
        """Reset ALL analysis state and UI to the startup blank slate.

        Called every time a new file is opened so there is zero carry-over
        from the previous recording.  Sidebar *parameter* widgets (channel,
        fs, thresholds, filters) are intentionally kept — users typically
        want the same settings for consecutive recordings from the same rig.

        IMPORTANT: must NOT call _init_state() — that method also zeroes out
        all widget-reference attributes (ent_channel, ent_fs, …) which are
        already live in the UI, causing AttributeError on the next snapshot.
        Only data variables are reset here.
        """
        # ── 1. Data-only state reset (no widget refs) ─────────────────────
        self._filepath             = None
        self._signal_raw           = None
        self._signal_raw_norm      = None
        self._signal_flt           = None
        self._time                 = None
        # _fs intentionally kept (same rig)
        self._rpeaks_ok            = None
        self._rpeaks_rej           = None
        self._all_cands            = None
        self._all_proms            = None
        self._thresh_amp           = 0.0
        self._results              = None
        self._epoch_df             = None
        self._rolling_hrv_df       = None
        self._analysis_t_start      = 0.0
        self._analysis_t_end        = 0.0
        if self._hover_after_id is not None:
            try:
                self.after_cancel(self._hover_after_id)
            except Exception:
                pass
        self._hover_samp       = None
        self._hover_samp_near  = False
        self._hover_after_id   = None
        self._arrhythmia_events    = []
        self._arrhythmia_tsv       = ""
        self._arr_selected_idx     = -1
        self._arr_nav_pos          = 0.0
        self._arr_win              = 3.0
        self._arr_edit_mode        = False
        self._sig_quality          = None
        self._artifact_report      = None
        self._ds_time              = None
        self._ds_sig               = None
        self._ds_sig_max           = None
        self._ds_sig_mid           = None
        self._ds_raw_sig           = None
        self._ds_raw_sig_max       = None
        self._ds_raw_sig_mid       = None
        self._manual_excluded      = set()
        self._rpeaks_manual_excl   = None
        self._manual_added         = set()
        self._rpeaks_manual_added  = None
        self._edit_mode            = False
        self._edit_undo            = []
        self._edit_redo            = []
        self._nav_pos              = 0.0
        self._signal_inverted      = False
        self._raw_only_loaded      = False
        self._thr_debounce_id      = None
        self._annotations          = []
        self._tsv_store            = {}
        self._wave_template        = None
        self._session_dirty        = False
        self._generation           = getattr(self, "_generation", 0) + 1  # invalidate async workers

        # ── 2. Result plots → placeholder ────────────────────────────────
        for slot in self._slots.values():
            try:
                slot._draw_fn = None
                slot._show_placeholder()
            except Exception as e:
                log.debug("slot placeholder reset failed: %s", e)

        # ── 3. KPI bar ────────────────────────────────────────────────────
        for lbl in self._kpi.values():
            try:
                lbl.configure(text="—", text_color=MUTED)
            except Exception as e:
                log.debug("KPI label reset failed: %s", e)

        # ── 4. Status labels ──────────────────────────────────────────────
        self._set_status("File loaded — run Detection", MUTED)
        self._set_progress(0, "")
        if self.lbl_file is not None:
            self.lbl_file.configure(text="Loading…", text_color=MUTED)  # type: ignore[union-attr]

        # ── 5. Sidebar detection status ───────────────────────────────────
        if self.lbl_npeaks is not None:
            self.lbl_npeaks.configure(  # type: ignore[union-attr]
                text="Run detection", text_color=MUTED)

        # ── 6. Disable per-tab on-demand buttons ─────────────────────────
        for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl",
                         "btn_run_arrhythmia", "btn_save_session"):
            w = getattr(self, btn_attr, None)
            if w is not None:
                try:
                    w.configure(state="disabled")
                except Exception as e:
                    log.debug("widget disable failed: %s", e)

        # ── 7. Per-tab status labels ──────────────────────────────────────
        if self.lbl_freq_status is not None:
            self.lbl_freq_status.configure(  # type: ignore[union-attr]
                text="  Run Core Analysis first", text_color=MUTED)
        if self.lbl_nonlin_status is not None:
            self.lbl_nonlin_status.configure(  # type: ignore[union-attr]
                text="  Run Core Analysis first", text_color=MUTED)
        if self.lbl_ivl_status is not None:
            self.lbl_ivl_status.configure(  # type: ignore[union-attr]
                text="  Run Core Analysis first", text_color=MUTED)
        if self.lbl_arrhythmia_status is not None:
            self.lbl_arrhythmia_status.configure(  # type: ignore[union-attr]
                text="  Run Core Analysis first", text_color=MUTED)
        if self.lbl_roll_status is not None:
            self.lbl_roll_status.configure(  # type: ignore[union-attr]
                text="  Run Core Analysis first", text_color=MUTED)

        # ── 8. Textboxes ──────────────────────────────────────────────────
        for tb_attr in ("txt_rr", "txt_td", "txt_fd"):
            tb = getattr(self, tb_attr, None)
            if tb is not None:
                try:
                    self._set_textbox(tb, "")
                except Exception as e:
                    log.debug("_set_textbox clear failed: %s", e)

        # ── 9. Arrhythmia event cards ─────────────────────────────────────
        if self._arr_card_widgets is not None:
            for w in self._arr_card_widgets:
                try:
                    w.destroy()
                except Exception:
                    pass
            self._arr_card_widgets.clear()
        if self.lbl_arr_event_title is not None:
            try:
                self.lbl_arr_event_title.configure(  # type: ignore[union-attr]
                    text="← Click on an episode", text_color=MUTED)
            except Exception as e:
                log.debug("lbl_arr_event_title reset failed: %s", e)

        # ── 9b. Interpretation cards ──────────────────────────────────────
        # Détruire TOUS les enfants du scroll (groupes + cartes) pour éviter
        # que des frames grises orphelines restent visibles après un nouveau fichier.
        # interpretation tab removed — no-op
        self._interp_cards = {}
        self._interp_ref_labels = {}

        # ── 10. Session UI ────────────────────────────────────────────────
        self._update_session_ui(has_session=False)

        # ── 11. Disconnect RR click handler ──────────────────────────────
        if self._rr_click_cid is not None:
            try:
                self._slots["rr"].canvas.mpl_disconnect(self._rr_click_cid)
            except Exception as e:
                log.debug("mpl_disconnect (rr_click_cid) failed: %s", e)
            self._rr_click_cid = None

        # ── 12. Switch to Detection tab so user lands in the right place ──
        try:
            self.tabs.set("Detection")
        except Exception as e:
            log.debug("tabs.set Detection (reset) failed: %s", e)

    def _try_restore_session(self, path: str) -> bool:
        """If a saved session exists for *path*, offer to restore it.

        Returns True if the session restoration was initiated (caller should
        skip _preview).  The actual restore runs in a background thread via
        _start_async so the UI stays responsive during signal reload + filtering.
        """
        state = load_session(path)
        if state is None:
            self._update_session_ui(has_session=False)
            return False

        saved_at = state.get("saved_at", "unknown time")
        n_beats  = state.get("n_beats", "?")
        answer   = messagebox.askyesno(
            "Session found",
            f"A saved analysis was found for this file:\n\n"
            f"  Saved:  {saved_at}\n"
            f"  Beats:  {n_beats}\n\n"
            f"Restore it?  (No = load raw signal — click Preview Detection to re-run)",
            parent=self,
        )
        if not answer:
            self._update_session_ui(has_session=True)
            return False

        # ── Restauration asynchrone — le reload + _compute_preview_bundle peut
        # prendre 3–10 s ──
        # _restore_state_from_session est réellement découpée en deux :
        #   _restore_session_worker() : partie lourde, pure — reload .mat +
        #       _compute_preview_bundle — AUCUNE écriture sur self ni sur les
        #       widgets Tkinter. Tourne dans le thread BG lancé par _start_async.
        #   _on_restore_session_done(): toutes les écritures sur self et les
        #       widgets (y compris _run_detection, _draw_detail, etc.).
        #       Tourne sur le thread principal via after(0, …).
        _state_snap = state
        _saved_at   = saved_at

        self._start_async(
            self.btn_preview, "Restoring…", "Restoring session…",
            lambda: self._restore_session_worker(_state_snap),
            lambda bundle: self._on_restore_session_done(bundle, _saved_at),
            pass_result=True,
        )
        return True

    def _restore_session_worker(self, state: dict) -> dict:
        """Background worker — MUST NOT write to self or touch any Tkinter widget.

        Reloads the raw signal from the original .mat file and re-runs the
        pure ``_compute_preview_bundle`` static method with the saved filter
        parameters. Returns a plain bundle; ``_on_restore_session_done``
        writes every field to ``self`` (and to widgets) on the main thread.

        Session restore previously ran entirely in the background thread
        (including dozens of widget .configure/.delete/.insert calls and
        _draw_detail()), which is unsafe — Tkinter/Tcl objects must only be
        touched from the main thread. This mirrors the _preview_worker /
        _on_preview_done split used elsewhere in the app.
        """
        if self._filepath is None:
            raise RuntimeError("No file loaded — cannot restore session.")

        fs = int(state["fs"])

        # ── Parse saved filter params via FilterParams.from_dict ────────
        # from_dict tolerates missing keys with safe defaults, so old session
        # files (saved before any new field was added) restore without error.
        fp = FilterParams.from_dict(state.get("filter_params", {}))

        try:
            sig_raw, detected_ch, _, detected_fs = load_mat_signal(
                self._filepath, fp.channel)
        except Exception as exc:
            raise RuntimeError(f"Could not reload signal from {self._filepath}: {exc}") from exc

        if detected_fs is not None:
            fs = int(detected_fs)

        # Apply the same time-crop that was active when the session was saved
        i0 = int(fp.t_start * fs) if fp.t_start > 0 else 0
        i1 = int(fp.t_end   * fs) if fp.t_end   > 0 else len(sig_raw)
        sig_raw = sig_raw[i0:i1]
        n_samples = len(sig_raw)

        # ── Re-run signal processing with saved filter params (pure/static) ──
        # fp.to_dict() produces the "notch" key (not "notch_filter") that
        # _compute_preview_bundle expects.
        prepare_params = fp.to_dict()
        signal_bundle = self._compute_preview_bundle(sig_raw, fs, prepare_params)

        results_raw  = state.get("results")
        epoch_raw    = state.get("epoch_df")
        rolling_raw  = state.get("rolling_hrv_df")

        return {
            "fs":               fs,
            "n_samples":        n_samples,
            "signal_raw":       sig_raw,
            "signal_bundle":    signal_bundle,
            "thresh_amp":       float(state.get("threshold", float(state.get("thresh_amp", 0.5)))),
            "manual_excluded":  set(int(x) for x in state.get("manual_excluded", [])),
            "manual_added":     set(int(x) for x in state.get("manual_added",    [])),
            "signal_inverted":  bool(state.get("signal_inverted", False)),
            "no_filter_mode":   bool(state.get("no_filter_mode",  fp.no_filter)),
            "results":          _deserialise_results(results_raw) if results_raw else None,
            "artifact_report":  state.get("artifact_report"),
            "epoch_df":         pd.DataFrame(epoch_raw)   if epoch_raw   is not None else None,
            "rolling_hrv_df":   pd.DataFrame(rolling_raw) if rolling_raw is not None else None,
            "annotations":      list(state.get("annotations", [])),
            "analysis_t_start": float(state.get("analysis_t_start", 0.0)),
            "analysis_t_end":   float(state.get("analysis_t_end",   0.0)),
            "exp_context":      state.get("exp_context", "telemetry_awake"),
            "nav_pos":          float(state.get("nav_pos", 0.0)),
            "edit_mode":        bool(state.get("edit_mode", False)),
            "current_tab":      state.get("current_tab", "Detection"),
        }

    def _on_restore_session_done(self, bundle: dict, saved_at: str) -> None:
        """Write all restored state to self and to widgets — main thread only.

        Counterpart to ``_restore_session_worker``. This is the ONLY place
        that writes session-restore state to ``self``/widgets, exactly as
        ``_on_preview_done`` is the sole writer after ``_preview_worker``.
        """
        # Clear figure cache on session restore
        self._figure_cache.clear()

        fs        = bundle["fs"]
        n_samples = bundle["n_samples"]
        self._fs  = fs

        self._signal_raw = bundle["signal_raw"]
        self._time       = np.arange(n_samples) / fs
        self._nav_pos    = 0.0
        self._ds_time         = None
        self._ds_sig          = None
        self._ds_sig_max      = None
        self._ds_sig_mid      = None
        self._ds_raw_sig      = None
        self._ds_raw_sig_max  = None
        self._ds_raw_sig_mid  = None

        sb = bundle["signal_bundle"]
        self._signal_raw_norm = sb["signal_raw_norm"]
        self._signal_flt      = sb["signal_flt"]
        self._all_cands       = sb["all_cands"]
        self._all_proms       = sb["all_proms"]
        self._no_filter_mode  = sb["no_filter_mode"]
        self._signal_inverted = sb.get("inverted", False)

        # ── Restore user-edited state (session values win over freshly
        # recomputed ones, matching the original restore order) ──────────
        self._thresh_amp      = bundle["thresh_amp"]
        self._manual_excluded = set(bundle["manual_excluded"])
        self._manual_added    = set(bundle["manual_added"])
        self._signal_inverted = bundle["signal_inverted"]
        self._no_filter_mode  = bundle["no_filter_mode"]

        # ── Restore analysis results (DataFrames already deserialised) ───
        self._results         = bundle["results"]
        self._artifact_report = bundle["artifact_report"]
        self._epoch_df        = bundle["epoch_df"]
        self._rolling_hrv_df  = bundle["rolling_hrv_df"]
        self._annotations     = list(bundle["annotations"])

        self._analysis_t_start = bundle["analysis_t_start"]
        self._analysis_t_end   = bundle["analysis_t_end"]
        # Sync analysis window entry widgets with restored values
        if self.ent_analysis_t0 is not None:
            self._batch_ui_update(self.ent_analysis_t0, state="normal")
            self.ent_analysis_t0.delete(0, "end")  # type: ignore[union-attr]
            if self._analysis_t_start > 0:
                self.ent_analysis_t0.insert(0, str(self._analysis_t_start))  # type: ignore[union-attr]
            self._batch_ui_update(self.ent_analysis_t0, state="normal")
        if self.ent_analysis_t1 is not None:
            self._batch_ui_update(self.ent_analysis_t1, state="normal")
            self.ent_analysis_t1.delete(0, "end")  # type: ignore[union-attr]
            if self._analysis_t_end > 0:
                self.ent_analysis_t1.insert(0, str(self._analysis_t_end))  # type: ignore[union-attr]
            self._batch_ui_update(self.ent_analysis_t1, state="normal")

        # ── Restore UI state ─────────────────────────────────────────────
        ctx = bundle["exp_context"]
        if ctx in EXPERIMENTAL_CONTEXTS:
            self._exp_context = ctx
            try:
                pass  # interpretation removed  # type: ignore[union-attr]
                pass  # interpretation removed
                pass  # interpretation removed
            except Exception as e:
                log.warning("Failed to restore context from session: %s", e)

        self._nav_pos = bundle["nav_pos"]
        self._sync_nav_pos_entry()

        saved_edit = bundle["edit_mode"]
        if saved_edit != self._edit_mode:
            self._toggle_edit_mode()

        # ── Restore threshold slider and re-apply detection ──────────────
        thr = self._thresh_amp
        if self.sl_thr is not None:
            try:
                self.sl_thr.set(float(thr))  # type: ignore[union-attr]
                self.ent_thr.delete(0, "end")  # type: ignore[union-attr]
                self.ent_thr.insert(0, f"{float(thr):.3f}")  # type: ignore[union-attr]
            except Exception as _exc:
                log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 7671, _exc)

        self._run_detection(thr)

        # ── Enable analysis buttons ──────────────────────────────────────
        self.btn_save_session.configure(state="normal")  # type: ignore[union-attr]
        for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl"):
            if getattr(self, btn_attr, None) is not None:
                getattr(self, btn_attr).configure(state="normal")

        n_beats = len(self._rpeaks_ok) if self._rpeaks_ok is not None else 0
        self._set_status(
            f"Session restored — {n_beats} beats  |  "
            f"{n_samples / fs:.0f} s recording  |  results ready", GREEN)

        # ── Render ───────────────────────────────────────────────────────
        self._draw_detail(self._nav_pos)
        self._update_ann_count()
        # Restore active tab
        try:
            saved_tab = bundle["current_tab"]
            self.tabs.set(saved_tab)
        except Exception as e:
            log.debug("Failed to restore tab from session: %s", e)
        if self._results is not None:
            self.after(100, self._draw_all_results)
            pass  # interpretation removed
            if self.lbl_freq_status is not None:
                fd = self._results.get("hrv_freq")
                has_freq = fd is not None and not (hasattr(fd, "empty") and fd.empty)
                self.lbl_freq_status.configure(  # type: ignore[union-attr]
                    text="  Loaded from session ✓" if has_freq
                    else "  Core done — click to compute LF / HF",
                    text_color=GREEN if has_freq else BLUE)
        self._update_kpis()
        self._update_session_ui(has_session=True, saved_at=saved_at)

    # ════════════════════════════════════════════════════════
    #  SESSION SAVE / RESTORE
    # ════════════════════════════════════════════════════════

    def _current_filter_params_dict(self) -> dict:
        """Return a serialisable filter-params dict, reading from widgets if available.

        Safe to call at any point — uses FilterParams defaults for any widget
        not yet built (e.g. during early startup or after a rebuild).
        """
        try:
            # Happy path: all widgets are built and readable
            fp = FilterParams.from_widgets(self)
            # Preserve the last-used no_filter state from self in case the
            # toggle widget is temporarily inconsistent during a rebuild.
            fp = dataclasses.replace(fp, no_filter=self._no_filter_mode)
            return fp.to_dict()
        except Exception:
            # Widget not yet built — return safe defaults carrying known state
            return FilterParams(no_filter=self._no_filter_mode).to_dict()

    def _safe_get_tab(self) -> str:
        """Return the current tab name, or 'Detection' if not yet built."""
        try:
            return self.tabs.get()
        except Exception:
            return "Detection"

    def _collect_session_state(self) -> dict:
        """Gather all serialisable app state into a flat dict.

        Signal arrays (signal_flt, signal_raw_norm, all_cands, all_proms) are
        intentionally NOT stored.  They are derived entirely from the .mat file
        and the filter parameters; _restore_session_worker re-runs
        _compute_preview_bundle to reconstruct them in < 2 s.  This keeps session
        files small (< 200 KB for a typical recording) regardless of recording length.

        Previously (v3) the session stored the full filtered signal as a Python
        list (~58 MB for 10 min at 2 kHz, ~346 MB for 1 h), making auto-save
        impractical and restore slow.
        """
        state: dict = {
            "saved_at":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "fs":              int(self._fs),
            "threshold":       float(self._thresh_amp),
            "manual_excluded": list(self._manual_excluded),
            "manual_added":    list(self._manual_added),
            "no_filter_mode":  self._no_filter_mode,
            "signal_inverted": self._signal_inverted,
            # Filter params — use FilterParams.from_widgets so the field list
            # is maintained in a single place.  Falls back to defaults for any
            # widget that hasn't been built yet (e.g. called during startup).
            "filter_params":   self._current_filter_params_dict(),
            # Beat count stored for the restore-dialog display only
            "n_beats":         len(self._rpeaks_ok) if self._rpeaks_ok is not None else 0,
            # UI state that must survive a session restore
            "exp_context":     self._exp_context,
            "nav_pos":         self._nav_pos,
            "edit_mode":       self._edit_mode,
            "current_tab":     self._safe_get_tab(),
            "analysis_t_start": self._analysis_t_start,
            "analysis_t_end":   self._analysis_t_end,
        }
        if self._results is not None:
            state["results"] = _serialise_results(self._results)
        if self._artifact_report is not None:
            state["artifact_report"] = self._artifact_report
        if self._epoch_df is not None and not self._epoch_df.empty:
            state["epoch_df"] = self._epoch_df.to_dict(orient="list")
        if self._rolling_hrv_df is not None and not self._rolling_hrv_df.empty:
            state["rolling_hrv_df"] = self._rolling_hrv_df.to_dict(orient="list")
        if self._annotations:
            state["annotations"] = self._annotations
        return state

    def _save_session(self) -> None:
        """Serialise full analysis state to a .ecgsession cache file and update registry."""
        if self._filepath is None:
            messagebox.showwarning("No file", "Open a file first.")
            return
        if self._signal_flt is None:
            messagebox.showwarning("Not ready", "Run Preview Detection first.")
            return
        try:
            self._set_status("Saving session…", MUTED)
            state    = self._collect_session_state()
            out_path = save_session(self._filepath, state)
            saved_at = state["saved_at"]
            self._session_dirty = False
            self._update_session_ui(has_session=True, saved_at=saved_at)
            # ── SQLite registry upsert ────────────────────────────────────
            if _DB_AVAILABLE:
                from session import _file_fingerprint
                _stats: dict = {}
                if self._results:
                    _rdf = self._results.get("rr_df")
                    _hrv = self._results.get("hrv_td")
                    if _rdf is not None and len(_rdf):
                        _stats["hr_mean"] = float(_rdf["HR_bpm"].mean())
                    if _hrv is not None and "HRV_SDNN" in _hrv.columns:
                        _stats["sdnn"]  = float(_hrv["HRV_SDNN"].values[0])
                    if _hrv is not None and "HRV_RMSSD" in _hrv.columns:
                        _stats["rmssd"] = float(_hrv["HRV_RMSSD"].values[0])
                    if self._time is not None and self._fs:
                        _stats["duration_s"] = float(len(self._time)) / self._fs
                    if self._rpeaks_ok is not None:
                        _stats["n_peaks"] = int(len(self._rpeaks_ok))
                upsert_recording(
                    filepath=self._filepath,
                    fingerprint=_file_fingerprint(self._filepath),
                    session_path=str(out_path),
                    stats=_stats,
                    notes=self._recording_notes,
                )
            self._set_status(f"Session saved — {Path(out_path).name}", GREEN)
        except Exception as exc:
            log.exception("_save_session failed")
            messagebox.showerror("Save failed", str(exc))

    def _delete_session(self) -> None:
        """Delete the session cache file for the current file."""
        if self._filepath is None:
            return
        deleted = delete_session(self._filepath)
        if deleted:
            self._update_session_ui(has_session=False)
            self._set_status("Session cache deleted.", MUTED)
        else:
            self._set_status("No session cache to delete.", MUTED)

    def _update_session_ui(self, has_session: bool,
                           saved_at: str = "") -> None:
        """Update the session info label and button states."""
        if self.lbl_session_info is None:
            return
        if has_session and saved_at:
            self.lbl_session_info.configure(  # type: ignore[union-attr]
                text=f"✓ Session saved  {saved_at}", text_color=GREEN)
        elif has_session:
            self.lbl_session_info.configure(  # type: ignore[union-attr]
                text="✓ Session file exists for this recording", text_color=GREEN)
        else:
            self.lbl_session_info.configure(  # type: ignore[union-attr]
                text="No session saved for this file", text_color=MUTED)
        # Update wave template info label
        if self.lbl_template_info is not None and self._wave_template is not None:
            wt = self._wave_template
            if wt.confirmed:
                self.lbl_template_info.configure(  # type: ignore[union-attr]
                    text=f"✓ Custom template active  (source={wt.source})",
                    text_color=GREEN)
            else:
                self.lbl_template_info.configure(  # type: ignore[union-attr]
                    text="Using default mouse landmarks", text_color=MUTED)

    # ════════════════════════════════════════════════════════
    #  WAVE TEMPLATE EDITOR
    # ════════════════════════════════════════════════════════

    def _open_wave_template_editor(self) -> None:
        """Open the interactive P/Q/R/S/T template editor.

        If a mean beat is available it is passed to the editor for display;
        otherwise the editor shows a synthetic reference trace.
        """
        beat_time = None
        mean_beat = None
        beat_sd   = None
        if self._results is not None:
            beat_time = self._results.get("beat_time")
            mean_beat = self._results.get("beat_template")
            beat_sd   = self._results.get("beat_sd")

        if self._wave_template is None:
            self._wave_template = WaveTemplate.load()

        editor = WaveTemplateMiniEditor(
            self, self._wave_template,
            beat_time=beat_time,
            mean_beat=mean_beat,
            beat_sd=beat_sd,
        )
        self.wait_window(editor)  # modal — blocks until closed

        # Refresh the label whether or not the user saved
        self._update_session_ui(
            has_session=load_session(self._filepath) is not None
            if self._filepath else False
        )
        if editor._saved:
            self._set_status(
                "Wave template updated — re-run Interval Delineation to apply changes.",
                GREEN)


    def _toggle_language(self) -> None:
        """Switch UI language between English and French and rebuild the UI."""
        try:
            from i18n import get_language, set_language
        except ImportError:
            return
        new_lang = "fr" if get_language() == "en" else "en"
        set_language(new_lang)
        # Rebuild so all labels refresh
        ui_state = self._snapshot_ui_state()
        self._rebuild_ui()
        self._restore_ui_state(ui_state)
        # Update the button label to show the OTHER language (the one we'd switch to)
        label = "FR" if new_lang == "en" else "EN"
        if self.btn_lang is not None:
            self.btn_lang.configure(text=label)
        self._set_status(f"Language switched to {'English' if new_lang == 'en' else 'Français'}", BLUE)

    def _open_params_dialog(self) -> None:
        """Open a dedicated floating parameters window with all settings clearly grouped."""
        win = ctk.CTkToplevel(self)
        win.title("⚙  Parameters")
        win.geometry("540x720")
        win.configure(fg_color=BG)
        win.resizable(True, True)
        win.grab_set()
        win.lift()

        scroll = ctk.CTkScrollableFrame(win, fg_color=BG,
                                        scrollbar_button_color=BORDER,
                                        scrollbar_button_hover_color=BORDER2)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)
        s = scroll
        px = dict(padx=SPACE_L)

        def _sec(title: str, color: str = BLUE) -> ctk.CTkFrame:
            hdr = ctk.CTkFrame(s, fg_color=PANEL, corner_radius=8)
            hdr.pack(fill="x", padx=SPACE_M, pady=(SPACE_M, SPACE_XS))
            ctk.CTkFrame(hdr, height=3, fg_color=color, corner_radius=2).pack(fill="x")
            ctk.CTkLabel(hdr, text=title, font=FONT_SIDEBAR_HDR,
                         text_color=color, anchor="w").pack(padx=SPACE_M, pady=(SPACE_S, SPACE_S), fill="x")
            body = ctk.CTkFrame(s, fg_color=CARD, corner_radius=6)
            body.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
            return body

        def _row_entry(parent, label: str, widget_attr: str) -> ctk.CTkEntry:
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", **px, pady=(SPACE_S, 0))
            row.columnconfigure(1, weight=1)
            ctk.CTkLabel(row, text=label, font=FONT_SMALL, text_color=MUTED,
                         anchor="w", width=160).grid(row=0, column=0, sticky="w")
            # Mirror value from existing sidebar widget
            src: Optional[ctk.CTkEntry] = getattr(self, widget_attr, None)
            default = src.get() if src is not None else ""
            ent = ctk.CTkEntry(row, font=FONT_LABEL, height=28, fg_color=BG,
                               border_color=BORDER2, text_color=TEXT)
            ent.insert(0, default)
            ent.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))
            return ent

        def _row_switch(parent, label: str, widget_attr: str) -> ctk.CTkSwitch:
            src: Optional[ctk.CTkSwitch] = getattr(self, widget_attr, None)
            is_on = bool(src.get()) if src is not None else False
            sw = ctk.CTkSwitch(parent, text=label, font=FONT_LABEL, text_color=MUTED,
                               progress_color=BLUE, button_color=BORDER2)
            if is_on:
                sw.select()
            sw.pack(**px, anchor="w", pady=(SPACE_S, SPACE_XS))
            return sw

        # ── FILE & SUBJECT ─────────────────────────────────────────────────
        f0 = _sec("📁  File & Subject", BLUE_DARK)
        dlg_channel = _row_entry(f0, "Channel name",    "ent_channel")
        dlg_subject = _row_entry(f0, "Subject ID",      "ent_subject")
        ctk.CTkFrame(f0, height=6, fg_color="transparent").pack()

        # ── SIGNAL ────────────────────────────────────────────────────────
        f1 = _sec("📡  Signal", ORANGE_DARK)
        dlg_fs      = _row_entry(f1, "Sampling rate (Hz)", "ent_fs")
        dlg_t_start = _row_entry(f1, "Start crop (s)",     "ent_t_start")
        dlg_t_end   = _row_entry(f1, "End crop (s)",       "ent_t_end")
        dlg_show_raw = _row_switch(f1, "Show raw signal (vs filtered)", "sw_show_raw")
        ctk.CTkFrame(f1, height=6, fg_color="transparent").pack()

        # ── FILTERS ───────────────────────────────────────────────────────
        f2 = _sec("🔧  Filters", PURPLE)
        dlg_no_filter = _row_switch(f2, "Raw signal — no DSP filters (recommended)", "sw_no_filter")
        dlg_notch     = _row_switch(f2, "Notch 50 Hz", "sw_notch")
        dlg_invert    = _row_switch(f2, "Invert signal polarity", "sw_invert_signal")
        dlg_hp = _row_entry(f2, "HP cut-off (Hz)",  "ent_lp")
        dlg_lp = _row_entry(f2, "LP cut-off (Hz)",  "ent_hp")

        # Clean method
        row_cm = ctk.CTkFrame(f2, fg_color="transparent")
        row_cm.pack(fill="x", **px, pady=(SPACE_S, SPACE_S))
        row_cm.columnconfigure(1, weight=1)
        ctk.CTkLabel(row_cm, text="NK2 clean method:", font=FONT_SMALL, text_color=MUTED,
                     anchor="w", width=160).grid(row=0, column=0, sticky="w")
        _clean_src: Optional[ctk.CTkComboBox] = getattr(self, "cb_clean", None)
        _clean_val = _clean_src.get() if _clean_src is not None else "neurokit"
        dlg_clean = ctk.CTkComboBox(row_cm, font=FONT_LABEL, height=28,
                                    fg_color=BG, border_color=BORDER2,
                                    button_color=BORDER2, text_color=TEXT,
                                    dropdown_fg_color=BG, dropdown_text_color=TEXT,
                                    values=["neurokit", "pantompkins1985",
                                            "elgendi2010", "hamilton2002", "biosppy"])
        dlg_clean.set(_clean_val)
        dlg_clean.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))
        ctk.CTkFrame(f2, height=6, fg_color="transparent").pack()

        # ── DETECTION ─────────────────────────────────────────────────────
        f3 = _sec("🔍  Detection", RED)
        dlg_minrr = _row_entry(f3, "Min R-R distance (ms)", "ent_minrr")

        # Detection method
        row_dm = ctk.CTkFrame(f3, fg_color="transparent")
        row_dm.pack(fill="x", **px, pady=(SPACE_S, SPACE_XS))
        row_dm.columnconfigure(1, weight=1)
        ctk.CTkLabel(row_dm, text="Detection method:", font=FONT_SMALL, text_color=MUTED,
                     anchor="w", width=160).grid(row=0, column=0, sticky="w")
        _dm_src: Optional[ctk.CTkComboBox] = getattr(self, "cb_det_method", None)
        _dm_val = _dm_src.get() if _dm_src is not None else "Auto (NeuroKit2)"
        dlg_det_method = ctk.CTkComboBox(
            row_dm, font=FONT_LABEL, height=28, fg_color=BG, border_color=BORDER2,
            button_color=BORDER2, text_color=TEXT, dropdown_fg_color=BG,
            dropdown_text_color=TEXT,
            values=["SG + Derivative (10 kHz)","Wavelet (CWT)", "Auto (NeuroKit2)", "Envelope Max"])
        dlg_det_method.set(_dm_val)
        dlg_det_method.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))

        dlg_sg_target_fs = _row_entry(f3, "SG target fs (Hz)", "ent_sg_target_fs")
        dlg_sg_window_ms = _row_entry(f3, "SG window (ms)",    "ent_sg_window_ms")
        ctk.CTkLabel(f3, text="SG+Deriv: downsample → Savitzky-Golay derivative → R detection\n"
                             "Wavelet: CWT bruit/QRS/J-wave séparés (pip install PyWavelets)\n"
                             "Envelope Max: maximum local — idéal signaux saturés (clipping ADC)",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=480, justify="left").pack(**px, pady=(0, SPACE_S), fill="x")

        # Threshold
        thr_val = float(self.sl_thr.get()) if self.sl_thr is not None else 0.5  # type: ignore[union-attr]
        row_thr = ctk.CTkFrame(f3, fg_color="transparent")
        row_thr.pack(fill="x", **px, pady=(SPACE_S, SPACE_S))
        row_thr.columnconfigure(1, weight=1)
        ctk.CTkLabel(row_thr, text="Threshold:", font=FONT_SMALL, text_color=MUTED,
                     anchor="w", width=160).grid(row=0, column=0, sticky="w")
        dlg_thr = ctk.CTkSlider(row_thr, from_=0, to=2,
                                 progress_color=RED, button_color=RED, fg_color=BORDER)
        dlg_thr.set(thr_val)
        dlg_thr.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))
        ctk.CTkFrame(f3, height=6, fg_color="transparent").pack()

        # ── ARTIFACTS ─────────────────────────────────────────────────────
        f4 = _sec("⚠️  Artifacts", ORANGE)
        dlg_artifact = _row_switch(f4, "Auto-correct on Full Analysis (OFF by default)", "sw_artifact")
        ctk.CTkLabel(
            f4,
            text="OFF recommended — use '🔍 Review Artifacts' button for full interactive control.\n"
                 "Auto-correct removes detected ectopic/non-physiological beats without review.",
            font=FONT_KPI_LABEL, text_color=LIGHT,
            anchor="w", wraplength=480, justify="left",
        ).pack(**px, pady=(0, SPACE_M), fill="x")

        # ── Buttons ───────────────────────────────────────────────────────
        ctk.CTkFrame(s, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_M, SPACE_S))
        btn_row = ctk.CTkFrame(s, fg_color="transparent")
        btn_row.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_M))

        def _apply():
            """Write all dialog values back to the sidebar widgets."""
            _write = lambda attr, val: (
                getattr(self, attr).delete(0, "end") or
                getattr(self, attr).insert(0, val)
            ) if getattr(self, attr, None) is not None else None

            _write("ent_channel",  dlg_channel.get())
            _write("ent_subject",  dlg_subject.get())
            _write("ent_fs",       dlg_fs.get())
            _write("ent_t_start",  dlg_t_start.get())
            _write("ent_t_end",    dlg_t_end.get())
            _write("ent_lp",       dlg_hp.get())
            _write("ent_hp",       dlg_lp.get())
            _write("ent_minrr",    dlg_minrr.get())

            # SG params
            _write("ent_sg_target_fs", dlg_sg_target_fs.get())
            _write("ent_sg_window_ms", dlg_sg_window_ms.get())

            # ComboBoxes
            if self.cb_clean is not None:
                self.cb_clean.set(dlg_clean.get())
            if self.cb_det_method is not None:
                self.cb_det_method.set(dlg_det_method.get())
                self._on_det_method_change(dlg_det_method.get())

            # Switches
            for sw_attr, dlg_sw in [
                ("sw_no_filter",    dlg_no_filter),
                ("sw_notch",        dlg_notch),
                ("sw_invert_signal",dlg_invert),
                ("sw_show_raw",     dlg_show_raw),
                ("sw_artifact",     dlg_artifact),
            ]:
                w = getattr(self, sw_attr, None)
                if w is None:
                    continue
                if dlg_sw.get():
                    w.select()
                else:
                    w.deselect()

            # Threshold slider
            if self.sl_thr is not None:
                self.sl_thr.set(float(dlg_thr.get()))  # type: ignore[union-attr]
                self.lbl_thr.configure(text=f"Sensitivity:  {dlg_thr.get():.3f}")

            self._on_no_filter_toggle()
            self._on_show_raw_toggle()
            self._set_status("Parameters applied ✓", GREEN)
            win.destroy()

        ctk.CTkButton(btn_row, text="✔  Apply & Close", command=_apply,
                      fg_color=GREEN, hover_color=GREEN_DARK, text_color="white",
                      font=FONT_BTN_PRIMARY, height=36, corner_radius=8).pack(
            side="left", padx=(0, SPACE_M))
        ctk.CTkButton(btn_row, text="✗  Cancel", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=36, corner_radius=8).pack(side="left")

    # ════════════════════════════════════════════════════════
    #  APPEARANCE / THEME
    # ════════════════════════════════════════════════════════

    def _open_theme_dialog(self) -> None:
        """Open the appearance settings dialog."""
        dlg = ThemeDialog(self)
        self.wait_window(dlg)

    def _snapshot_ui_state(self) -> dict:
        """Capture all editable widget values before a UI rebuild."""
        s: dict = {}
        for attr in ("ent_channel", "ent_subject", "ent_fs", "ent_t_start",
                     "ent_t_end", "ent_lp", "ent_hp", "ent_minrr",
                     "ent_epoch", "ent_overlap", "ent_thr", "ent_window",
                     "ent_sg_target_fs", "ent_sg_window_ms"):
            try:
                s[attr] = getattr(self, attr).get()
            except Exception:
                s[attr] = ""
        for attr in ("sw_show_raw", "sw_no_filter", "sw_notch",
                     "sw_artifact", "sw_epoch"):
            try:
                s[attr] = bool(getattr(self, attr).get())
            except Exception:
                s[attr] = False
        try:
            s["sl_thr"] = float(self.sl_thr.get())  # type: ignore[union-attr]
        except Exception:
            s["sl_thr"] = 0.5
        try:
            s["cb_clean"] = self.cb_clean.get()  # type: ignore[union-attr]
        except Exception:
            s["cb_clean"] = "neurokit"
        try:
            s["cb_det_method"] = self.cb_det_method.get() if self.cb_det_method is not None else "SG + Derivative (10 kHz)"
        except Exception:
            s["cb_det_method"] = "SG + Derivative (10 kHz)"
        try:
            s["adv_filters_open"] = getattr(self, "_adv_filters_open", False)
        except Exception:
            s["adv_filters_open"] = False
        s["exp_context"] = self._exp_context
        try:
            s["current_tab"] = self.tabs.get()
        except Exception:
            s["current_tab"] = "Detection"
        s["edit_mode"] = self._edit_mode
        s["nav_pos"]   = self._nav_pos
        s["dark_mode"] = self._dark_mode
        return s

    def _restore_ui_state(self, s: dict) -> None:
        """Restore widget values captured before a UI rebuild."""
        for attr in ("ent_channel", "ent_subject", "ent_fs", "ent_t_start",
                     "ent_t_end", "ent_lp", "ent_hp", "ent_minrr",
                     "ent_epoch", "ent_overlap", "ent_thr", "ent_window",
                     "ent_sg_target_fs", "ent_sg_window_ms"):
            val = s.get(attr)
            if val is None:
                continue
            try:
                w = getattr(self, attr, None)
                if w is None:
                    continue
                w.delete(0, "end")
                w.insert(0, str(val))
            except Exception as _exc:
                log.debug("_restore_ui_state entry %s: %s", attr, _exc)

        sw_map = {
            "sw_show_raw": "sw_show_raw",
            "sw_no_filter": "sw_no_filter",
            "sw_notch": "sw_notch",
            "sw_artifact": "sw_artifact",
            "sw_epoch": "sw_epoch",
        }
        for key, attr in sw_map.items():
            try:
                w = getattr(self, attr)
                if s.get(key):
                    w.select()
                else:
                    w.deselect()
            except Exception as _exc:
                log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 7929, _exc)

        try:
            self.sl_thr.set(float(s.get("sl_thr", 0.5)))  # type: ignore[union-attr]
        except Exception as _exc:
            log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 7934, _exc)
        try:
            self.cb_clean.set(s.get("cb_clean", "neurokit"))  # type: ignore[union-attr]
        except Exception as _exc:
            log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 7938, _exc)
        try:
            if self.cb_det_method is not None:
                dm = s.get("cb_det_method", "SG + Derivative (10 kHz)")
                self.cb_det_method.set(dm)
                self._on_det_method_change(dm)
        except Exception as _exc:
            log.debug("cb_det_method restore: %s", _exc)
        # Restore advanced-filter panel state
        try:
            was_open = bool(s.get("adv_filters_open", False))
            if was_open != getattr(self, "_adv_filters_open", False):
                self._btn_adv_flt.invoke()   # toggles the sub-section
        except Exception as _exc:
            log.debug("adv_filters restore: %s", _exc)
        try:
            self.tabs.set(s.get("current_tab", "Detection"))
        except Exception as _exc:
            log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 7942, _exc)
        # Re-apply non-widget state
        self._nav_pos  = float(s.get("nav_pos",  0.0))

        # Restore experimental context
        ctx = s.get("exp_context", "telemetry_awake")
        if ctx in EXPERIMENTAL_CONTEXTS:
            self._exp_context = ctx
            try:
                pass  # interpretation removed
                pass  # interpretation removed
                pass  # interpretation removed
            except Exception as e:
                log.warning("Failed to restore context from session: %s", e)
        self._dark_mode = bool(s.get("dark_mode", THEME.is_dark))
        saved_edit = s.get("edit_mode", False)
        if bool(saved_edit) != self._edit_mode:
            self._toggle_edit_mode()

    def _rebuild_ui(self) -> None:
        """Destroy and recreate the entire widget tree with the current theme.

        All application data (signal, results, peaks) is preserved.
        Widget values (threshold, filter settings, etc.) are snapshot and
        restored.  Matplotlib figures are closed before destruction to avoid
        memory leaks.
        """
        # Capture widget values
        ui_state = self._snapshot_ui_state()

        # Close all matplotlib figures before destroying their canvas widgets
        for slot in self._slots.values():
            try:
                plt.close(slot.fig)
            except Exception as _exc:
                log.debug("%s at %s:%d — %s", type(_exc).__name__, __name__, 7966, _exc)

        # Destroy the existing UI
        for child in self.winfo_children():
            try:
                child.destroy()
            except Exception:
                pass

        # Reset widget registries
        self._slots = {}
        self._kpi   = {}

        # Rebuild from scratch using updated globals
        self._build()

        # Restore widget values
        self._restore_ui_state(ui_state)

        # Restore file info label
        if self._filepath:
            if self.lbl_file is not None:
                self.lbl_file.configure(  # type: ignore[union-attr]
                    text=os.path.basename(self._filepath), text_color=GREEN)
        if self._rpeaks_ok is not None:
            n = len(self._rpeaks_ok)
            if self.lbl_npeaks is not None:
                self.lbl_npeaks.configure(text=f"Peaks detected: {n}", text_color=GREEN)  # type: ignore[union-attr]

        # Restore session / template labels
        # Guard: _wave_template may not exist if _init_state never ran
        # (e.g. theme dialog opened on a fresh launch before file load)
        if self._wave_template is None:
            self._wave_template = WaveTemplate.load()
        if not self._session_dirty:
            self._session_dirty = False
        has_session = bool(
            self._filepath and load_session(self._filepath) is not None)
        self._update_session_ui(has_session=has_session)

        # Enable action buttons if data is ready
        if self._signal_flt is not None:
            if self.btn_save_session is not None:
                self.btn_save_session.configure(state="normal")  # type: ignore[union-attr]
        if self._results is not None:
            for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl",
                              "btn_save_session"):
                btn = getattr(self, btn_attr, None)
                if btn is not None:
                    btn.configure(state="normal")

        # Repaint signal and results
        apply_theme_config(THEME)
        if self._signal_flt is not None:
            self._draw_detail(self._nav_pos)
        if self._results is not None:
            self.after(80, self._draw_all_results)
            pass  # interpretation removed
        self._update_kpis()

    # ════════════════════════════════════════════════════════
    #  DARK MODE (legacy — kept for backward compat)
    # ════════════════════════════════════════════════════════

    def _toggle_dark(self) -> None:
        self._dark_mode = not self._dark_mode
        ctk.set_appearance_mode("dark" if self._dark_mode else "light")
        apply_plot_theme(self._dark_mode)
        # Update tk.Frame backgrounds in all CanvasSlots (not managed by CTk)
        for slot in self._slots.values():
            try:
                slot.frame.configure(bg=PLOT["bg"])
                slot._cv_frame.configure(bg=PLOT["bg"])
                slot.fig.patch.set_facecolor(PLOT["bg"])
            except AttributeError as _slot_exc:
                log.debug("slot theme update: %s", _slot_exc)
        if self._signal_flt is not None:
            self._draw_detail()
        if self._results is not None:
            self._draw_all_results()

    # ════════════════════════════════════════════════════════
    #  CLIPBOARD / TXT SAVE
    # ════════════════════════════════════════════════════════

    def _copy_summary(self) -> None:
        self.export_ctrl.copy_summary()

    def _save_summary_txt(self) -> None:
        self.export_ctrl.save_summary_txt()

    # ════════════════════════════════════════════════════════
    #  DRAG AND DROP  (tkinterdnd2 optional)
    # ════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════
    #  1. KEYBOARD SHORTCUTS
    # ════════════════════════════════════════════════════════

    def _bind_keyboard_shortcuts(self) -> None:
        """Register global keyboard shortcuts."""
        bindings: "list[tuple[str, Any]]" = [
            ("<space>",       lambda e: self._preview()),
            ("<Control-r>",   lambda e: self._run_analysis()),
            ("<Control-o>",   lambda e: self._open_file()),
            ("<Control-s>",   lambda e: self._save_session()),
            ("<Control-e>",   lambda e: self._export_excel()),
            ("<Control-w>",   lambda e: self._export_rr_csv()),
            ("<Control-b>",   lambda e: self._open_batch_dialog()),
            ("<Control-m>",   lambda e: self._open_compare_segments()),
            ("<F1>",          lambda e: self._show_shortcuts_help()),
        ]
        for seq, cb in bindings:
            try:
                self.bind(seq, cb)
            except Exception as exc:
                log.debug("bind %s: %s", seq, exc)

    def _show_shortcuts_help(self) -> None:
        shortcuts = (
            "Space          Preview Detection\n"
            "Ctrl+R         Run Full Analysis\n"
            "Ctrl+O         Open .mat file\n"
            "Ctrl+S         Save session\n"
            "Ctrl+E         Export Excel\n"
            "Ctrl+W         Export RR intervals CSV\n"
            "Ctrl+B         Batch processing\n"
            "Ctrl+M         Compare segments\n"
            "Ctrl+Z / Y     Undo / Redo peak edits\n"
            "←  / →         Navigate tachogram\n"
            "F1             This help\n"
        )
        messagebox.showinfo("Keyboard shortcuts", shortcuts)
    # ════════════════════════════════════════════════════════
    #  3. EXPORT RR INTERVALS AS CSV
    # ════════════════════════════════════════════════════════

    def _export_rr_csv(self) -> None:
        """Export RR intervals to a lightweight CSV (no Excel dependency)."""
        self.export_ctrl.export_rr_csv()

    # ════════════════════════════════════════════════════════
    #  4. SIGNAL QUALITY BADGE (persistent header)
    # ════════════════════════════════════════════════════════

    def _update_quality_badge(self) -> None:
        """Update the persistent quality badge in the KPI bar."""
        if not hasattr(self, "_lbl_quality_badge"):
            return
        beat_corr = (self._results or {}).get("beat_corr")
        if beat_corr is None or len(beat_corr) == 0:
            self._lbl_quality_badge.configure(text="", fg_color="transparent")
            return
        mean_r = float(np.nanmean(beat_corr))
        n_bad  = int(np.sum(beat_corr < 0.90))
        pct    = 100.0 * n_bad / max(len(beat_corr), 1)
        if mean_r >= 0.95:
            col, label = GREEN,   f"● Excellent  {mean_r:.3f}"
        elif mean_r >= 0.90:
            col, label = GREEN,   f"● Good  {mean_r:.3f}"
        elif mean_r >= 0.80:
            col, label = ORANGE,  f"● Fair  {mean_r:.3f}  ({pct:.0f}% low)"
        else:
            col, label = RED,     f"● Poor  {mean_r:.3f}  ({pct:.0f}% low)"
        self._lbl_quality_badge.configure(text=label, fg_color=col,
                                           text_color="white")

    # ════════════════════════════════════════════════════════
    #  5. DRAG-AND-DROP (fixed wiring)
    # ════════════════════════════════════════════════════════

    def _setup_dnd(self) -> None:
        try:
            self.drop_target_register("DND_Files")   # provided by tkinterdnd2 at runtime  # type: ignore[attr-defined]
            self.dnd_bind("<<Drop>>", self._on_drop)  # provided by tkinterdnd2 at runtime  # type: ignore[attr-defined]
        except Exception:
            pass   # tkinterdnd2 not installed — drag-drop silently disabled

    def _on_drop(self, event) -> None:
        """Handle drag-and-drop .mat file."""
        raw = getattr(event, "data", "").strip()
        # tkinterdnd2 wraps paths with spaces in {braces}
        path = raw.strip("{}")
        if path.lower().endswith(".mat") and os.path.exists(path):
            self._load_path(path)
        else:
            # Try splitting multiple files and load the first .mat
            for part in raw.replace("{", "").replace("}", "").split():
                if part.lower().endswith(".mat") and os.path.exists(part):
                    self._load_path(part)
                    break

    # ════════════════════════════════════════════════════════
    #  6. BATCH PROCESSING DIALOG
    # ════════════════════════════════════════════════════════

    def _open_batch_dialog(self) -> None:
        """Open the batch processing window."""
        win = ctk.CTkToplevel(self)
        win.title("Batch Processing")
        win.geometry("760x560")
        win.configure(fg_color=BG)
        win.grab_set(); win.lift()

        # ── Header ────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(win, fg_color=PANEL, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="⚡  Batch Processing",
                     font=FONT_BTN_PRIMARY, text_color=TEXT,
                     anchor="w").pack(side="left", padx=SPACE_L, pady=SPACE_M)
        ctk.CTkButton(hdr, text="✗ Close", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=28).pack(side="right", padx=SPACE_L)

        body = ctk.CTkFrame(win, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=SPACE_L, pady=SPACE_M)

        # ── File list ─────────────────────────────────────────────────────
        fl = ctk.CTkFrame(body, fg_color=CARD, corner_radius=8)
        fl.pack(fill="both", expand=True)
        fl_hdr = ctk.CTkFrame(fl, fg_color="transparent")
        fl_hdr.pack(fill="x", padx=SPACE_M, pady=(SPACE_M, SPACE_S))
        ctk.CTkLabel(fl_hdr, text="Files to process",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").pack(side="left")
        ctk.CTkButton(fl_hdr, text="+ Add folder", height=26,
                      fg_color=BLUE_DARK, hover_color=BLUE, text_color="white",
                      font=FONT_SMALL, corner_radius=5,
                      command=lambda: _add_folder()).pack(side="right")
        ctk.CTkButton(fl_hdr, text="+ Add files", height=26,
                      fg_color=BLUE_DARK, hover_color=BLUE, text_color="white",
                      font=FONT_SMALL, corner_radius=5,
                      command=lambda: _add_files()).pack(side="right", padx=(0, SPACE_S))

        file_box = ctk.CTkTextbox(fl, fg_color=BG, text_color=TEXT,
                                  font=FONT_MONO, height=180)
        file_box.pack(fill="both", expand=True, padx=SPACE_M, pady=(0, SPACE_M))

        # ── Settings row ──────────────────────────────────────────────────
        cfg = ctk.CTkFrame(body, fg_color="transparent")
        cfg.pack(fill="x", pady=(SPACE_M, SPACE_S))

        for lbl, attr, default, w in [
            ("Channel:", "bc_channel", "ECG", 80),
            ("Output folder:", "bc_outdir", str(Path.home() / "ECG_batch"), 220),
            ("Workers:", "bc_workers", "4", 44),
        ]:
            ctk.CTkLabel(cfg, text=lbl, font=FONT_SMALL,
                         text_color=MUTED).pack(side="left", padx=(0, SPACE_XS))
            ent = ctk.CTkEntry(cfg, width=w, height=28, font=FONT_LABEL,
                               fg_color=BG, border_color=BORDER2, text_color=TEXT)
            ent.insert(0, default)
            ent.pack(side="left", padx=(0, SPACE_L))
            setattr(self, f"_batch_{attr}", ent)

        ctk.CTkButton(cfg, text="📁 Browse", height=28, width=70,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_SMALL,
                      command=lambda: (
                          _d := filedialog.askdirectory(),
                          self._batch_bc_outdir.delete(0, "end") or
                          self._batch_bc_outdir.insert(0, _d) if _d else None
                      )).pack(side="left")

        # ── Progress area ─────────────────────────────────────────────────
        prog_frame = ctk.CTkFrame(body, fg_color="transparent")
        prog_frame.pack(fill="x", pady=(SPACE_S, SPACE_S))
        prog_bar = ctk.CTkProgressBar(prog_frame, height=8, mode="determinate",
                                      progress_color=BLUE)
        prog_bar.set(0)
        prog_bar.pack(fill="x", pady=(0, SPACE_S))
        lbl_prog = ctk.CTkLabel(prog_frame, text="", font=FONT_SMALL,
                                text_color=MUTED, anchor="w")
        lbl_prog.pack(fill="x")

        run_btn = ctk.CTkButton(body, text="▶▶  Start Batch",
                                fg_color=GREEN, hover_color=GREEN_DARK,
                                text_color="white",
                                font=FONT_BTN_PRIMARY, height=34, corner_radius=8)
        run_btn.pack(fill="x", pady=(SPACE_S, 0))

        _filepaths: "list[str]" = []
        _bp: "Any" = None

        def _add_files():
            paths = filedialog.askopenfilenames(
                filetypes=[("MATLAB", "*.mat"), ("All", "*.*")],
                title="Select .mat files")
            for p in paths:
                if p not in _filepaths:
                    _filepaths.append(p)
                    file_box.insert("end", p + "\n")

        def _add_folder():
            d = filedialog.askdirectory(title="Select folder containing .mat files")
            if not d:
                return
            import glob
            for p in sorted(glob.glob(os.path.join(d, "*.mat"))):
                if p not in _filepaths:
                    _filepaths.append(p)
                    file_box.insert("end", p + "\n")

        def _start():
            nonlocal _bp
            if not _filepaths:
                messagebox.showwarning("No files", "Add .mat files first.")
                return
            out_dir = self._batch_bc_outdir.get().strip()  # type: ignore[union-attr]
            ch      = self._batch_bc_channel.get().strip()  # type: ignore[union-attr]
            try:
                n_workers = int(self._batch_bc_workers.get())  # type: ignore[union-attr]
            except Exception:
                n_workers = 2
            vlf, lf, hf = self._get_freq_bands()
            params = self._snapshot_params()
            params.update({
                "channel":   ch or "ECG",
                "threshold": float(self.sl_thr.get()) if self.sl_thr else 0.5,  # type: ignore[union-attr]
                "vlf_band":  list(vlf), "lf_band": list(lf), "hf_band": list(hf),
                "subject":   self.ent_subject.get().strip() or Path(_filepaths[0]).stem,
            })
            run_btn.configure(state="disabled", text="Running…")
            n = len(_filepaths)
            prog_bar.set(0)

            def _cb(done: int, total: int, stem: str) -> None:
                win.after(0, lambda: (
                    prog_bar.set(done / total),
                    lbl_prog.configure(
                        text=f"{done}/{total}  —  {stem}  {'✓' if done==total else '…'}")
                ))

            from batch import BatchProcessor
            _bp = BatchProcessor(_filepaths, params, out_dir,
                                  progress_cb=_cb, max_workers=n_workers)
            import threading
            def _worker():
                results = _bp.run()
                ok  = sum(1 for r in results if r.get("ok"))
                win.after(0, lambda: (
                    run_btn.configure(state="normal", text="▶▶  Start Batch"),
                    lbl_prog.configure(
                        text=f"Done — {ok}/{n} succeeded  →  {out_dir}",
                        text_color=GREEN if ok == n else ORANGE),
                    messagebox.showinfo("Batch done",
                        f"{ok}/{n} files processed successfully.\n"
                        f"Summary → {os.path.join(out_dir, '_batch_summary.xlsx')}")
                ))
            threading.Thread(target=_worker, daemon=True).start()

        run_btn.configure(command=_start)

    # ════════════════════════════════════════════════════════
    #  7. RECORDING NOTES (per-file, saved in SQLite)
    # ════════════════════════════════════════════════════════

    def _open_notes_dialog(self) -> None:
        """Open a small dialog for free-text notes on the current recording."""
        if not self._filepath:
            messagebox.showinfo("Notes", "Load a file first.")
            return
        current = get_notes(self._filepath) if _DB_AVAILABLE else self._recording_notes

        win = ctk.CTkToplevel(self)
        win.title(f"Notes — {os.path.basename(self._filepath)}")
        win.geometry("500x340")
        win.configure(fg_color=BG)
        win.grab_set(); win.lift()

        ctk.CTkLabel(win, text="📝  Experiment notes",
                     font=FONT_SIDEBAR_HDR, text_color=TEXT,
                     anchor="w").pack(padx=SPACE_L, pady=(SPACE_L, SPACE_S), fill="x")
        ctk.CTkLabel(win, text="Saved per recording (animal ID, drug, dose, time…)",
                     font=FONT_KPI_LABEL, text_color=MUTED, anchor="w").pack(
            padx=SPACE_L, pady=(0, SPACE_S), fill="x")
        txt = ctk.CTkTextbox(win, fg_color=CARD, text_color=TEXT,
                             font=FONT_LABEL, border_width=1,
                             border_color=BORDER2, height=180)
        txt.pack(fill="both", expand=True, padx=SPACE_L, pady=(0, SPACE_M))
        txt.insert("1.0", current)

        def _save():
            notes = txt.get("1.0", "end").strip()
            self._recording_notes = notes
            if _DB_AVAILABLE and self._filepath:
                set_notes(self._filepath, notes)
            self._set_status("Notes saved ✓", GREEN)
            win.destroy()

        btn_row = ctk.CTkFrame(win, fg_color="transparent")
        btn_row.pack(fill="x", padx=SPACE_L, pady=(0, SPACE_L))
        ctk.CTkButton(btn_row, text="✔  Save", command=_save,
                      fg_color=GREEN, hover_color=GREEN_DARK, text_color="white",
                      font=FONT_BTN_PRIMARY, height=30, corner_radius=6).pack(side="left")
        ctk.CTkButton(btn_row, text="✗ Cancel", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=30, corner_radius=6).pack(
            side="left", padx=(SPACE_M, 0))

    # ════════════════════════════════════════════════════════
    #  8. ANNOTATION TIMELINE (event markers on RR plot)
    # ════════════════════════════════════════════════════════

    def _open_annotation_dialog(self) -> None:
        """Add / manage time annotations shown on the tachogram and plots."""
        win = ctk.CTkToplevel(self)
        win.title("Annotation Timeline")
        win.geometry("620x480")
        win.configure(fg_color=BG)
        win.grab_set(); win.lift()

        _COLORS = [ORANGE_DARK, BLUE, RED, GREEN_DARK, PURPLE, TEAL]
        dur = float(len(self._signal_flt)) / self._fs if self._signal_flt is not None and self._fs else 0.0

        ctk.CTkLabel(win, text="⏱  Event annotations",
                     font=FONT_CARD_TITLE, text_color=TEXT,
                     anchor="w").pack(padx=SPACE_L, pady=(SPACE_L, SPACE_XS), fill="x")
        ctk.CTkLabel(win,
                     text="Markers appear as coloured vertical lines on the RR tachogram and all time-domain plots.",
                     font=FONT_KPI_LABEL, text_color=MUTED, anchor="w",
                     wraplength=590).pack(padx=SPACE_L, pady=(0, SPACE_M), fill="x")

        # Existing annotations list
        list_frame = ctk.CTkScrollableFrame(win, fg_color=BG, height=200)
        list_frame.pack(fill="both", expand=True, padx=SPACE_M, pady=(0, SPACE_M))

        def _refresh_list() -> None:
            for w in list_frame.winfo_children():
                w.destroy()
            for i, ann in enumerate(self._annotations):
                row = ctk.CTkFrame(list_frame, fg_color=CARD, corner_radius=5)
                row.pack(fill="x", pady=(0, SPACE_XS))
                col = ann.get("color", ORANGE_DARK)
                ctk.CTkFrame(row, width=6, fg_color=col,
                             corner_radius=3).pack(side="left", fill="y", padx=(0, SPACE_S))
                ctk.CTkLabel(row, text=ann.get("label", "Event"),
                             font=FONT_LABEL, text_color=TEXT, anchor="w").pack(
                    side="left", padx=(0, SPACE_M))
                ctk.CTkLabel(row,
                             text=f"t={ann['t_start']:.1f} – {ann['t_end']:.1f} s",
                             font=FONT_SMALL, text_color=MUTED).pack(side="left")
                ctk.CTkButton(row, text="✗", width=24, height=24,
                              fg_color=BORDER, hover_color=RED, text_color=MUTED,
                              font=FONT_SMALL,
                              command=lambda ii=i: (
                                  self._annotations.pop(ii),
                                  _refresh_list(),
                                  self._redraw_annotations()
                              )).pack(side="right", padx=SPACE_S)
            if not self._annotations:
                ctk.CTkLabel(list_frame, text="No annotations yet",
                             font=FONT_SMALL, text_color=MUTED).pack(pady=SPACE_L)

        _refresh_list()

        # Add new annotation
        add_card = ctk.CTkFrame(win, fg_color=CARD, corner_radius=8)
        add_card.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
        ctk.CTkLabel(add_card, text="Add annotation",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").pack(padx=SPACE_M, pady=(SPACE_M, SPACE_S), fill="x")
        row1 = ctk.CTkFrame(add_card, fg_color="transparent")
        row1.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
        ctk.CTkLabel(row1, text="Label:", font=FONT_SMALL,
                     text_color=MUTED, width=44).pack(side="left")
        ent_lbl = ctk.CTkEntry(row1, width=140, height=26, font=FONT_LABEL,
                               fg_color=BG, border_color=BORDER2, text_color=TEXT,
                               placeholder_text="Drug injection")
        ent_lbl.pack(side="left", padx=(0, SPACE_M))
        ctk.CTkLabel(row1, text="Start (s):", font=FONT_SMALL,
                     text_color=MUTED, width=56).pack(side="left")
        ent_ts = ctk.CTkEntry(row1, width=68, height=26, font=FONT_LABEL,
                              fg_color=BG, border_color=BORDER2, text_color=TEXT)
        ent_ts.pack(side="left", padx=(0, SPACE_M))
        ctk.CTkLabel(row1, text="End (s):", font=FONT_SMALL,
                     text_color=MUTED, width=50).pack(side="left")
        ent_te = ctk.CTkEntry(row1, width=68, height=26, font=FONT_LABEL,
                              fg_color=BG, border_color=BORDER2, text_color=TEXT)
        ent_te.insert(0, str(int(dur)))
        ent_te.pack(side="left")

        # Colour picker
        _col_var = tk.StringVar(value=_COLORS[0])
        row2 = ctk.CTkFrame(add_card, fg_color="transparent")
        row2.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_M))
        ctk.CTkLabel(row2, text="Colour:", font=FONT_SMALL,
                     text_color=MUTED, width=44).pack(side="left")
        for c in _COLORS:
            b = tk.Button(row2, bg=c, width=2, relief="flat", cursor="hand2",
                          command=lambda cc=c: _col_var.set(cc))
            b.pack(side="left", padx=SPACE_XS, pady=SPACE_XS)

        def _add():
            try:
                ts = float(ent_ts.get())
                te = float(ent_te.get())
            except ValueError:
                messagebox.showwarning("Invalid", "Enter numeric start/end times.")
                return
            if ts > te:
                ts, te = te, ts
            self._annotations.append({
                "label":   ent_lbl.get().strip() or "Event",
                "t_start": ts, "t_end": te,
                "color":   _col_var.get(),
            })
            _refresh_list()
            self._redraw_annotations()

        ctk.CTkButton(row2, text="+ Add", height=26, width=70,
                      fg_color=GREEN, hover_color=GREEN_DARK, text_color="white",
                      font=FONT_SMALL, corner_radius=5,
                      command=_add).pack(side="right")

        ctk.CTkButton(win, text="Close", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=28).pack(pady=(0, SPACE_M))

    def _redraw_annotations(self) -> None:
        """Re-render the RR tachogram to reflect updated annotations."""
        if self._results is not None:
            try:
                self._plot_rr(self._results)
            except Exception as exc:
                log.debug("_redraw_annotations: %s", exc)

    # ════════════════════════════════════════════════════════
    #  9. COMPARE SEGMENTS — statistical test (Wilcoxon)
    # ════════════════════════════════════════════════════════

    @staticmethod
    def _wilcoxon_test(a: "np.ndarray", b: "np.ndarray") -> "tuple[float, str]":
        """Mann-Whitney U test for two independent RR series.

        Returns (p_value, interpretation_string).
        Falls back gracefully if scipy is unavailable.
        """
        try:
            from scipy.stats import mannwhitneyu
            if len(a) < 5 or len(b) < 5:
                return float("nan"), "n<5"
            result: Any = mannwhitneyu(a, b, alternative="two-sided")
            p_val = getattr(result, "pvalue", None)
            p = float(p_val if p_val is not None else result[1])
            interp = ("**** p<0.0001" if p < 0.0001 else
                      "*** p<0.001"   if p < 0.001  else
                      "** p<0.01"     if p < 0.01   else
                      "* p<0.05"      if p < 0.05   else
                      "ns")
            return p, interp
        except Exception:
            return float("nan"), "—"

    # ════════════════════════════════════════════════════════
    #  10. ARRHYTHMIA EPISODE PDF (one strip per episode)
    # ════════════════════════════════════════════════════════

    def _export_arrhythmia_pdf(self) -> None:
        """Export a PDF with one annotated ECG strip per arrhythmia episode."""
        self.export_ctrl.export_arrhythmia_pdf()

    # ════════════════════════════════════════════════════════
    #  11. LIVE THEME TOGGLE (no restart needed)
    # ════════════════════════════════════════════════════════

    def _toggle_dark_live(self) -> None:
        """Switch dark ↔ light and rebuild the UI without restarting."""
        from theme import THEME, apply_theme_config
        THEME.is_dark = not THEME.is_dark
        apply_theme_config(THEME)
        THEME.save()
        ui_state = self._snapshot_ui_state()
        self._rebuild_ui()
        self._restore_ui_state(ui_state)
        mode = "Dark" if THEME.is_dark else "Light"
        self._set_status(f"Theme: {mode} mode applied", BLUE)

    # ════════════════════════════════════════════════════════
    #  SIDEBAR WIDGET HELPERS
    # ════════════════════════════════════════════════════════


    def _sidebar_sep(self, parent) -> None:
        ctk.CTkFrame(parent, height=1, fg_color=BORDER).pack(
            fill="x", padx=SPACE_L, pady=SPACE_M)

    def _sidebar_hdr(self, parent, text: str) -> None:
        ctk.CTkLabel(parent, text=text, font=FONT_SIDEBAR_HDR,
                     text_color=RED, anchor="w").pack(padx=SPACE_L, fill="x", pady=(SPACE_M, SPACE_XS))

    def _sidebar_entry(self, parent, label: str, attr: str, default: str, pad: dict) -> None:
        """Add a labelled entry and bind it to self.ent_<attr>."""
        ctk.CTkLabel(parent, text=label, font=FONT_SMALL,
                     text_color=MUTED, anchor="w").pack(**pad, fill="x")
        e = ctk.CTkEntry(parent, font=FONT_LABEL, height=30, fg_color=BG,
                          border_color=BORDER2, text_color=TEXT)
        e.insert(0, default)
        e.pack(**pad, fill="x", pady=(SPACE_XS, SPACE_M))
        setattr(self, f"ent_{attr}", e)

    def _sidebar_entry_row(self, parent, pad: dict, items: list[tuple]) -> None:
        """Render a horizontal row of (label, attr, default) entry pairs."""
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(**pad, fill="x", pady=(0, SPACE_S))
        for i, (label, attr, default) in enumerate(items):
            col = ctk.CTkFrame(row, fg_color="transparent")
            col.pack(side="left", fill="x", expand=True,
                     padx=(0, SPACE_S) if i < len(items) - 1 else 0)
            ctk.CTkLabel(col, text=label, font=FONT_SMALL, text_color=MUTED).pack(anchor="w")
            e = ctk.CTkEntry(col, font=FONT_LABEL, height=28, fg_color=BG,
                              border_color=BORDER2, text_color=TEXT)
            e.insert(0, default)
            e.pack(fill="x")
            setattr(self, f"ent_{attr}", e)

    def _switch(self, parent, label: str, pad: dict, default_on: bool = False) -> ctk.CTkSwitch:
        sw = ctk.CTkSwitch(parent, text=label, font=FONT_LABEL,
                            text_color=MUTED, progress_color=BLUE,
                            button_color=BORDER2)
        if default_on:
            sw.select()
        sw.pack(**pad, anchor="w", pady=(0, SPACE_S))
        return sw

    def _btn(self, parent, text: str, command, pad: dict,
             fg: str = BORDER, h: int = 28, bold: bool = False) -> ctk.CTkButton:
        is_dark = fg in (BLUE, RED)
        _h = max(24, int(h * THEME.font_scale))
        btn = ctk.CTkButton(
            parent, text=text, command=command,
            fg_color=fg,
            hover_color=(BLUE_HOVER if fg == BLUE else RED_DARK if fg == RED else BORDER2),
            text_color="white" if is_dark else MUTED,
            font=FONT_BTN_PRIMARY if bold else FONT_BTN_SEC, height=_h, corner_radius=8,
        )
        btn.pack(**pad, fill="x", pady=(0, SPACE_S))
        return btn

    def _reset_params(self) -> None:
        """Reset all filter/detection parameters to MouseECG species defaults."""
        fp = FilterParams()   # constructed with all dataclass defaults

        _fields: list[tuple[str, str]] = [
            ("ent_channel", fp.channel),
            ("ent_fs",      str(fp.fs)),
            ("ent_t_start", str(fp.t_start)),
            ("ent_t_end",   str(fp.t_end)),
            ("ent_lp",      str(fp.lp)),
            ("ent_hp",      str(int(fp.hp))),
            ("ent_minrr",   str(int(fp.min_rr_ms))),  # physiological floor
            # peak_distance_ms stored on self._peak_distance_ms (no dedicated widget)
            ("ent_sg_target_fs", str(fp.sg_target_fs)),
            ("ent_sg_window_ms", "20"),
        ]
        for attr, val in _fields:
            w = getattr(self, attr, None)
            if w is None:
                continue
            w.delete(0, "end")
            w.insert(0, val)

        # Threshold slider + entry
        try:
            self.sl_thr.set(fp.thresh)  # type: ignore[union-attr]
            self.lbl_thr.configure(text=f"Sensitivity:  {fp.thresh:.2f}")
        except Exception as e:
            log.debug("sl_thr/lbl_thr restore failed: %s", e)
        try:
            self.ent_thr.delete(0, "end")  # type: ignore[union-attr]
            self.ent_thr.insert(0, str(fp.thresh))  # type: ignore[union-attr]
        except Exception as e:
            log.debug("ent_thr restore failed: %s", e)

        # Epoch entries
        try:
            self.ent_epoch.delete(0, "end")
            self.ent_epoch.insert(0, str(int(MouseECG.EPOCH_DEFAULT_S)))
            self.ent_overlap.delete(0, "end")
            self.ent_overlap.insert(0, "0")
        except Exception as e:
            log.debug("epoch/overlap entry restore failed: %s", e)

        # Switches — match FilterParams defaults
        _sw_defaults: list[tuple[str, bool]] = [
            ("sw_notch",         fp.notch_filter),
            ("sw_artifact",      fp.artifact_correction),
            ("sw_no_filter",     fp.no_filter),
            ("sw_epoch",         fp.auto_epochs),
            ("sw_invert_signal", fp.invert_signal),
        ]
        for attr, on in _sw_defaults:
            w = getattr(self, attr, None)
            if w is None:
                continue
            w.select() if on else w.deselect()

        # ComboBox
        try:
            self.cb_clean.set(fp.clean_method)  # type: ignore[union-attr]
        except Exception as e:
            log.debug("cb_clean restore failed: %s", e)
        # Detection method reset
        try:
            if self.cb_det_method is not None:
                self.cb_det_method.set("SG + Derivative (10 kHz)")  # type: ignore[union-attr]
                self._on_det_method_change("SG + Derivative (10 kHz)")
        except Exception as e:
            log.debug("cb_det_method reset failed: %s", e)

        # Collapse advanced filters on reset
        try:
            if getattr(self, "_adv_filters_open", False):
                self._btn_adv_flt.invoke()
        except Exception as e:
            log.debug("_btn_adv_flt invoke failed: %s", e)

        self._set_status("Parameters reset to mouse ECG defaults ✓", GREEN)

    # ════════════════════════════════════════════════════════
    #  GENERAL HELPERS
    # ════════════════════════════════════════════════════════

    def _textbox(self, parent, h: int = 180, padx: int = 0, expand: bool = False):
        kwargs = dict(font=FONT_BODY, fg_color="transparent", text_color=TEXT,
                      border_width=0, scrollbar_button_color=BORDER,
                      scrollbar_button_hover_color=BORDER2)
        if h > 0:
            kwargs["height"] = h
        tb = ctk.CTkTextbox(parent, **kwargs)  # type: ignore[arg-type]
        if h < 0 or expand:
            tb.pack(fill="both", expand=True, padx=padx, pady=(0, SPACE_S))
        else:
            tb.pack(fill="x", padx=padx, pady=(0, SPACE_S))
        return tb

    def _set_textbox(self, widget, text: str,
                     tsv: "str | None" = None) -> None:
        """Update a CTkTextbox with *text* and optionally store *tsv* for
        'Copy as TSV' (clipboard format that pastes into Excel cell-by-cell).
        """
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", text)
        widget.configure(state="disabled")
        wid = id(widget)
        if tsv is not None:
            self._tsv_store[wid] = tsv
            self._bind_copy_menu(widget)
        elif wid in self._tsv_store:
            del self._tsv_store[wid]

    def _bind_copy_menu(self, widget) -> None:
        """Attach a right-click context menu with copy options to *widget*."""
        def _show_menu(event):
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(
                label="📋  Copy as TSV  (paste into Excel)",
                command=lambda: self._copy_tsv(widget))
            menu.add_command(
                label="📄  Copy as plain text",
                command=lambda: self._copy_plain(widget))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
        # Bind to the underlying tk widget inside CTkTextbox
        try:
            inner = widget._textbox  # CTkTextbox wraps a tk.Text
            inner.bind("<Button-3>", _show_menu, add=True)
        except AttributeError:
            widget.bind("<Button-3>", _show_menu, add=True)

    def _copy_tsv(self, widget) -> None:
        """Copy the TSV version of this widget's data to the clipboard."""
        tsv = self._tsv_store.get(id(widget))
        if not tsv:
            self._copy_plain(widget)
            return
        self.clipboard_clear()
        self.clipboard_append(tsv)
        self._set_status("Copied as TSV — paste into Excel ✓", GREEN)

    def _copy_plain(self, widget) -> None:
        """Copy the plain-text content of *widget* to the clipboard."""
        try:
            text = widget.get("1.0", "end")
        except Exception:
            text = ""
        self.clipboard_clear()
        self.clipboard_append(text)
        self._set_status("Copied as plain text ✓", GREEN)

    def _set_status(self, text: str, color: str = MUTED) -> None:
        self.lbl_status.configure(text=text, text_color=color)

    def _widget_float(self, widget: "ctk.CTkEntry", default: float = 0.0) -> float:
        """Read a float from a CTk Entry widget.  Returns *default* on any error.

        Use this when the input is definitely a widget.  Misspelled attribute
        names produce an AttributeError rather than returning a silent 0.0,
        making bugs in callers visible during development.
        """
        try:
            v = float(widget.get())
            if not np.isfinite(v):
                log.debug("_widget_float: non-finite value from %r — using %s", widget, default)
                return default
            return v
        except (ValueError, TypeError, AttributeError) as exc:
            log.debug("_widget_float: could not parse widget %r — using %s: %s",
                      widget, default, exc)
            return default

    def _safe_float(self, widget_or_value: "Any", default: float = 0.0) -> float:
        """Safely extract a float from a CTk widget or a raw scalar value.

        Kept for backward compatibility.  Prefer _widget_float() when the
        caller knows the input is a widget.
        Returns *default* if the value is missing, empty, or non-numeric.
        """
        try:
            raw = widget_or_value.get() if hasattr(widget_or_value, "get") else widget_or_value
            v   = float(raw)
            if not np.isfinite(v):
                log.debug("_safe_float: non-finite value %r — using default %s", raw, default)
                return default
            return v
        except (ValueError, TypeError) as exc:
            log.debug("_safe_float: could not parse %r — using default %s: %s",
                      widget_or_value, default, exc)
            return default

    @staticmethod
    def _safe_df_val(df: "pd.DataFrame | None", col: str, decimals: int = 3) -> str:
        """Return a formatted scalar from a DataFrame cell, or '—' on any error."""
        try:
            if df is None or col not in df.columns:
                return "—"
            v = float(df[col].values[0])
            return f"{v:.{decimals}f}" if np.isfinite(v) else "—"
        except Exception:
            return "—"

    @staticmethod
    def _df_to_tsv(df: "pd.DataFrame | None") -> str:
        """Convert a DataFrame to tab-separated values for Excel paste.

        First row = header (column names), subsequent rows = values.
        Pastes into Excel so each metric is in its own column.
        """
        if df is None or df.empty:
            return ""
        rows = ["	".join(str(c).replace("HRV_", "") for c in df.columns)]
        for _, row in df.iterrows():
            cells = []
            for v in row:
                try:
                    fv = float(v)
                    cells.append(f"{fv:.5g}" if np.isfinite(fv) else "")
                except (TypeError, ValueError):
                    cells.append(str(v) if v is not None else "")
            rows.append("	".join(cells))
        return "\n".join(rows)

    @staticmethod
    def _describe_to_tsv(df: "pd.DataFrame | None") -> str:
        """Convert a describe()-style DataFrame to TSV for Excel.

        Output: stat name in col A, then one column per metric.
        e.g.  stat\tPR_ms\tQRS_ms\tQT_ms
              mean\t42.1\t12.3\t68.4
        """
        if df is None or df.empty:
            return ""
        rows = ["stat	" + "	".join(str(c) for c in df.columns)]
        for stat, row in df.iterrows():
            cells = [str(stat)]
            for v in row:
                try:
                    fv = float(v)
                    cells.append(f"{fv:.5g}" if np.isfinite(fv) else "")
                except (TypeError, ValueError):
                    cells.append(str(v) if v is not None else "")
            rows.append("	".join(cells))
        return "\n".join(rows)

    @staticmethod
    def _kv_to_tsv(pairs: "list[tuple[str, float]]") -> str:
        """Convert a list of (name, value) pairs to two-column TSV."""
        header = "Metric	Value"
        rows   = [f"{k}	{v:.5g}" if np.isfinite(float(v)) else f"{k}	"
                  for k, v in pairs]
        return "\n".join([header] + rows)

    @staticmethod
    def _df_to_text(df: "pd.DataFrame | None") -> str:
        """Render every finite numeric column of a NeuroKit2 HRV DataFrame as readable text.

        Format uses dot-leaders for visual alignment without requiring a fixed-width font:
            SDNN ............  12.35
            RMSSD ...........   8.27
        """
        if df is None or df.empty:
            return "  (not computed)"
        rows = []
        for col in df.columns:
            try:
                v = float(df[col].values[0])
                if np.isfinite(v):
                    name   = col.replace("HRV_", "")
                    # Adaptive precision: integers → 0 dp, small values → 4 dp
                    if abs(v) >= 100:
                        fmt = f"{v:.1f}"
                    elif abs(v) >= 1:
                        fmt = f"{v:.3f}"
                    else:
                        fmt = f"{v:.5f}"
                    # Dot-leader padding to column 30
                    dots = "·" * max(2, 30 - len(name))
                    rows.append(f"  {name} {dots}  {fmt:>10}")
            except Exception as exc:
                log.debug("_df_to_text skip '%s': %s", col, exc)
        return "\n".join(rows) or "  (no finite values)"


# ════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()

    if not NK_AVAILABLE:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Missing dependency",
            "Install NeuroKit2:\n  pip install neurokit2",
        )
        root.destroy()
    else:
        app = ECGApp()
        app.mainloop()
