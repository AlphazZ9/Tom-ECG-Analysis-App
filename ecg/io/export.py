# -*- coding: utf-8 -*-
"""
ecg.io.export
-------------
Excel (openpyxl) and GraphPad Prism (.pzfx) exporters.
Also holds ParamInfo -- metadata for the Interpretation tab parameters.
"""
from __future__ import annotations

import io
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, cast

import dataclasses
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ecg.core.models import AnalysisResults, MouseECG

log = logging.getLogger("ecg")

@dataclasses.dataclass(frozen=True)
class ParamInfo:
    """Metadata for one analysis parameter displayed in the Interprétation tab.

    Using a frozen dataclass instead of a plain 6-tuple means:
    - Fields are named and IDE-navigable (.label, .unit, etc.)
    - Wrong number of arguments raises TypeError at module import time
    - The dict type annotation is precise: dict[str, ParamInfo]
    """
    label:        str
    unit:         str
    ref_lo:       float
    ref_hi:       float
    description:  str
    significance: str


# Knowledge base: every parameter with its meaning, units, normal range (mouse)
# and interpretation thresholds.
PARAM_INFO: dict[str, ParamInfo] = {
    # ── HR / RR ─────────────────────────────────────────────────────────
    "HR_mean": ParamInfo(
        label="Fréquence cardiaque moyenne", unit="bpm",
        ref_lo=MouseECG.HR_MIN_BPM, ref_hi=MouseECG.HR_MAX_BPM,
        description="Nombre de battements par minute, calculé comme 60 000 / RR moyen.",
        significance=(
            "Indicateur direct de l'état autonomique global. Une FC basse (< 400 bpm) "
            "peut refléter une anesthésie profonde, une défaillance cardiaque ou une "
            "bradycardie vagale. Une FC élevée (> 700 bpm) suggère un stress, une douleur "
            "ou une activation sympathique forte. La valeur de repos typique chez la souris "
            "éveillée non contrainte est 450–600 bpm."
        ),
    ),
    "RR_mean": ParamInfo(
        label="Intervalle RR moyen", unit="ms",
        ref_lo=60_000 / MouseECG.HR_MAX_BPM, ref_hi=60_000 / MouseECG.HR_MIN_BPM,
        description="Durée moyenne entre deux pics R consécutifs (inverse de la FC).",
        significance=(
            "Le RR moyen reflète le tonus chronotrope. Un allongement indique une "
            "dominance vagale (parasympathique) ; un raccourcissement indique une "
            "activation sympathique. Sert de référence pour normaliser les indices de VRC."
        ),
    ),
    "RR_SDNN": ParamInfo(
        label="SDNN", unit="ms",
        ref_lo=2.0, ref_hi=25.0,
        description="Écart-type de tous les intervalles RR sur l'enregistrement.",
        significance=(
            "Indice global de la variabilité de la fréquence cardiaque (VFC). "
            "Intègre toutes les composantes (sympathique, parasympathique, respiratoire). "
            "Valeurs typiques chez la souris éveillée : 5–20 ms. "
            "Une SDNN basse (< 3 ms) suggère une rigidité autonomique ou un stress "
            "chronique ; une SDNN très haute peut indiquer des artefacts résiduels."
        ),
    ),
    "RR_RMSSD": ParamInfo(
        label="RMSSD", unit="ms",
        ref_lo=1.5, ref_hi=20.0,
        description="Racine carrée de la moyenne des différences au carré entre RR consécutifs.",
        significance=(
            "Indice de la modulation parasympathique à court terme (tonus vagal). "
            "Corrèle fortement avec la puissance HF du spectre. "
            "La RMSSD est le marqueur de VFC recommandé pour les enregistrements courts "
            "car elle n'est pas influencée par les tendances de FC à long terme. "
            "Diminuée dans les pathologies cardiovasculaires, le stress et l'insuffisance cardiaque."
        ),
    ),
    "RR_pNN6": ParamInfo(
        label="pNN6", unit="%",
        ref_lo=0.0, ref_hi=60.0,
        description=(
            f"Pourcentage des intervalles RR successifs différant de plus de "
            f"{MouseECG.PNN_THRESHOLD} ms (seuil adapté à la souris)."
        ),
        significance=(
            f"Equivalent souris du pNN50 humain. Chez la souris (RR moyen ~120 ms), "
            f"le seuil de 50 ms représente ~42 % du RR moyen — trop élevé pour détecter "
            f"la modulation vagale normale (±5–10 ms). Le seuil de "
            f"{MouseECG.PNN_THRESHOLD} ms (~5 % du RR moyen) est physiologiquement "
            f"calibré pour la souris. Diminué en cas de rigidité autonomique ou stress "
            f"chronique. Référence : Thireau et al. 2008, Am J Physiol 294:H977."
        ),
    ),
    "LF_pct": ParamInfo(
        label="Puissance LF", unit="%",
        ref_lo=10.0, ref_hi=60.0,
        description="Part de la puissance spectrale dans la bande basse fréquence (0.4–1.5 Hz).",
        significance=(
            "Traditionnellement associée à la modulation sympathique ET parasympathique. "
            "Chez la souris, la bande LF capture principalement les oscillations "
            "baroréflexes. Une dominance LF élevée peut indiquer un tonus sympathique "
            "augmenté ou une réponse de stress. Interprétation controversée seule ; "
            "utiliser en conjonction avec HF et le ratio LF/HF."
        ),
    ),
    "HF_pct": ParamInfo(
        label="Puissance HF", unit="%",
        ref_lo=20.0, ref_hi=70.0,
        description=(
            "Part de la puissance spectrale dans la bande haute fréquence (1.5–5.0 Hz). "
            "Correspond à la fréquence respiratoire de la souris (2–4 Hz)."
        ),
        significance=(
            "Marqueur de la modulation parasympathique respiratoire (arythmie sinusale "
            "respiratoire). Une HF élevée indique un bon tonus vagal. "
            "Une HF basse est associée au stress, à la douleur, aux pathologies "
            "cardiaques et à l'inflammation. C'est la mesure fréquentielle la plus "
            "spécifique du système nerveux autonome parasympathique."
        ),
    ),
    "LFHF": ParamInfo(
        label="Ratio LF/HF", unit="—",
        ref_lo=0.5, ref_hi=4.0,
        description="Rapport entre la puissance LF et la puissance HF.",
        significance=(
            "Utilisé comme indice de la balance sympatho-vagale. Un ratio élevé (> 3) "
            "suggère une dominance sympathique ; un ratio bas (< 0.5) une dominance "
            "parasympathique. Fortement critiqué dans la littérature récente comme "
            "marqueur ambigu. À interpréter avec les valeurs absolues de LF et HF "
            "et le contexte expérimental."
        ),
    ),
    "SD1": ParamInfo(
        label="SD1 (Poincaré)", unit="ms",
        ref_lo=1.0, ref_hi=15.0,
        description="Demi-axe court de l'ellipse de Poincaré — variabilité instantanée battement-à-battement.",
        significance=(
            "Équivaut mathématiquement à RMSSD / √2. Mesure la dispersion à court terme "
            "des intervalles RR sur le diagramme de Poincaré. Reflète la modulation "
            "parasympathique rapide. Diminuée dans les insuffisances cardiaques, "
            "la neuropathie autonomique et sous anesthésie."
        ),
    ),
    "SD2": ParamInfo(
        label="SD2 (Poincaré)", unit="ms",
        ref_lo=3.0, ref_hi=30.0,
        description="Demi-axe long de l'ellipse de Poincaré — variabilité à long terme.",
        significance=(
            "Capture les oscillations lentes de la FC (LF, VLF, tendances circadiennes). "
            "SD2 > SD1 est la norme ; un ratio SD1/SD2 proche de 1 peut indiquer "
            "une perte de la dynamique à long terme de la VFC."
        ),
    ),
    "SampEn": ParamInfo(
        label="Entropie d'échantillon", unit="—",
        ref_lo=0.5, ref_hi=2.5,
        description="Mesure de la complexité et de l'imprévisibilité de la série RR (SampEn).",
        significance=(
            "Une entropie élevée indique un système cardiovasculaire sain et adaptatif, "
            "capable de répondre à des perturbations diverses. Une entropie basse reflète "
            "une régularité pathologique — observée dans l'insuffisance cardiaque, "
            "le diabète autonomique et le vieillissement. Insensible à la longueur "
            "de l'enregistrement (contrairement à ApEn)."
        ),
    ),
    "DFA_a1": ParamInfo(
        label="DFA α1 (court terme)", unit="—",
        ref_lo=0.75, ref_hi=1.25,
        description="Exposant de fluctuation sans tendance à court terme (4–16 battements).",
        significance=(
            "α1 ≈ 1.0 indique une dynamique fractal saine (corrélations à longue portée). "
            "α1 < 0.75 suggère une anti-corrélation (possible arythmie). "
            "α1 > 1.5 indique un comportement de marche aléatoire ou une tendance "
            "non-stationnaire. Marqueur pronostique fort dans les maladies cardiovasculaires "
            "chez l'humain ; applicable chez la souris pour caractériser le remodelage cardiaque."
        ),
    ),
    "PR_ms": ParamInfo(
        label="Intervalle PR", unit="ms",
        ref_lo=MouseECG.PR_NORMAL[0], ref_hi=MouseECG.PR_NORMAL[1],
        description=(
            "Durée entre le début de l'onde P et le début du complexe QRS "
            "(point Q quand disponible, sinon pic R en fallback)."
        ),
        significance=(
            "Reflète le temps de conduction auriculo-ventriculaire (nœud AV + faisceau de His). "
            "La mesure utilise le pic Q comme fin de l'intervalle quand la délinéation NK2 le "
            "détecte, ce qui correspond à la définition clinique standard. En l'absence de pic Q "
            "détecté, le pic R est utilisé comme proxy (légère surestimation de ~5–8 ms). "
            "Allongement (> 55 ms chez la souris) : bloc AV du 1er degré — souvent fonctionnel "
            "sous anesthésie vagolytique ou pathologique (fibrose, myocardite). "
            "Raccourcissement : syndrome de pré-excitation (Wolff-Parkinson-White) rare chez la souris."
        ),
    ),
    "QRS_ms": ParamInfo(
        label="Durée du complexe QRS", unit="ms",
        ref_lo=MouseECG.QRS_NORMAL[0], ref_hi=MouseECG.QRS_NORMAL[1],
        description="Durée de la dépolarisation ventriculaire (approximation pic Q → pic S).",
        significance=(
            "Le QRS de la souris est très court (8–15 ms typique) comparé à l'humain "
            "(80–120 ms). La mesure pic-à-pic (Q→S) sous-estime légèrement la durée réelle "
            "(≈ 2–4 ms) par rapport à la définition clinique début-à-fin. "
            "Un élargissement (> 20 ms) suggère un bloc de branche, "
            "une hypertrophie ventriculaire ou une ischémie myocardique. "
            "C'est le paramètre le plus sensible au remodelage structurel ventriculaire."
        ),
    ),
    "QT_ms": ParamInfo(
        label="Intervalle QT", unit="ms",
        ref_lo=MouseECG.QT_NORMAL[0], ref_hi=MouseECG.QT_NORMAL[1],
        description="Durée totale de la systole ventriculaire (dépolarisation + repolarisation).",
        significance=(
            "L'allongement du QT (> 90 ms chez la souris) est associé à un risque "
            "de torsade de pointe et de fibrillation ventriculaire — marqueur de "
            "cardiotoxicité médicamenteuse et de channelopathies (LQTS). "
            "Le QT se raccourcit normalement quand la FC augmente."
        ),
    ),
    "QTc_ms": ParamInfo(
        label="Intervalle QTc (Mitchell)", unit="ms",
        ref_lo=MouseECG.QTC_NORMAL[0], ref_hi=MouseECG.QTC_NORMAL[1],
        description=(
            "QT corrigé pour la fréquence cardiaque via la formule de Mitchell : "
            "QTc = QT / RR_s^(1/3)  (correction cubique)."
        ),
        significance=(
            "La correction de Mitchell (1998) est validée spécifiquement pour la souris. "
            "La formule de Bazett (QT / √RR) sur-corrige fortement aux FC élevées "
            "(> 600 bpm) : facteur ≈ 3,4× à 700 bpm, rendant toute comparaison impossible. "
            "Mitchell réduit ce biais : facteur ≈ 2,3× à 700 bpm. "
            "Le QTc est l'indice standard dans les études de sécurité cardiaque "
            "(cardiotoxicité, criblage de composés thérapeutiques). "
            "Référence : Mitchell et al. (1998) Am J Physiol 274:H747 ; "
            "Kmecova & Klimas (2010) Eur J Pharmacol."
        ),
    ),
}


# ════════════════════════════════════════════════════════════
#  EXCEL EXPORTER
# ════════════════════════════════════════════════════════════


class ExcelExporter:
    """Pure utility — builds formatted openpyxl workbooks with no GUI dependency.

    All methods are static/class-level.  ECGApp.build_excel_workbook() delegates
    here, passing only the data it needs.  This makes the export logic testable
    without instantiating the full application.

    Sheet layout
    ─────────────────────────────────────────────────────────
    Cover          — subject info, analysis date, file metadata
    ECG_Signal     — Time, Raw, Filtered (downsampled ≤ 50 k rows),
                     Is_RPeak column (1/0)
    HR_Statistics  — mean/min/max/SD HR, N beats
    RR_Timeseries  — beat-by-beat Time, RR, HR
    HRV_Time       — NeuroKit2 time-domain metrics
    HRV_Frequency  — frequency-domain metrics
    HRV_NonLinear  — non-linear metrics (SD1/SD2, SampEn …)
    ECG_Intervals  — PR / QRS / QT / QTc per beat
    Epochs         — epoch-level HR / SDNN / RMSSD (if computed)
    """

    # ── Style helpers ─────────────────────────────────────────────────

    @staticmethod
    def header_font() -> "Font":
        return Font(name="Calibri", bold=True, color="FFFFFF", size=11)

    @staticmethod
    def header_fill(hex_color: str = "1565C0") -> "PatternFill":
        return PatternFill("solid", fgColor=hex_color)

    @staticmethod
    def alt_fill(even: bool) -> "PatternFill":
        color = "EFF3FB" if even else "FFFFFF"
        return PatternFill("solid", fgColor=color)

    @staticmethod
    def peak_fill() -> "PatternFill":
        """Highlight fill for R-peak rows in the ECG_Signal sheet."""
        return PatternFill("solid", fgColor="FFF3E0")  # light amber

    @staticmethod
    def thin_border() -> "Border":
        side = Side(style="thin", color="D0D0D0")
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def autofit_columns(ws: "OpenpyxlWorksheet",
                        min_w: int = 10, max_w: int = 42) -> None:
        """Set each column width to fit its widest cell."""
        for col_cells in ws.columns:
            width = min_w
            for cell in col_cells:
                if cell.value is not None:
                    width = max(width, min(max_w, len(str(cell.value)) + 4))
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width  # type: ignore[arg-type]

    @classmethod
    def write_dataframe(
        cls,
        ws:           "OpenpyxlWorksheet",
        df:           "pd.DataFrame",
        header_color: str      = "1565C0",
        peak_col:     "str | None" = None,
        peak_set:     "set | None" = None,
        start_row:    int      = 1,
    ) -> None:
        """Write a DataFrame into *ws* with full formatting."""
        hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor=header_color)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
        data_font = Font(name="Calibri", size=10)
        num_align = Alignment(horizontal="right")
        border    = cls.thin_border()

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = border

        ws.row_dimensions[start_row].height = 18

        for row_idx, (_, series) in enumerate(df.iterrows(), start=1):
            is_peak = (peak_col is not None and peak_set is not None
                       and series.get(peak_col, None) in peak_set)
            fill = cls.peak_fill() if is_peak else cls.alt_fill(row_idx % 2 == 0)

            for col_idx, (col_name, value) in enumerate(series.items(), start=1):
                xrow = start_row + row_idx
                cell = ws.cell(row=xrow, column=col_idx)
                cell.font   = data_font
                cell.fill   = fill
                cell.border = border

                if isinstance(value, float):
                    if pd.isna(value):
                        cell.value = None
                    else:
                        # Store the full-precision value; only the display
                        # format is rounded. Writing round(value, 4) into the
                        # cell used to permanently truncate the exported
                        # number itself -- re-reading the .xlsx (e.g. with
                        # pandas) got the rounded figure, not the original.
                        cell.value     = float(value)  # type: ignore[assignment]
                        cell.alignment = num_align
                        cell.number_format = (
                            "0.000000E+00"
                            if abs(value) >= 1e6 or (abs(value) < 0.001 and value != 0)
                            else "0.0000"
                        )
                elif isinstance(value, (int, np.integer)):
                    cell.value     = int(value)  # type: ignore[assignment]
                    cell.alignment = num_align
                else:
                    cell.value = value

        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)  # type: ignore[assignment]

    @classmethod
    def build_workbook(
        cls,
        results:     "AnalysisResults",
        signal_flt:  "Optional[np.ndarray]",
        signal_raw:  "Optional[np.ndarray]",
        time:        "Optional[np.ndarray]",
        rpeaks_ok:   "Optional[np.ndarray]",
        fs:          int,
        filepath:    "Optional[str]",
        subject:     str,
        sig_quality: "Optional[int]",
        epoch_df:    "Optional[pd.DataFrame]",
    ) -> "Workbook":
        """Build and return a formatted openpyxl Workbook."""
        if "hr" not in results:
            raise ValueError(
                "AnalysisResults is missing the 'hr' key -- run the core "
                "analysis before building the export workbook."
            )
        hr = results["hr"]  # type: ignore[typeddict-item]
        wb = Workbook()

        # ── COVER sheet ──────────────────────────────────────────────────
        ws_cover = wb.active
        assert ws_cover is not None
        ws_cover.title = "Cover"
        ws_cover.sheet_view.showGridLines = False

        title_font = Font(name="Calibri", bold=True, size=20, color="1565C0")
        sub_font   = Font(name="Calibri", bold=True, size=12, color="333333")
        val_font   = Font(name="Calibri", size=11)
        label_fill = PatternFill("solid", fgColor="EFF3FB")

        ws_cover.column_dimensions["A"].width = 28
        ws_cover.column_dimensions["B"].width = 52
        ws_cover["A1"].value = "ECG Analysis Report"
        ws_cover["A1"].font  = title_font
        ws_cover.row_dimensions[1].height = 32

        dur_str = f"{float(time[-1]):.1f} s" if time is not None and len(time) else "—"
        cover_rows = [
            ("Subject",        subject),
            ("File",           os.path.basename(filepath or "—")),
            ("Analysis date",  datetime.now().strftime("%Y-%m-%d  %H:%M:%S")),
            ("Sampling rate",  f"{fs} Hz"),
            ("Duration",       dur_str),
            ("N beats",        str(hr["n"])),
            ("HR mean",        f"{hr['mean']:.1f} bpm"),
            ("HR range",       f"{hr['min']:.0f} – {hr['max']:.0f} bpm"),
            ("Signal quality", f"{sig_quality}%" if sig_quality is not None else "—"),
        ]
        for i, (label, value) in enumerate(cover_rows, start=3):
            lbl = ws_cover.cell(row=i, column=1, value=label)
            val = ws_cover.cell(row=i, column=2, value=value)
            lbl.font = sub_font;  lbl.fill = label_fill
            val.font = val_font;  val.fill = label_fill

        # ── ECG_SIGNAL sheet ─────────────────────────────────────────────
        ws_sig = cast(OpenpyxlWorksheet, wb.create_sheet("ECG_Signal"))
        MAX_ROWS = 50_000
        n      = len(time) if time is not None else 0
        step   = max(1, n // MAX_ROWS)
        idx    = np.arange(0, n, step)

        rpeaks_sorted = (np.sort(np.asarray(rpeaks_ok))
                          if rpeaks_ok is not None and len(rpeaks_ok) else np.array([], dtype=int))
        if step > 1 and len(rpeaks_sorted):
            # Each exported row now stands in for the whole sample window
            # [idx[k], idx[k]+step) that got collapsed into it. Checking
            # "is idx[k] itself a peak sample" (the old exact-index test)
            # missed virtually every peak once step > 1, since a detected
            # R-peak's sample index is essentially never an exact multiple
            # of step. Flag the row instead if ANY peak sample falls in
            # the window it represents.
            starts = np.searchsorted(rpeaks_sorted, idx, side="left")
            ends   = np.searchsorted(rpeaks_sorted, idx + step, side="left")
            is_rpeak_col = (ends > starts).astype(int).tolist()
        else:
            rp_set = set(rpeaks_sorted.tolist())
            is_rpeak_col = [1 if i in rp_set else 0 for i in idx]

        sig_df = pd.DataFrame({
            "Time_s":   time[idx]    if time       is not None else [],
            "Raw_norm": signal_raw[idx] if signal_raw is not None else np.full(len(idx), np.nan),
            "Filtered": signal_flt[idx] if signal_flt is not None else np.full(len(idx), np.nan),
            "Is_RPeak": is_rpeak_col,
        })

        cls.write_dataframe(ws_sig, sig_df, header_color="37474F",
                            peak_col="Is_RPeak", peak_set={1})
        cls.autofit_columns(ws_sig)
        for col, w in [("A", 14), ("B", 14), ("C", 14), ("D", 11)]:
            ws_sig.column_dimensions[col].width = w

        if step > 1:
            note_row = len(sig_df) + 3
            ws_sig.cell(row=note_row, column=1,
                        value=f"Note: signal downsampled 1:{step} for export "
                              f"({n:,} → {len(idx):,} rows).")
            ws_sig.cell(row=note_row, column=1).font = Font(italic=True, color="888888", size=9)

        # ── HR_STATISTICS sheet ──────────────────────────────────────────
        ws_hr = cast(OpenpyxlWorksheet, wb.create_sheet("HR_Statistics"))
        hr_df = pd.DataFrame([{
            "HR_mean_bpm": round(hr["mean"], 2),
            "HR_min_bpm":  round(hr["min"],  2),
            "HR_max_bpm":  round(hr["max"],  2),
            "HR_std_bpm":  round(hr["std"],  2),
            "N_beats":     hr["n"],
            "Duration_s":  round(float(time[-1]), 1) if time is not None and len(time) else None,
            "Fs_Hz":       fs,
        }])
        cls.write_dataframe(ws_hr, hr_df, header_color="2E7D32")
        cls.autofit_columns(ws_hr)

        # ── Metric DataFrames ────────────────────────────────────────────
        sheet_specs: "list[tuple[str, Optional[pd.DataFrame], str]]" = [
            ("RR_Timeseries", results.get("rr_df"),       "E65100"),
            ("HRV_Time",      results.get("hrv_time"),    "1565C0"),
            ("HRV_Frequency", results.get("hrv_freq"),    "6A1B9A"),
            ("HRV_NonLinear", results.get("hrv_nonlin"),  "00695C"),
            ("ECG_Intervals", results.get("intervals"),   "AD1457"),
        ]
        for name, df, color in sheet_specs:
            if df is None or df.empty:
                continue
            ws = cast(OpenpyxlWorksheet, wb.create_sheet(name))
            cls.write_dataframe(ws, df, header_color=color)
            cls.autofit_columns(ws)

        # ── EPOCHS sheet (optional) ──────────────────────────────────────
        if epoch_df is not None and not epoch_df.empty:
            ws_ep = cast(OpenpyxlWorksheet, wb.create_sheet("Epochs"))
            cls.write_dataframe(ws_ep, epoch_df, header_color="455A64")
            cls.autofit_columns(ws_ep)

        return wb

    @classmethod
    def add_annotations_sheet(cls, wb, annotations: "list[dict]") -> None:
        """Append an Annotations sheet to an existing workbook."""
        if not annotations:
            return
        ann_df = pd.DataFrame([{
            "Start_s":    a["t_start"],
            "End_s":      a["t_end"],
            "Duration_s": round(a["t_end"] - a["t_start"], 4),
            "Label":      a.get("label", ""),
            "Color":      a.get("color", ""),
        } for a in annotations])
        ws = cast(OpenpyxlWorksheet, wb.create_sheet("Annotations"))
        cls.write_dataframe(ws, ann_df, header_color="4A148C")
        cls.autofit_columns(ws)


# ── Predefined annotation colours ─────────────────────────────────────────────
# Fixed swatch options for user-created annotations -- intentionally NOT tied
# to the live UI theme (an annotation colour a user picked shouldn't change
# meaning if they later switch the app's appearance preset). Four of these
# used to be imported from theme.py (coincidentally matching its palette at
# the time), which pulled a customtkinter import (and its module-level
# ctk.set_appearance_mode()/set_default_color_theme() calls) into every
# process that imports export.py -- including batch.py's subprocess workers,
# which are documented as having no Tkinter/CTk dependency. Hardcoded here
# to match the previous values exactly and keep that contract true.
ANNOTATION_COLORS: list[tuple[str, str]] = [
    ("#E65100", "Orange"),
    ("#1A56DB", "Blue"),
    ("#AD1457", "Pink"),
    ("#2E7D32", "Green"),
    ("#6A1B9A", "Purple"),
    ("#00838F", "Teal"),
    ("#F9A825", "Yellow"),
    ("#37474F", "Dark grey"),
]



class PrismExporter:
    """Pure utility — exports ECG analysis results to GraphPad Prism .pzfx format.

    Uses the official `pzfx` Python library (pip install pzfx) which produces
    correctly structured XML accepted by Prism 5–10 without modification.

    Tables exported
    ───────────────
    XY        — RR intervals & HR tachogram, ECG intervals timeline,
                Rolling HRV timeline, Epoch HRV
    Column    — HRV time-domain summary, frequency-domain summary,
                non-linear summary  (with named row titles)
    OneWay    — ECG intervals per beat (PR / QRS / QT / QTc distributions)
    """

    @staticmethod
    def _require_pzfx() -> None:
        try:
            import pzfx  # noqa: F401
        except ImportError:
            raise RuntimeError(
                "The 'pzfx' package is required for GraphPad Prism export.\n"
                "Install it with: pip install pzfx"
            )

    @classmethod
    def build_and_write(
        cls,
        path:            "str | Path",
        results:         "AnalysisResults",
        rolling_hrv_df:  "Optional[pd.DataFrame]" = None,
        epoch_df:        "Optional[pd.DataFrame]"  = None,
        subject:         str = "",
        context_label:   str = "",
        arrhythmia_events: "Optional[list]" = None,
        beat_corr: "Optional[np.ndarray]" = None,
        beat_peak_times: "Optional[np.ndarray]" = None,
        segment_a: "Optional[dict]" = None,
        segment_b: "Optional[dict]" = None,
        fs: "Optional[float]" = None,
        recording_notes: str = "",
    ) -> int:
        """Build all tables and write directly to *path*.

        Returns the number of tables written.
        Raises RuntimeError if pzfx is not installed.
        """
        cls._require_pzfx()
        from pzfx import write_pzfx

        tables:    "dict[str, pd.DataFrame]" = {}
        row_names: "list[bool]"              = []
        x_cols:    "list[str | int]"         = []

        def _add(name: str, df: "pd.DataFrame",
                 x_col: "str | int" = 0,
                 row_name: bool = False) -> None:
            tables[name]    = df
            x_cols.append(x_col)
            row_names.append(row_name)

        # ── 1. RR intervals & HR tachogram (XY) ─────────────────────────
        rr_df = results.get("rr_df")
        if rr_df is not None and not rr_df.empty:
            df = rr_df[["Time_s", "RR_ms", "HR_bpm"]].copy()
            df.columns = pd.Index(["Time (s)", "RR (ms)", "HR (bpm)"])
            _add("RR Intervals & HR", df, x_col="Time (s)")  # type: ignore[arg-type]

        # ── 2. HRV time-domain summary (Column with row titles) ──────────
        hrv_t = results.get("hrv_time")
        if hrv_t is not None and not hrv_t.empty:
            key_map = [
                ("HRV_MeanNN",   "MeanNN (ms)"),
                ("HRV_SDNN",     "SDNN (ms)"),
                ("HRV_RMSSD",    "RMSSD (ms)"),
                ("HRV_pNN6",     "pNN6 (%)"),
                ("HRV_pNN50",    "pNN50 (%)"),
                ("HRV_CVNN",     "CVNN"),
                ("HRV_MedianNN", "MedianNN (ms)"),
                ("HRV_MadNN",    "MadNN (ms)"),
                ("HRV_IQRNN",    "IQRNN (ms)"),
            ]
            rows = {
                label: float(hrv_t[col].values[0])
                for col, label in key_map
                if col in hrv_t.columns
                and pd.notna(hrv_t[col].values[0])
                and pd.api.types.is_float_dtype(hrv_t[col])
            }
            if rows:
                df = pd.DataFrame({"Value": rows.values()},
                                  index=pd.Index(list(rows.keys())))
                _add("HRV — Time Domain", df, x_col=0, row_name=True)

        # ── 3. HRV frequency-domain summary (Column) ────────────────────
        hrv_f = results.get("hrv_freq")
        if hrv_f is not None and not hrv_f.empty:
            key_map_f = [
                ("HRV_LF",    "LF (ms²)"),
                ("HRV_HF",    "HF (ms²)"),
                ("HRV_VLF",   "VLF (ms²)"),
                ("HRV_TP",    "Total Power (ms²)"),
                ("HRV_LFn",   "LF norm (%)"),
                ("HRV_HFn",   "HF norm (%)"),
                ("HRV_LF_HF", "LF/HF ratio"),
                ("HRV_LFpeak","LF peak (Hz)"),
                ("HRV_HFpeak","HF peak (Hz)"),
            ]
            rows_f = {
                label: float(hrv_f[col].values[0])
                for col, label in key_map_f
                if col in hrv_f.columns and pd.notna(hrv_f[col].values[0])
            }
            if rows_f:
                df = pd.DataFrame({"Value": rows_f.values()},
                                  index=pd.Index(list(rows_f.keys())))
                _add("HRV — Frequency Domain", df, x_col=0, row_name=True)

        # ── 4. HRV non-linear summary (Column) ──────────────────────────
        hrv_nl = results.get("hrv_nonlin")
        if hrv_nl is not None and not hrv_nl.empty:
            key_map_nl = [
                ("HRV_SD1",        "SD1 (ms)"),
                ("HRV_SD2",        "SD2 (ms)"),
                ("HRV_SD1SD2",     "SD1/SD2"),
                ("HRV_S",          "S (ms²)"),
                ("HRV_SampEn",     "SampEn"),
                ("HRV_ApEn",       "ApEn"),
                ("HRV_DFA_alpha1", "DFA α1"),
                ("HRV_DFA_alpha2", "DFA α2"),
                ("HRV_PI",         "Porta Index"),
            ]
            rows_nl = {
                label: float(hrv_nl[col].values[0])
                for col, label in key_map_nl
                if col in hrv_nl.columns and pd.notna(hrv_nl[col].values[0])
            }
            if rows_nl:
                df = pd.DataFrame({"Value": rows_nl.values()},
                                  index=pd.Index(list(rows_nl.keys())))
                _add("HRV — Non-Linear", df, x_col=0, row_name=True)

        # ── 5. ECG intervals per beat (OneWay — distributions) ───────────
        ivl = results.get("intervals")
        if ivl is not None and not ivl.empty:
            col_map = [
                ("PR_ms",  "PR (ms)"),
                ("QRS_ms", "QRS (ms)"),
                ("QT_ms",  "QT (ms)"),
                ("QTc_ms", "QTc (ms)"),
            ]
            ivl_cols = {
                label: ivl[col].dropna().values
                for col, label in col_map
                if col in ivl.columns and ivl[col].notna().sum() > 0
            }
            if ivl_cols:
                # OneWay: pad shorter columns with NaN so DataFrame is rectangular
                max_len = max(len(v) for v in ivl_cols.values())
                df = pd.DataFrame({
                    k: pd.Series(
                        np.pad(np.asarray(v, dtype=float),
                               (0, max_len - len(v)),
                               constant_values=np.nan)
                    )
                    for k, v in ivl_cols.items()
                })
                _add("ECG Intervals per Beat", df, x_col=0, row_name=False)

        # ── 6. ECG intervals timeline (XY) ──────────────────────────────
        if ivl is not None and not ivl.empty and "Time_s" in ivl.columns:
            tl_cols = {
                label: ivl[col].fillna(np.nan).values
                for col, label in [
                    ("PR_ms",  "PR (ms)"), ("QRS_ms", "QRS (ms)"),
                    ("QT_ms",  "QT (ms)"), ("QTc_ms", "QTc (ms)"),
                ]
                if col in ivl.columns
            }
            if tl_cols:
                tl_df = pd.DataFrame({"Time (s)": ivl["Time_s"].values, **tl_cols})
                _add("ECG Intervals — Timeline", tl_df, x_col="Time (s)")

        # ── 7. Rolling HRV timeline (XY) ────────────────────────────────
        if rolling_hrv_df is not None and not rolling_hrv_df.empty:
            df_r = rolling_hrv_df.copy()
            rename = {"t_mid": "Time (s)", "HR": "HR (bpm)",
                      "SDNN": "SDNN (ms)", "RMSSD": "RMSSD (ms)", "pNN6": "pNN6 (%)"}
            df_r = df_r.rename(columns={k: v for k, v in rename.items()
                                        if k in df_r.columns})
            _add("Rolling HRV Timeline", df_r, x_col="Time (s)")

        # ── 8. Epoch HRV (XY) ───────────────────────────────────────────
        if epoch_df is not None and not epoch_df.empty:
            df_e = epoch_df.copy()
            df_e.insert(0, "Time (s)",
                        ((df_e["Epoch_start_s"] + df_e["Epoch_end_s"]) / 2).round(2))
            rename_e = {"HR_mean": "HR (bpm)", "MeanNN": "MeanNN (ms)",
                        "SDNN": "SDNN (ms)", "RMSSD": "RMSSD (ms)"}
            df_e = df_e.rename(columns={k: v for k, v in rename_e.items()
                                        if k in df_e.columns})
            # Keep only numeric columns
            keep = ["Time (s)"] + [c for c in df_e.columns
                                   if c not in ("Epoch_start_s", "Epoch_end_s",
                                                "N_beats", "Time (s)")
                                   and pd.api.types.is_numeric_dtype(df_e[c])]
            _add("Epoch HRV", df_e[keep], x_col="Time (s)")  # type: ignore[arg-type]

        if not tables:
            raise ValueError("No data available for Prism export. "
                             "Run the full analysis first.")

        # Build notes info with subject/context metadata
        notes = pd.DataFrame({
            "Name":  ["Notes", "Subject", "Context"],
            "Value": [f"ECG Analysis TOM v6 — {subject}", subject, context_label],
        })

        write_pzfx(
            tables,
            str(path),
            row_names=row_names,
            x_col=x_cols,
            n_digits=4,
            notes={"Project Info": notes},
        )
        return len(tables)

