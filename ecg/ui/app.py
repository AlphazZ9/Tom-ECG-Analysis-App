# -*- coding: utf-8 -*-
"""
ecg.ui.app
----------
ECGApp -- the main application window (CustomTkinter CTk).
Imports all other ecg.* modules and wires them together.
"""
from __future__ import annotations

import logging
import os
import threading
import time
import traceback
from datetime import datetime
from typing import Any, Callable, Optional

import customtkinter as ctk  # type: ignore[import-untyped]
import tkinter as tk
from tkinter import filedialog, messagebox

import matplotlib
import matplotlib.figure
import matplotlib.ticker
from matplotlib.figure import Figure
if __name__ == "__main__":
    matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import numpy as np
# numpy 2.x may expose only trapezoid; ensure np.trapz alias exists.
if not hasattr(np, "trapz"):
    if hasattr(np, "trapezoid"):
        np.trapz = np.trapezoid  # type: ignore[attr-defined]
    else:
        def _trapz(y, x=None, dx=1.0, axis=-1):
            y_arr = np.asanyarray(y)
            if x is None:
                n = y_arr.shape[axis]
                x = np.arange(n, dtype=float) * dx
            else:
                x = np.asanyarray(x)
            # Broadcasting rules apply; use numpy.diff and slicing
            slice_all = [slice(None)] * y_arr.ndim
            slice1 = slice_all.copy(); slice2 = slice_all.copy()
            slice1[axis] = slice(1, None)
            slice2[axis] = slice(0, -1)
            dx_arr = np.diff(x)
            y_avg = (y_arr[tuple(slice1)] + y_arr[tuple(slice2)]) / 2.0
            return np.sum(y_avg * dx_arr, axis=axis)
        np.trapz = _trapz  # type: ignore[attr-defined]
import pandas as pd
from PIL import Image, ImageTk
from openpyxl import Workbook

# ── ecg.core ──────────────────────────────────────────────────────────────────
from ecg.core.models import (
    ArrhythmiaEvent, MouseECG, FilterParams, EXPERIMENTAL_CONTEXTS,
)
from ecg.core.detection import (
    detect_rr_artifacts, apply_artifact_decisions,
)
from ecg.core.wave_template import WaveTemplate
from ecg.ui.state import SignalState, DetectionState, AnalysisState, UIState, SessionState
from ecg.ui.navigation_controller import NavigationController
from ecg.ui.export_controller import ExportController
from ecg.ui.plot_controller import PlotController
from ecg.ui.session_controller import SessionController
from ecg.ui.detection_controller import DetectionController
from ecg.ui.signal_controller import SignalController
from ecg.ui.analysis_controller import AnalysisController

# ── ecg.io ────────────────────────────────────────────────────────────────────
from ecg.io.loaders import list_channels
from ecg.io.session import load_session
from ecg.io.db import (
    _DB_AVAILABLE, get_notes, set_notes,
    recent_recordings,
)

# ── ecg.ui ────────────────────────────────────────────────────────────────────
from ecg.ui.theme import (
    THEME, apply_theme_config,
    NK_AVAILABLE, APP_ICON_PATH,
    BG, PANEL, CARD, BORDER, BORDER2, TEXT, MUTED, LIGHT, PLOT,
    RED, BLUE, GREEN, ORANGE,
    BLUE_DARK, BLUE_HOVER, BLUE_MID, BLUE_DEEP,
    PURPLE, PURPLE_DARK, PINK, TEAL, TEAL_DARK,
    GREEN_DARK, ORANGE_DARK,
    RED_DARK,
    COLOR_PRIMARY, COLOR_PRIMARY_HOVER, COLOR_SUCCESS, COLOR_SUCCESS_HOVER,
    COLOR_WARNING, COLOR_WARNING_HOVER, COLOR_DANGER, COLOR_DANGER_HOVER,
    COLOR_SECONDARY, COLOR_SECONDARY_HOVER,
    SPACE_XS, SPACE_S, SPACE_M, SPACE_L,
    FONT_TITLE, FONT_SECTION_HDR, FONT_LABEL, FONT_SMALL, FONT_BODY, FONT_MONO,
    FONT_KPI_VALUE, FONT_KPI_LABEL, FONT_BTN_PRIMARY, FONT_BTN_SEC,
    FONT_SIDEBAR_HDR,
    FONT_MICRO, FONT_HINT, FONT_BADGE, FONT_SUBSECTION, FONT_CARD_TITLE,
)
from ecg.ui.widgets import make_stat_tile, make_quality_gauge
from ecg.ui.plots import CanvasSlot, style_axes
from ecg.ui.dialogs import (
    ThemeDialog, ArtifactReviewDialog,
    AnnotationManagerDialog, PacingPeriodManagerDialog,
)
from ecg.ui.wave_editor import WaveTemplateMiniEditor
from ecg.ui.sidebar import CollapsibleSection, IntervalVerifierPanel

log = logging.getLogger("ecg")

# SPACE_XS/S/M/L (padx/pady spacing scale) now live in ecg.ui.theme, the
# single source of truth — see the "Layout spacing tokens" section there.


class ECGApp(ctk.CTk):

    def __init__(self) -> None:
        super().__init__()
        self.title("ECG Analysis")
        self.geometry("1920x1080")
        self.minsize(1200, 750)
        # Open in fullscreen/maximized state
        self.after(100, lambda: self.state("zoomed"))  # Windows fullscreen
        self.configure(fg_color=PANEL)

        # Forward declarations for widgets built during _build() — lets Pylance
        # know these attributes exist even though they're assigned later.
        # (declaration moved to _init_state)
        self.ent_subject:  ctk.CTkEntry
        self.ent_fs:       ctk.CTkEntry
        self.ent_epoch:    ctk.CTkEntry
        self.ent_overlap:  ctk.CTkEntry
        self.txt_td:       ctk.CTkTextbox
        self.txt_fd:       ctk.CTkTextbox

        self._load_icon()
        self._init_state()
        self._build()
        self.after(200, self._setup_dnd)
        self._bind_keyboard_shortcuts()

    # ─── Startup helpers ──────────────────────────────────────

    def _on_window_resize(self, event) -> None:
        """Adjust sidebar width proportionally when the window is resized.

        Debounced to 80 ms so rapid drag-resize events don't flood the layout
        engine with redundant configure calls.
        """
        if event.widget is not self:
            return
        if self._thr_debounce_id is not None:
            try:
                self.after_cancel(self._thr_debounce_id)
            except Exception:
                pass
        new_w = max(220, min(340, int(event.width * 0.20)))
        self._thr_debounce_id = self.after(
            80, lambda w=new_w: self._apply_resize(w)
        )

    def _apply_resize(self, new_w: int) -> None:
        """Apply the debounced sidebar width update."""
        self._thr_debounce_id = None
        try:
            self.sidebar.configure(width=new_w)
        except Exception as e:
            log.debug("sidebar.configure width failed: %s", e)

    def _load_icon(self) -> None:
        if not APP_ICON_PATH:
            return
        try:
            # Pillow 9+ uses Image.Resampling.LANCZOS; older versions use Image.LANCZOS
            _resample = (getattr(getattr(Image, "Resampling", None), "LANCZOS", None)
                         or getattr(Image, "LANCZOS", Image.BICUBIC))  # type: ignore[attr-defined]
            img = Image.open(APP_ICON_PATH).resize((256, 256), _resample)
            # PhotoImage is accepted at runtime; Tkinter stubs are incomplete
            self.iconphoto(True, ImageTk.PhotoImage(img))  # type: ignore[arg-type]
        except Exception as exc:
            log.debug("App icon not loaded: %s", exc)

    def _init_state(self) -> None:
        """Initialise all instance variables that hold application state.

        Non-widget state lives in typed dataclasses from state.py (`self.signal`,
        `self.detection`, `self.analysis`, `self.ui`, `self.session`). The
        "Legacy state shims" property block right after this method keeps every
        existing `self._xxx` access (inside this class and in dialogs.py /
        wave_editor.py / models.py) working unchanged. New code should read/write
        the dataclasses directly.
        """
        self.signal = SignalState()
        self.detection = DetectionState()
        self.analysis = AnalysisState()
        self.ui = UIState()
        self.session = SessionState()

        self.nav_ctrl = NavigationController(self)
        self.export_ctrl = ExportController(self)
        self.plot_ctrl = PlotController(self)
        self.session_ctrl = SessionController(self)
        self.detection_ctrl = DetectionController(self)
        self.signal_ctrl = SignalController(self)
        self.analysis_ctrl = AnalysisController(self)

        # Widget registries (populated in _build)
        self._slots:       dict[str, CanvasSlot]   = {}
        self._kpi:         dict[str, ctk.CTkLabel] = {}
        # Compact top-bar mirrors of a few _kpi values (hr_mean/n_beats/dur) --
        # a SEPARATE dict since these are distinct widgets living in the top
        # bar, updated alongside (not instead of) the stat-panel tiles by the
        # same update_kpis() call.
        self._topbar_vals: dict[str, ctk.CTkLabel] = {}

        # ── Forward declarations for widgets created in _build ────────────────
        # Declared here so that methods called before the UI is fully constructed
        # (e.g. _collect_session_state called during startup restore) get a
        # predictable None rather than raising AttributeError, and so that
        # hasattr(self, ...) guards can be replaced with "is not None" checks.
        #
        # Top toolbar (built in _build_toolbar)
        self.ent_project_name:  "Optional[ctk.CTkEntry]"   = None
        self.lbl_topbar_project:"Optional[ctk.CTkLabel]"   = None
        self.lbl_topbar_file:   "Optional[ctk.CTkLabel]"   = None
        self.quality_gauge:     "Optional[ctk.CTkFrame]"   = None
        # Sidebar / session controls
        self.btn_save_session:  "Optional[ctk.CTkButton]"  = None
        self.lbl_session_info:  "Optional[ctk.CTkLabel]"   = None
        self.lbl_template_info: "Optional[ctk.CTkLabel]"   = None
        # Per-module on-demand analysis buttons + status labels
        self.btn_run_freq:      "Optional[ctk.CTkButton]"  = None
        self.btn_run_nonlin:    "Optional[ctk.CTkButton]"  = None
        self.btn_run_ivl:       "Optional[ctk.CTkButton]"  = None
        self.btn_run_arrhythmia:"Optional[ctk.CTkButton]"  = None
        self.frm_ivl_nav:       "Optional[tk.Frame]"       = None
        self._ivl_verifier:     "Optional[IntervalVerifierPanel]" = None
        self.lbl_freq_status:   "Optional[ctk.CTkLabel]"   = None
        self.lbl_nonlin_status: "Optional[ctk.CTkLabel]"   = None
        self.lbl_ivl_status:    "Optional[ctk.CTkLabel]"   = None
        # Filter / signal controls (widgets used in _collect_session_state)
        self.ent_channel:       "Optional[ctk.CTkEntry]"   = None
        self.ent_t_start:       "Optional[ctk.CTkEntry]"   = None
        self.ent_t_end:         "Optional[ctk.CTkEntry]"   = None
        self.ent_analysis_t0:   "Optional[ctk.CTkEntry]"   = None
        self.ent_analysis_t1:   "Optional[ctk.CTkEntry]"   = None
        self.lbl_analysis_window: "Optional[ctk.CTkLabel]" = None
        self.ent_lp:            "Optional[ctk.CTkEntry]"   = None
        self.ent_hp:            "Optional[ctk.CTkEntry]"   = None
        self.ent_minrr:         "Optional[ctk.CTkEntry]"   = None
        self.sw_notch:          "Optional[ctk.CTkSwitch]"  = None
        self.sw_invert_signal:  "Optional[ctk.CTkSwitch]"  = None
        self.sw_filter_preview: "Optional[ctk.CTkSwitch]"  = None
        self.cb_clean:          "Optional[ctk.CTkComboBox]" = None
        self.sl_thr:            "Optional[ctk.CTkSlider]"  = None
        self.ent_thr:           "Optional[ctk.CTkEntry]"   = None
        self.sw_permissive:     "Optional[ctk.CTkSwitch]"  = None
        self.btn_annotations:   "Optional[ctk.CTkButton]"  = None
        self.lbl_ann_count:     "Optional[ctk.CTkLabel]"   = None
        self.btn_toggle_rrhr:   "Optional[ctk.CTkButton]"  = None
        self.btn_toggle_left_panel:  "Optional[ctk.CTkButton]" = None
        self.btn_toggle_right_panel: "Optional[ctk.CTkButton]" = None
        self.lbl_panel_ann_count:    "Optional[ctk.CTkLabel]" = None
        self.lbl_panel_pacing_count: "Optional[ctk.CTkLabel]" = None
        self.btn_copy_rr:       "Optional[ctk.CTkButton]"  = None
        self.btn_copy_ivl:      "Optional[ctk.CTkButton]"  = None
        self.btn_copy_epochs:   "Optional[ctk.CTkButton]"  = None
        self.cb_det_method:     "Optional[ctk.CTkComboBox]"  = None
        self.ent_sg_target_fs:  "Optional[ctk.CTkEntry]"     = None
        self.ent_sg_window_ms:  "Optional[ctk.CTkEntry]"     = None
        self._sg_frame:         "Optional[ctk.CTkFrame]"     = None
        self.cb_qtc_formula:    "Optional[ctk.CTkComboBox]"  = None
        self.cb_freq_band:      "Optional[ctk.CTkComboBox]"  = None   # HRV band preset
        self.lbl_ml_status:     "Optional[ctk.CTkLabel]"     = None
        self.sw_verified_training: "Optional[ctk.CTkSwitch]" = None
        self.btn_train_ml:      "Optional[ctk.CTkButton]"    = None
        self.btn_save_for_training: "Optional[ctk.CTkButton]" = None
        # Sidebar / detection tab widgets (forward-declared)
        self.lbl_npeaks:           "Optional[ctk.CTkLabel]"              = None
        self.btn_review_art:       "Optional[ctk.CTkButton]"             = None
        self.lbl_file:             "Optional[ctk.CTkLabel]"              = None
        self.lbl_context_subtitle: "Optional[ctk.CTkLabel]"              = None
        self.lbl_arrhythmia_status:"Optional[ctk.CTkLabel]"              = None
        self.lbl_roll_status:      "Optional[ctk.CTkLabel]"              = None
        self.lbl_arr_event_title:  "Optional[ctk.CTkLabel]"              = None
        self._arr_card_widgets:    "list[ctk.CTkBaseClass]"              = []
        self._filter_advanced_group: "Optional[ctk.CTkFrame]"            = None
        self.lbl_filter_summary:  "Optional[ctk.CTkLabel]"              = None
        self._interp_scroll:       "Optional[ctk.CTkScrollableFrame]"    = None
        # Navigation bar widgets (forward-declared so _reset_for_new_file can update them)
        self.ent_nav_pos:       "Optional[ctk.CTkEntry]"    = None
        self.lbl_sig_duration:  "Optional[ctk.CTkLabel]"    = None
        # Intervals tab widgets
        self.ent_max_beats:     "Optional[ctk.CTkEntry]"    = None  # removed from UI but kept for compat
        # Interpretation removed
        self._interp_cards:      "dict[str, tuple[ctk.CTkLabel, ctk.CTkLabel, str]]" = {}
        self._interp_ref_labels: "dict[str, ctk.CTkLabel]" = {}
        # HRV unified tab internal state
        self._hrv_subframes:     "dict[str, ctk.CTkFrame]" = {}
        self._hrv_seg:           "Optional[ctk.CTkSegmentedButton]" = None
        self._hrv_content_area:  "Optional[ctk.CTkFrame]"           = None
        # Summary tab label widgets (populated in _build_tab_summary)
        self._sum_quality_vals:  "dict[str, ctk.CTkLabel]" = {}
        self._sum_metric_vals:   "dict[str, ctk.CTkLabel]" = {}

    # ─── Legacy state shims ───────────────────────────────────────────────
    # Backward-compatible properties for the old flat self._xxx attributes,
    # now held in state.py dataclasses. Every existing method body in this
    # class, plus the sibling modules that reach into an ECGApp instance
    # (dialogs.py: _rebuild_ui/_annotations/_time/_session_dirty/_draw_detail/
    # _update_ann_count; wave_editor.py: _wave_template; models.py:
    # _peak_distance_ms/_safe_float), keep working unchanged through these.
    # New/extracted code should read and write self.signal/.detection/
    # .analysis/.ui/.session directly instead of these shims.

    # -- SignalState --
    @property
    def _filepath(self) -> "Optional[str]":
        return self.signal.filepath
    @_filepath.setter
    def _filepath(self, value: "Optional[str]") -> None:
        self.signal.filepath = value

    @property
    def _signal_raw(self) -> "Optional[np.ndarray]":
        return self.signal.raw
    @_signal_raw.setter
    def _signal_raw(self, value: "Optional[np.ndarray]") -> None:
        self.signal.raw = value

    @property
    def _signal_raw_norm(self) -> "Optional[np.ndarray]":
        return self.signal.raw_norm
    @_signal_raw_norm.setter
    def _signal_raw_norm(self, value: "Optional[np.ndarray]") -> None:
        self.signal.raw_norm = value

    @property
    def _signal_flt(self) -> "Optional[np.ndarray]":
        return self.signal.filtered
    @_signal_flt.setter
    def _signal_flt(self, value: "Optional[np.ndarray]") -> None:
        self.signal.filtered = value

    @property
    def _time(self) -> "Optional[np.ndarray]":
        return self.signal.time
    @_time.setter
    def _time(self, value: "Optional[np.ndarray]") -> None:
        self.signal.time = value

    @property
    def _fs(self) -> int:
        return self.signal.fs
    @_fs.setter
    def _fs(self, value: int) -> None:
        self.signal.fs = value

    @property
    def _peak_distance_ms(self) -> float:
        return self.signal.peak_distance_ms
    @_peak_distance_ms.setter
    def _peak_distance_ms(self, value: float) -> None:
        self.signal.peak_distance_ms = value

    @property
    def _raw_only_loaded(self) -> bool:
        return self.signal.raw_only_loaded
    @_raw_only_loaded.setter
    def _raw_only_loaded(self, value: bool) -> None:
        self.signal.raw_only_loaded = value

    @property
    def _no_filter_mode(self) -> bool:
        return self.signal.no_filter_mode
    @_no_filter_mode.setter
    def _no_filter_mode(self, value: bool) -> None:
        self.signal.no_filter_mode = value

    @property
    def _signal_inverted(self) -> bool:
        return self.signal.inverted
    @_signal_inverted.setter
    def _signal_inverted(self, value: bool) -> None:
        self.signal.inverted = value

    # -- DetectionState --
    @property
    def _rpeaks_ok(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_ok
    @_rpeaks_ok.setter
    def _rpeaks_ok(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_ok = value

    @property
    def _rpeaks_rej(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_rej
    @_rpeaks_rej.setter
    def _rpeaks_rej(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_rej = value

    @property
    def _all_cands(self) -> "Optional[np.ndarray]":
        return self.detection.all_candidates
    @_all_cands.setter
    def _all_cands(self, value: "Optional[np.ndarray]") -> None:
        self.detection.all_candidates = value

    @property
    def _all_proms(self) -> "Optional[np.ndarray]":
        return self.detection.all_prominences
    @_all_proms.setter
    def _all_proms(self, value: "Optional[np.ndarray]") -> None:
        self.detection.all_prominences = value

    @property
    def _thresh_amp(self) -> float:
        return self.detection.thresh_amp
    @_thresh_amp.setter
    def _thresh_amp(self, value: float) -> None:
        self.detection.thresh_amp = value

    @property
    def _sig_quality(self) -> "Optional[int]":
        return self.detection.sig_quality
    @_sig_quality.setter
    def _sig_quality(self, value: "Optional[int]") -> None:
        self.detection.sig_quality = value

    @property
    def _manual_excluded(self) -> "set[int]":
        return self.detection.manual_excluded
    @_manual_excluded.setter
    def _manual_excluded(self, value: "set[int]") -> None:
        self.detection.manual_excluded = value

    @property
    def _rpeaks_manual_excl(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_manual_excl
    @_rpeaks_manual_excl.setter
    def _rpeaks_manual_excl(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_manual_excl = value

    @property
    def _manual_added(self) -> "set[int]":
        return self.detection.manual_added
    @_manual_added.setter
    def _manual_added(self, value: "set[int]") -> None:
        self.detection.manual_added = value

    @property
    def _rpeaks_manual_added(self) -> "Optional[np.ndarray]":
        return self.detection.rpeaks_manual_added
    @_rpeaks_manual_added.setter
    def _rpeaks_manual_added(self, value: "Optional[np.ndarray]") -> None:
        self.detection.rpeaks_manual_added = value

    @property
    def _edit_mode(self) -> bool:
        return self.detection.edit_mode
    @_edit_mode.setter
    def _edit_mode(self, value: bool) -> None:
        self.detection.edit_mode = value

    @property
    def _edit_free_placement(self) -> bool:
        return self.detection.edit_free_placement
    @_edit_free_placement.setter
    def _edit_free_placement(self, value: bool) -> None:
        self.detection.edit_free_placement = value

    @property
    def _edit_undo(self) -> "list[tuple[frozenset, frozenset]]":
        return self.detection.edit_undo
    @_edit_undo.setter
    def _edit_undo(self, value: "list[tuple[frozenset, frozenset]]") -> None:
        self.detection.edit_undo = value

    @property
    def _edit_redo(self) -> "list[tuple[frozenset, frozenset]]":
        return self.detection.edit_redo
    @_edit_redo.setter
    def _edit_redo(self, value: "list[tuple[frozenset, frozenset]]") -> None:
        self.detection.edit_redo = value

    @property
    def _EDIT_UNDO_LIMIT(self) -> int:
        return DetectionState.EDIT_UNDO_LIMIT

    @property
    def _hover_samp(self) -> "Optional[int]":
        return self.detection.hover_samp
    @_hover_samp.setter
    def _hover_samp(self, value: "Optional[int]") -> None:
        self.detection.hover_samp = value

    @property
    def _hover_samp_near(self) -> bool:
        return self.detection.hover_samp_near
    @_hover_samp_near.setter
    def _hover_samp_near(self, value: bool) -> None:
        self.detection.hover_samp_near = value

    # -- AnalysisState --
    @property
    def _results(self) -> "Optional[dict]":
        return self.analysis.results
    @_results.setter
    def _results(self, value: "Optional[dict]") -> None:
        self.analysis.results = value

    @property
    def _epoch_df(self) -> "Optional[pd.DataFrame]":
        return self.analysis.epoch_df
    @_epoch_df.setter
    def _epoch_df(self, value: "Optional[pd.DataFrame]") -> None:
        self.analysis.epoch_df = value

    @property
    def _rolling_hrv_df(self) -> "Optional[pd.DataFrame]":
        return self.analysis.rolling_hrv_df
    @_rolling_hrv_df.setter
    def _rolling_hrv_df(self, value: "Optional[pd.DataFrame]") -> None:
        self.analysis.rolling_hrv_df = value

    @property
    def _arrhythmia_events(self) -> "list":
        return self.analysis.arrhythmia_events
    @_arrhythmia_events.setter
    def _arrhythmia_events(self, value: "list") -> None:
        self.analysis.arrhythmia_events = value

    @property
    def _arrhythmia_tsv(self) -> str:
        return self.analysis.arrhythmia_tsv
    @_arrhythmia_tsv.setter
    def _arrhythmia_tsv(self, value: str) -> None:
        self.analysis.arrhythmia_tsv = value

    @property
    def _arr_selected_idx(self) -> int:
        return self.analysis.arr_selected_idx
    @_arr_selected_idx.setter
    def _arr_selected_idx(self, value: int) -> None:
        self.analysis.arr_selected_idx = value

    @property
    def _arr_nav_pos(self) -> float:
        return self.analysis.arr_nav_pos
    @_arr_nav_pos.setter
    def _arr_nav_pos(self, value: float) -> None:
        self.analysis.arr_nav_pos = value

    @property
    def _arr_win(self) -> float:
        return self.analysis.arr_win
    @_arr_win.setter
    def _arr_win(self, value: float) -> None:
        self.analysis.arr_win = value

    @property
    def _arr_edit_mode(self) -> bool:
        return self.analysis.arr_edit_mode
    @_arr_edit_mode.setter
    def _arr_edit_mode(self, value: bool) -> None:
        self.analysis.arr_edit_mode = value

    @property
    def _last_seg_a(self) -> "Optional[dict]":
        return self.analysis.last_seg_a
    @_last_seg_a.setter
    def _last_seg_a(self, value: "Optional[dict]") -> None:
        self.analysis.last_seg_a = value

    @property
    def _last_seg_b(self) -> "Optional[dict]":
        return self.analysis.last_seg_b
    @_last_seg_b.setter
    def _last_seg_b(self, value: "Optional[dict]") -> None:
        self.analysis.last_seg_b = value

    @property
    def _artifact_report(self) -> "Optional[dict]":
        return self.analysis.artifact_report
    @_artifact_report.setter
    def _artifact_report(self, value: "Optional[dict]") -> None:
        self.analysis.artifact_report = value

    @property
    def _artifact_candidates(self) -> "list[dict]":
        return self.analysis.artifact_candidates
    @_artifact_candidates.setter
    def _artifact_candidates(self, value: "list[dict]") -> None:
        self.analysis.artifact_candidates = value

    @property
    def _exp_context(self) -> str:
        return self.analysis.exp_context
    @_exp_context.setter
    def _exp_context(self, value: str) -> None:
        self.analysis.exp_context = value

    @property
    def _analysis_t_start(self) -> float:
        return self.analysis.t_start
    @_analysis_t_start.setter
    def _analysis_t_start(self, value: float) -> None:
        self.analysis.t_start = value

    @property
    def _analysis_t_end(self) -> float:
        return self.analysis.t_end
    @_analysis_t_end.setter
    def _analysis_t_end(self, value: float) -> None:
        self.analysis.t_end = value

    @property
    def _annotations(self) -> "list[dict]":
        return self.analysis.annotations
    @_annotations.setter
    def _annotations(self, value: "list[dict]") -> None:
        self.analysis.annotations = value

    @property
    def _pacing_periods(self) -> "list[dict]":
        return self.analysis.pacing_periods
    @_pacing_periods.setter
    def _pacing_periods(self, value: "list[dict]") -> None:
        self.analysis.pacing_periods = value

    @property
    def _wave_template(self) -> "Optional[WaveTemplate]":
        return self.analysis.wave_template
    @_wave_template.setter
    def _wave_template(self, value: "Optional[WaveTemplate]") -> None:
        self.analysis.wave_template = value

    # -- UIState --
    @property
    def _nav_pos(self) -> float:
        return self.ui.nav_pos
    @_nav_pos.setter
    def _nav_pos(self, value: float) -> None:
        self.ui.nav_pos = value

    @property
    def _show_raw(self) -> bool:
        return self.ui.show_raw
    @_show_raw.setter
    def _show_raw(self, value: bool) -> None:
        self.ui.show_raw = value

    @property
    def _filter_preview_on(self) -> bool:
        return self.ui.filter_preview_on
    @_filter_preview_on.setter
    def _filter_preview_on(self, value: bool) -> None:
        self.ui.filter_preview_on = value

    @property
    def _thr_debounce_id(self) -> "str | None":
        return self.ui.thr_debounce_id
    @_thr_debounce_id.setter
    def _thr_debounce_id(self, value: "str | None") -> None:
        self.ui.thr_debounce_id = value

    @property
    def _hover_motion_cid(self) -> "Optional[int]":
        return self.ui.hover_motion_cid
    @_hover_motion_cid.setter
    def _hover_motion_cid(self, value: "Optional[int]") -> None:
        self.ui.hover_motion_cid = value

    @property
    def _hover_after_id(self) -> "str | None":
        return self.ui.hover_after_id
    @_hover_after_id.setter
    def _hover_after_id(self, value: "str | None") -> None:
        self.ui.hover_after_id = value

    @property
    def _beat_nav_cid(self) -> "Optional[int]":
        return self.ui.beat_nav_cid
    @_beat_nav_cid.setter
    def _beat_nav_cid(self, value: "Optional[int]") -> None:
        self.ui.beat_nav_cid = value

    @property
    def _rr_click_cid(self) -> "Optional[int]":
        return self.ui.rr_click_cid
    @_rr_click_cid.setter
    def _rr_click_cid(self, value: "Optional[int]") -> None:
        self.ui.rr_click_cid = value

    @property
    def _hrv_current_view(self) -> str:
        return self.ui.hrv_current_view
    @_hrv_current_view.setter
    def _hrv_current_view(self, value: str) -> None:
        self.ui.hrv_current_view = value

    @property
    def _ctx_keys(self) -> "list":
        return self.ui.ctx_keys
    @_ctx_keys.setter
    def _ctx_keys(self, value: "list") -> None:
        self.ui.ctx_keys = value

    @property
    def _operation_start_time(self) -> "Optional[float]":
        return self.ui.operation_start_time
    @_operation_start_time.setter
    def _operation_start_time(self, value: "Optional[float]") -> None:
        self.ui.operation_start_time = value

    @property
    def _ui_update_batch(self) -> "list[tuple]":
        return self.ui.ui_update_batch
    @_ui_update_batch.setter
    def _ui_update_batch(self, value: "list[tuple]") -> None:
        self.ui.ui_update_batch = value

    @property
    def _figure_cache(self) -> "dict":
        return self.ui.figure_cache
    @_figure_cache.setter
    def _figure_cache(self, value: "dict") -> None:
        self.ui.figure_cache = value

    @property
    def _tsv_store(self) -> "dict[int, str]":
        return self.ui.tsv_store
    @_tsv_store.setter
    def _tsv_store(self, value: "dict[int, str]") -> None:
        self.ui.tsv_store = value

    # -- SessionState --
    @property
    def _recent(self) -> "list[str]":
        return self.session.recent_files
    @_recent.setter
    def _recent(self, value: "list[str]") -> None:
        self.session.recent_files = value

    @property
    def _session_dirty(self) -> bool:
        return self.session.dirty
    @_session_dirty.setter
    def _session_dirty(self, value: bool) -> None:
        self.session.dirty = value

    @property
    def _recording_notes(self) -> str:
        return self.session.recording_notes
    @_recording_notes.setter
    def _recording_notes(self, value: str) -> None:
        self.session.recording_notes = value

    def _batch_ui_update(self, widget: ctk.CTkBaseClass, **kwargs) -> None:
        """Batch UI updates to reduce flicker during session restore."""
        self._ui_update_batch.append((widget, kwargs))

    def _flush_ui_updates(self) -> None:
        """Apply all batched UI updates at once."""
        for widget, kwargs in self._ui_update_batch:
            try:
                widget.configure(**kwargs)
            except Exception as exc:
                log.debug("Batch UI update failed: %s", exc)
        self._ui_update_batch.clear()

    def _current_ref(self, key: str) -> "tuple[float, float]":
        """Return (lo, hi) reference range for the given metric key from current experimental context."""
        return self.analysis_ctrl.current_ref(key)

    # ════════════════════════════════════════════════════════
    #  UI CONSTRUCTION
    # ════════════════════════════════════════════════════════

    def _build(self) -> None:
        self.update_idletasks()
        # Packed on self, side="top", BEFORE the sidebar -- this is what makes
        # it span the full window width rather than being confined to main's
        # width the way the old KPI bar was (sidebar's own side="left" would
        # otherwise claim the left edge first).
        self._build_toolbar()
        w = self.winfo_width() or 1480
        _sidebar_w = max(220, min(340, int(w * 0.20)))
        self.sidebar = ctk.CTkFrame(self, width=_sidebar_w, fg_color=PANEL, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self.bind("<Configure>", self._on_window_resize)
        # Keyboard navigation shortcuts (active when focus is on the main window)
        self.bind("<Left>",  lambda e: self._kb_navigate(-1))
        self.bind("<Right>", lambda e: self._kb_navigate(+1))
        self._build_sidebar()
        # Apply the default state immediately so FILTER SETTINGS widgets are
        # greyed out at startup (Filtering is OFF by default).
        self.after(50, self._on_filtering_toggle)

        # Packed side="right" BEFORE main, for the same reason the sidebar is
        # packed before main -- main's expand=True must not claim the space
        # first.
        self._build_right_panel()

        # Stored as self.main (not a local var) so _toggle_left_panel()/
        # _toggle_right_panel() can re-pack a collapsed panel with
        # before=self.main -- see those methods for why that's required.
        # fg_color=PANEL (not BG) so the workspace reads as the same grey as
        # the sidebar/right_panel it sits between, instead of a subtly
        # different shade that shows as a seam down both edges.
        self.main = ctk.CTkFrame(self, fg_color=PANEL, corner_radius=0)
        self.main.pack(side="left", fill="both", expand=True)
        self._build_tabs(self.main)
        # Global keyboard shortcuts
        self.bind("<Control-z>", self._undo_edit)
        self.bind("<Control-Z>", self._undo_edit)
        self.bind("<Control-y>", self._redo_edit)
        self.bind("<Control-Y>", self._redo_edit)

        # Apply persisted collapse state (defaults False -- both expanded,
        # matching how the toggle buttons were just constructed). Route
        # through the toggle methods themselves rather than duplicating
        # their pack_forget()/button-text logic here -- each toggle flips
        # the UIState bool from its current (just-initialized "expanded")
        # value, so only fire it when the persisted value says "collapsed".
        if self.ui.left_panel_collapsed:
            self.ui.left_panel_collapsed = False
            self._toggle_left_panel()
        if self.ui.right_panel_collapsed:
            self.ui.right_panel_collapsed = False
            self._toggle_right_panel()

    # ─── Sidebar ──────────────────────────────────────────────

    def _build_sidebar(self) -> None:
        """Compact sidebar: minimal always-visible workflow + scrollable settings."""
        px = dict(padx=SPACE_L)

        # ══════════════════════════════════════════════════════
        #  FIXED TOP — title, file, workflow actions, threshold
        # ══════════════════════════════════════════════════════
        top = ctk.CTkFrame(self.sidebar, fg_color=PANEL, corner_radius=0)
        top.pack(fill="x", side="top")

        # Title + version
        title_row = ctk.CTkFrame(top, fg_color="transparent")
        title_row.pack(fill="x", padx=SPACE_L, pady=(SPACE_M, SPACE_XS))
        ctk.CTkLabel(title_row, text="ECG Analysis",
                     font=FONT_TITLE, text_color=TEXT,
                     anchor="w").pack(side="left")
        ctk.CTkLabel(title_row, text="v6",
                     font=FONT_HINT, text_color=LIGHT).pack(side="right")

        # File label
        self.lbl_file = ctk.CTkLabel(
            top, text="No file loaded", font=FONT_SMALL,
            text_color=MUTED, wraplength=260, anchor="w", justify="left")
        self.lbl_file.pack(**px, pady=(0, SPACE_XS), fill="x")

        # Open button + Channels / Recent on same row
        open_row = ctk.CTkFrame(top, fg_color="transparent")
        open_row.pack(fill="x", expand=True, **px, pady=(0, SPACE_XS))
        open_row.rowconfigure(0, weight=1)
        open_row.columnconfigure(0, weight=3)
        open_row.columnconfigure(1, weight=1)
        open_row.columnconfigure(2, weight=1)
        ctk.CTkButton(open_row, text="Open .mat file",
                      fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
                      font=FONT_BTN_PRIMARY, height=30, corner_radius=8,
                      command=self._open_file).grid(row=0, column=0, sticky="ewns", padx=(0, SPACE_XS))
        ctk.CTkButton(open_row, text="Channel", height=10,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, corner_radius=6,
                      command=self._show_channels).grid(row=0, column=1, sticky="ewns", padx=(0, SPACE_XS))
        ctk.CTkButton(open_row, text="Recent", height=10,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, corner_radius=6,
                      command=self._open_recent).grid(row=0, column=2, sticky="ewns")

        ctk.CTkFrame(top, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_L, pady=(SPACE_XS, SPACE_XS))

        # Peak count status
        self.lbl_npeaks = ctk.CTkLabel(
            top, text="Peaks detected: —",
            font=FONT_CARD_TITLE, text_color=BLUE, anchor="w")
        self.lbl_npeaks.pack(**px, fill="x", pady=(0, SPACE_XS))

        # Method/Analysis-window/Status/Threshold relocated to the right
        # panel's DETECTION accordion section (_build_detection_section()).

        # ══════════════════════════════════════════════════════
        #  SCROLLABLE SETTINGS
        # ══════════════════════════════════════════════════════
        scroll = ctk.CTkScrollableFrame(
            self.sidebar, fg_color=PANEL,
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=BORDER2)
        scroll.pack(fill="both", expand=True, side="top")
        s = scroll
        fpx = dict(padx=SPACE_L)

        # ── SIGNAL ────────────────────────────────────────────
        sec_sig = CollapsibleSection(s, "SIGNAL", initially_open=False)
        f = sec_sig.frame
        # Project name: a label spanning the whole working session (many
        # recordings), not per-recording metadata like Subject ID below --
        # like every other sidebar entry, reset_for_new_file() never touches
        # it, so it naturally survives opening a different recording in the
        # same session with no special-case logic needed.
        self._sidebar_entry(f, "Project name", "project_name", "", fpx)
        self.ent_project_name.bind("<FocusOut>", self._on_project_name_change)  # type: ignore[union-attr]
        self.ent_project_name.bind("<Return>", self._on_project_name_change)  # type: ignore[union-attr]
        self._sidebar_entry_row(f, fpx, [
            ("Channel", "channel", "ECG"),
            ("Subject ID", "subject", "subject_01"),
        ])
        self._sidebar_entry(f, "Sampling rate (Hz)", "fs",
                            str(MouseECG.FS_DEFAULT), fpx)
        self.lbl_fs_source = ctk.CTkLabel(
            f, text="", font=FONT_KPI_LABEL, text_color=MUTED,
            anchor="w", wraplength=230)
        self.lbl_fs_source.pack(**fpx, fill="x", pady=(0, SPACE_XS))
        self._sidebar_entry_row(f, fpx, [
            ("Crop start (s)", "t_start", "0"),
            ("Crop end (s)",   "t_end",   "0"),
        ])

        # ── DETECTION ─────────────────────────────────────────
        self._build_detection_section(s)

        # ── FILTERS ───────────────────────────────────────────
        # sw_filtering ON = DSP filtering is applied (band-pass/notch/
        # cleaning); OFF (default) = raw signal, matching the app's
        # raw-by-default behaviour on file load. DISPLAY & PREVIEW controls
        # (show raw, invert, live preview) always stay interactive -- pure
        # display toggles, or apply regardless of sw_filtering. FILTER
        # SETTINGS (notch/band-pass/cleaning) stay visible but greyed out
        # when filtering is off, so the user can see/set them in advance --
        # see on_filtering_toggle in detection_controller.py.
        sec_flt = CollapsibleSection(s, "FILTERS", initially_open=False)
        f = sec_flt.frame
        self.sw_filtering = self._switch(f, "Filtering", fpx, default_on=False)
        self.sw_filtering.configure(command=self._on_filtering_toggle)
        ctk.CTkLabel(f, text="Off analyzes the raw signal. On applies the settings below.",
                     font=FONT_HINT, text_color=MUTED,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_XS))
        self.lbl_filter_summary = ctk.CTkLabel(
            f, text="Processing\n•  Raw signal", font=FONT_HINT, text_color=MUTED,
            anchor="w", justify="left", wraplength=230)
        self.lbl_filter_summary.pack(**fpx, fill="x", pady=(0, SPACE_S))

        ctk.CTkLabel(f, text="DISPLAY & PREVIEW", font=FONT_SUBSECTION,
                     text_color=MUTED, anchor="w").pack(**fpx, fill="x", pady=(0, SPACE_XS))
        self.sw_show_raw = self._switch(
            f, "Show raw signal (vs filtered)", fpx, default_on=True)
        self.sw_show_raw.configure(command=self._on_show_raw_toggle)
        self.sw_invert_signal = self._switch(
            f, "⟳  Invert signal (polarity)", fpx, default_on=False)
        # A plain parameter -- doesn't auto-run the full pipeline on toggle.
        # Picked up on the next "Preview Detection" click, or immediately in
        # the before/after overlay if "Preview filtering" is on, exactly
        # like every other control in this section.
        self.sw_invert_signal.configure(command=self._refresh_filter_preview)
        ctk.CTkLabel(f, text="Useful if R peaks appear negative",
                     font=FONT_HINT, text_color=MUTED,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.sw_filter_preview = self._switch(f, "👁  Preview filtering", fpx, default_on=False)
        self.sw_filter_preview.configure(command=self._on_filter_preview_toggle)
        ctk.CTkLabel(f, text="Overlays the filtered signal on the visible window,\n"
                             "using the settings below — no need to run Preview Detection.",
                     font=FONT_HINT, text_color=MUTED,
                     anchor="w", justify="left", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))

        ctk.CTkFrame(f, height=1, fg_color=BORDER).pack(
            fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkLabel(f, text="FILTER SETTINGS", font=FONT_SUBSECTION,
                     text_color=MUTED, anchor="w").pack(**fpx, fill="x", pady=(0, SPACE_XS))
        self._filter_advanced_group = ctk.CTkFrame(f, fg_color="transparent")
        self._filter_advanced_group.pack(**fpx, fill="x")
        g = self._filter_advanced_group
        self.sw_notch = self._switch(g, "Notch 50 Hz", dict(padx=0), default_on=False)
        self.sw_notch.configure(command=self._refresh_filter_preview)
        ctk.CTkLabel(g, text="Band-pass filter", font=FONT_SMALL,
                     text_color=MUTED, anchor="w").pack(anchor="w", pady=(SPACE_S, 0))
        # Labelled by frequency position (Low/High), matching filtering.py's
        # own bandpass(lo, hi) naming -- not by filter-type jargon (HP/LP),
        # which technically read correctly (HP corner = the low edge, LP
        # corner = the high edge) but was confusing at a glance.
        self._sidebar_entry_row(g, dict(padx=0), [
            ("Low cutoff (Hz)",  "lp", str(MouseECG.BP_LO_HZ)),
            ("High cutoff (Hz)", "hp", str(int(MouseECG.BP_HI_HZ))),
        ])
        for _ent in (self.ent_lp, self.ent_hp):
            if _ent is not None:
                _ent.bind("<Return>",   lambda _e: self._refresh_filter_preview())
                _ent.bind("<FocusOut>", lambda _e: self._refresh_filter_preview())
        ctk.CTkLabel(g, text="Signal cleaning:", font=FONT_SMALL,
                     text_color=MUTED, anchor="w").pack(anchor="w", pady=(SPACE_XS, 0))
        self.cb_clean = ctk.CTkComboBox(
            g, font=FONT_LABEL, height=28, fg_color=BG,
            border_color=BORDER2, button_color=BORDER2, text_color=TEXT,
            dropdown_fg_color=BG, dropdown_text_color=TEXT,
            values=["neurokit", "pantompkins1985", "elgendi2010", "hamilton2002", "biosppy"],
            command=lambda _v: self._refresh_filter_preview())
        self.cb_clean.set("neurokit")
        self.cb_clean.pack(fill="x", pady=(SPACE_XS, SPACE_S))

        # ── ARTIFACTS ─────────────────────────────────────────
        self._build_artifacts_section(s)

        # ── ANNOTATIONS ───────────────────────────────────────
        self._build_annotations_section(s)

        # ── BOTTOM BUTTONS ────────────────────────────────────
        # Session info + cache-clearing sit here (not in a named accordion
        # section) since neither the left panel's spec (Signal/Detection/
        # Filters/Artifacts/Annotations) nor the right panel's (Statistics/
        # Exports/ML Detector Saving) has room for them -- this area is
        # already the app's catch-all for cross-cutting maintenance actions.
        # Save Session itself lives in the top toolbar's File ops chip
        # (self.btn_save_session, built in _build_toolbar()).
        ctk.CTkFrame(s, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        self.lbl_session_info = ctk.CTkLabel(
            s, text="No session saved for this file", font=FONT_HINT,
            text_color=MUTED, anchor="w", wraplength=230, justify="left")
        self.lbl_session_info.pack(padx=SPACE_M, pady=(0, SPACE_S), fill="x")
        self._btn(s, "🗑  Clear Session Cache", self._delete_session, dict(padx=SPACE_M), variant="secondary", h=26)
        ctk.CTkFrame(s, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        ctk.CTkButton(
            s, text="⚖  Compare Segments",
            command=self._open_compare_segments,
            fg_color=TEAL, hover_color=TEAL_DARK, text_color="white",
            font=FONT_SIDEBAR_HDR, height=28, corner_radius=8,
        ).pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkButton(
            s, text="⚙  Parameters",
            command=self._open_params_dialog,
            fg_color=BLUE_DARK, hover_color=BLUE, text_color="white",
            font=FONT_SIDEBAR_HDR, height=28, corner_radius=8,
        ).pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkButton(
            s, text="↺  Reset to Mouse ECG Defaults",
            command=self._reset_params,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            font=FONT_HINT, height=24, corner_radius=5,
        ).pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkLabel(s, text="F1 — keyboard shortcuts",
                     font=FONT_MICRO, text_color=LIGHT, anchor="center",
                     cursor="hand2").pack(pady=(0, SPACE_S))




    # ════════════════════════════════════════════════════════
    #  FREQUENCY BAND HELPER
    # ════════════════════════════════════════════════════════

    def _get_freq_bands(self) -> "tuple[tuple, tuple, tuple]":
        """Return (vlf, lf, hf) tuples from the selected band preset.

        Reads ``cb_freq_band`` if it exists; falls back to Mouse Thireau defaults.
        """
        return self.analysis_ctrl.get_freq_bands()

    # ════════════════════════════════════════════════════════
    #  EXPORT FIGURES (clean PNG, no spike markers)
    # ════════════════════════════════════════════════════════

    def _export_figures(self) -> None:
        """Export curated publication-ready PNG figures to a chosen folder."""
        self.export_ctrl.export_figures()

    # ════════════════════════════════════════════════════════
    #  COMPARE SEGMENTS
    # ════════════════════════════════════════════════════════

    def _open_compare_segments(self) -> None:
        """Open the Segment Comparison window.

        Two configurable time windows (A / B) → side-by-side metric table
        + superimposed RR tachogram.  All stats run on a background thread.
        """
        if self._signal_flt is None or self._rpeaks_ok is None or self._fs is None:
            messagebox.showwarning("No data", "Load a file and run Core Analysis first.")
            return

        sig    = self._signal_flt
        rp     = self._rpeaks_ok.copy()
        fs     = self._fs
        dur    = float(len(sig)) / fs
        vlf, lf, hf = self._get_freq_bands()
        sub    = self.ent_subject.get().strip() if self.ent_subject else "subject"

        win = ctk.CTkToplevel(self)
        win.title("Compare Segments")
        win.geometry("1060x720")
        win.configure(fg_color=PANEL)
        win.resizable(True, True)

        hdr = ctk.CTkFrame(win, fg_color=PANEL, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="⚖  Segment Comparison",
                     font=FONT_KPI_VALUE, text_color=TEXT,
                     anchor="w").pack(side="left", padx=SPACE_L, pady=SPACE_M)
        ctk.CTkLabel(hdr,
                     text=f"{sub}  ·  {dur:.1f} s  ·  {len(rp)} peaks",
                     font=FONT_SMALL, text_color=MUTED).pack(side="left", padx=(0, SPACE_L))
        ctk.CTkButton(hdr, text="✗  Close", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=28, corner_radius=6).pack(
            side="right", padx=SPACE_L, pady=SPACE_M)

        seg_frame = ctk.CTkFrame(win, fg_color="transparent")
        seg_frame.pack(fill="x", padx=SPACE_L, pady=(SPACE_M, SPACE_S))

        seg_entries: "list[dict]" = []
        _colors = [BLUE, ORANGE_DARK]
        _names  = ["Segment A", "Segment B"]

        for i, (name, col) in enumerate(zip(_names, _colors)):
            card = ctk.CTkFrame(seg_frame, fg_color=CARD, corner_radius=8,
                                border_width=2, border_color=col)
            card.pack(side="left", fill="x", expand=True,
                      padx=(0, 8 if i == 0 else 0))
            ctk.CTkLabel(card, text=name, font=FONT_SIDEBAR_HDR,
                         text_color=col, anchor="w").pack(padx=SPACE_L, pady=(SPACE_M, SPACE_S), fill="x")
            row = ctk.CTkFrame(card, fg_color="transparent")
            row.pack(fill="x", padx=SPACE_L, pady=(0, SPACE_M))
            ctk.CTkLabel(row, text="From (s):", font=FONT_SMALL,
                         text_color=MUTED, width=62).pack(side="left")
            lo_e = ctk.CTkEntry(row, width=80, height=28, font=FONT_LABEL,
                                fg_color=BG, border_color=col, text_color=TEXT)
            lo_e.insert(0, str(int(dur * i / 2)))
            lo_e.pack(side="left", padx=(0, SPACE_M))
            ctk.CTkLabel(row, text="To (s):", font=FONT_SMALL,
                         text_color=MUTED, width=46).pack(side="left")
            hi_e = ctk.CTkEntry(row, width=80, height=28, font=FONT_LABEL,
                                fg_color=BG, border_color=col, text_color=TEXT)
            hi_e.insert(0, str(int(dur * (i + 1) / 2)))
            hi_e.pack(side="left", padx=(0, SPACE_M))
            lbl_e = ctk.CTkEntry(row, width=120, height=28, font=FONT_LABEL,
                                 fg_color=BG, border_color=BORDER2, text_color=TEXT,
                                 placeholder_text="Label (optional)")
            lbl_e.insert(0, name)
            lbl_e.pack(side="left")
            seg_entries.append({"lo": lo_e, "hi": hi_e, "lbl": lbl_e, "color": col})

        ctrl = ctk.CTkFrame(win, fg_color="transparent")
        ctrl.pack(fill="x", padx=SPACE_L, pady=(0, SPACE_S))
        lbl_status = ctk.CTkLabel(ctrl, text="", font=FONT_SMALL, text_color=MUTED)
        lbl_status.pack(side="right", padx=(SPACE_M, 0))
        progress_bar = ctk.CTkProgressBar(ctrl, height=8, mode="indeterminate",
                                          progress_color=TEAL)
        progress_bar.pack(side="right", padx=(SPACE_M, 0))
        progress_bar.pack_forget()

        run_btn = ctk.CTkButton(ctrl, text="▶  Compare",
                                fg_color=TEAL, hover_color=TEAL_DARK,
                                text_color="white",
                                font=FONT_BTN_PRIMARY, height=34, corner_radius=8)
        run_btn.pack(side="left")

        results_frame = ctk.CTkFrame(win, fg_color="transparent")
        results_frame.pack(fill="both", expand=True, padx=SPACE_L, pady=(0, SPACE_M))

        table_card = ctk.CTkFrame(results_frame, fg_color=PANEL, corner_radius=8)
        table_card.pack(side="left", fill="both", expand=True, padx=(0, SPACE_M))
        ctk.CTkLabel(table_card, text="Metric comparison",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").pack(padx=SPACE_M, pady=(SPACE_M, SPACE_S), fill="x")
        lbl_mw = ctk.CTkLabel(table_card, text="", font=FONT_MICRO,
                              text_color=MUTED, anchor="w")
        lbl_mw.pack(padx=SPACE_M, pady=(0, SPACE_XS), fill="x")
        tbl = ctk.CTkScrollableFrame(table_card, fg_color=BG, height=380,
                                     scrollbar_button_color=BORDER)
        tbl.pack(fill="both", expand=True, padx=SPACE_S, pady=(0, SPACE_M))
        for ci in range(4):
            tbl.grid_columnconfigure(ci, weight=1)

        plot_card = ctk.CTkFrame(results_frame, fg_color=PANEL, corner_radius=8)
        plot_card.pack(side="right", fill="both", expand=True)
        ctk.CTkLabel(plot_card, text="RR tachograms (superimposed)",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").pack(padx=SPACE_M, pady=(SPACE_M, SPACE_XS), fill="x")
        from ecg.ui.plots import CanvasSlot as _CS
        plot_slot = _CS(plot_card, 8, 4, toolbar=False, yscale_bar=True)

        _METRICS: "list[tuple[str, str, str]]" = [
            ("n_beats",   "Beats",           ""),
            ("duration_s","Duration",         "s"),
            ("hr_mean",   "Mean HR",          "bpm"),
            ("hr_sd",     "HR SD",            "bpm"),
            ("hr_min",    "Min HR",           "bpm"),
            ("hr_max",    "Max HR",           "bpm"),
            ("rr_mean",   "Mean RR",          "ms"),
            ("rr_sd",     "RR SD",            "ms"),
            ("rr_cv",     "RR CV",            "%"),
            ("sdnn",      "SDNN",             "ms"),
            ("rmssd",     "RMSSD",            "ms"),
            ("pnn6",      "pNN6",             "%"),
            ("lf_nu",     "LF (n.u.)",        "%"),
            ("hf_nu",     "HF (n.u.)",        "%"),
            ("lf_hf",     "LF/HF",            ""),
            ("sd1",       "Poincaré SD1",      "ms"),
            ("sd2",       "Poincaré SD2",      "ms"),
            ("sd_ratio",  "SD1/SD2",           ""),
        ]

        def _fmt(v: "Any") -> str:
            try:
                fv = float(v)
                if not np.isfinite(fv):
                    return "—"
                if abs(fv) >= 100: return f"{fv:.1f}"
                if abs(fv) >= 10:  return f"{fv:.2f}"
                return f"{fv:.3f}"
            except Exception:
                return str(v) if v else "—"

        def _populate_table(sa: dict, sb: dict) -> None:
            for w in tbl.winfo_children():
                w.destroy()
            for ci, (txt, col) in enumerate([
                ("Metric", TEXT), (sa["label"], _colors[0]),
                (sb["label"], _colors[1]), ("Δ%", MUTED),
            ]):
                ctk.CTkLabel(tbl, text=txt, font=FONT_BADGE,
                             text_color=col, anchor="w").grid(
                    row=0, column=ci, sticky="ew", padx=SPACE_S, pady=(SPACE_XS, SPACE_S))
            for ri, (key, name, unit) in enumerate(_METRICS, start=1):
                va = float(sa.get(key, float("nan")))
                vb = float(sb.get(key, float("nan")))
                bg = CARD if ri % 2 == 0 else BG
                lbl_txt = f"{name}  {unit}" if unit else name
                ctk.CTkLabel(tbl, text=lbl_txt, font=FONT_KPI_LABEL,
                             text_color=MUTED, fg_color=bg,
                             anchor="w").grid(row=ri, column=0, sticky="ew", padx=SPACE_S, pady=SPACE_XS)
                ctk.CTkLabel(tbl, text=_fmt(va), font=FONT_KPI_LABEL,
                             text_color=TEXT, fg_color=bg,
                             anchor="e").grid(row=ri, column=1, sticky="ew", padx=SPACE_S, pady=SPACE_XS)
                ctk.CTkLabel(tbl, text=_fmt(vb), font=FONT_KPI_LABEL,
                             text_color=TEXT, fg_color=bg,
                             anchor="e").grid(row=ri, column=2, sticky="ew", padx=SPACE_S, pady=SPACE_XS)
                if np.isfinite(va) and np.isfinite(vb) and abs(va) > 1e-9:
                    delta = 100.0 * (vb - va) / abs(va)
                    ctk.CTkLabel(tbl, text=f"{'▲' if delta>0 else '▼'} {abs(delta):.1f}%",
                                 font=FONT_MICRO, text_color=MUTED, fg_color=bg,
                                 anchor="e").grid(row=ri, column=3, sticky="ew", padx=SPACE_S, pady=SPACE_XS)

        def _populate_plot(sa: dict, sb: dict) -> None:
            _t_a, _r_a, _la, _ca = sa["t_rr"], sa["rr_ms"], sa["label"], _colors[0]
            _t_b, _r_b, _lb, _cb = sb["t_rr"], sb["rr_ms"], sb["label"], _colors[1]
            def _draw(fig):
                ax = fig.add_subplot(111)
                style_axes(ax)
                if len(_t_a) and len(_r_a):
                    ax.plot(_t_a - _t_a[0], _r_a, color=_ca, lw=1.0, alpha=0.85, label=_la)
                    ax.axhline(float(np.mean(_r_a)), color=_ca, lw=1.0, ls="--", alpha=0.5)
                if len(_t_b) and len(_r_b):
                    ax.plot(_t_b - _t_b[0], _r_b, color=_cb, lw=1.0, alpha=0.85, label=_lb)
                    ax.axhline(float(np.mean(_r_b)), color=_cb, lw=1.0, ls="--", alpha=0.5)
                ax.set_xlabel("Time within segment (s)", fontsize=9)
                ax.set_ylabel("RR (ms)", fontsize=9)
                ax.set_title("RR intervals — segments superimposed", loc="left", fontsize=9)
                ax.legend(framealpha=0, fontsize=8)
            plot_slot.update(_draw)

        def _run_compare() -> None:
            try:
                lo_a = float(seg_entries[0]["lo"].get()); hi_a = float(seg_entries[0]["hi"].get())
                la   = seg_entries[0]["lbl"].get().strip() or "Segment A"
                lo_b = float(seg_entries[1]["lo"].get()); hi_b = float(seg_entries[1]["hi"].get())
                lb   = seg_entries[1]["lbl"].get().strip() or "Segment B"
            except ValueError:
                lbl_status.configure(text="⚠  Invalid time values", text_color=RED)
                return
            for lo, hi, lbl_s in [(lo_a, hi_a, "A"), (lo_b, hi_b, "B")]:
                if lo >= hi:
                    lbl_status.configure(text=f"⚠  Segment {lbl_s}: start ≥ end", text_color=RED)
                    return
                if lo < 0 or hi > dur:
                    lbl_status.configure(
                        text=f"⚠  Segment {lbl_s}: outside recording ({dur:.1f} s)",
                        text_color=RED)
                    return
            run_btn.configure(state="disabled")
            progress_bar.pack(side="right", padx=(SPACE_M, 0))
            progress_bar.start()
            lbl_status.configure(text="  Computing…", text_color=ORANGE)

            import threading as _th
            def _worker():
                from ecg.core.analysis import compute_segment_stats
                sa = compute_segment_stats(sig, rp, fs, lo_a, hi_a, la,
                                            lf_band=lf, hf_band=hf)
                sb = compute_segment_stats(sig, rp, fs, lo_b, hi_b, lb,
                                            lf_band=lf, hf_band=hf)
                # Whole-distribution significance test between the two
                # segments' RR-interval series (not per-metric -- Mann-
                # Whitney compares two samples, and the per-row table above
                # is mostly single summary statistics, not resampleable
                # distributions).
                mw_p, mw_interp = self.analysis_ctrl.mannwhitney_test(
                    np.asarray(sa.get("rr_ms", []), dtype=float),
                    np.asarray(sb.get("rr_ms", []), dtype=float))
                win.after(0, lambda: _on_done(sa, sb, mw_p, mw_interp))

            def _on_done(sa: dict, sb: dict, mw_p: float, mw_interp: str) -> None:
                run_btn.configure(state="normal")
                progress_bar.stop(); progress_bar.pack_forget()
                # Persist for Prism export
                self._last_seg_a = sa
                self._last_seg_b = sb
                errs = [s["error"] for s in (sa, sb) if s.get("error")]
                if errs:
                    lbl_status.configure(text=f"⚠  {errs[0]}", text_color=RED)
                else:
                    lbl_status.configure(
                        text=f"✓  {sa['label']}: {sa['n_beats']} beats  ·  "
                             f"{sb['label']}: {sb['n_beats']} beats",
                        text_color=GREEN)
                    export_btn.configure(state="normal")   # unlock Export button
                mw_txt = (f"p = {mw_p:.4g}" if np.isfinite(mw_p) else "n/a")
                lbl_mw.configure(
                    text=f"RR-interval distributions (Mann-Whitney U):  "
                         f"{mw_txt}  ({mw_interp})")
                _populate_table(sa, sb)
                _populate_plot(sa, sb)

            _th.Thread(target=_worker, daemon=True).start()

        # ── Export function ───────────────────────────────────────────────
        _METRICS_EXPORT: "list[tuple[str, str, str]]" = [
            ("n_beats",   "Beats",             ""),
            ("duration_s","Duration",           "s"),
            ("hr_mean",   "Mean HR",            "bpm"),
            ("hr_sd",     "HR SD",              "bpm"),
            ("hr_min",    "Min HR",             "bpm"),
            ("hr_max",    "Max HR",             "bpm"),
            ("rr_mean",   "Mean RR",            "ms"),
            ("rr_sd",     "RR SD",              "ms"),
            ("rr_cv",     "RR CV",              "%"),
            ("sdnn",      "SDNN",               "ms"),
            ("rmssd",     "RMSSD",              "ms"),
            ("pnn6",      "pNN6",               "%"),
            ("lf_nu",     "LF n.u.",            "%"),
            ("hf_nu",     "HF n.u.",            "%"),
            ("lf_hf",     "LF/HF",              ""),
            ("sd1",       "Poincaré SD1",       "ms"),
            ("sd2",       "Poincaré SD2",       "ms"),
            ("sd_ratio",  "SD1/SD2",            ""),
        ]

        def _export_comparison() -> None:
            sa = self._last_seg_a
            sb = self._last_seg_b
            if sa is None or sb is None:
                messagebox.showwarning("No results",
                                       "Run the comparison first.", parent=win)
                return

            sub  = self.ent_subject.get().strip() if self.ent_subject else "subject"
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"{sub}_comparison_{ts}"

            path = filedialog.asksaveasfilename(
                parent=win,
                title="Export segment comparison",
                defaultextension=".xlsx",
                initialfile=f"{default_name}.xlsx",
                filetypes=[
                    ("Excel workbook", "*.xlsx"),
                    ("CSV",            "*.csv"),
                    ("All files",      "*.*"),
                ],
            )
            if not path:
                return

            la = sa["label"]
            lb = sb["label"]

            # ── Build metric comparison DataFrame ─────────────────────────
            rows = []
            for key, name, unit in _METRICS_EXPORT:
                va = sa.get(key, float("nan"))
                vb = sb.get(key, float("nan"))
                try:
                    va = float(va) if va is not None else float("nan")
                    vb = float(vb) if vb is not None else float("nan")
                except (TypeError, ValueError):
                    va = vb = float("nan")
                delta = (100.0 * (vb - va) / abs(va)
                         if np.isfinite(va) and np.isfinite(vb) and abs(va) > 1e-9
                         else float("nan"))
                rows.append({
                    "Metric":      f"{name}  ({unit})" if unit else name,
                    la:            va,
                    lb:            vb,
                    "Δ% (A→B)":    delta,
                })
            metrics_df = pd.DataFrame(rows)

            # ── RR series per segment ─────────────────────────────────────
            rr_a = pd.DataFrame({
                "Time (s)": np.asarray(sa["t_rr"], dtype=float),
                "RR (ms)":  np.asarray(sa["rr_ms"], dtype=float),
            })
            rr_b = pd.DataFrame({
                "Time (s)": np.asarray(sb["t_rr"], dtype=float),
                "RR (ms)":  np.asarray(sb["rr_ms"], dtype=float),
            })

            # ── Write ─────────────────────────────────────────────────────
            if path.lower().endswith(".csv"):
                metrics_df.to_csv(path, index=False, float_format="%.4f")
                lbl_status.configure(
                    text=f"✓  CSV saved → {os.path.basename(path)}",
                    text_color=GREEN)
            else:
                try:
                    import openpyxl
                    from openpyxl.styles import (Font, PatternFill,
                                                  Alignment, Border, Side)
                    from openpyxl.utils import get_column_letter
                except ImportError:
                    messagebox.showerror("Missing dependency",
                                         "openpyxl required: pip install openpyxl",
                                         parent=win)
                    return

                wb = openpyxl.Workbook()

                # ── Sheet 1: Metric comparison ────────────────────────────
                ws = wb.active
                assert ws is not None
                ws.title = "Comparison"

                # Header style
                hdr_fill = PatternFill("solid", fgColor="1A1A2E")
                col_fills = {
                    la: PatternFill("solid", fgColor="1565C0"),   # blue
                    lb: PatternFill("solid", fgColor="E65100"),   # orange
                    "Δ% (A→B)": PatternFill("solid", fgColor="2E2E2E"),
                }
                hdr_font = Font(bold=True, color="FFFFFF", size=10)
                thin = Side(style="thin", color="333333")
                border = Border(left=thin, right=thin,
                                top=thin, bottom=thin)

                headers = list(metrics_df.columns)
                for ci, h in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=ci, value=h)
                    cell.font = hdr_font
                    cell.fill = col_fills.get(h, hdr_fill)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border

                for ri, row_d in enumerate(metrics_df.itertuples(index=False), start=2):
                    for ci, val in enumerate(row_d, start=1):
                        cell = ws.cell(row=ri, column=ci, value=val)
                        cell.border = border
                        cell.alignment = Alignment(
                            horizontal="right" if ci > 1 else "left")
                        # Shade alternating rows
                        if ri % 2 == 0:
                            cell.fill = PatternFill("solid", fgColor="1E1E2E")
                        # Colour-code Δ% column
                        if headers[ci-1] == "Δ% (A→B)" and isinstance(val, float):
                            if np.isfinite(val):
                                cell.font = Font(
                                    color="66BB6A" if val > 0 else "EF5350",
                                    bold=True)

                # Auto-fit column widths
                for col in ws.columns:
                    if not col:
                        continue
                    first_cell = col[0]
                    if first_cell.column is None:
                        continue
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value), default=8)
                    ws.column_dimensions[
                        get_column_letter(first_cell.column)].width = min(
                        max_len + 3, 30)

                # ── Sheet 2: RR series ────────────────────────────────────
                def _rr_sheet(name: str, rr_df: pd.DataFrame,
                              fill_color: str) -> None:
                    ws2 = wb.create_sheet(name)
                    hf2 = PatternFill("solid", fgColor=fill_color)
                    for ci, col in enumerate(rr_df.columns, start=1):
                        c = ws2.cell(row=1, column=ci, value=col)
                        c.font = Font(bold=True, color="FFFFFF")
                        c.fill = hf2
                        c.border = border
                    for ri, row_vals in enumerate(rr_df.itertuples(index=False),
                                                   start=2):
                        for ci, v in enumerate(row_vals, start=1):
                            ws2.cell(row=ri, column=ci, value=round(float(v), 4))
                    ws2.column_dimensions["A"].width = 14
                    ws2.column_dimensions["B"].width = 12

                _rr_sheet(f"RR — {la}", rr_a, "1565C0")
                _rr_sheet(f"RR — {lb}", rr_b, "E65100")

                # ── Sheet 3: Summary stats ────────────────────────────────
                ws3 = wb.create_sheet("Summary")
                ws3.cell(row=1, column=1, value="Subject").font = Font(bold=True)
                ws3.cell(row=1, column=2, value=sub)
                ws3.cell(row=2, column=1, value="Export date").font = Font(bold=True)
                ws3.cell(row=2, column=2,
                         value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                ws3.cell(row=3, column=1, value="LF band (Hz)").font = Font(bold=True)
                ws3.cell(row=3, column=2, value=f"{lf[0]:.2f}–{lf[1]:.2f}")
                ws3.cell(row=4, column=1, value="HF band (Hz)").font = Font(bold=True)
                ws3.cell(row=4, column=2, value=f"{hf[0]:.2f}–{hf[1]:.2f}")
                for col in ws3.columns:
                    ws3.column_dimensions[
                        get_column_letter(col[0].column)].width = 22

                wb.save(path)
                lbl_status.configure(
                    text=f"✓  Exported → {os.path.basename(path)}",
                    text_color=GREEN)

            # ── Also save the plot as PNG next to the xlsx ─────────────────
            png_path = os.path.splitext(path)[0] + "_plot.png"
            try:
                export_fig = Figure(
                    figsize=(10, 4), dpi=200,
                    facecolor=PLOT.get("bg", "#1A1A2E"), layout="constrained")
                if hasattr(export_fig, "set_constrained_layout_pads"):
                    # Was previously called with left=/right=/top=/bottom= --
                    # not real parameters of this method (only w_pad, h_pad,
                    # wspace, hspace are), so this raised TypeError on every
                    # single call, silently caught by the except block below
                    # at DEBUG level: this companion PNG has never actually
                    # been written. Reuse CanvasSlot's own pad constants
                    # instead of a third hardcoded copy of the numbers.
                    getattr(export_fig, "set_constrained_layout_pads")(
                        w_pad=CanvasSlot._CL_PAD, h_pad=CanvasSlot._CL_PAD,
                        wspace=CanvasSlot._CL_SPACE, hspace=CanvasSlot._CL_SPACE)
                if plot_slot._draw_fn is not None:
                    plot_slot._draw_fn(export_fig)
                    export_fig.savefig(
                        png_path, dpi=200,
                        facecolor=export_fig.get_facecolor(),
                        bbox_inches="tight")
                    plt.close(export_fig)
                    lbl_status.configure(
                        text=lbl_status.cget("text") + f"  +  {os.path.basename(png_path)}",
                        text_color=GREEN)
            except Exception as exc:
                log.debug("comparison PNG export: %s", exc)

        run_btn.configure(command=_run_compare)

        # Export button — disabled until comparison has run
        export_btn = ctk.CTkButton(
            ctrl, text="📊  Export",
            command=_export_comparison,
            fg_color=GREEN_DARK, hover_color=GREEN, text_color="white",
            font=FONT_BTN_PRIMARY, height=34, corner_radius=8,
            state="disabled",
        )
        export_btn.pack(side="left", padx=(SPACE_M, 0))

        ctk.CTkLabel(win,
                     text=f"Tip: equal durations give a fair comparison  ·  "
                          f"LF {lf[0]:.2f}–{lf[1]:.2f} Hz  HF {hf[0]:.2f}–{hf[1]:.2f} Hz",
                     font=FONT_MICRO, text_color=LIGHT, anchor="w").pack(
            fill="x", padx=SPACE_L, pady=(0, SPACE_S))

    # ─── KPI bar ──────────────────────────────────────────────

    def _build_toolbar(self) -> None:
        """Primary instrument-style toolbar spanning the FULL window width.

        Packed on self (not inside main -- see _build()), before the
        sidebar/main split, so it isn't confined to main's width. Consolidates
        what used to be split across the old KPI/top bar AND the sidebar's
        numbered workflow buttons (Preview/Run) into one real toolbar, per
        the Adaptive Workbench brief: Open/Save/Export/Detect Peaks/Analyze/
        Settings, Undo/Redo, project name, progress indicator.

        btn_preview/btn_run/btn_save_session are RELOCATED here (same
        attribute names, same CTkButton objects) rather than recreated --
        session_controller.py/signal_controller.py/analysis_controller.py all
        reference them by name for busy-state text swaps ("Loading…",
        "Analysing…", "Saving…", "Restoring…") and must keep working
        unchanged.
        """
        bar = ctk.CTkFrame(self, fg_color=PANEL, corner_radius=0)
        bar.pack(side="top", fill="x")
        ctk.CTkFrame(bar, height=1, fg_color=BORDER).pack(side="bottom", fill="x")

        # Progress row -- child of `bar` itself (not the inline row below) so
        # that _start_async's existing self._prog_row.pack(side="bottom",
        # fill="x", ...) call spans the full toolbar width, unchanged. Not
        # packed here -- _start_async/_stop_progress show/hide it dynamically.
        self._prog_row = ctk.CTkFrame(bar, fg_color="transparent", height=20)
        self.progress = ctk.CTkProgressBar(
            self._prog_row, height=4, mode="determinate",
            progress_color=COLOR_PRIMARY, fg_color=BORDER, corner_radius=2,
        )
        self.progress.set(0)
        self.progress.pack(side="left", fill="x", expand=True, padx=(SPACE_M, SPACE_S))
        self.lbl_progress = ctk.CTkLabel(
            self._prog_row, text="", font=FONT_HINT,
            text_color=MUTED, width=260, anchor="w",
        )
        self.lbl_progress.pack(side="left", padx=(0, SPACE_M))

        row = ctk.CTkFrame(bar, fg_color="transparent", height=44)
        row.pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        row.pack_propagate(False)

        # ── Left panel toggle ──────────────────────────────────────────────
        self.btn_toggle_left_panel = ctk.CTkButton(
            row, text="⟨", width=32, height=34,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            font=FONT_BTN_SEC, corner_radius=8,
            command=self._toggle_left_panel,
        )
        self.btn_toggle_left_panel.pack(side="left", padx=(0, SPACE_S))
        self._bind_hover_tip(self.btn_toggle_left_panel, "Hide/show the left panel")

        # ── File ops chip: Open / Save ────────────────────────────────────
        # Export is reached solely from the right panel's EXPORTS section now
        # (was previously duplicated here too via a toolbar dropdown).
        file_chip = self._toolbar_chip(row)
        file_chip.pack(side="left", padx=(0, SPACE_M))

        ctk.CTkButton(
            file_chip, text="Open", width=84, height=34,
            fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
            font=FONT_BTN_PRIMARY, corner_radius=8,
            command=self._open_file,
        ).pack(side="left", padx=(SPACE_S, SPACE_XS))

        self.btn_save_session = ctk.CTkButton(
            file_chip, text="💾  Save", command=self._save_session,
            fg_color=GREEN, hover_color=GREEN_DARK, text_color="white",
            font=FONT_CARD_TITLE, height=34, corner_radius=8,
            state="disabled")
        self.btn_save_session.pack(side="left", padx=(0, SPACE_S))

        # ── Analysis actions chip: Detect Peaks / Analyze ─────────────────
        analysis_chip = self._toolbar_chip(row)
        analysis_chip.pack(side="left", padx=(0, SPACE_M))

        self.btn_preview = ctk.CTkButton(
            analysis_chip, text="Detect Peaks",
            command=self._preview,
            fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
            font=FONT_BTN_PRIMARY, height=34, corner_radius=8)
        self.btn_preview.pack(side="left", padx=(SPACE_S, SPACE_XS))

        run_color = GREEN if NK_AVAILABLE else BORDER2
        self.btn_run = ctk.CTkButton(
            analysis_chip, text="Analyze",
            command=self._run_analysis,
            fg_color=run_color, hover_color=GREEN_DARK if NK_AVAILABLE else BORDER2,
            text_color="white", font=FONT_BTN_PRIMARY, height=34,
            corner_radius=8, state="normal" if NK_AVAILABLE else "disabled")
        self.btn_run.pack(side="left", padx=(0, SPACE_S))
        if not NK_AVAILABLE:
            self._bind_hover_tip(self.btn_run, "pip install neurokit2 to enable", ORANGE)

        # ── Edit chip: Undo / Redo toolbar mirrors ────────────────────────
        edit_chip = self._toolbar_chip(row)
        edit_chip.pack(side="left", padx=(0, SPACE_M))
        self.btn_toolbar_undo = ctk.CTkButton(
            edit_chip, text="↩ Undo", width=88, height=34, font=FONT_BTN_SEC,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8, state="disabled", command=self._undo_edit)
        self.btn_toolbar_undo.pack(side="left", padx=(SPACE_S, SPACE_XS))
        self.btn_toolbar_redo = ctk.CTkButton(
            edit_chip, text="↪ Redo", width=88, height=34, font=FONT_BTN_SEC,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8, state="disabled", command=self._redo_edit)
        self.btn_toolbar_redo.pack(side="left", padx=(0, SPACE_S))

        # ── Identity: project name + file name ────────────────────────────
        left = ctk.CTkFrame(row, fg_color="transparent")
        left.pack(side="left")
        self.lbl_topbar_project = ctk.CTkLabel(
            left, text="—", font=FONT_SIDEBAR_HDR, text_color=TEXT, anchor="w")
        self.lbl_topbar_project.pack(side="left")
        ctk.CTkLabel(left, text="   ", font=FONT_HINT, text_color=LIGHT).pack(side="left")
        self.lbl_topbar_file = ctk.CTkLabel(
            left, text="No file loaded", font=FONT_HINT, text_color=MUTED, anchor="w")
        self.lbl_topbar_file.pack(side="left")

        # ── Right: quality gauge + badge + Settings/Theme ─────────────────
        right = ctk.CTkFrame(row, fg_color="transparent")
        right.pack(side="right")

        self.quality_gauge = make_quality_gauge(right, score=None, compact=True)
        self.quality_gauge.pack(side="left", padx=(0, SPACE_M))
        # lbl_quality now IS the gauge's own caption label -- update_signal_
        # quality()'s existing lbl_quality.configure(...) call needs no
        # changes; its output is immediately superseded by the more
        # descriptive Excellent/Good/Medium/Poor caption update_quality_
        # gauge() sets right after it (see detection_controller.py).
        self.lbl_quality = self.quality_gauge.score_label

        # Quality badge — updated by _update_quality_badge() after analysis
        self._lbl_quality_badge = ctk.CTkLabel(
            right, text="", font=FONT_BADGE,
            text_color="white", fg_color="transparent",
            corner_radius=6, width=120, height=22, anchor="center")
        self._lbl_quality_badge.pack(side="left", padx=(0, SPACE_M))

        # Settings is reached solely from the left sidebar's "⚙ Parameters"
        # bottom button now (was previously duplicated here too).
        ctk.CTkButton(right, text="Theme", width=76, height=34,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      font=FONT_BTN_SEC, command=self._open_theme_dialog,
                      corner_radius=8).pack(side="left", padx=(0, SPACE_S))
        ctk.CTkButton(right, text="☀/☾", width=52, height=34,
                      fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
                      font=FONT_BTN_SEC, corner_radius=8,
                      command=self._toggle_dark_live,
                      ).pack(side="left", padx=(0, SPACE_S))

        # ── Right panel toggle ─────────────────────────────────────────────
        self.btn_toggle_right_panel = ctk.CTkButton(
            right, text="⟩", width=32, height=34,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            font=FONT_BTN_SEC, corner_radius=8,
            command=self._toggle_right_panel,
        )
        self.btn_toggle_right_panel.pack(side="left")
        self._bind_hover_tip(self.btn_toggle_right_panel, "Hide/show the right panel")

    # ─── Right accordion panel ───────────────────────────────────

    def _build_right_panel(self) -> None:
        """Fixed-width accordion panel on the right edge of the window.

        Sibling of self.sidebar (not nested in `main`) -- packed side="right"
        before `main` is created so main's fill="both", expand=True doesn't
        claim the space first, mirroring how the sidebar itself is packed
        before main. Wrapped in a scrollable frame from the start since later
        sub-phases add more sections here.

        Order: STATISTICS / EXPORTS / ML DETECTOR SAVING. Detection/Artifacts/
        Annotations moved to the left sidebar (see _build_sidebar()).
        """
        self.right_panel = ctk.CTkFrame(self, width=280, fg_color=PANEL, corner_radius=0)
        self.right_panel.pack(side="right", fill="y")
        self.right_panel.pack_propagate(False)
        ctk.CTkFrame(self.right_panel, width=1, fg_color=BORDER).pack(side="left", fill="y")

        scroll = ctk.CTkScrollableFrame(self.right_panel, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        self._build_stat_section(scroll)
        self._build_exports_section(scroll)
        self._build_ml_detector_section(scroll)

    def _build_detection_section(self, parent) -> None:
        """DETECTION accordion section (left sidebar).

        Consolidates content that had no independent collapse state before
        (the sidebar's always-visible Method/Analysis-window/Status/Threshold
        block, plus its own Min R-R entry and ADVANCED's SG Options
        sub-group) into one open-by-default section.
        """
        fpx = dict(padx=SPACE_L)

        # ══════════════════════════════════════════════════════
        #  DETECTION
        # ══════════════════════════════════════════════════════
        sec_det = CollapsibleSection(parent, "DETECTION", initially_open=True)
        f = sec_det.frame

        # ── Method ──────────────────────────────────────────
        det_card = ctk.CTkFrame(f, fg_color=CARD, corner_radius=8,
                                border_width=1, border_color=BORDER)
        det_card.pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        det_card_row = ctk.CTkFrame(det_card, fg_color="transparent")
        det_card_row.pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        det_card_row.columnconfigure(1, weight=1)
        ctk.CTkLabel(det_card_row, text="Method",
                     font=FONT_SIDEBAR_HDR, text_color=TEXT,
                     anchor="w").grid(row=0, column=0, sticky="w", padx=(0, SPACE_M))
        self.cb_det_method = ctk.CTkComboBox(
            det_card_row, font=FONT_LABEL, height=26,
            fg_color=BG, border_color=BLUE, button_color=BLUE,
            text_color=TEXT, dropdown_fg_color=BG, dropdown_text_color=TEXT,
            values=["Auto (NeuroKit2)", "SG + Derivative (10 kHz)", "Wavelet (CWT)", "Envelope Max", "ML Detector"],
            command=self._on_det_method_change)
        self.cb_det_method.set("SG + Derivative (10 kHz)")
        self.cb_det_method.grid(row=0, column=1, sticky="ew")

        # ── Analysis window (inline, compact) ──────────────────
        aw_frame = ctk.CTkFrame(f, fg_color=CARD, corner_radius=8,
                                border_width=1, border_color=BORDER)
        aw_frame.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        aw_hdr = ctk.CTkFrame(aw_frame, fg_color="transparent")
        aw_hdr.pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        ctk.CTkLabel(aw_hdr, text="Analysis window", font=FONT_SUBSECTION,
                     text_color=TEXT, anchor="w").pack(side="left")
        ctk.CTkLabel(aw_hdr, text="optional", font=FONT_KPI_LABEL,
                     text_color=LIGHT).pack(side="right")
        aw_entries = ctk.CTkFrame(aw_frame, fg_color="transparent")
        aw_entries.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        aw_entries.columnconfigure(1, weight=1)
        aw_entries.columnconfigure(3, weight=1)
        ctk.CTkLabel(aw_entries, text="From", font=FONT_SMALL, text_color=MUTED,
                     width=36, anchor="w").grid(row=0, column=0)
        self.ent_analysis_t0 = ctk.CTkEntry(
            aw_entries, height=24, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT,
            corner_radius=5, placeholder_text="0 s")
        self.ent_analysis_t0.grid(row=0, column=1, sticky="ew", padx=(SPACE_XS, SPACE_M))
        ctk.CTkLabel(aw_entries, text="To", font=FONT_SMALL, text_color=MUTED,
                     width=24, anchor="w").grid(row=0, column=2)
        self.ent_analysis_t1 = ctk.CTkEntry(
            aw_entries, height=24, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT,
            corner_radius=5, placeholder_text="end")
        self.ent_analysis_t1.grid(row=0, column=3, sticky="ew", padx=(SPACE_XS, 0))
        aw_btns = ctk.CTkFrame(aw_frame, fg_color="transparent")
        aw_btns.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_XS))
        ctk.CTkButton(aw_btns, text="Apply", width=70, height=26,
                      fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
                      font=FONT_SMALL, corner_radius=5,
                      command=self._apply_analysis_window).pack(side="left", padx=(0, SPACE_S))
        ctk.CTkButton(aw_btns, text="Full", width=56, height=26,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_SMALL, corner_radius=5,
                      command=self._reset_analysis_window).pack(side="left")
        self.lbl_analysis_window = ctk.CTkLabel(
            aw_btns, text="", font=FONT_KPI_LABEL, text_color=MUTED, anchor="e")
        self.lbl_analysis_window.pack(side="right", fill="x", expand=True)

        ctk.CTkFrame(f, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))

        # ── Status ──────────────────────────────────────────
        self.lbl_status = ctk.CTkLabel(
            f, text="Ready", font=FONT_SMALL, text_color=MUTED,
            anchor="w", wraplength=230, justify="left")
        self.lbl_status.pack(**fpx, pady=(SPACE_XS, SPACE_XS), fill="x")

        # ── Threshold — always visible, prominent ─────────────
        thr_card = ctk.CTkFrame(f, fg_color=CARD, corner_radius=8,
                                border_width=1, border_color=BORDER)
        thr_card.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
        thr_top_row = ctk.CTkFrame(thr_card, fg_color="transparent")
        thr_top_row.pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        self.lbl_thr = ctk.CTkLabel(
            thr_top_row, text="Threshold:  0.50",
            font=FONT_CARD_TITLE, text_color=TEXT, anchor="w")
        self.lbl_thr.pack(side="left")
        ctk.CTkLabel(thr_top_row, text="strict ↑  /  sensitive ↓",
                     font=FONT_KPI_LABEL, text_color=LIGHT, anchor="e").pack(side="right")
        self.sl_thr = ctk.CTkSlider(
            thr_card, from_=0.01, to=2.0,  # type: ignore
            progress_color=RED, button_color=RED,
            button_hover_color="#FF5252", fg_color=BORDER,
            height=20, command=self._on_threshold_slide)
        self.sl_thr.set(0.50)
        self.sl_thr.pack(fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        thr_bot = ctk.CTkFrame(thr_card, fg_color="transparent")
        thr_bot.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
        thr_bot.columnconfigure(1, weight=1)
        ctk.CTkLabel(thr_bot, text="Exact:", font=FONT_SMALL,
                     text_color=MUTED).grid(row=0, column=0, padx=(0, SPACE_S))
        self.ent_thr = ctk.CTkEntry(thr_bot, height=24, font=FONT_LABEL,
                                     fg_color=PANEL, border_color=BORDER2, text_color=TEXT)
        self.ent_thr.insert(0, "0.50")
        self.ent_thr.grid(row=0, column=1, sticky="ew")
        self.ent_thr.bind("<Return>",   self._on_threshold_entry)
        self.ent_thr.bind("<FocusOut>", self._on_threshold_entry)

        # ── Min R-R ──────────────────────────────────────────
        self._sidebar_entry(f, "Min R-R physio (ms)", "minrr",
                            str(int(MouseECG.MIN_RR_MS)), fpx)
        ctk.CTkLabel(f, text="SG+Deriv: downsample → Savitzky-Golay derivative\n"
                             "Wavelet: CWT bruit/QRS/J-wave séparés (pywt requis)\n"
                             "Envelope Max: maximum local — idéal signaux saturés (clipping ADC)",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))

        # ── Savitzky-Golay options ──────────────────────────
        ctk.CTkFrame(f, height=1, fg_color=BORDER).pack(
            fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        ctk.CTkLabel(f, text="Savitzky-Golay options", font=FONT_SUBSECTION,
                     text_color=MUTED, anchor="w").pack(**fpx, fill="x", pady=(0, SPACE_XS))
        self._sg_frame = ctk.CTkFrame(f, fg_color="transparent")
        self._sidebar_entry_row(self._sg_frame, fpx, [
            ("Target fs (Hz)", "sg_target_fs", "10000"),
            ("SG window (ms)", "sg_window_ms",  "20"),
        ])
        # Sync initial visibility to the Method combobox's default -- DETECTION
        # defaults open (unlike the old ADVANCED home for this frame, which
        # defaulted closed), so an un-synced _sg_frame would show an empty
        # "Savitzky-Golay options" header on first launch even though Method
        # defaults to "SG + Derivative (10 kHz)".
        self._on_det_method_change(self.cb_det_method.get())

    def _build_artifacts_section(self, parent) -> None:
        """ARTIFACTS accordion section (left sidebar), relocated verbatim
        from the old combined Detection/Artifacts/ML-Detector builder."""
        fpx = dict(padx=SPACE_L)
        sec_art = CollapsibleSection(parent, "ARTIFACTS", initially_open=False)
        f = sec_art.frame
        self.btn_review_art = ctk.CTkButton(
            f, text="🔍  Review Artifacts",
            command=self._open_artifact_review,
            fg_color=ORANGE, hover_color=ORANGE_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(30, int(34 * THEME.font_scale)),
            corner_radius=8, state="disabled")
        self.btn_review_art.pack(**fpx, fill="x", pady=(SPACE_S, SPACE_XS))
        ctk.CTkLabel(f, text="Detect + review every artifact. Run Preview first.",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230, justify="left").pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.sw_artifact = self._switch(
            f, "Auto-correct on Full Analysis", fpx, default_on=False)
        ctk.CTkLabel(f, text="OFF by default — use Review for full control",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=230).pack(**fpx, fill="x", pady=(0, SPACE_S))

    def _build_ml_detector_section(self, parent) -> None:
        """ML DETECTOR SAVING accordion section (right panel), relocated
        verbatim from the old combined Detection/Artifacts/ML-Detector
        builder -- same content/bindings, header relabeled per the user's
        request."""
        fpx = dict(padx=SPACE_L)
        sec_ml = CollapsibleSection(parent, "ML DETECTOR SAVING", initially_open=False)
        f = sec_ml.frame
        self.lbl_ml_status = ctk.CTkLabel(
            f, text="No trained model yet", font=FONT_HINT,
            text_color=MUTED, anchor="w", wraplength=230, justify="left")
        self.lbl_ml_status.pack(**fpx, fill="x", pady=(SPACE_S, SPACE_XS))
        self.sw_verified_training = self._switch(
            f, "✓ Verified for training", fpx, default_on=False,
            command=self._on_verified_training_toggle)
        ctk.CTkLabel(
            f, text="Marks this recording's corrected R-peaks as clean "
                    "training data. Takes effect on Save Session.",
            font=FONT_KPI_LABEL, text_color=LIGHT,
            anchor="w", wraplength=230, justify="left").pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.btn_save_for_training = ctk.CTkButton(
            f, text="📥  Save for Training  (no session save)",
            command=self._save_for_training_only,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            font=FONT_BTN_SEC, height=max(28, int(30 * THEME.font_scale)),
            corner_radius=8)
        self.btn_save_for_training.pack(**fpx, fill="x", pady=(0, SPACE_S))
        ctk.CTkLabel(
            f, text="Caches this recording's peaks as training data right "
                    "away, without writing a .ecgsession file or export stats.",
            font=FONT_KPI_LABEL, text_color=LIGHT,
            anchor="w", wraplength=230, justify="left").pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.btn_train_ml = ctk.CTkButton(
            f, text="🤖  Train / Retrain Model…",
            command=self._open_ml_training_dialog,
            fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(30, int(34 * THEME.font_scale)),
            corner_radius=8)
        self.btn_train_ml.pack(**fpx, fill="x", pady=(0, SPACE_S))
        self.after(0, self.refresh_ml_status)

    def _build_exports_section(self, parent) -> None:
        """EXPORTS accordion section (right panel): the export-format
        buttons + Notes, moved verbatim from the old ADVANCED section's
        "Export & Notes" group. This is now the sole entry point for
        exports -- the toolbar's old "Export ▾" dropdown duplicated the
        same 7 commands and was removed."""
        fpx = dict(padx=SPACE_L)
        sec = CollapsibleSection(parent, "EXPORTS", initially_open=False)
        f = sec.frame
        self._btn(f, "📊  Export Excel",              self._export_excel,      fpx, variant="secondary", h=28)
        self._btn(f, "📄  Export RR CSV  (Ctrl+W)",   self._export_rr_csv,     fpx, variant="secondary", h=28)
        self._btn(f, "🖼  Export Figures  (PNG)",      self._export_figures,    fpx, variant="secondary", h=28)
        self._btn(f, "📦  Export ZIP  (Excel+Figs)",  self._export_zip,        fpx, variant="secondary", h=28)
        self._btn(f, "📄  PDF Report  (1 page)",      self._export_pdf_report, fpx, variant="secondary", h=28)
        self._btn(f, "🔬  Export Abnormal Events PDF", self._export_arrhythmia_pdf, fpx, variant="secondary", h=28)
        self._btn(f, "🔬  Export GraphPad Prism",     self._export_prism,      fpx, variant="secondary", h=28)
        ctk.CTkFrame(f, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_S))
        self._btn(f, "📝  Notes (this recording)",    self._open_notes_dialog, fpx, variant="secondary", h=28)

    def _build_stat_section(self, parent) -> None:
        """Categorized statistics: Heart Rate / RR Intervals / HRV / Signal /
        Quality, as a vertically-stacked accordion section in the right
        panel. All values are still driven by update_kpis()
        (plot_controller.py), fanned out to tiles via make_stat_tile()
        (ecg/ui/widgets.py) exactly as before. Tiles are arranged 2-per-row
        within each category box (rather than one full-width column) to keep
        this section compact now that EXPORTS and ML DETECTOR SAVING also
        need to fit below it in the same panel.
        """
        sec = CollapsibleSection(parent, "STATISTICS", initially_open=True)

        # (key, label, unit, hero) -- unit is baked into the tile once at
        # construction (static), hero=True gets the one FONT_KPI_HERO use.
        categories = [
            ("Heart Rate", ORANGE_DARK, [
                ("hr_mean",  "HR Mean",  "bpm", True),
                ("hr_range", "HR Range", "",    False),
            ]),
            ("RR Intervals", BLUE_MID, [
                ("rr_mean", "Mean RR", "ms", False),
                ("n_beats", "N Beats", "",   False),
            ]),
            ("HRV", TEAL, [
                ("sdnn",  "SDNN",  "ms", False),
                ("rmssd", "RMSSD", "ms", False),
                ("pnn50", "pNN6",  "%",  False),
            ]),
            ("Signal", BLUE_DEEP, [
                ("dur",     "Duration",         "s", False),
                ("sq_corr", "Mean correlation", "",  False),
            ]),
            ("Quality", GREEN_DARK, [
                ("sq_score",    "Overall score",       "%", False),
                ("sq_badbeats", "Beats < 0.90 corr.",  "",  False),
                ("sq_noisy_time", "Time < 0.90 corr.", "%", False),
                ("sq_artifact", "Auto-corrected",       "", False),
            ]),
        ]

        for cat_name, accent, tiles in categories:
            box = ctk.CTkFrame(sec.frame, fg_color=CARD, corner_radius=8,
                                border_width=1, border_color=BORDER)
            box.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
            ctk.CTkFrame(box, fg_color=accent, height=3, corner_radius=2).pack(
                fill="x", padx=SPACE_S, pady=(SPACE_XS, 0))
            ctk.CTkLabel(box, text=cat_name.upper(), font=FONT_MICRO,
                         text_color=MUTED, anchor="w").pack(
                fill="x", padx=SPACE_S, pady=(SPACE_XS, SPACE_XS))
            # 2-per-row: chunk tiles into pairs, each pair sharing a row frame.
            # An odd tile out (HRV/Quality have 3) ends up alone on its row.
            for i in range(0, len(tiles), 2):
                row = ctk.CTkFrame(box, fg_color="transparent")
                row.pack(fill="x", padx=SPACE_S, pady=(0, SPACE_XS))
                for key, label, unit, hero in tiles[i:i + 2]:
                    tile = make_stat_tile(row, label, "—", unit=unit, hero=hero, card=False)
                    tile.pack(side="left", fill="both", expand=True, padx=(0, SPACE_XS))
                    # The right panel is much narrower than the old full-width
                    # strip -- long values (e.g. sq_artifact's "N (a dup · b
                    # non-physio · c ectopic)" breakdown) need to wrap instead
                    # of clipping against the tile's now-halved width.
                    tile.value_label.configure(wraplength=95, justify="left")
                    self._kpi[key] = tile.value_label

    def _build_annotations_section(self, parent) -> None:
        """Compact ANNOTATIONS section: live counts + buttons that open the
        existing AnnotationManagerDialog/PacingPeriodManagerDialog popups
        unchanged. Kept alongside (not instead of) the Detection tab's own
        toolbar chips -- those are Detection-tab-only, this is visible from
        every tab.
        """
        sec = CollapsibleSection(parent, "ANNOTATIONS", initially_open=False)
        f = sec.frame
        fpx = dict(padx=SPACE_L)

        # Initial text set directly from current data (not via
        # _update_ann_count()/_update_pacing_count()) -- those also touch
        # self.lbl_ann_count (the Detection tab's toolbar badge), which
        # during _rebuild_ui() isn't rebuilt yet at this point in _build()'s
        # sequence (_build_tabs() runs after _build_right_panel()) and would
        # still be the just-destroyed pre-rebuild widget. _rebuild_ui()
        # refreshes both surfaces itself once _build() has fully completed.
        self.lbl_panel_ann_count = ctk.CTkLabel(
            f, text=f"{len(self._annotations)} annotations",
            font=FONT_LABEL, text_color=TEXT, anchor="w")
        self.lbl_panel_ann_count.pack(**fpx, fill="x", pady=(SPACE_S, SPACE_XS))
        self._btn(f, "📍  Manage Annotations…", self._open_annotations, fpx,
                  variant="secondary", h=28)

        ctk.CTkFrame(f, height=1, fg_color=BORDER).pack(
            fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))

        self.lbl_panel_pacing_count = ctk.CTkLabel(
            f, text=f"{len(self._pacing_periods)} pacing periods",
            font=FONT_LABEL, text_color=TEXT, anchor="w")
        self.lbl_panel_pacing_count.pack(**fpx, fill="x", pady=(SPACE_S, SPACE_XS))
        self._btn(f, "⏱  Manage Pacing Periods…", self._open_pacing_periods, fpx,
                  variant="secondary", h=28)

    # ─── Tabs ─────────────────────────────────────────────────

    def _build_tabs(self, parent) -> None:
        self.tabs = ctk.CTkTabview(
            parent, fg_color=PANEL,
            segmented_button_fg_color=PANEL,
            segmented_button_selected_color=BLUE,
            segmented_button_selected_hover_color=BLUE_HOVER,
            segmented_button_unselected_color=CARD,
            segmented_button_unselected_hover_color=BORDER,
            text_color=TEXT, text_color_disabled=MUTED,
        )
        self.tabs.pack(fill="both", expand=True)
        for name in ["📈 Detection", "💓 HRV", "📏 Intervals",
                     "〰 Beat Template", "⚠ Abnormal Events", "📋 Summary"]:
            self.tabs.add(name)

        self._build_tab_detection()
        self._build_tab_hrv_unified()
        self._build_tab_intervals()
        self._build_tab_beat_template()
        self._build_tab_arrhythmias()
        self._build_tab_summary()

    def _build_tab_detection(self) -> None:
        t = self.tabs.tab("📈 Detection")

        # Layout (top → bottom, all pack):
        #   nav bar     (fill=x, no expand) — time navigation controls
        #   thin separator
        #   detail toolbar (fill=x, no expand)
        #   detail plot    (fill=both, expand=True)

        # ── Navigation bar ────────────────────────────────────────────────
        # Instrument-style layout: 3 grouped chips inside the flat PANEL bar
        # -- Transport | Position & Zoom | Readout (docked right).
        nav = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0, height=34)
        nav.pack(side="top", fill="x")
        nav.pack_propagate(False)

        # ── Transport chip: jump/step controls ──────────────────────────
        transport_chip = self._toolbar_chip(nav)
        transport_chip.pack(side="left", padx=(SPACE_M, SPACE_S))

        btn_nav_reset = ctk.CTkButton(
            transport_chip, text="⏮", width=36, height=28, font=FONT_LABEL,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            corner_radius=8, command=self._nav_reset)
        btn_nav_reset.pack(side="left", padx=(SPACE_S, SPACE_XS))
        self._bind_hover_tip(btn_nav_reset, "Jump to start")

        btn_nav_big_back = ctk.CTkButton(
            transport_chip, text="◀◀", width=44, height=28, font=FONT_LABEL,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            corner_radius=8, command=lambda: self._navigate_big(-1))
        btn_nav_big_back.pack(side="left", padx=(0, SPACE_S))
        self._bind_hover_tip(btn_nav_big_back, "Big step back")

        btn_nav_back = ctk.CTkButton(
            transport_chip, text="◀", width=36, height=28, font=FONT_LABEL,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            corner_radius=8, command=lambda: self._navigate(-1))
        btn_nav_back.pack(side="left", padx=(0, SPACE_M))
        self._bind_hover_tip(btn_nav_back, "Step back")

        btn_nav_fwd = ctk.CTkButton(
            transport_chip, text="▶", width=36, height=28, font=FONT_LABEL,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            corner_radius=8, command=lambda: self._navigate(+1))
        btn_nav_fwd.pack(side="left", padx=(0, SPACE_S))
        self._bind_hover_tip(btn_nav_fwd, "Step forward")

        btn_nav_big_fwd = ctk.CTkButton(
            transport_chip, text="▶▶", width=44, height=28, font=FONT_LABEL,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            corner_radius=8, command=lambda: self._navigate_big(+1))
        btn_nav_big_fwd.pack(side="left", padx=(0, SPACE_XS))
        self._bind_hover_tip(btn_nav_big_fwd, "Big step forward")

        btn_nav_end = ctk.CTkButton(
            transport_chip, text="⏭", width=36, height=28, font=FONT_LABEL,
            fg_color=BORDER, hover_color=BORDER2, text_color=TEXT,
            corner_radius=8, command=self._nav_end)
        btn_nav_end.pack(side="left", padx=(0, SPACE_S))
        self._bind_hover_tip(btn_nav_end, "Jump to end")

        # ── Position & Zoom chip: current position + window size ────────
        pos_chip = self._toolbar_chip(nav)
        pos_chip.pack(side="left", padx=(0, SPACE_S))

        lbl_nav_pos = ctk.CTkLabel(pos_chip, text="t =", font=FONT_SMALL,
                     text_color=MUTED)
        lbl_nav_pos.pack(side="left", padx=(SPACE_S, SPACE_XS))
        self.ent_nav_pos = ctk.CTkEntry(
            pos_chip, width=72, height=28, font=FONT_LABEL,
            fg_color=CARD, border_color=BORDER2, text_color=TEXT,
            corner_radius=6,
            placeholder_text="0.000")
        self.ent_nav_pos.pack(side="left", padx=(0, SPACE_XS))  # type: ignore[union-attr]
        ctk.CTkLabel(pos_chip, text="s", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        _pos_tip = "Visible window start time — type a value and press Go (or Enter) to jump there"
        self._bind_hover_tip(lbl_nav_pos, _pos_tip)
        self._bind_hover_tip(self.ent_nav_pos, _pos_tip)
        ctk.CTkButton(pos_chip, text="Go", width=48, height=28, font=FONT_SMALL,
                      fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
                      corner_radius=8,
                      command=self._nav_goto).pack(side="left", padx=(SPACE_S, SPACE_L))
        # Bind Enter key on the position field
        self.ent_nav_pos.bind("<Return>", lambda _e: self._nav_goto())  # type: ignore[union-attr]

        ctk.CTkLabel(pos_chip, text="Window:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(0, SPACE_S))
        self.ent_window = ctk.CTkEntry(pos_chip, width=48, height=28, font=FONT_LABEL,
                                       fg_color=CARD, border_color=BORDER2, text_color=TEXT,
                                       corner_radius=6)
        self.ent_window.insert(0, "2")
        self.ent_window.pack(side="left", padx=(0, SPACE_XS))
        self._bind_hover_tip(self.ent_window, "Visible window duration (s)")
        ctk.CTkLabel(pos_chip, text="s", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(0, SPACE_S))

        # ── Readout chip: duration, docked to the bar's right edge ──────
        readout_chip = self._toolbar_chip(nav)
        readout_chip.pack(side="right", padx=(SPACE_S, SPACE_M))
        self.lbl_sig_duration = ctk.CTkLabel(
            readout_chip, text="", font=FONT_MONO, text_color=MUTED, anchor="w")
        self.lbl_sig_duration.pack(side="left", padx=SPACE_S)  # type: ignore[union-attr]

        # Separator
        tk.Frame(t, height=1, bg=BORDER).pack(side="top", fill="x", padx=SPACE_M, pady=SPACE_XS)

        # Detail toolbar (fixed height)
        hdr = ctk.CTkFrame(t, fg_color="transparent", height=32)
        hdr.pack(side="top", fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_XS))
        hdr.pack_propagate(False)

        ctk.CTkLabel(hdr, text="SIGNAL ECG", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left", anchor="w")

        ctk.CTkFrame(hdr, width=1, fg_color=BORDER).pack(side="left", fill="y", padx=(SPACE_L, SPACE_S), pady=SPACE_XS)

        # ── Edit tools chip ───────────────────────────────────────────────
        edit_chip = self._toolbar_chip(hdr)
        edit_chip.pack(side="left", padx=(0, SPACE_M))
        self.btn_edit_mode = ctk.CTkButton(
            edit_chip, text="Edit Peaks", width=96, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._toggle_edit_mode,
        )
        self.btn_edit_mode.pack(side="left", padx=(SPACE_S, SPACE_XS))
        self.lbl_edit_hint = ctk.CTkLabel(
            edit_chip,
            text="L-click: exclude/restore   R-click: add   Ctrl+Z: undo",
            font=FONT_HINT, text_color=ORANGE,
        )
        self.btn_undo_edit = ctk.CTkButton(
            edit_chip, text="↩ Undo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._undo_edit,
        )
        self.btn_undo_edit.pack(side="left", padx=(SPACE_XS, 0))
        self.btn_redo_edit = ctk.CTkButton(
            edit_chip, text="↪ Redo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._redo_edit,
        )
        self.btn_redo_edit.pack(side="left", padx=(SPACE_XS, SPACE_S))
        self.btn_clear_excl = ctk.CTkButton(
            edit_chip, text="Clear Edits", width=100, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._clear_manual_exclusions,
        )
        self.btn_clear_excl.pack(side="left", padx=(0, SPACE_S))

        # ── Free Placement chip (bypass proximity guard) ──────────────────
        fp_chip = self._toolbar_chip(hdr)
        fp_chip.pack(side="left", padx=(0, SPACE_M))
        self.btn_free_placement = ctk.CTkButton(
            fp_chip, text="Free Placement", width=118, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._toggle_free_placement,
        )
        self.btn_free_placement.pack(side="left", padx=(SPACE_S, SPACE_XS))
        _fp_tip = ctk.CTkLabel(
            fp_chip, text="?", width=28, height=28,
            font=FONT_CARD_TITLE, text_color=MUTED,
            fg_color=BORDER, corner_radius=14,
        )
        _fp_tip.pack(side="left", padx=(0, SPACE_S))
        self._bind_hover_tip(
            _fp_tip,
            "Free Placement: right-click adds a peak at the exact clicked position — "
            "no snapping, no proximity guard, works even on top of existing peaks.")

        # ── Annotations chip ───────────────────────────────────────────────
        ann_chip = self._toolbar_chip(hdr)
        ann_chip.pack(side="left", padx=(0, SPACE_M))
        self.btn_annotations = ctk.CTkButton(
            ann_chip, text="Annotations", width=110, height=28, font=FONT_SMALL,
            fg_color=PURPLE_DARK, hover_color=PURPLE, text_color="white",
            corner_radius=8,
            command=self._open_annotations,
        )
        self.btn_annotations.pack(side="left", padx=(SPACE_S, SPACE_XS))  # type: ignore[union-attr]
        self.lbl_ann_count = ctk.CTkLabel(
            ann_chip, text="", font=FONT_HINT, text_color=MUTED, anchor="w")
        self.lbl_ann_count.pack(side="left", padx=(0, SPACE_S))  # type: ignore[union-attr]

        # ── RR/HR/Quality strip toggle ───────────────────────────────────────
        # "Pacing Periods" chip removed here (redundant with the ANNOTATIONS
        # panel section's "Manage Pacing Periods…" button, the sole remaining
        # entry point) -- this slot now hosts the show/hide toggle for the
        # RR/HR/Quality sub-plot strip below the detail plot.
        rrhr_chip = self._toolbar_chip(hdr)
        rrhr_chip.pack(side="left")
        _rrhr_on = self.ui.rrhr_strip_visible
        self.btn_toggle_rrhr = ctk.CTkButton(
            rrhr_chip, text="RR/HR/Quality", width=118, height=28, font=FONT_SMALL,
            fg_color=BLUE if _rrhr_on else BORDER,
            hover_color=BLUE_HOVER if _rrhr_on else BORDER2,
            text_color="white" if _rrhr_on else MUTED,
            corner_radius=8,
            command=self._toggle_rrhr_strip,
        )
        self.btn_toggle_rrhr.pack(side="left", padx=(SPACE_S, SPACE_S))  # type: ignore[union-attr]

        # Full-recording scrubber strip, directly above the detail plot
        # (Audacity/Premiere-style position indicator). Fixed height
        # (pack_propagate(False), mirroring nav/hdr) rather than fill=both.
        # A simple flat line + draggable viewport cursor (see draw_overview())
        # needs much less vertical space than the old envelope-plot minimap did.
        ov_frame = tk.Frame(t, bg=PLOT["bg"], bd=0, highlightthickness=0, height=28)
        ov_frame.pack(side="top", fill="x", padx=SPACE_XS, pady=(0, SPACE_XS))
        ov_frame.pack_propagate(False)
        self._slots["overview"] = CanvasSlot(ov_frame, 14, 0.35, toolbar=False)
        self._slots["overview"].canvas.mpl_connect(
            "button_press_event", self._on_overview_click)
        self._slots["overview"].canvas.mpl_connect(
            "motion_notify_event", self._on_overview_motion)
        self._slots["overview"].canvas.mpl_connect(
            "button_release_event", self._on_overview_release)

        # Plot area — detail trace (~70%) + synced RR/HR strip (~30%),
        # grid-managed so both share the remaining space proportionally
        # (same grid_rowconfigure(weight=N) idiom already used by
        # _build_hrv_view_rr's RR-tachogram/stats split).
        self.plot_area = ctk.CTkFrame(t, fg_color="transparent")
        self.plot_area.pack(side="top", fill="both", expand=True)
        self.plot_area.grid_rowconfigure(0, weight=7)
        self.plot_area.grid_rowconfigure(1, weight=3 if self.ui.rrhr_strip_visible else 0)
        self.plot_area.grid_columnconfigure(0, weight=1)

        det_frame = tk.Frame(self.plot_area, bg=PLOT["bg"], bd=0, highlightthickness=0)
        det_frame.grid(row=0, column=0, sticky="nsew", padx=SPACE_XS, pady=(0, SPACE_XS))
        self._slots["detail"] = CanvasSlot(det_frame, 12, 6.5, toolbar=True)

        self._slots["detail"].canvas.mpl_connect(
            "scroll_event", self._on_detail_scroll)
        self._slots["detail"].canvas.mpl_connect(
            "button_press_event", self._on_detail_click)
        self._hover_motion_cid = self._slots["detail"].canvas.mpl_connect(
            "motion_notify_event", self._on_detail_motion)

        # RR/HR strip — synced to the same nav window as the detail plot,
        # via ECGApp._draw_detail()'s fan-out (see draw_detail_rrhr()).
        # Show/hide is toggled by btn_toggle_rrhr in the hdr toolbar above.
        self.subplot_frame = tk.Frame(self.plot_area, bg=PLOT["bg"], bd=0, highlightthickness=0)
        self.subplot_frame.grid(row=1, column=0, sticky="nsew", padx=SPACE_XS, pady=(0, SPACE_XS))
        if not self.ui.rrhr_strip_visible:
            self.subplot_frame.grid_remove()
        self._slots["detail_rrhr"] = CanvasSlot(self.subplot_frame, 12, 2.8, toolbar=False)

    def _build_tab_hrv_unified(self) -> None:
        """Unified HRV tab with internal segmented navigation.

        Merges: RR / HR  |  Temporel  |  Fréquentiel  |  Non-linéaire  |  Epochs  |  Glissant
        """
        t = self.tabs.tab("💓 HRV")
        t.grid_rowconfigure(1, weight=1)
        t.grid_columnconfigure(0, weight=1)

        # ── Top: segmented sub-tab selector ───────────────────────────────
        _HRV_VIEWS = ["RR / HR", "Time Domain", "Frequency", "Non-linear", "Epochs", "Rolling"]
        self._hrv_subframes: "dict[str, ctk.CTkFrame]" = {}

        seg_bar = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0)
        seg_bar.grid(row=0, column=0, sticky="ew")
        self._hrv_seg = ctk.CTkSegmentedButton(
            seg_bar,
            values=_HRV_VIEWS,
            font=FONT_SMALL,
            fg_color=BORDER,
            selected_color=BLUE,
            selected_hover_color=BLUE_HOVER,
            unselected_color=BORDER,
            unselected_hover_color=BORDER2,
            text_color=TEXT,
            text_color_disabled=MUTED,
            command=self._on_hrv_view_change,
        )
        self._hrv_seg.pack(padx=SPACE_M, pady=SPACE_S, anchor="w")  # type: ignore[union-attr]
        self._hrv_seg.set("RR / HR")  # type: ignore[union-attr]

        # ── Content area — one frame per view, pack/pack_forget ───────────
        content_area = ctk.CTkFrame(t, fg_color="transparent")
        content_area.grid(row=1, column=0, sticky="nsew")
        content_area.grid_rowconfigure(0, weight=1)
        content_area.grid_columnconfigure(0, weight=1)
        self._hrv_content_area = content_area

        for view in _HRV_VIEWS:
            f = ctk.CTkFrame(content_area, fg_color="transparent")
            f.grid_rowconfigure(0, weight=1)
            f.grid_columnconfigure(0, weight=1)
            self._hrv_subframes[view] = f

        # ── Build each sub-view content ───────────────────────────────────
        self._build_hrv_view_rr(self._hrv_subframes["RR / HR"])
        self._build_hrv_view_temporel(self._hrv_subframes["Time Domain"])
        self._build_hrv_view_freq(self._hrv_subframes["Frequency"])
        self._build_hrv_view_nonlin(self._hrv_subframes["Non-linear"])
        self._build_hrv_view_epochs(self._hrv_subframes["Epochs"])
        self._build_hrv_view_rolling(self._hrv_subframes["Rolling"])

        # Show initial view
        self._hrv_subframes["RR / HR"].pack(fill="both", expand=True)
        self._hrv_current_view = "RR / HR"

    def _on_hrv_view_change(self, view: str) -> None:
        """Switch the visible HRV sub-view."""
        if not hasattr(self, "_hrv_subframes"):
            return
        for name, frame in self._hrv_subframes.items():
            if name == view:
                frame.pack(fill="both", expand=True)
            else:
                frame.pack_forget()
        self._hrv_current_view = view

    def _build_hrv_view_rr(self, parent: ctk.CTkFrame) -> None:
        """RR tachogram + histogram — formerly the 'RR / HR' tab."""
        # Ratio 4:1 (était 3:2) — les deux tachogrammes empilés (RR + HR)
        # sont le contenu principal de l'onglet et bénéficient d'un maximum
        # de hauteur ; le tableau de stats + histogramme du bas est compact
        # et n'a pas besoin d'autant d'espace que ce que 2/5 lui donnait.
        parent.grid_rowconfigure(0, weight=4)
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        rr_frame = ctk.CTkFrame(parent, fg_color="transparent")
        rr_frame.grid(row=0, column=0, sticky="nsew", padx=SPACE_S, pady=(SPACE_S, SPACE_XS))
        rr_frame.grid_rowconfigure(0, weight=1)
        rr_frame.grid_columnconfigure(0, weight=1)
        self._slots["rr"] = CanvasSlot(rr_frame, 14, 5.0, toolbar=False)

        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(SPACE_XS, SPACE_S))
        row.grid_rowconfigure(0, weight=1)
        row.grid_columnconfigure(0, weight=1)
        row.grid_columnconfigure(1, weight=2)

        # Stats textbox (narrow)
        stats_card = ctk.CTkFrame(row, fg_color=PANEL, corner_radius=0)
        stats_card.grid(row=0, column=0, sticky="nsew", padx=(0, SPACE_S))
        stats_card.grid_rowconfigure(0, weight=0)
        stats_card.grid_rowconfigure(1, weight=1)
        stats_card.grid_columnconfigure(0, weight=1)
        _rr_btn = ctk.CTkButton(stats_card, text="📋 Copy for Excel", height=22,
                                font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                                text_color="white",
                                command=lambda: self._copy_tsv(self.txt_rr))
        _rr_btn.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, 0))
        self.txt_rr = ctk.CTkTextbox(stats_card, font=FONT_MONO, fg_color=PANEL,
                                     text_color=TEXT, border_width=0)
        self.txt_rr.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(SPACE_XS, SPACE_S))

        # Histogram
        hist_card = ctk.CTkFrame(row, fg_color=PANEL, corner_radius=0)
        hist_card.grid(row=0, column=1, sticky="nsew")
        hist_card.grid_rowconfigure(0, weight=1)
        hist_card.grid_columnconfigure(0, weight=1)
        self._slots["rr_hist"] = CanvasSlot(hist_card, 7, 4.5, toolbar=False)

    def _build_hrv_view_temporel(self, parent: ctk.CTkFrame) -> None:
        """Time-domain HRV — text tables (computed in Core Analysis)."""
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        inner = ctk.CTkFrame(parent, fg_color=PANEL, corner_radius=6)
        inner.grid(row=0, column=0, sticky="nsew", padx=SPACE_M, pady=SPACE_M)
        inner.grid_rowconfigure(1, weight=1)
        inner.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(inner, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_XS))
        ctk.CTkLabel(hdr, text="HRV TIME DOMAIN", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left")
        ctk.CTkButton(hdr, text="📋 Copy for Excel", height=22,
                      font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                      text_color="white", width=140,
                      command=lambda: self._copy_tsv(self.txt_td)
                      ).pack(side="right")
        self.txt_td = ctk.CTkTextbox(inner, font=FONT_MONO, fg_color=PANEL,
                                     text_color=TEXT, border_width=0)
        self.txt_td.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_M))

    def _build_hrv_view_freq(self, parent: ctk.CTkFrame) -> None:
        """Frequency-domain HRV — PSD + radar + tables."""
        parent.grid_rowconfigure(0, weight=0)
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        # Action bar
        bar = ctk.CTkFrame(parent, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        self.btn_run_freq = ctk.CTkButton(
            bar, text="⚡  Freq HRV",
            command=self._run_freq,
            fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=5,
        )
        self.btn_run_freq.pack(side="left")  # type: ignore[union-attr]
        self.lbl_freq_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_freq_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        # Content: text left, PSD+radar right
        content = ctk.CTkFrame(parent, fg_color="transparent")
        content.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(SPACE_XS, SPACE_S))
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=2)

        left = ctk.CTkFrame(content, fg_color=PANEL, corner_radius=6)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, SPACE_S))
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)
        _hdr = ctk.CTkFrame(left, fg_color="transparent")
        _hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        ctk.CTkLabel(_hdr, text="FREQUENCY DOMAIN", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left")
        ctk.CTkButton(_hdr, text="📋 Copy for Excel", height=20,
                      font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                      text_color="white", width=140,
                      command=lambda: self._copy_tsv(self.txt_fd)
                      ).pack(side="right")
        self.txt_fd = ctk.CTkTextbox(left, font=FONT_MONO, fg_color=PANEL,
                                     text_color=TEXT, border_width=0)
        self.txt_fd.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_S))

        right = ctk.CTkFrame(content, fg_color="transparent")
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)
        for row_i, slot_key in enumerate(["psd", "radar"]):
            card = ctk.CTkFrame(right, fg_color=PANEL, corner_radius=0)
            # Espacement symétrique entre les deux graphiques empilés (PSD/radar) :
            # avant, le bas avait pady=0 → les deux cartes semblaient "collées"
            # côté inférieur alors qu'il y avait de l'air côté supérieur.
            card.grid(row=row_i, column=0, sticky="nsew",
                      pady=(0, SPACE_S) if row_i == 0 else (SPACE_S, 0))
            card.grid_rowconfigure(0, weight=1)
            card.grid_columnconfigure(0, weight=1)
            self._slots[slot_key] = CanvasSlot(card, 8, 4.5, toolbar=False)

    def _build_hrv_view_nonlin(self, parent: ctk.CTkFrame) -> None:
        """Non-linear HRV — Poincaré + SampEn/DFA table."""
        parent.grid_rowconfigure(0, weight=0)
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        bar = ctk.CTkFrame(parent, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        self.btn_run_nonlin = ctk.CTkButton(
            bar, text="⚡  Non-linear HRV",
            command=self._run_nonlinear,
            fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=5,
        )
        self.btn_run_nonlin.pack(side="left")  # type: ignore[union-attr]
        self.lbl_nonlin_status = ctk.CTkLabel(
            bar, text="  SampEn + DFA can take 30 s+ on long recordings",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_nonlin_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        content = ctk.CTkFrame(parent, fg_color="transparent")
        content.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(SPACE_XS, SPACE_S))
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=2)

        left = ctk.CTkFrame(content, fg_color=PANEL, corner_radius=6)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, SPACE_S))
        left.grid_rowconfigure(0, weight=0)
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)
        _nl_hdr = ctk.CTkFrame(left, fg_color="transparent")
        _nl_hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
        ctk.CTkLabel(_nl_hdr, text="NON-LINEAR", font=FONT_SIDEBAR_HDR,
                     text_color=MUTED).pack(side="left")
        ctk.CTkButton(_nl_hdr, text="📋 Copy for Excel", height=20,
                      font=FONT_HINT, fg_color=BLUE, hover_color=BLUE_HOVER,
                      text_color="white", width=140,
                      command=lambda: self._copy_tsv(self.txt_nl)
                      ).pack(side="right")
        self.txt_nl = ctk.CTkTextbox(left, font=FONT_MONO, fg_color=PANEL,
                                     text_color=TEXT, border_width=0)
        self.txt_nl.grid(row=1, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_S))

        right = ctk.CTkFrame(content, fg_color=PANEL, corner_radius=0)
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(0, weight=1)
        right.grid_columnconfigure(0, weight=1)
        self._slots["poincare"] = CanvasSlot(right, 8, 8, toolbar=False)

    def _build_hrv_view_epochs(self, parent: ctk.CTkFrame) -> None:
        """Epoch HRV analysis — formerly the 'Epochs' tab."""
        parent.grid_rowconfigure(1, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        hdr = ctk.CTkFrame(parent, fg_color="transparent")
        hdr.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))
        ctk.CTkLabel(hdr, text="HRV PER EPOCH",
                     font=FONT_SECTION_HDR, text_color=MUTED).pack(side="left")
        self.btn_compute_epochs = ctk.CTkButton(
            hdr, text="⟳  Compute epochs", command=self._compute_epochs,
            fg_color=BLUE, hover_color=BLUE_DEEP, text_color="white",
            font=FONT_SMALL, height=28, corner_radius=5)
        self.btn_compute_epochs.pack(side="right")
        self.lbl_epoch_count = ctk.CTkLabel(
            hdr, text="", font=FONT_SMALL, text_color=BLUE)
        self.lbl_epoch_count.pack(side="right", padx=(0, SPACE_M))

        # Settings row
        settings_row = ctk.CTkFrame(parent, fg_color="transparent")
        settings_row.grid(row=1, column=0, sticky="ew", padx=SPACE_M, pady=(0, SPACE_S))
        for lbl_t, attr_n, default in [
            ("Epoch (s)", "epoch", str(int(MouseECG.EPOCH_DEFAULT_S))),
            ("Overlap (s)", "overlap", "0"),
        ]:
            ctk.CTkLabel(settings_row, text=lbl_t, font=FONT_SMALL,
                         text_color=MUTED).pack(side="left", padx=(0, SPACE_S))
            ent = ctk.CTkEntry(settings_row, width=58, height=26,
                               font=FONT_LABEL, fg_color=BG,
                               border_color=BORDER2, text_color=TEXT)
            ent.insert(0, default)
            ent.pack(side="left", padx=(0, SPACE_L))
            setattr(self, f"ent_{attr_n}", ent)
        self.sw_epoch = self._switch(settings_row, "Auto-run after analysis",
                                     dict(padx=0))
        self.sw_epoch.pack(side="left")
        self.lbl_epoch_info = ctk.CTkLabel(
            settings_row, text="", font=FONT_HINT, text_color=MUTED)
        self.lbl_epoch_info.pack(side="right", padx=(0, SPACE_S))

        parent.grid_rowconfigure(1, weight=0)
        parent.grid_rowconfigure(2, weight=0)
        parent.grid_rowconfigure(3, weight=0)
        parent.grid_rowconfigure(4, weight=1)

        # copy bar
        ep_bar = ctk.CTkFrame(parent, fg_color="transparent")
        ep_bar.grid(row=2, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, 0))
        ctk.CTkLabel(ep_bar, text="Epoch table:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self.btn_copy_epochs = ctk.CTkButton(
            ep_bar, text="📋  Copy for Excel", height=24,
            font=FONT_SMALL, fg_color=BLUE, hover_color=BLUE_HOVER,
            text_color="white", width=160,
            command=lambda: self._copy_tsv(self.txt_epochs))
        self.btn_copy_epochs.pack(side="right", padx=(0, SPACE_XS))  # type: ignore[union-attr]

        ep_sf = ctk.CTkFrame(parent, fg_color=PANEL, corner_radius=0)
        ep_sf.grid(row=3, column=0, sticky="nsew", padx=SPACE_M, pady=(SPACE_S, SPACE_S))
        ep_sf.grid_rowconfigure(0, weight=1)
        ep_sf.grid_columnconfigure(0, weight=1)
        self._slots["epochs"] = CanvasSlot(ep_sf, 14, 5.5, toolbar=False)

        epochs_tb = ctk.CTkTextbox(parent, font=FONT_BODY, fg_color="transparent",
                                   text_color=TEXT, border_width=0,
                                   scrollbar_button_color=BORDER,
                                   scrollbar_button_hover_color=BORDER2,
                                   height=160)
        epochs_tb.grid(row=4, column=0, sticky="ew", padx=SPACE_M, pady=(0, SPACE_S))
        self.txt_epochs = epochs_tb

    def _build_hrv_view_rolling(self, parent: ctk.CTkFrame) -> None:
        """Rolling / sliding-window HRV — formerly the 'Rolling HRV' tab."""
        parent.grid_columnconfigure(0, weight=1)

        bar = ctk.CTkFrame(parent, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))

        ctk.CTkLabel(bar, text="Window (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self.ent_roll_win = ctk.CTkEntry(
            bar, width=58, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_roll_win.insert(0, "30")
        self.ent_roll_win.pack(side="left", padx=(SPACE_S, SPACE_L))

        ctk.CTkLabel(bar, text="Step (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self.ent_roll_step = ctk.CTkEntry(
            bar, width=48, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_roll_step.insert(0, "5")
        self.ent_roll_step.pack(side="left", padx=(SPACE_S, SPACE_L))

        self.btn_roll_compute = ctk.CTkButton(
            bar, text="⟳  Compute", command=self._compute_rolling_hrv,
            fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white",
            font=FONT_BTN_SEC, height=28, corner_radius=5)
        self.btn_roll_compute.pack(side="left", padx=(SPACE_L, 0))

        self.lbl_roll_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_roll_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        # Metrics checkboxes get their own row -- HR/SDNN/RMSSD/pNN6 plus
        # SD1/SD2/LF_nu/HF_nu/LF_HF (task: surface compute_segment_stats'
        # richer metric set as continuous trends, not just inside the
        # Compare Segments modal) no longer fit the single window/step/
        # compute row without wrapping badly.
        metrics_bar = ctk.CTkFrame(parent, fg_color="transparent")
        metrics_bar.grid(row=1, column=0, sticky="ew", padx=SPACE_M, pady=(0, SPACE_S))
        ctk.CTkLabel(metrics_bar, text="Metrics:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self._roll_metrics: "dict[str, ctk.CTkCheckBox]" = {}
        for metric, default in [("HR", True), ("SDNN", True), ("RMSSD", True),
                                 ("pNN6", False), ("SD1", False), ("SD2", False),
                                 ("LF_nu", False), ("HF_nu", False), ("LF_HF", False)]:
            cb = ctk.CTkCheckBox(metrics_bar, text=metric, font=FONT_SMALL,
                                 text_color=MUTED, fg_color=BLUE,
                                 checkmark_color="white",
                                 border_color=BORDER2, width=16)
            if default:
                cb.select()
            cb.pack(side="left", padx=(SPACE_S, 0))
            self._roll_metrics[metric] = cb

        parent.grid_rowconfigure(1, weight=0)
        parent.grid_rowconfigure(2, weight=0)
        parent.grid_rowconfigure(3, weight=1)
        parent.grid_rowconfigure(4, weight=0)

        # copy bar
        roll_bar = ctk.CTkFrame(parent, fg_color="transparent")
        roll_bar.grid(row=2, column=0, sticky="ew", padx=SPACE_M, pady=(0, SPACE_S))
        ctk.CTkLabel(roll_bar, text="Window table:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left")
        self.btn_copy_rolling = ctk.CTkButton(
            roll_bar, text="📋  Copy for Excel", height=24,
            font=FONT_SMALL, fg_color=BLUE, hover_color=BLUE_HOVER,
            text_color="white", width=160,
            command=lambda: self._copy_tsv(self.txt_rolling))
        self.btn_copy_rolling.pack(side="right", padx=(0, SPACE_XS))

        plot_frame = ctk.CTkFrame(parent, fg_color=PANEL, corner_radius=0)
        plot_frame.grid(row=3, column=0, sticky="nsew", padx=SPACE_M, pady=(0, SPACE_S))
        plot_frame.grid_rowconfigure(0, weight=1)
        plot_frame.grid_columnconfigure(0, weight=1)
        self._slots["rolling_hrv"] = CanvasSlot(plot_frame, 14, 6.0, toolbar=False)

        rolling_tb = ctk.CTkTextbox(parent, font=FONT_BODY, fg_color="transparent",
                                    text_color=TEXT, border_width=0,
                                    scrollbar_button_color=BORDER,
                                    scrollbar_button_hover_color=BORDER2,
                                    height=160)
        rolling_tb.grid(row=4, column=0, sticky="ew", padx=SPACE_M, pady=(0, SPACE_S))
        self.txt_rolling = rolling_tb

    def _build_tab_arrhythmias(self) -> None:
        """Build the Abnormal Events tab: event cards (left) + ECG viewer (right)."""
        t = self.tabs.tab("⚠ Abnormal Events")
        t.grid_rowconfigure(1, weight=1)
        t.grid_columnconfigure(0, weight=1)

        # ── Row 0 : Action bar ────────────────────────────────
        bar = ctk.CTkFrame(t, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))

        self.btn_run_arrhythmia = ctk.CTkButton(
            bar, text="⚡  Classify",
            command=self._run_arrhythmia_analysis,
            fg_color=RED, hover_color=RED_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=8,
        )
        self.btn_run_arrhythmia.pack(side="left")  # type: ignore[union-attr]

        # ── Baseline window ──
        ctk.CTkLabel(bar, text="Baseline (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(SPACE_L, SPACE_XS))
        self.ent_arr_baseline = ctk.CTkEntry(
            bar, width=48, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_baseline.insert(0, "30")
        self.ent_arr_baseline.pack(side="left", padx=(0, SPACE_XS))
        _tip = ctk.CTkLabel(bar, text="(pre-stimulus)", font=FONT_KPI_LABEL,
                            text_color=MUTED)
        _tip.pack(side="left", padx=(0, SPACE_M))

        # ── Brady/tachy threshold ──
        ctk.CTkLabel(bar, text="ΔHR threshold (%):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(0, SPACE_XS))
        self.ent_arr_brady_pct = ctk.CTkEntry(
            bar, width=44, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_brady_pct.insert(0, "20")
        self.ent_arr_brady_pct.pack(side="left", padx=(0, SPACE_M))

        # ── Minimum beats ──
        ctk.CTkLabel(bar, text="Min duration (beats):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="left", padx=(0, SPACE_XS))
        self.ent_arr_min_beats = ctk.CTkEntry(
            bar, width=44, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_min_beats.insert(0, "10")
        self.ent_arr_min_beats.pack(side="left", padx=(0, SPACE_M))

        self.lbl_arrhythmia_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_arrhythmia_status.pack(side="left", padx=SPACE_M)  # type: ignore[union-attr]

        ctk.CTkButton(bar, text="📋 Copy TSV",
                      command=self._copy_arrhythmia_tsv,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=28, corner_radius=5,
                      ).pack(side="right")

        # ── Row 1 : Split body ────────────────────────────────
        body = tk.Frame(t, bg=BG, bd=0, highlightthickness=0)
        body.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(0, SPACE_S))
        body.grid_rowconfigure(0, weight=1)
        body.grid_columnconfigure(0, weight=0)   # event list — fixed
        body.grid_columnconfigure(1, weight=1)   # ECG viewer — elastic

        # ── Left : scrollable event list ─────────────────────
        list_outer = ctk.CTkFrame(body, fg_color=PANEL, corner_radius=6, width=280)
        list_outer.grid(row=0, column=0, sticky="ns", padx=(0, SPACE_S))
        list_outer.grid_propagate(False)
        list_outer.grid_rowconfigure(1, weight=1)
        list_outer.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(list_outer, text="DETECTED EPISODES",
                     font=FONT_SUBSECTION, text_color=MUTED,
                     anchor="w").grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))

        self._arr_event_scroll = ctk.CTkScrollableFrame(
            list_outer, fg_color="transparent",
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=BORDER2,
        )
        self._arr_event_scroll.grid(row=1, column=0, sticky="nsew", padx=SPACE_XS, pady=(0, SPACE_S))
        self._arr_event_scroll.grid_columnconfigure(0, weight=1)
        self._arr_card_widgets = []

        # ── Right : ECG detail panel ──────────────────────────
        right = tk.Frame(body, bg=BG, bd=0, highlightthickness=0)
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)

        # Edit / nav controls bar
        ebar = ctk.CTkFrame(right, fg_color=PANEL, corner_radius=0)
        ebar.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, SPACE_XS))

        self.lbl_arr_event_title = ctk.CTkLabel(
            ebar, text="← Click on an episode",
            font=FONT_SIDEBAR_HDR, text_color=MUTED, anchor="w")
        self.lbl_arr_event_title.pack(side="left", padx=SPACE_M, pady=SPACE_S)  # type: ignore[union-attr]

        # Separator
        ctk.CTkFrame(ebar, width=1, fg_color=BORDER).pack(
            side="left", fill="y", padx=(SPACE_S, SPACE_S), pady=SPACE_S)

        self.btn_arr_edit = ctk.CTkButton(
            ebar, text="Edit Peaks", width=96, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            command=self._toggle_arr_edit_mode,
        )
        self.btn_arr_edit.pack(side="left", padx=SPACE_XS)

        self.btn_arr_undo = ctk.CTkButton(
            ebar, text="↩ Undo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._undo_edit,
        )
        self.btn_arr_undo.pack(side="left", padx=SPACE_XS)
        self.btn_arr_redo = ctk.CTkButton(
            ebar, text="↪ Redo", width=72, height=28, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            corner_radius=8,
            state="disabled", command=self._redo_edit,
        )
        self.btn_arr_redo.pack(side="left", padx=SPACE_XS)

        self.lbl_arr_edit_hint = ctk.CTkLabel(
            ebar, text="L-click: exclude/restore   R-click: add",
            font=FONT_HINT, text_color=ORANGE,
        )
        # Nav: ← → buttons + window size
        self.btn_arr_prev = ctk.CTkButton(
            ebar, text="◀", width=32, height=26, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            command=lambda: self._arr_navigate(-1),
        )
        self.btn_arr_prev.pack(side="right", padx=SPACE_XS)
        self.btn_arr_next = ctk.CTkButton(
            ebar, text="▶", width=32, height=26, font=FONT_SMALL,
            fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
            command=lambda: self._arr_navigate(+1),
        )
        self.btn_arr_next.pack(side="right", padx=SPACE_XS)
        ctk.CTkLabel(ebar, text="Window (s):", font=FONT_SMALL,
                     text_color=MUTED).pack(side="right", padx=(0, SPACE_XS))
        self.ent_arr_win = ctk.CTkEntry(
            ebar, width=48, height=26, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, text_color=TEXT)
        self.ent_arr_win.insert(0, "3")
        self.ent_arr_win.pack(side="right", padx=(0, SPACE_S))

        # ECG canvas
        ecg_frame = tk.Frame(right, bg=PLOT["bg"], bd=0, highlightthickness=0)
        ecg_frame.grid(row=1, column=0, sticky="nsew")
        self._slots["arr_detail"] = CanvasSlot(ecg_frame, 14, 5.0, toolbar=False)

        # Wire up click handler on the canvas
        def _wire_arr_click(slot=self._slots["arr_detail"]):
            slot.canvas.mpl_connect("button_press_event", self._on_arr_detail_click)
            slot.canvas.mpl_connect("scroll_event",       self._on_arr_scroll)
        self.after(200, _wire_arr_click)

    def _run_arrhythmia_analysis(self) -> None:
        self.analysis_ctrl.run_arrhythmia_analysis()

    # ── Event card builder ────────────────────────────────────

    def _build_arrhythmia_card(
        self, idx: int, ev: "ArrhythmiaEvent",
        sev_colors: dict, kind_icons: dict,
    ) -> None:
        self.analysis_ctrl.build_arrhythmia_card(idx, ev, sev_colors, kind_icons)

    # ── Event selection & ECG viewer ─────────────────────────

    def _select_arrhythmia_event(self, idx: int) -> None:
        """Highlight selected card and load the ECG window for this event."""
        self.nav_ctrl.select_arrhythmia_event(idx)

    def _draw_arr_detail(self) -> None:
        """Draw ECG strip for the selected arrhythmia event, with editable R peaks."""
        self.plot_ctrl.draw_arr_detail()
    # ── Edit mode toggle for arrhythmia tab ──────────────────

    def _toggle_arr_edit_mode(self) -> None:
        self.analysis_ctrl.toggle_arr_edit_mode()

    # ── Click handler (mirrors _on_detail_click) ─────────────

    def _on_arr_detail_click(self, event) -> None:
        """Left-click: toggle exclusion.  Right-click: add/remove peak."""
        self.analysis_ctrl.on_arr_detail_click(event)

    # ── Scroll zoom on arrhythmia ECG ────────────────────────

    def _on_arr_scroll(self, event) -> None:
        self.analysis_ctrl.on_arr_scroll(event)

    # ── ◀ ▶ navigation ───────────────────────────────────────

    def _arr_navigate(self, direction: int) -> None:
        self.nav_ctrl.arr_navigate(direction)

    # ── sync undo/redo buttons in arrhythmia tab ─────────────

    def _update_undo_btns(self) -> None:
        """Update all undo/redo button instances (Detection + Abnormal Events tabs)."""
        self.detection_ctrl.update_undo_btns()

    def _copy_arrhythmia_tsv(self) -> None:
        self.analysis_ctrl.copy_arrhythmia_tsv()

    def _build_tab_intervals(self) -> None:
        """Interval delineation tab: P/Q/R/S/T measurement per beat."""
        t = self.tabs.tab("📏 Intervals")

        # ── Action bar ─────────────────────────────────────────────────────
        bar = ctk.CTkFrame(t, fg_color="transparent")
        bar.pack(side="top", fill="x", padx=SPACE_M, pady=(SPACE_M, SPACE_S))

        self.btn_run_ivl = ctk.CTkButton(
            bar, text="⚡  Delineate waves",
            command=self._run_intervals,
            fg_color=ORANGE, hover_color=ORANGE_DARK, text_color="white",
            font=FONT_BTN_PRIMARY, height=max(28, int(32 * THEME.font_scale)),
            corner_radius=8, state="disabled",
        )
        self.btn_run_ivl.pack(side="left")

        self.lbl_ivl_status = ctk.CTkLabel(
            bar, text="  Run Core Analysis first",
            font=FONT_SMALL, text_color=MUTED, anchor="w")
        self.lbl_ivl_status.pack(side="left", padx=SPACE_M)

        # QTc formula selector
        ctk.CTkLabel(bar, text="QTc formula:", font=FONT_SMALL,
                     text_color=MUTED).pack(side="right", padx=(0, SPACE_S))
        self.cb_qtc_formula = ctk.CTkComboBox(
            bar, width=140, height=28, font=FONT_LABEL,
            fg_color=BG, border_color=BORDER2, button_color=BORDER2,
            text_color=TEXT, dropdown_fg_color=BG, dropdown_text_color=TEXT,
            values=["Mitchell (∛RR)", "Bazett (√RR)", "Hodges (linear HR)"],
            command=self._on_qtc_formula_change,
        )
        self.cb_qtc_formula.set("Mitchell (∛RR)")
        self.cb_qtc_formula.pack(side="right", padx=(0, SPACE_M))

        # Permissive bounds toggle
        self.sw_permissive = self._switch(bar, "Permissive bounds", dict(padx=0))
        self.sw_permissive.pack(side="right", padx=(0, SPACE_L))

        # ── Interval verifier nav bar (populated by _launch_interval_verifier) ──
        self.frm_ivl_nav = tk.Frame(t, bg=PANEL, bd=0, highlightthickness=0)
        self.frm_ivl_nav.pack(side="top", fill="x")

        # ── Body: violin plots (left) + annotated beat strip (right) ─────────
        body = tk.Frame(t, bg=BG, bd=0, highlightthickness=0)
        body.pack(side="top", fill="both", expand=True, padx=SPACE_S, pady=(0, SPACE_S))

        # Use a PanedWindow for resizable split
        paned = tk.PanedWindow(body, orient=tk.HORIZONTAL,
                               bg=BORDER, sashwidth=4, sashrelief="flat",
                               handlesize=0)
        paned.pack(fill="both", expand=True)

        left_inner = tk.Frame(paned, bg=PANEL, bd=0, highlightthickness=0)
        right_inner = tk.Frame(paned, bg=PANEL, bd=0, highlightthickness=0)
        paned.add(left_inner, minsize=180, stretch="always")
        paned.add(right_inner, minsize=300, stretch="always")
        # Set initial split after widget is mapped
        def _set_sash(event=None):
            try:
                paned.sash_place(0, max(200, paned.winfo_width() // 3), 0)
            except Exception:
                pass
        paned.bind("<Map>", _set_sash)

        self._slots["intervals"] = CanvasSlot(left_inner, 6, 6, toolbar=False)
        self._slots["intervals_ecg"] = CanvasSlot(right_inner, 11, 6, toolbar=False)


    def _build_tab_beat_template(self) -> None:
        """Beat template tab: mean beat ± SD and morphology distributions."""
        t = self.tabs.tab("〰 Beat Template")
        t.grid_rowconfigure(1, weight=3)
        t.grid_rowconfigure(2, weight=2)
        t.grid_columnconfigure(0, weight=1)

        # ── Header bar ──────────────────────────────────────────────────────
        bar = ctk.CTkFrame(t, fg_color="transparent")
        bar.grid(row=0, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_M, SPACE_S))
        ctk.CTkLabel(bar, text="BEAT TEMPLATE  —  mean ± SD  ·  morphology",
                     font=FONT_SIDEBAR_HDR, text_color=MUTED).pack(side="left")
        ctk.CTkButton(
            bar, text="⚙  Landmarks",
            command=self._open_wave_template_editor,
            fg_color=PURPLE, hover_color=PURPLE_DARK, text_color="white",
            font=FONT_BTN_SEC, height=28, corner_radius=6,
        ).pack(side="right")
        self.lbl_template_info = ctk.CTkLabel(
            bar, text="", font=FONT_HINT, text_color=MUTED, anchor="e")
        self.lbl_template_info.pack(side="right", padx=(0, SPACE_M))  # type: ignore[union-attr]

        # ── Mean beat plot ──────────────────────────────────────────────────
        beat_card = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0)
        beat_card.grid(row=1, column=0, sticky="nsew", padx=SPACE_S, pady=(0, SPACE_S))
        beat_card.grid_rowconfigure(0, weight=1)
        beat_card.grid_columnconfigure(0, weight=1)
        self._slots["beat"] = CanvasSlot(beat_card, 14, 5.0, toolbar=False)

        # ── Distributions: amplitude + correlation ──────────────────────────
        dist_card = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0)
        dist_card.grid(row=2, column=0, sticky="nsew", padx=SPACE_S, pady=(0, SPACE_S))
        dist_card.grid_rowconfigure(0, weight=1)
        dist_card.grid_columnconfigure(0, weight=1)
        self._slots["beat_dist"] = CanvasSlot(dist_card, 14, 4.0, toolbar=False)


    # ════════════════════════════════════════════════════════
    #  QTC FORMULA
    # ════════════════════════════════════════════════════════

    def _qtc_formula(self) -> str:
        """Return 'mitchell' or 'bazett' from the Intervals tab combo selector."""
        return self.analysis_ctrl.qtc_formula()

    def _on_qtc_formula_change(self, _choice: str = "") -> None:
        """Re-compute QTc with the selected formula and refresh interval plots."""
        self.analysis_ctrl.on_qtc_formula_change(_choice)

    def _update_interval_plots(self) -> None:
        """Redraw the interval violin plots after a QTc formula change."""
        self.analysis_ctrl.update_interval_plots()

    # ════════════════════════════════════════════════════════
    #  ROLLING HRV
    # ════════════════════════════════════════════════════════

    def _compute_rolling_hrv(self) -> None:
        """Compute sliding-window HRV and render the timeline plot."""
        self.analysis_ctrl.compute_rolling_hrv()

    def _build_tab_summary(self) -> None:
        """Summary tab: a curated verdict, not a wall of duplicate plots.

        Every plot mirrored here used to also exist full-size in its own tab
        (HRV > Frequency/Non-linear/Rolling, Intervals, Beat Template) --
        showing it again at ~1/3 the size with the same font/legend/tick
        density as the full-tab version just looked cramped. The global KPI
        bar above the tabs already shows the headline numbers, so this tab
        no longer duplicates them either. What's left: a Signal Quality
        panel (surfacing numbers the app already computed but never showed
        anywhere -- mean beat-to-template correlation, % of low-quality
        beats, artifact-correction counts), the two densest kept visuals
        (RR tachogram, Poincare), the RR-asymmetry breakdown (unique to this
        tab, not a duplicate), a plain metrics table with no physiological
        reference-range judgments, and the detailed text report.
        """
        t = self.tabs.tab("📋 Summary")
        t.grid_rowconfigure(0, weight=0)   # verdict banner
        t.grid_rowconfigure(1, weight=0)   # action bar
        t.grid_rowconfigure(2, weight=1)   # scrollable body
        t.grid_columnconfigure(0, weight=1)

        # ── Verdict banner ───────────────────────────────────────────────────
        banner = ctk.CTkFrame(t, fg_color=PANEL, corner_radius=0, height=44)
        banner.grid(row=0, column=0, sticky="ew")
        banner.pack_propagate(False)
        self.lbl_sum_verdict = ctk.CTkLabel(
            banner, text="Run analysis to see the signal-quality summary.",
            font=FONT_KPI_LABEL, text_color=MUTED, anchor="w")
        self.lbl_sum_verdict.pack(side="left", padx=SPACE_L, pady=SPACE_S)
        ctk.CTkFrame(t, height=1, fg_color=BORDER).grid(row=0, column=0, sticky="sew")

        # ── Action bar ────────────────────────────────────────────────────────
        btn_row = ctk.CTkFrame(t, fg_color="transparent", height=40)
        btn_row.grid(row=1, column=0, sticky="ew", padx=SPACE_M, pady=(SPACE_S, 0))
        btn_row.pack_propagate(False)
        for label, cmd in [
            ("📋  Copy Report",      self._copy_summary),
            ("💾  Save .txt",        self._save_summary_txt),
        ]:
            ctk.CTkButton(btn_row, text=label, command=cmd,
                          fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                          font=FONT_SMALL, height=28, corner_radius=5).pack(
                side="left", padx=(0, SPACE_S))
        self.lbl_epoch_info = ctk.CTkLabel(btn_row, text="", font=FONT_SMALL,
                                            text_color=MUTED)
        self.lbl_epoch_info.pack(side="right")

        # ── Scrollable body ───────────────────────────────────────────────────
        outer_scroll = ctk.CTkScrollableFrame(
            t, fg_color=PANEL,
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=BORDER2,
        )
        outer_scroll.grid(row=2, column=0, sticky="nsew", padx=0, pady=(SPACE_S, 0))
        outer_scroll.grid_columnconfigure(0, weight=1)
        self._sum_scroll = outer_scroll

        # ── Layout helpers ────────────────────────────────────────────────────
        def _section(parent, title: str, color: str = MUTED,
                     subtitle: str = "") -> ctk.CTkFrame:
            """Titled section with accent stripe, returns the content frame."""
            hdr = ctk.CTkFrame(parent, fg_color="transparent")
            hdr.pack(fill="x", padx=SPACE_M, pady=(SPACE_L, SPACE_S))
            stripe = ctk.CTkFrame(hdr, height=3, fg_color=color, corner_radius=2)
            stripe.pack(fill="x", pady=(0, SPACE_S))
            title_row = ctk.CTkFrame(hdr, fg_color="transparent")
            title_row.pack(fill="x")
            ctk.CTkLabel(title_row, text=title.upper(),
                         font=FONT_SIDEBAR_HDR,
                         text_color=color, anchor="w").pack(side="left")
            if subtitle:
                ctk.CTkLabel(title_row, text=f"  {subtitle}",
                             font=FONT_KPI_LABEL, text_color=MUTED,
                             anchor="w").pack(side="left", padx=(SPACE_S, 0))
            body = ctk.CTkFrame(parent, fg_color="transparent")
            body.pack(fill="x", padx=SPACE_S, pady=(SPACE_XS, 0))
            return body

        def _row(parent, specs: "list[tuple[str,int,int]]") -> None:
            """specs = [(key, height_px, weight), ...]  — one horizontal row."""
            row_frame = ctk.CTkFrame(parent, fg_color="transparent")
            row_frame.pack(fill="x", pady=(0, SPACE_S))
            for i, (key, h, w) in enumerate(specs):
                pad_right = i < len(specs) - 1
                card = ctk.CTkFrame(row_frame, fg_color=PANEL, corner_radius=6, height=h)
                px = (0, 4) if pad_right else (0, 0)
                card.pack(side="left", fill="both", expand=True, padx=px)
                card.pack_propagate(False)
                inner = tk.Frame(card, bg=PLOT["bg"], bd=0, highlightthickness=0)
                inner.pack(fill="both", expand=True, padx=SPACE_XS, pady=SPACE_XS)
                self._slots[key] = CanvasSlot(inner, 10 * w, h / 100, toolbar=False)

        # ━━━ SECTION 1 — Signal Quality ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec_q = _section(outer_scroll, "✓  Signal Quality", GREEN_DARK,
                         subtitle="how much to trust this recording, not the physiology itself")
        stat_row = ctk.CTkFrame(sec_q, fg_color=CARD, corner_radius=6)
        stat_row.pack(fill="x", pady=(0, SPACE_S))
        stat_grid = ctk.CTkFrame(stat_row, fg_color="transparent")
        stat_grid.pack(fill="x", padx=SPACE_L, pady=SPACE_M)
        for col in range(5):
            stat_grid.grid_columnconfigure(col, weight=1, uniform="sq")
        _SQ_DEFS = [
            ("sq_score",   "Overall score"),
            ("sq_corr",    "Mean template correlation"),
            ("sq_badbeats","Beats below 0.90 corr."),
            ("sq_noisy_time", "Time below 0.90 corr."),
            ("sq_artifact","Auto-corrected"),
        ]
        self._sum_quality_vals: "dict[str, ctk.CTkLabel]" = {}
        for i, (key, label) in enumerate(_SQ_DEFS):
            cell = ctk.CTkFrame(stat_grid, fg_color="transparent")
            cell.grid(row=0, column=i, sticky="ew", padx=(0 if i == 0 else SPACE_M, 0))
            ctk.CTkLabel(cell, text=label, font=FONT_KPI_LABEL,
                         text_color=MUTED, anchor="w").pack(anchor="w")
            val_lbl = ctk.CTkLabel(cell, text="—", font=FONT_KPI_VALUE,
                                   text_color=TEXT, anchor="w")
            val_lbl.pack(anchor="w")
            self._sum_quality_vals[key] = val_lbl
        _row(sec_q, [("sum_quality_time", 180, 1)])

        # ━━━ SECTION 2 — Rythme cardiaque ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec1 = _section(outer_scroll, "❤  Heart Rate & Rhythm", ORANGE_DARK,
                        subtitle="densest single views — full plots live in their own tabs")
        _row(sec1, [("sum_rr", 260, 3), ("sum_poincare", 260, 2)])
        _row(sec1, [("sum_asymmetry", 220, 1)])

        # ━━━ SECTION 3 — Metrics table ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec2 = _section(outer_scroll, "▤  Metrics", BLUE_DARK,
                        subtitle="values only — no reference-range judgment")
        metrics_card = ctk.CTkFrame(sec2, fg_color=CARD, corner_radius=6)
        metrics_card.pack(fill="x", pady=(0, SPACE_S))
        metrics_body = ctk.CTkFrame(metrics_card, fg_color="transparent")
        metrics_body.pack(fill="x", padx=SPACE_L, pady=SPACE_M)

        _METRIC_GROUPS = [
            ("Rate",                  ORANGE_DARK, [("hr_mean", "HR mean", "bpm"),
                                                     ("hr_range", "HR range", "bpm")]),
            ("Time-domain HRV",       BLUE_DARK,   [("sdnn", "SDNN", "ms"),
                                                     ("rmssd", "RMSSD", "ms"),
                                                     ("pnn6", "pNN6", "%")]),
            ("Frequency / non-linear", "#37474F",  [("lf_hf", "LF/HF", ""),
                                                     ("sampen", "SampEn", ""),
                                                     ("dfa1", "DFA α1", "")]),
            ("Intervals (delineation)", PINK,      [("pr", "PR", "ms"),
                                                     ("qrs", "QRS", "ms"),
                                                     ("qtc", "QTc", "ms")]),
        ]
        self._sum_metric_vals: "dict[str, ctk.CTkLabel]" = {}
        for gi, (gtitle, gcolor, rows) in enumerate(_METRIC_GROUPS):
            ctk.CTkLabel(metrics_body, text=gtitle.upper(), font=FONT_SUBSECTION,
                         text_color=gcolor, anchor="w").pack(
                anchor="w", pady=(SPACE_S if gi else 0, SPACE_XS))
            for key, name, unit in rows:
                r = ctk.CTkFrame(metrics_body, fg_color="transparent")
                r.pack(fill="x")
                ctk.CTkLabel(r, text=name, font=FONT_LABEL, text_color=TEXT,
                             anchor="w", width=160).pack(side="left")
                val_lbl = ctk.CTkLabel(r, text="—", font=FONT_LABEL,
                                       text_color=TEXT, anchor="e", width=80)
                val_lbl.pack(side="left")
                if unit:
                    ctk.CTkLabel(r, text=unit, font=FONT_MICRO, text_color=MUTED,
                                 anchor="w").pack(side="left", padx=(SPACE_XS, 0))
                self._sum_metric_vals[key] = val_lbl

        linkout = ctk.CTkFrame(sec2, fg_color=CARD, corner_radius=6)
        linkout.pack(fill="x", pady=(0, SPACE_S))
        ctk.CTkLabel(linkout, text="Full plots for every metric above live in their own tabs.",
                     font=FONT_SMALL, text_color=MUTED, anchor="w").pack(
            side="left", padx=SPACE_M, pady=SPACE_S)
        ctk.CTkLabel(linkout, text="HRV · Intervals · Beat Template · Abnormal Events  →",
                     font=FONT_SMALL, text_color=BLUE, anchor="e").pack(
            side="right", padx=SPACE_M, pady=SPACE_S)

        # ━━━ SECTION 4 — Rapport texte ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        sec3 = _section(outer_scroll, "📝  Detailed Report", MUTED)
        txt_card = ctk.CTkFrame(sec3, fg_color=PANEL, corner_radius=6)
        txt_card.pack(fill="x", pady=(0, SPACE_L))
        self.txt_sum = ctk.CTkTextbox(txt_card, font=FONT_MONO, fg_color=PANEL,
                                      text_color=TEXT, border_width=0, height=380)
        self.txt_sum.pack(fill="both", expand=True, padx=SPACE_M, pady=SPACE_M)

    # ════════════════════════════════════════════════════════
    #  SIGNAL PIPELINE  (detection engine)
    # ════════════════════════════════════════════════════════

    def _snapshot_params(self) -> dict:
        """Read every widget value on the main thread and return a plain dict snapshot.

        Delegates to ``FilterParams.from_widgets`` so that the field list is
        maintained in exactly one place (the FilterParams dataclass).  Any new
        parameter added to FilterParams is automatically picked up here and in
        _collect_session_state / _restore_session_worker.
        """
        return self.session_ctrl.snapshot_params()

    def _compute_preview_bundle(
        self,
        sig_raw:     np.ndarray,
        fs:          int,
        params:      dict,
        progress_cb: "Optional[Callable[[int, str], None]]" = None,
    ) -> dict:
        """Pure computation — no access to *self*, no Tkinter calls.

        Filters, normalises, fixes polarity, and finds all R-peak candidates.
        Returns a plain dict that ``_on_preview_done`` writes to ``self.*``
        atomically on the main thread.

        This separation is the thread-safety contract:
        * Background workers call this method and return the bundle.
        * Only the main thread writes instance variables.
        """
        return self.signal_ctrl.compute_preview_bundle(sig_raw, fs, params, progress_cb)

    def _prepare_signal(
        self,
        params: dict,
        progress_cb: "Optional[Callable[[int, str], None]]" = None,
    ) -> None:
        """Filter, normalise, and fix polarity. Writes results to self.

        Thin wrapper around the pure ``_compute_preview_bundle`` for callers
        that already have signal data on self and are running on the main
        thread. Not currently called anywhere in the app — session restore
        now calls ``_compute_preview_bundle`` directly from
        ``_restore_session_worker`` (a background thread) and writes the
        result to self from ``_on_restore_session_done`` (main thread) —
        kept for any future main-thread caller that wants the write-to-self
        convenience.

        The background preview path uses ``_compute_preview_bundle`` directly
        and writes to self only from ``_on_preview_done`` (main thread).
        """
        self.signal_ctrl.prepare_signal(params, progress_cb)

    def _run_detection(self, thresh: float | None = None) -> int:
        """Apply current threshold to pre-computed candidates.

        Parameters
        ----------
        thresh : float | None
            If supplied, use this value directly (safe from a background
            thread).  If None, read ``self.sl_thr`` — only safe to call
            from the main thread.

        Returns the number of accepted peaks.  Fast — no signal processing.
        All Tkinter widget *writes* are marshalled through ``after(0, …)``
        so this method is safe to call from either thread.
        """
        return self.detection_ctrl.run_detection(thresh)

    def _update_signal_quality(self, accepted: np.ndarray) -> None:
        """Compute a 0–100 quality score and update the KPI label.

        Quality is based on:
        1. Beat morphology (primary): mean beat-to-template correlation from
           the last analysis run.  High correlation = clean, consistent QRS.
           Falls back to RR regularity if no template analysis has been run yet.
        2. Detection completeness (secondary): ratio of detected to expected
           beats, clipped to [0.5, 1.5] so it modulates but never dominates.

        The previous formula (1 - rr_cv) was unreliable because a healthy mouse
        at high HR during stress has a low rr_cv not from noise but from genuine
        sympathetic activation.
        """
        self.detection_ctrl.update_signal_quality(accepted)

    def _on_project_name_change(self, event=None) -> None:
        """Mirror the sidebar 'Project name' entry into the top bar.

        Deliberately independent of update_kpis() (which only repaints once
        analysis results exist) -- typing a project name should show up in
        the top bar immediately, not only after running analysis. Also
        called from _rebuild_ui() after _restore_ui_state() so a typed value
        survives a theme toggle's destroy/rebuild cycle.
        """
        if self.ent_project_name is None or self.lbl_topbar_project is None:
            return
        name = self.ent_project_name.get().strip()  # type: ignore[union-attr]
        self.lbl_topbar_project.configure(text=name if name else "—")  # type: ignore[union-attr]

    def _on_det_method_change(self, choice: str) -> None:
        """Show/hide SG options frame based on selected detection method."""
        self.detection_ctrl.on_det_method_change(choice)

    # ── ML detector: verified-for-training switch ────────────
    def _on_verified_training_toggle(self) -> None:
        self.session_ctrl.on_verified_training_toggle()

    def _save_for_training_only(self) -> None:
        self.session_ctrl.save_for_training_only()

    def _open_ml_training_dialog(self) -> None:
        """Open the Train/Retrain dialog for the ML R-peak detector."""
        from ecg.ui.dialogs import MLTrainingDialog
        dlg = MLTrainingDialog(self)
        self.wait_window(dlg)
        self.refresh_ml_status()

    def refresh_ml_status(self) -> None:
        """Update the ML-detector sidebar status label.

        Shows verified-file/sample counts and, if a model is trained, its
        hold-out accuracy/F1 from the metadata sidecar. Called on startup,
        after Save Session, and after the training dialog closes.
        """
        if self.lbl_ml_status is None:
            return
        from ecg.core.ml_detector import training_data_summary, MLPeakModel
        summary = training_data_summary()
        model = MLPeakModel.load()
        if model is not None:
            meta = model.meta
            text = (f"Model trained on {meta.get('n_training_files', '?')} file(s), "
                    f"{meta.get('n_training_samples', '?')} samples — "
                    f"acc={meta.get('holdout_accuracy', 0):.2f} "
                    f"f1={meta.get('holdout_f1', 0):.2f}\n"
                    f"{summary['n_files']} file(s) currently verified.")
            color = GREEN
        else:
            text = (f"No trained model yet. {summary['n_files']} file(s) verified "
                    f"({summary['n_samples']} samples).")
            color = MUTED
        self.lbl_ml_status.configure(text=text, text_color=color)  # type: ignore[union-attr]

    # ── Filtering master toggle ───────────────────────────────

    def _on_filtering_toggle(self) -> None:
        """Grey out the notch/band-pass/cleaning group when Filtering is off."""
        self.detection_ctrl.on_filtering_toggle()

    _CLEAN_METHOD_LABELS = {
        "neurokit":        "NeuroKit",
        "pantompkins1985": "Pan-Tompkins",
        "elgendi2010":     "Elgendi",
        "hamilton2002":    "Hamilton",
        "biosppy":         "BioSPPy",
    }

    def _update_filter_summary(self) -> None:
        """Refresh the read-only 'Processing' summary from live widget values.

        Mirrors exactly what compute_preview_bundle() will actually do to
        the signal -- so it only reflects FILTER SETTINGS (band-pass/notch/
        cleaning), not the DISPLAY & PREVIEW group (show raw/invert/preview),
        which don't change what gets analysed.
        """
        if self.lbl_filter_summary is None:
            return
        if self.sw_filtering is None or not bool(self.sw_filtering.get()):
            lines = ["Processing", "•  Raw signal"]
        else:
            lp = self._safe_float(self.ent_lp, MouseECG.BP_LO_HZ)
            hp = self._safe_float(self.ent_hp, MouseECG.BP_HI_HZ)
            lines = ["Processing", f"•  Band-pass: {lp:g}–{hp:g} Hz"]
            if self.sw_notch is not None and bool(self.sw_notch.get()):
                lines.append("•  Notch: 50 Hz")
            clean_val = self.cb_clean.get() if self.cb_clean is not None else "neurokit"
            clean_label = self._CLEAN_METHOD_LABELS.get(clean_val, clean_val.title())
            lines.append(f"•  {clean_label} cleaning")
        self.lbl_filter_summary.configure(text="\n".join(lines))

    # ── Raw / Filtered toggle ─────────────────────────────────

    def _on_show_raw_toggle(self) -> None:
        """Switch the overview and detail plots between raw and filtered signals.

        The raw signal is normalised (zero-mean, unit-variance) to match
        the amplitude scale of the filtered signal so that peak markers
        remain visually coherent regardless of which view is active.
        No re-processing is needed — both arrays are pre-computed.
        """
        self.detection_ctrl.on_show_raw_toggle()

    # ── Overview click-to-navigate ────────────────────────────

    def _on_overview_click(self, event) -> None:
        """Stub — overview removed."""
        self.nav_ctrl.on_overview_click(event)

    def _on_overview_scroll(self, event) -> None:
        """Stub — overview removed."""
        self.nav_ctrl.on_overview_scroll(event)

    def _on_overview_motion(self, event) -> None:
        """Drag-to-scrub while the mouse button is held on the minimap strip."""
        self.nav_ctrl.on_overview_motion(event)

    def _on_overview_release(self, event) -> None:
        """End minimap drag-to-scrub."""
        self.nav_ctrl.on_overview_release(event)

    # ── Detail scroll-wheel zoom ──────────────────────────────

    def _on_detail_scroll(self, event) -> None:
        """Zoom the detail view's x-axis in/out centred on the cursor position.

        Each scroll tick zooms by a factor of 1.25 (in) or 0.8 (out).
        After zooming, ``_nav_pos`` and the window-entry widget are updated
        to reflect the new visible range.

        The y-axis is intentionally unchanged — vertical zoom is handled by
        the matplotlib toolbar's Zoom-to-rectangle tool.
        """
        self.nav_ctrl.on_detail_scroll(event)

    # ── Undo / Redo for manual peak edits ────────────────────

    def _push_edit_undo(self) -> None:
        """Snapshot state before a destructive edit action."""
        self.detection_ctrl.push_edit_undo()

    def _undo_edit(self, _event=None) -> None:
        """Ctrl+Z — restore previous peak-edit state."""
        self.detection_ctrl.undo_edit(_event)

    def _redo_edit(self, _event=None) -> None:
        """Ctrl+Y — rétablir après undo."""
        self.detection_ctrl.redo_edit(_event)

    def _apply_edit_state(self) -> None:
        self.detection_ctrl.apply_edit_state()

    # ── Manual peak exclusion ─────────────────────────────────

    def _toggle_edit_mode(self) -> None:
        """Toggle the click-to-exclude edit mode on/off."""
        self.detection_ctrl.toggle_edit_mode()

    def _toggle_free_placement(self) -> None:
        """Toggle free-placement mode: bypass proximity constraint when adding peaks.

        When active, right-clicking adds a peak at the local max *regardless* of
        how close it is to an existing peak.  This is useful for very high-rate
        signals or for correcting closely-spaced double-peaks.

        Note: edit mode must be active for this to have any effect.
        """
        self.detection_ctrl.toggle_free_placement()

    def _clear_manual_exclusions(self) -> None:
        """Re-include all manually excluded peaks, remove all manually added peaks."""
        self.detection_ctrl.clear_manual_exclusions()

    def _on_detail_motion(self, event) -> None:
        """Track mouse position in edit mode and compute the preview peak position.

        In normal mode: snaps to the local maximum within ±tol_samp and shows
        an orange/red marker if it would land too close to an existing peak.
        In free placement mode: the preview follows the cursor exactly (no
        snapping) and is always shown in the "ok" colour.

        Redraws are throttled to 30 ms (≈33 fps) via after().
        """
        self.detection_ctrl.on_detail_motion(event)

    def _flush_hover_redraw(self) -> None:
        """Execute the throttled hover redraw on the main thread."""
        self.detection_ctrl.flush_hover_redraw()

    def _on_detail_click(self, event) -> None:
        """Edit-mode click handler for the detail view.

        Left-click  (button 1) near an existing peak → toggle exclusion
        Right-click (button 3) anywhere              → add peak at local max,
                                                       or remove if clicking
                                                       a manually-added peak

        Only active when ``_edit_mode`` is True.
        """
        self.detection_ctrl.on_detail_click(event)

    # ── Threshold slider / entry callbacks ────────────────────
    def _on_threshold_slide(self, value: float) -> None:
        """Called continuously while the slider is being dragged.

        Widget label and entry are updated immediately for visual feedback.
        Detection and redraws are debounced (80 ms) so rapid drag events do
        not flood the rendering pipeline — especially important for long
        recordings where apply_threshold() + two canvas draws take ~50 ms.
        """
        self.detection_ctrl.on_threshold_slide(value)

    def _apply_threshold_ui(self, value: float) -> None:
        """Run detection and refresh plots — called after debounce delay.

        Always executes on the main thread (scheduled via after()), so it is
        safe to read the slider and write widgets directly.
        """
        self.detection_ctrl.apply_threshold_ui(value)

    def _on_threshold_entry(self, event=None) -> None:
        """Called when the user types a value in the exact-threshold entry.

        Applies immediately — no debounce — since this is a deliberate commit.
        """
        self.detection_ctrl.on_threshold_entry(event)

    # ════════════════════════════════════════════════════════
    #  ACTIONS  (file, preview, run)
    # ════════════════════════════════════════════════════════

    def _open_file(self) -> None:
        path = filedialog.askopenfilename(
            title="Open Spike2 .mat",
            filetypes=[("MATLAB files", "*.mat"), ("All", "*.*")],
        )
        if path:
            self._load_path(path)

    def _show_channels(self) -> None:
        if not self._filepath:
            messagebox.showwarning("No file", "Open a .mat file first.")
            return
        try:
            messagebox.showinfo("Channels", list_channels(self._filepath))
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def _load_raw_only(self) -> None:
        """Load the file and display ONLY the raw signal.

        No bandpass/notch/NK-clean, no polarity correction, no detection —
        just the samples read from disk, windowed by the Analysis window
        fields if set, and z-score normalised purely for display scale.

        This is what runs automatically when a file is opened.  The user
        must click '1 ▶ Preview Detection' to run the actual DSP + detection
        pipeline (see _preview / _preview_worker).
        """
        self.signal_ctrl.load_raw_only()

    def _load_raw_worker(self, params: dict) -> dict:
        """Background worker — loads the file, returns the RAW signal only.

        Deliberately mirrors only the first few steps of _preview_worker
        (file read + time-window crop). It stops before any filtering,
        polarity correction, or detection call.
        """
        return self.signal_ctrl.load_raw_worker(params)

    def _on_raw_load_done(self, bundle: dict) -> None:
        """Write raw-only state on the main thread and draw the raw trace.

        Mirrors the bookkeeping parts of _on_preview_done (channel/fs
        feedback, KPI reset, status message) but explicitly leaves every
        detection/filtering field at None — signal_flt, all_cands,
        rpeaks_ok, thresh_amp, etc. — so the rest of the app's existing
        "not previewed yet" guards (which already check for None) behave
        correctly until the user clicks Preview Detection.
        """
        self.signal_ctrl.on_raw_load_done(bundle)

    def _preview(self) -> None:
        """Load, filter, and detect peaks — fast, no HRV."""
        self.signal_ctrl.preview()

    def _preview_worker(self, params: dict) -> dict:
        """Background worker — MUST NOT write to self.

        Loads and processes the signal, then returns a plain data bundle.
        All instance-variable writes happen in ``_on_preview_done`` on the
        main thread, preventing data races with Tk resize/redraw callbacks.
        """
        return self.signal_ctrl.preview_worker(params)

    def _apply_detected_fs(self, fs: float) -> None:
        """Update the fs entry and source label on the main thread."""
        self.signal_ctrl.apply_detected_fs(fs)

    def _on_preview_done(self, bundle: dict) -> None:
        """Atomically write all signal state on the main thread, then draw.

        This is the ONLY place that should assign signal/peak instance variables
        after a preview.  Because it runs via after(0, …) (scheduled by
        _start_async after the background worker finishes), it is guaranteed to
        execute on the Tk main thread with no concurrent background writes.
        """
        self.signal_ctrl.on_preview_done(bundle)

    def _windowed_peaks(self) -> "Optional[np.ndarray]":
        """Return a copy of _rpeaks_ok filtered to the current analysis window.

        If no window is set (both bounds = 0), returns the full array.
        Returns None if _rpeaks_ok is None.

        This is the single source of truth for all analysis methods
        (_run_freq, _run_nonlinear, _run_intervals, _run_arrhythmia_analysis,
        _compute_epochs, _compute_rolling_hrv) — they all call this instead of
        doing ``self._rpeaks_ok.copy()`` directly.
        """
        return self.signal_ctrl.windowed_peaks()

    def _apply_analysis_window(self) -> None:
        """Read the analysis window entries and store in _analysis_t_start/_end.

        Updates the feedback label with the peak count inside the window.
        Does NOT re-run detection or analysis — the window is applied on
        the next Core Analysis run.
        """
        self.signal_ctrl.apply_analysis_window()

    def _reset_analysis_window(self) -> None:
        """Reset analysis window to full signal."""
        self.signal_ctrl.reset_analysis_window()

    def _run_analysis(self) -> None:
        self.analysis_ctrl.run_analysis()

    def _analysis_worker(self, params: dict) -> dict:
        """Background worker — MUST NOT write to self.

        Runs core analysis and optional artifact correction, then returns a
        plain bundle.  ``_on_analysis_done`` writes all results to self on
        the main thread, preventing races with Tk draw callbacks.
        """
        return self.analysis_ctrl.analysis_worker(params)

    def _on_analysis_done(self, bundle: dict) -> None:
        self.analysis_ctrl.on_analysis_done(bundle)


    # ── Per-module on-demand analysis ────────────────────────

    def _open_artifact_review(self) -> None:
        """Detect artifact candidates and open the interactive review dialog."""
        if self._signal_flt is None or self._rpeaks_ok is None or len(self._rpeaks_ok) < 4:
            messagebox.showwarning("Not ready",
                                   "Run Preview Detection first to load peaks.")
            return

        rp  = self._rpeaks_ok.copy()
        fs  = self._fs
        sig = self._signal_flt
        rr_min = float(self._safe_float(self.ent_minrr, MouseECG.RR_MIN_MS))

        self._set_status("Detecting artifacts…", MUTED)
        self.btn_review_art.configure(state="disabled", text="Detecting…")  # type: ignore[union-attr]
        self.update_idletasks()

        try:
            candidates = detect_rr_artifacts(
                rp, fs,
                rr_min_ms    = rr_min,
                rr_max_ms    = MouseECG.RR_MAX_MS,
                window_beats = 11,
                dev_threshold= 0.20,
                signal       = sig,
            )
        except Exception as exc:
            messagebox.showerror("Detection error", str(exc))
            self.btn_review_art.configure(state="normal", text="🔍  Review Artifacts")  # type: ignore[union-attr]
            return

        self.btn_review_art.configure(state="normal", text="🔍  Review Artifacts")  # type: ignore[union-attr]

        n = len(candidates)
        if n == 0:
            self._set_status("No artifacts detected — signal looks clean ✓", GREEN)
            messagebox.showinfo("No artifacts",
                                "No artifact candidates found with the current settings.\n\n"
                                "If you suspect issues, try lowering the Min R-R distance "
                                "or adjusting the sensitivity threshold.")
            return

        self._set_status(f"{n} artifact candidates found — opening review…", ORANGE)

        # Count by type for display
        counts = {}
        for c in candidates:
            counts[c["type"]] = counts.get(c["type"], 0) + 1
        detail = "  ·  ".join(f"{v} {k}" for k, v in counts.items())
        self._set_status(f"Reviewing {n} candidates ({detail})", ORANGE)

        dlg = ArtifactReviewDialog(self, sig, rp, fs, candidates, rr_min_ms=rr_min)
        self.wait_window(dlg)   # blocks until dialog closes

        result = dlg.get_result()
        if result is None:
            self._set_status("Artifact review cancelled", MUTED)
            return

        corrected, report = apply_artifact_decisions(rp, result)
        removed = report["n_in"] - report["n_out"]

        self._rpeaks_ok       = corrected
        self._artifact_report = report
        self._artifact_candidates = result

        # Clear any manual exclusions that overlapped removed peaks
        if removed > 0:
            removed_samples = {c["sample"] for c in result if c["decision"] == "remove"}
            self._manual_excluded -= removed_samples

        art_str = (f"Artifact review: −{removed} beats "
                   f"(non-physio={report['n_nonphysio']}  "
                   f"ectopic={report['n_ectopic']}  "
                   f"dup={report['n_duplicate']}  "
                   f"kept={report['n_kept']})")

        if removed > 0:
            # Peaks changed → previous HRV metrics are stale.  Discard them so
            # the user cannot export results that don't match the corrected peaks.
            self._results = None
            self._epoch_df = None
            stale_note = "  ⚠ Re-run Core Analysis to update HRV metrics."
            self._set_status(art_str + stale_note, ORANGE)
            # Visual warning on every analysis tab
            warn_text = "⚠  Peaks changed after artifact review — re-run Core Analysis"
            for attr in ("lbl_freq_status", "lbl_nonlin_status", "lbl_ivl_status"):
                lbl = getattr(self, attr, None)
                if lbl is not None:
                    lbl.configure(text=warn_text, text_color=ORANGE)
            for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl"):
                btn = getattr(self, btn_attr, None)
                if btn is not None:
                    btn.configure(state="disabled")
        else:
            self._set_status(art_str, GREEN)

        # Refresh the overview / detail with the cleaned peaks
        self._draw_detail()
        color = GREEN if len(corrected) > 10 else RED
        self.lbl_npeaks.configure(  # type: ignore[union-attr]
            text=f"Peaks detected: {len(corrected)}  (after review)",
            text_color=color)

    def _run_freq(self) -> None:
        """Compute frequency-domain HRV in background, then render."""
        self.analysis_ctrl.run_freq()

    def _run_nonlinear(self) -> None:
        """Compute non-linear HRV in background, then render."""
        self.analysis_ctrl.run_nonlinear()

    def _open_annotations(self) -> None:
        """Open the annotation manager dialog."""
        AnnotationManagerDialog(self)

    def _update_ann_count(self) -> None:
        """Refresh the annotation count badge in the toolbar and the
        right panel's ANNOTATIONS section summary."""
        if self.lbl_ann_count is not None:
            n = len(self._annotations)
            self.lbl_ann_count.configure(  # type: ignore[union-attr]
                text=f"{n}" if n else "",
                text_color=ORANGE if n else MUTED)
        if self.lbl_panel_ann_count is not None:
            n = len(self._annotations)
            self.lbl_panel_ann_count.configure(text=f"{n} annotations")  # type: ignore[union-attr]

    def _open_pacing_periods(self) -> None:
        """Open the pacing/stimulation period manager dialog."""
        PacingPeriodManagerDialog(self)

    def _update_pacing_count(self) -> None:
        """Refresh the pacing-period count in the ANNOTATIONS panel section's
        summary (the toolbar's redundant "Pacing Periods" chip was removed --
        this dialog is now reached solely from that panel section)."""
        if self.lbl_panel_pacing_count is not None:
            n = len(self._pacing_periods)
            self.lbl_panel_pacing_count.configure(text=f"{n} pacing periods")  # type: ignore[union-attr]

    def _toggle_rrhr_strip(self) -> None:
        """Show/hide the RR/HR/Quality sub-plot strip below the detail plot.

        Hiding it lets the detail plot reclaim row 1's space (weight=0 makes
        row 0 the only expanding row); showing it restores the 70/30 split.
        """
        self.ui.rrhr_strip_visible = not self.ui.rrhr_strip_visible
        if self.ui.rrhr_strip_visible:
            self.subplot_frame.grid()
            self.plot_area.grid_rowconfigure(1, weight=3)
            self.btn_toggle_rrhr.configure(
                fg_color=BLUE, hover_color=BLUE_HOVER, text_color="white")
        else:
            self.subplot_frame.grid_remove()
            self.plot_area.grid_rowconfigure(1, weight=0)
            self.btn_toggle_rrhr.configure(
                fg_color=BORDER, hover_color=BORDER2, text_color=MUTED)

    def _toggle_left_panel(self) -> None:
        """Hide/show the left sidebar entirely (simple hide/show, not a rail).

        Collapsing un-packs the panel so the center workspace immediately
        reclaims its width. Restoring it MUST pass before=self.main --
        self.sidebar/self.right_panel are deliberately packed before
        self.main in _build() so main's expand=True doesn't claim the full
        cavity first; a bare .pack() after .pack_forget() re-appends the
        slave to the END of the master's pack order (i.e. after main, which
        has already claimed all the space), silently producing a permanent
        zero-width panel. before=self.main reinserts it at the correct
        position regardless of call order.
        """
        self.ui.left_panel_collapsed = not self.ui.left_panel_collapsed
        if self.ui.left_panel_collapsed:
            self.sidebar.pack_forget()
            self.btn_toggle_left_panel.configure(text="⟩")
        else:
            self.sidebar.pack(side="left", fill="y", before=self.main)
            self.btn_toggle_left_panel.configure(text="⟨")

    def _toggle_right_panel(self) -> None:
        """Hide/show the right panel entirely -- see _toggle_left_panel()'s
        docstring for why before=self.main is required on restore."""
        self.ui.right_panel_collapsed = not self.ui.right_panel_collapsed
        if self.ui.right_panel_collapsed:
            self.right_panel.pack_forget()
            self.btn_toggle_right_panel.configure(text="⟨")
        else:
            self.right_panel.pack(side="right", fill="y", before=self.main)
            self.btn_toggle_right_panel.configure(text="⟩")

    def _run_intervals(self) -> None:
        """Compute interval delineation in background, then launch verifier."""
        self.analysis_ctrl.run_intervals()

    def _launch_interval_verifier(
        self,
        df:       "pd.DataFrame",
        beat_mat: "np.ndarray",
        beat_time:"np.ndarray",
    ) -> None:
        """Instantiate IntervalVerifierPanel in the intervals tab.

        Called from _run_intervals _done callback (main thread only).
        The verifier renders into the intervals_ecg CanvasSlot and places
        its navigation bar into frm_ivl_nav.
        """
        self.analysis_ctrl.launch_interval_verifier(df, beat_mat, beat_time)

    def _start_async_result(
        self,
        button: ctk.CTkButton,
        busy_label: str,
        worker: Callable[[], Any],
        on_done: Callable[[Any], None],
        ) -> None:
        """Convenience wrapper: like _start_async but forwards the worker return value.

        Calls _start_async with pass_result=True so that on_done receives the
        value returned by worker().  All threading, progress, and error handling
        are identical to _start_async.
        """
        self._start_async(
            button,
            busy_label,
            "",          # no separate status message — caller sets its own
            worker,
            on_done,
            pass_result=True,
        )

    # ── Async helper ──────────────────────────────────────────

    def _set_progress(self, pct: int, msg: str) -> None:
        """Update the deterministic progress bar + stage label (main thread only).

        Does NOT call update_idletasks() — that would block the event loop and
        cause the UI to freeze mid-analysis.  The progress bar is updated
        asynchronously by the normal Tk event loop.
        """
        frac = max(0.0, min(1.0, pct / 100.0))
        self.progress.set(frac)
        self._last_prog_pct = pct
        self._last_prog_msg = msg
        self._render_progress_label()

    def _render_progress_label(self) -> None:
        """Render lbl_progress from the last known (pct, msg) + a time-left estimate.

        Called on every real progress_cb checkpoint (via _set_progress) --
        with the fake heartbeat pulse removed, this no longer ticks between
        checkpoints, so the ETA shown is the best estimate as of the last
        real checkpoint, not a live-updating countdown.
        """
        pct = getattr(self, "_last_prog_pct", 0)
        msg = getattr(self, "_last_prog_msg", "")
        start = getattr(self, "_async_start_time", None)
        suffix = ""
        if start is not None and pct >= 1:
            elapsed = time.time() - start
            eta = elapsed * (100 - pct) / pct
            suffix = f"   (~{eta:.0f}s left)"
        self.lbl_progress.configure(text=f"{pct}%  {msg}{suffix}")

    def _start_async(
        self,
        button: ctk.CTkButton,
        btn_busy_label: str,
        status_msg: str,
        worker: Callable[[], Any],
        on_done: Callable[..., None],
        original_label: "str | None" = None,
        pass_result: bool = False,
    ) -> None:
        """Disable *button*, show the progress bar, and run *worker* in a thread.

        On success, *on_done* is scheduled on the main thread via ``after(0, …)``.
        On failure, an error dialog is shown and the button is re-enabled.
        The original button text is restored in both cases.

        Parameters
        ----------
        button          : Button to disable while busy.
        btn_busy_label  : Text shown on the button while running.
        status_msg      : Status bar text shown while running (pass "" to skip).
        worker          : Callable executed in the background thread.
                          Must not access Tkinter widgets directly.
        on_done         : Callable executed on the main thread on success.
                          If pass_result=True it receives worker()'s return value.
        original_label  : Button text to restore; auto-detected if None.
        pass_result     : If True, worker() return value is passed to on_done(result).
                          If False (default), on_done is called with no arguments.
        """
        if getattr(self, "_async_busy", False):
            # self.progress/self.lbl_progress/_last_prog_pct/_async_start_time
            # are shared instance state with no per-operation isolation --
            # letting two _start_async calls run concurrently would interleave
            # their writes to those attributes and produce chaotic, seemingly
            # random progress-bar/label motion. Refuse the second call instead.
            self._set_status("Another operation is already running — please wait", ORANGE)
            return
        original_text = original_label or button.cget("text")
        button.configure(state="disabled", text=btn_busy_label)
        if status_msg:
            self._set_status(status_msg, ORANGE)
        self.progress.set(0)
        self._prog_row.pack(side="bottom", fill="x", padx=SPACE_M, pady=(SPACE_XS, SPACE_S))
        self._async_start_time = time.time()
        self._last_prog_pct = 0
        self._last_prog_msg = ""
        self._async_busy = True
        self._render_progress_label()

        def _thread_target() -> None:
            try:
                result = worker()
                if pass_result:
                    self.after(0, lambda r=result: _finish_success(r))
                else:
                    self.after(0, lambda: _finish_success(None))
            except Exception as exc:
                tb = traceback.format_exc()
                self.after(0, lambda e=exc, t=tb: _finish_error(e, t))

        def _finish_success(result: Any) -> None:
            self._async_busy = False
            self._stop_progress(button, original_text)
            if pass_result:
                on_done(result)
            else:
                on_done()  # noqa: pass_result=False branch — on_done takes no args

        def _finish_error(exc: Exception, tb: str) -> None:
            self._async_busy = False
            self._stop_progress(button, original_text)
            self._set_status(f"Error: {exc}", RED)
            messagebox.showerror("Error", f"{exc}\n\n{tb}")

        threading.Thread(target=_thread_target, daemon=True).start()

    def _stop_progress(self, button: ctk.CTkButton, original_label: str) -> None:
        self.progress.set(1.0)
        self._prog_row.pack_forget()
        self.lbl_progress.configure(text="")
        self._async_start_time = None
        button.configure(state="normal", text=original_label)

    # ════════════════════════════════════════════════════════
    #  DRAWING — overview & detail
    # ════════════════════════════════════════════════════════

    def _draw_overview(self) -> None:
        """Stub — overview removed; detail view is the sole signal display."""
        self.plot_ctrl.draw_overview()

    def _compute_filter_preview_segment(
        self, t_start: float, t_end: float,
    ) -> "Optional[tuple[np.ndarray, np.ndarray, np.ndarray]]":
        """Compute a live-filtered preview of the signal over [t_start, t_end].

        Operates ONLY on the visible window (+ a short margin to absorb
        filtfilt edge transients) using the CURRENT filter widget values
        (HP/LP cutoffs, notch, clean method) — never on the full recording.
        This is intentionally cheap: it never touches ``self._signal_flt``
        or any detection state, so it's safe to recompute on every redraw
        without affecting Preview Detection / Run Full Analysis results.

        Returns (t_slice, raw_slice_norm, filt_slice_norm), or None if the
        raw signal isn't loaded yet or the window is too short to filter.
        """
        return self.signal_ctrl.compute_filter_preview_segment(t_start, t_end)

    def _on_filter_preview_toggle(self) -> None:
        """Toggle the before/after filter overlay and redraw."""
        self.signal_ctrl.on_filter_preview_toggle()

    def _refresh_filter_preview(self) -> None:
        """Recompute the filter-preview overlay with current widget values.

        Bound to HP/LP entry <Return>/<FocusOut> and notch/clean-method
        changes — the preview segment isn't auto-reactive to keystrokes,
        only re-evaluated on these discrete commit events, matching how
        the rest of the sidebar (Preview Detection button) already works.
        Also the single choke point for refreshing the "Processing" summary,
        since every FILTER SETTINGS widget is already wired to this method.
        """
        self.signal_ctrl.refresh_filter_preview()
        self._update_filter_summary()

    def _draw_detail(self, t_start: float | None = None) -> None:
        """Draw the time-windowed detail view with peak markers.

        The active signal (raw or filtered) is drawn at full opacity; the
        other signal is drawn as a ghost at 20 % opacity so the filtering
        effect is always visible.  Peak markers are computed from the filtered
        signal regardless of mode — they are not re-detected on toggle.

        Also handles two extra states, both display-only (no detection state
        is touched):
        • Raw-only (just opened, Preview Detection not yet run): signal_flt
          is None, so only the raw trace is drawn, with no peaks/threshold.
        • Filter preview overlay (self._filter_preview_on): an on-the-fly
          filtered version of the visible window, computed from the current
          filter widget values, overlaid on the raw trace so the user can
          judge filter settings before committing to Preview Detection.
        """
        self.plot_ctrl.draw_detail(t_start)
        self.plot_ctrl.draw_overview()
        self.plot_ctrl.draw_detail_rrhr(t_start)

    def _kb_navigate(self, direction: int) -> None:
        """Keyboard left/right arrow navigation — only active on Detection tab."""
        self.nav_ctrl.kb_navigate(direction)

    def _navigate(self, direction: int) -> None:
        """Shift the detail view left/right by 80 % of the window width."""
        self.nav_ctrl.navigate(direction)

    def _navigate_big(self, direction: int) -> None:
        """Jump by 10× the current window width."""
        self.nav_ctrl.navigate_big(direction)

    def _nav_reset(self) -> None:
        self.nav_ctrl.nav_reset()

    def _nav_end(self) -> None:
        """Jump to the end of the signal."""
        self.nav_ctrl.nav_end()

    def _nav_goto(self) -> None:
        """Jump to the time entered in the position field."""
        self.nav_ctrl.nav_goto()

    def _sync_nav_pos_entry(self) -> None:
        """Update the position entry widget to reflect _nav_pos."""
        self.nav_ctrl.sync_nav_pos_entry()

    # ════════════════════════════════════════════════════════
    #  RESULT PLOTS
    # ════════════════════════════════════════════════════════

    def _run_plot_chain(
        self,
        tasks: list,
        on_complete: "Optional[Callable[[], None]]" = None,
        auto_epochs: bool = False,
    ) -> None:
        """Run a list of (label, fn) plot tasks sequentially via after() chain."""
        self.plot_ctrl.run_plot_chain(tasks, on_complete=on_complete, auto_epochs=auto_epochs)

    def _draw_core_results(
        self,
        on_complete: "Optional[Callable[[], None]]" = None,
        auto_epochs: bool = False,
    ) -> None:
        """Render only the fast core plots (RR, Beat, Summary, Poincaré).

        Called immediately after core analysis.  Freq / non-linear / intervals
        are rendered separately when their per-tab buttons are clicked.
        """
        self.plot_ctrl.draw_core_results(on_complete=on_complete, auto_epochs=auto_epochs)

    def _draw_all_results(
        self,
        on_complete: "Optional[Callable[[], None]]" = None,
        auto_epochs: bool = False,
    ) -> None:
        """Render ALL result plots (used by export and legacy callers)."""
        self.plot_ctrl.draw_all_results(on_complete=on_complete, auto_epochs=auto_epochs)

    def _plot_rr(self, r: dict) -> None:
        """Plot RR tachogram, HR trace, and RR distribution histogram.

        Drastic RR changes are detected and shown as orange/red markers on
        the tachogram.  Clicking any point navigates to that beat in Detection.
        Right-clicking jumps specifically to the nearest spike.
        """
        self.plot_ctrl.plot_rr(r)

    def _plot_hrv_tables(self, r: dict) -> None:
        """Populate time-domain and frequency-domain HRV text boxes."""
        self.plot_ctrl.plot_hrv_tables(r)

    def _plot_psd(self, r: dict) -> None:
        """Welch PSD with mouse-specific VLF / LF / HF band shading.

        RR intervals are resampled to a uniform time grid using a cubic spline
        before computing the Welch periodogram.  Cubic (vs linear) resampling
        preserves spectral shape and avoids the artificial high-frequency power
        that linear interpolation introduces.

        Mouse-specific design choices
        ─────────────────────────────
        • Interpolation rate: 20 Hz  (Nyquist = 10 Hz >> HF ceiling of 5 Hz)
        • nperseg: aims for ≥ 0.02 Hz resolution — enough to separate
          VLF (0–0.4), LF (0.4–1.5) and HF (1.5–5.0) bands cleanly.
          Formula: nperseg = max(256, min(fs_interp / 0.02, N // 2))
          e.g. 20 / 0.02 = 1000, so for long recordings nperseg = 1000.
        • noverlap: 75 % of nperseg (Welch variance reduction)
        • Window: Hann (default scipy) — good sidelobe suppression
        """
        self.plot_ctrl.plot_psd(r)

    def _plot_radar(self, r: dict) -> None:
        """Normalised HRV spider / radar chart."""
        self.plot_ctrl.plot_radar(r)

    def _plot_nonlinear(self, r: dict) -> None:
        """Poincaré plot and non-linear HRV metric table."""
        self.plot_ctrl.plot_nonlinear(r)


    def _plot_intervals_ecg(self, r: dict) -> None:
        """ECG beat strip annotated with P / Q / R / S / T landmarks.

        Design
        ------
        • X-axis is relative time from R peak (ms) — always centred at 0.
        • 3 beats are selected with the most complete wave annotation.
        • Plus a 4th "anatomy" reference panel on the right.
        • R_peak_s is read directly from the DataFrame (no index-mapping guesses).
        """
        self.plot_ctrl.plot_intervals_ecg(r)

    def _plot_intervals(self, r: dict) -> None:
        """Violin + box plot for PR / QRS / QT / QTc intervals."""
        self.plot_ctrl.plot_intervals(r)

    def _plot_beat_template(self, r: dict) -> None:
        """Average beat template, ±1 SD band, and amplitude / morphology distributions.

        All heavy numpy work (beat matrix, SD, per-beat correlations) was pre-computed
        in analyse_core() on the background thread.  This function only renders.
        """
        self.plot_ctrl.plot_beat_template(r)

    def _plot_summary(self, r: dict) -> None:
        """Populate the Summary tab: KPI cards, all plots, and text report."""
        self.plot_ctrl.plot_summary(r)

    # ── KPI bar update ────────────────────────────────────────

    def _reset_result_plots(self) -> None:
        """Clear stored draw_fn on every result-plot slot.

        Prevents stale draw functions from a previous file replaying
        on window resize after a new file is loaded.
        """
        self.plot_ctrl.reset_result_plots()

    def _reset_tab_status_labels(self) -> None:
        """Reset per-tab status labels and disable action buttons.

        Called on new file load so labels from the previous analysis
        (e.g. "Done LF=42%") don't persist after loading a new file.
        """
        self.plot_ctrl.reset_tab_status_labels()

    def _reset_kpis(self) -> None:
        """Reset all KPI labels to dash when results are invalidated."""
        self.plot_ctrl.reset_kpis()

    def _update_kpis(self) -> None:
        self.plot_ctrl.update_kpis()

    # ════════════════════════════════════════════════════════
    #  EPOCH ANALYSIS
    # ════════════════════════════════════════════════════════

    def _compute_epochs(self) -> None:
        """Compute epoch-level HRV in a background thread to keep the UI responsive.

        nk.hrv_time() is called once per epoch.  On long recordings with many
        short epochs this adds up to several seconds — enough to freeze the UI
        noticeably.  The calculation is therefore moved to a daemon thread via
        _start_async_result, exactly like _run_freq / _run_nonlinear.
        """
        self.analysis_ctrl.compute_epochs()

    # ════════════════════════════════════════════════════════
    #  EXPORT
    # ════════════════════════════════════════════════════════

    def _build_excel_workbook(self) -> "Workbook":
        """Build a formatted openpyxl Workbook from the current results."""
        return self.export_ctrl.build_excel_workbook()

    def _write_excel(self, destination) -> None:
        """Write the formatted workbook to *destination* (path or BytesIO)."""
        self.export_ctrl.write_excel(destination)

    def _export_excel(self) -> None:
        self.export_ctrl.export_excel()

    def _export_zip(self) -> None:
        self.export_ctrl.export_zip()

    def _export_pdf_report(self) -> None:
        """Generate a one-page PDF summary: ECG strip + KPI table + interpretation."""
        self.export_ctrl.export_pdf_report()

    def _export_prism(self) -> None:
        """Export all analysis results to a GraphPad Prism .pzfx file."""
        self.export_ctrl.export_prism()


    def _add_recent(self, path: str) -> None:
        self.session_ctrl.add_recent(path)

    def _open_recent(self) -> None:
        """Show recent recordings from SQLite registry with summary stats."""
        db_rows = recent_recordings(limit=20) if _DB_AVAILABLE else []
        db_paths = {r["filepath"] for r in db_rows}
        extra = [p for p in self._recent if p not in db_paths and os.path.exists(p)]

        if not db_rows and not extra:
            messagebox.showinfo("Recent files", "No recent recordings yet.")
            return

        win = ctk.CTkToplevel(self)
        win.title("Recent recordings")
        win.geometry("740x480")
        win.configure(fg_color=PANEL)
        win.grab_set(); win.lift()

        hdr = ctk.CTkFrame(win, fg_color=PANEL, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Recent recordings", font=FONT_CARD_TITLE,
                     text_color=TEXT, anchor="w").pack(side="left", padx=SPACE_L, pady=SPACE_M)

        scroll = ctk.CTkScrollableFrame(win, fg_color=PANEL)
        scroll.pack(fill="both", expand=True, padx=SPACE_M, pady=SPACE_M)

        def _entry(path: str, hr: str = "—", sdnn: str = "—",
                   dur: str = "—", notes: str = "") -> None:
            if not os.path.exists(path):
                return
            card = ctk.CTkFrame(scroll, fg_color=CARD, corner_radius=6)
            card.pack(fill="x", pady=(0, SPACE_S))
            top = ctk.CTkFrame(card, fg_color="transparent")
            top.pack(fill="x", padx=SPACE_M, pady=(SPACE_S, SPACE_XS))
            ctk.CTkButton(top, text=os.path.basename(path), anchor="w",
                          fg_color="transparent", hover_color=BORDER,
                          text_color=BLUE, font=FONT_SMALL,
                          command=lambda p=path: (win.destroy(), self._load_path(p))
                          ).pack(side="left", fill="x", expand=True)
            ctk.CTkLabel(top, text=f"HR {hr}  SDNN {sdnn}  {dur}",
                         font=FONT_KPI_LABEL, text_color=MUTED).pack(side="right")
            if notes:
                ctk.CTkLabel(card, text=f"📝 {notes[:90]}",
                             font=FONT_KPI_LABEL, text_color=LIGHT,
                             anchor="w").pack(padx=SPACE_M, pady=(0, SPACE_S), fill="x")

        for r in db_rows:
            _entry(
                r["filepath"],
                hr=f"{r['hr_mean']:.0f} bpm" if r.get("hr_mean") else "—",
                sdnn=f"{r['sdnn']:.1f} ms"   if r.get("sdnn") else "—",
                dur=f"{r['duration_s']:.0f} s" if r.get("duration_s") else "",
                notes=r.get("notes", ""),
            )
        for p in extra:
            _entry(p)



    def _load_path(self, path: str) -> None:
        self.signal_ctrl.load_path(path)

    def _reset_for_new_file(self) -> None:
        """Reset ALL analysis state and UI to the startup blank slate.

        Called every time a new file is opened so there is zero carry-over
        from the previous recording.  Sidebar *parameter* widgets (channel,
        fs, thresholds, filters) are intentionally kept — users typically
        want the same settings for consecutive recordings from the same rig.

        IMPORTANT: must NOT call _init_state() — that method also zeroes out
        all widget-reference attributes (ent_channel, ent_fs, …) which are
        already live in the UI, causing AttributeError on the next snapshot.
        Only data variables are reset here.
        """
        self.signal_ctrl.reset_for_new_file()

    def _try_restore_session(self, path: str) -> bool:
        """If a saved session exists for *path*, offer to restore it.

        Returns True if the session restoration was initiated (caller should
        skip _preview).  The actual restore runs in a background thread via
        _start_async so the UI stays responsive during signal reload + filtering.
        """
        return self.session_ctrl.try_restore_session(path)

    def _restore_session_worker(self, state: dict) -> dict:
        """Background worker — MUST NOT write to self or touch any Tkinter widget.

        Reloads the raw signal from the original .mat file and re-runs the
        pure ``_compute_preview_bundle`` static method with the saved filter
        parameters. Returns a plain bundle; ``_on_restore_session_done``
        writes every field to ``self`` (and to widgets) on the main thread.

        Session restore previously ran entirely in the background thread
        (including dozens of widget .configure/.delete/.insert calls and
        _draw_detail()), which is unsafe — Tkinter/Tcl objects must only be
        touched from the main thread. This mirrors the _preview_worker /
        _on_preview_done split used elsewhere in the app.
        """
        return self.session_ctrl.restore_session_worker(state)

    def _on_restore_session_done(self, bundle: dict, saved_at: str) -> None:
        """Write all restored state to self and to widgets — main thread only.

        Counterpart to ``_restore_session_worker``. This is the ONLY place
        that writes session-restore state to ``self``/widgets, exactly as
        ``_on_preview_done`` is the sole writer after ``_preview_worker``.
        """
        self.session_ctrl.on_restore_session_done(bundle, saved_at)

    # ════════════════════════════════════════════════════════
    #  SESSION SAVE / RESTORE
    # ════════════════════════════════════════════════════════

    def _current_filter_params_dict(self) -> dict:
        """Return a serialisable filter-params dict, reading from widgets if available.

        Safe to call at any point — uses FilterParams defaults for any widget
        not yet built (e.g. during early startup or after a rebuild).
        """
        return self.session_ctrl.current_filter_params_dict()

    def _safe_get_tab(self) -> str:
        """Return the current tab name, or 'Detection' if not yet built."""
        return self.session_ctrl.safe_get_tab()

    def _collect_session_state(self) -> dict:
        """Gather all serialisable app state into a flat dict.

        Signal arrays (signal_flt, signal_raw_norm, all_cands, all_proms) are
        intentionally NOT stored.  They are derived entirely from the .mat file
        and the filter parameters; _restore_session_worker re-runs
        _compute_preview_bundle to reconstruct them in < 2 s.  This keeps session
        files small (< 200 KB for a typical recording) regardless of recording length.

        Previously (v3) the session stored the full filtered signal as a Python
        list (~58 MB for 10 min at 2 kHz, ~346 MB for 1 h), making auto-save
        impractical and restore slow.
        """
        return self.session_ctrl.collect_session_state()

    def _save_session(self) -> None:
        """Serialise full analysis state to a .ecgsession cache file and update registry."""
        self.session_ctrl.save_session()

    def _delete_session(self) -> None:
        """Delete the session cache file for the current file."""
        self.session_ctrl.delete_session()

    def _update_session_ui(self, has_session: bool,
                           saved_at: str = "") -> None:
        """Update the session info label and button states."""
        self.session_ctrl.update_session_ui(has_session, saved_at=saved_at)

    # ════════════════════════════════════════════════════════
    #  WAVE TEMPLATE EDITOR
    # ════════════════════════════════════════════════════════

    def _open_wave_template_editor(self) -> None:
        """Open the interactive P/Q/R/S/T template editor.

        If a mean beat is available it is passed to the editor for display;
        otherwise the editor shows a synthetic reference trace.
        """
        beat_time = None
        mean_beat = None
        beat_sd   = None
        if self._results is not None:
            beat_time = self._results.get("beat_time")
            mean_beat = self._results.get("beat_template")
            beat_sd   = self._results.get("beat_sd")

        if self._wave_template is None:
            self._wave_template = WaveTemplate.load()

        editor = WaveTemplateMiniEditor(
            self, self._wave_template,
            beat_time=beat_time,
            mean_beat=mean_beat,
            beat_sd=beat_sd,
        )
        self.wait_window(editor)  # modal — blocks until closed

        # Refresh the label whether or not the user saved
        self._update_session_ui(
            has_session=load_session(self._filepath) is not None
            if self._filepath else False
        )
        if editor._saved:
            self._set_status(
                "Wave template updated — re-run Interval Delineation to apply changes.",
                GREEN)


    def _open_params_dialog(self) -> None:
        """Open a dedicated floating parameters window with all settings clearly grouped."""
        win = ctk.CTkToplevel(self)
        win.title("⚙  Parameters")
        win.geometry("540x720")
        win.configure(fg_color=PANEL)
        win.resizable(True, True)
        win.grab_set()
        win.lift()

        scroll = ctk.CTkScrollableFrame(win, fg_color=PANEL,
                                        scrollbar_button_color=BORDER,
                                        scrollbar_button_hover_color=BORDER2)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)
        s = scroll
        px = dict(padx=SPACE_L)

        def _sec(title: str, color: str = BLUE) -> ctk.CTkFrame:
            hdr = ctk.CTkFrame(s, fg_color=PANEL, corner_radius=8)
            hdr.pack(fill="x", padx=SPACE_M, pady=(SPACE_M, SPACE_XS))
            ctk.CTkFrame(hdr, height=3, fg_color=color, corner_radius=2).pack(fill="x")
            ctk.CTkLabel(hdr, text=title, font=FONT_SIDEBAR_HDR,
                         text_color=color, anchor="w").pack(padx=SPACE_M, pady=(SPACE_S, SPACE_S), fill="x")
            body = ctk.CTkFrame(s, fg_color=CARD, corner_radius=6)
            body.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_S))
            return body

        def _row_entry(parent, label: str, widget_attr: str) -> ctk.CTkEntry:
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", **px, pady=(SPACE_S, 0))
            row.columnconfigure(1, weight=1)
            ctk.CTkLabel(row, text=label, font=FONT_SMALL, text_color=MUTED,
                         anchor="w", width=160).grid(row=0, column=0, sticky="w")
            # Mirror value from existing sidebar widget
            src: Optional[ctk.CTkEntry] = getattr(self, widget_attr, None)
            default = src.get() if src is not None else ""
            ent = ctk.CTkEntry(row, font=FONT_LABEL, height=28, fg_color=BG,
                               border_color=BORDER2, text_color=TEXT)
            ent.insert(0, default)
            ent.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))
            return ent

        def _row_switch(parent, label: str, widget_attr: str) -> ctk.CTkSwitch:
            src: Optional[ctk.CTkSwitch] = getattr(self, widget_attr, None)
            is_on = bool(src.get()) if src is not None else False
            sw = ctk.CTkSwitch(parent, text=label, font=FONT_LABEL, text_color=MUTED,
                               progress_color=BLUE, button_color=BORDER2)
            if is_on:
                sw.select()
            sw.pack(**px, anchor="w", pady=(SPACE_S, SPACE_XS))
            return sw

        # ── FILE & SUBJECT ─────────────────────────────────────────────────
        f0 = _sec("📁  File & Subject", BLUE_DARK)
        dlg_channel = _row_entry(f0, "Channel name",    "ent_channel")
        dlg_subject = _row_entry(f0, "Subject ID",      "ent_subject")
        ctk.CTkFrame(f0, height=6, fg_color="transparent").pack()

        # ── SIGNAL ────────────────────────────────────────────────────────
        f1 = _sec("📡  Signal", ORANGE_DARK)
        dlg_fs      = _row_entry(f1, "Sampling rate (Hz)", "ent_fs")
        dlg_t_start = _row_entry(f1, "Start crop (s)",     "ent_t_start")
        dlg_t_end   = _row_entry(f1, "End crop (s)",       "ent_t_end")
        dlg_show_raw = _row_switch(f1, "Show raw signal (vs filtered)", "sw_show_raw")
        ctk.CTkFrame(f1, height=6, fg_color="transparent").pack()

        # ── FILTERS ───────────────────────────────────────────────────────
        f2 = _sec("🔧  Filters", PURPLE)
        dlg_filtering = _row_switch(f2, "Filtering", "sw_filtering")
        dlg_notch     = _row_switch(f2, "Notch 50 Hz", "sw_notch")
        dlg_invert    = _row_switch(f2, "Invert signal polarity", "sw_invert_signal")
        # Labelled Low/High to match the sidebar's FILTERS section -- see
        # the comment there on why (matches filtering.py's bandpass(lo, hi)
        # naming, avoids HP/LP filter-type jargon at a glance).
        dlg_hp = _row_entry(f2, "Low cutoff (Hz)",  "ent_lp")
        dlg_lp = _row_entry(f2, "High cutoff (Hz)", "ent_hp")

        # Clean method
        row_cm = ctk.CTkFrame(f2, fg_color="transparent")
        row_cm.pack(fill="x", **px, pady=(SPACE_S, SPACE_S))
        row_cm.columnconfigure(1, weight=1)
        ctk.CTkLabel(row_cm, text="NK2 clean method:", font=FONT_SMALL, text_color=MUTED,
                     anchor="w", width=160).grid(row=0, column=0, sticky="w")
        _clean_src: Optional[ctk.CTkComboBox] = getattr(self, "cb_clean", None)
        _clean_val = _clean_src.get() if _clean_src is not None else "neurokit"
        dlg_clean = ctk.CTkComboBox(row_cm, font=FONT_LABEL, height=28,
                                    fg_color=BG, border_color=BORDER2,
                                    button_color=BORDER2, text_color=TEXT,
                                    dropdown_fg_color=BG, dropdown_text_color=TEXT,
                                    values=["neurokit", "pantompkins1985",
                                            "elgendi2010", "hamilton2002", "biosppy"])
        dlg_clean.set(_clean_val)
        dlg_clean.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))
        ctk.CTkFrame(f2, height=6, fg_color="transparent").pack()

        # ── DETECTION ─────────────────────────────────────────────────────
        f3 = _sec("🔍  Detection", RED)
        dlg_minrr = _row_entry(f3, "Min R-R distance (ms)", "ent_minrr")

        # Detection method
        row_dm = ctk.CTkFrame(f3, fg_color="transparent")
        row_dm.pack(fill="x", **px, pady=(SPACE_S, SPACE_XS))
        row_dm.columnconfigure(1, weight=1)
        ctk.CTkLabel(row_dm, text="Detection method:", font=FONT_SMALL, text_color=MUTED,
                     anchor="w", width=160).grid(row=0, column=0, sticky="w")
        _dm_src: Optional[ctk.CTkComboBox] = getattr(self, "cb_det_method", None)
        _dm_val = _dm_src.get() if _dm_src is not None else "Auto (NeuroKit2)"
        dlg_det_method = ctk.CTkComboBox(
            row_dm, font=FONT_LABEL, height=28, fg_color=BG, border_color=BORDER2,
            button_color=BORDER2, text_color=TEXT, dropdown_fg_color=BG,
            dropdown_text_color=TEXT,
            values=["SG + Derivative (10 kHz)","Wavelet (CWT)", "Auto (NeuroKit2)", "Envelope Max"])
        dlg_det_method.set(_dm_val)
        dlg_det_method.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))

        dlg_sg_target_fs = _row_entry(f3, "SG target fs (Hz)", "ent_sg_target_fs")
        dlg_sg_window_ms = _row_entry(f3, "SG window (ms)",    "ent_sg_window_ms")
        ctk.CTkLabel(f3, text="SG+Deriv: downsample → Savitzky-Golay derivative → R detection\n"
                             "Wavelet: CWT bruit/QRS/J-wave séparés (pip install PyWavelets)\n"
                             "Envelope Max: maximum local — idéal signaux saturés (clipping ADC)",
                     font=FONT_KPI_LABEL, text_color=LIGHT,
                     anchor="w", wraplength=480, justify="left").pack(**px, pady=(0, SPACE_S), fill="x")

        # Threshold
        thr_val = float(self.sl_thr.get()) if self.sl_thr is not None else 0.5  # type: ignore[union-attr]
        row_thr = ctk.CTkFrame(f3, fg_color="transparent")
        row_thr.pack(fill="x", **px, pady=(SPACE_S, SPACE_S))
        row_thr.columnconfigure(1, weight=1)
        ctk.CTkLabel(row_thr, text="Threshold:", font=FONT_SMALL, text_color=MUTED,
                     anchor="w", width=160).grid(row=0, column=0, sticky="w")
        dlg_thr = ctk.CTkSlider(row_thr, from_=0, to=2,
                                 progress_color=RED, button_color=RED, fg_color=BORDER)
        dlg_thr.set(thr_val)
        dlg_thr.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))
        ctk.CTkFrame(f3, height=6, fg_color="transparent").pack()

        # ── ARTIFACTS ─────────────────────────────────────────────────────
        f4 = _sec("⚠️  Artifacts", ORANGE)
        dlg_artifact = _row_switch(f4, "Auto-correct on Full Analysis (OFF by default)", "sw_artifact")
        ctk.CTkLabel(
            f4,
            text="OFF recommended — use '🔍 Review Artifacts' button for full interactive control.\n"
                 "Auto-correct removes detected ectopic/non-physiological beats without review.",
            font=FONT_KPI_LABEL, text_color=LIGHT,
            anchor="w", wraplength=480, justify="left",
        ).pack(**px, pady=(0, SPACE_M), fill="x")

        # ── EXPERIMENTAL CONTEXT ─────────────────────────────────────────────
        # Reference ranges shown throughout the app (Time Domain status,
        # radar chart, Epochs/Rolling bands, Intervals bands, PDF export) all
        # come from EXPERIMENTAL_CONTEXTS[exp_context] -- there was previously
        # no live UI to actually choose among them (only session-file restore
        # could change it), so every analysis silently used the
        # telemetry_awake default regardless of the real recording condition.
        f5 = _sec("🧪  Experimental Context", TEAL)
        _CTX_LABELS = {key: ranges.label for key, ranges in EXPERIMENTAL_CONTEXTS.items()}
        _CTX_KEYS_BY_LABEL = {v: k for k, v in _CTX_LABELS.items()}
        row_ctx = ctk.CTkFrame(f5, fg_color="transparent")
        row_ctx.pack(fill="x", **px, pady=(SPACE_S, SPACE_S))
        row_ctx.columnconfigure(1, weight=1)
        ctk.CTkLabel(row_ctx, text="Context", font=FONT_SMALL, text_color=MUTED,
                     anchor="w", width=160).grid(row=0, column=0, sticky="w")
        dlg_ctx = ctk.CTkComboBox(row_ctx, font=FONT_LABEL, height=28,
                                  fg_color=BG, border_color=BORDER2,
                                  button_color=BORDER2, text_color=TEXT,
                                  dropdown_fg_color=BG, dropdown_text_color=TEXT,
                                  values=list(_CTX_LABELS.values()))
        current_label = _CTX_LABELS.get(self.analysis.exp_context, _CTX_LABELS["telemetry_awake"])
        dlg_ctx.set(current_label)
        dlg_ctx.grid(row=0, column=1, sticky="ew", padx=(SPACE_S, 0))

        lbl_ctx_desc = ctk.CTkLabel(
            f5, text=EXPERIMENTAL_CONTEXTS.get(self.analysis.exp_context, EXPERIMENTAL_CONTEXTS["telemetry_awake"]).description,
            font=FONT_KPI_LABEL, text_color=LIGHT, anchor="w",
            wraplength=480, justify="left")
        lbl_ctx_desc.pack(**px, pady=(0, SPACE_S), fill="x")

        def _on_ctx_change(_choice: str = "") -> None:
            key = _CTX_KEYS_BY_LABEL.get(dlg_ctx.get())
            ctx = EXPERIMENTAL_CONTEXTS.get(key) if key else None
            lbl_ctx_desc.configure(text=ctx.description if ctx else "")
        dlg_ctx.configure(command=_on_ctx_change)

        def _open_custom_context() -> None:
            from ecg.ui.dialogs import CustomContextDialog
            def _on_saved() -> None:
                # Refresh the combo's label/value list in case the user
                # renamed "Custom", and select it now that it's been saved.
                new_labels = {key: ranges.label for key, ranges in EXPERIMENTAL_CONTEXTS.items()}
                _CTX_LABELS.clear(); _CTX_LABELS.update(new_labels)
                _CTX_KEYS_BY_LABEL.clear()
                _CTX_KEYS_BY_LABEL.update({v: k for k, v in new_labels.items()})
                dlg_ctx.configure(values=list(_CTX_LABELS.values()))
                dlg_ctx.set(_CTX_LABELS["custom"])
                _on_ctx_change()
            CustomContextDialog(win, on_saved=_on_saved)

        ctk.CTkButton(f5, text="✎  Edit Custom Context…", command=_open_custom_context,
                     fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                     font=FONT_BTN_SEC, height=28, corner_radius=6
                     ).pack(**px, pady=(0, SPACE_M), anchor="w")

        # ── Buttons ───────────────────────────────────────────────────────
        ctk.CTkFrame(s, height=1, fg_color=BORDER).pack(fill="x", padx=SPACE_M, pady=(SPACE_M, SPACE_S))
        btn_row = ctk.CTkFrame(s, fg_color="transparent")
        btn_row.pack(fill="x", padx=SPACE_M, pady=(0, SPACE_M))

        def _apply():
            """Write all dialog values back to the sidebar widgets."""
            _write = lambda attr, val: (
                getattr(self, attr).delete(0, "end") or
                getattr(self, attr).insert(0, val)
            ) if getattr(self, attr, None) is not None else None

            _write("ent_channel",  dlg_channel.get())
            _write("ent_subject",  dlg_subject.get())
            _write("ent_fs",       dlg_fs.get())
            _write("ent_t_start",  dlg_t_start.get())
            _write("ent_t_end",    dlg_t_end.get())
            _write("ent_lp",       dlg_hp.get())
            _write("ent_hp",       dlg_lp.get())
            _write("ent_minrr",    dlg_minrr.get())

            # SG params
            _write("ent_sg_target_fs", dlg_sg_target_fs.get())
            _write("ent_sg_window_ms", dlg_sg_window_ms.get())

            # ComboBoxes
            if self.cb_clean is not None:
                self.cb_clean.set(dlg_clean.get())
            if self.cb_det_method is not None:
                self.cb_det_method.set(dlg_det_method.get())
                self._on_det_method_change(dlg_det_method.get())

            # Experimental context -- no sidebar widget to mirror (there is
            # none), so this writes straight to analysis state. Panels pick
            # it up next time they're (re)computed, same as every other
            # Parameters-dialog change here.
            ctx_key = _CTX_KEYS_BY_LABEL.get(dlg_ctx.get())
            if ctx_key:
                self.analysis.exp_context = ctx_key

            # Switches
            for sw_attr, dlg_sw in [
                ("sw_filtering",    dlg_filtering),
                ("sw_notch",        dlg_notch),
                ("sw_invert_signal",dlg_invert),
                ("sw_show_raw",     dlg_show_raw),
                ("sw_artifact",     dlg_artifact),
            ]:
                w = getattr(self, sw_attr, None)
                if w is None:
                    continue
                if dlg_sw.get():
                    w.select()
                else:
                    w.deselect()

            # Threshold slider
            if self.sl_thr is not None:
                self.sl_thr.set(float(dlg_thr.get()))  # type: ignore[union-attr]
                self.lbl_thr.configure(text=f"Sensitivity:  {dlg_thr.get():.3f}")

            self._on_filtering_toggle()
            self._on_show_raw_toggle()
            self._set_status("Parameters applied ✓", GREEN)
            win.destroy()

        ctk.CTkButton(btn_row, text="✔  Apply & Close", command=_apply,
                      fg_color=GREEN, hover_color=GREEN_DARK, text_color="white",
                      font=FONT_BTN_PRIMARY, height=36, corner_radius=8).pack(
            side="left", padx=(0, SPACE_M))
        ctk.CTkButton(btn_row, text="✗  Cancel", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=36, corner_radius=8).pack(side="left")

    # ════════════════════════════════════════════════════════
    #  APPEARANCE / THEME
    # ════════════════════════════════════════════════════════

    def _open_theme_dialog(self) -> None:
        """Open the appearance settings dialog."""
        dlg = ThemeDialog(self)
        self.wait_window(dlg)

    def _snapshot_ui_state(self) -> dict:
        """Capture all editable widget values before a UI rebuild."""
        return self.session_ctrl.snapshot_ui_state()

    def _restore_ui_state(self, s: dict) -> None:
        """Restore widget values captured before a UI rebuild."""
        self.session_ctrl.restore_ui_state(s)

    def _rebuild_ui(self) -> None:
        """Destroy and recreate the entire widget tree with the current theme.

        All application data (signal, results, peaks) is preserved.
        Widget values (threshold, filter settings, etc.) are snapshot and
        restored.  Matplotlib figures are closed before destruction to avoid
        memory leaks.
        """
        # Capture widget values
        ui_state = self._snapshot_ui_state()

        # Close all matplotlib figures before destroying their canvas widgets
        for slot in self._slots.values():
            try:
                plt.close(slot.fig)
            except Exception as _exc:
                log.debug("plt.close(slot.fig) failed: %s", _exc, exc_info=True)

        # Destroy the existing UI
        for child in self.winfo_children():
            try:
                child.destroy()
            except Exception:
                pass

        # Reset widget registries
        self._slots = {}
        self._kpi   = {}

        # Rebuild from scratch using updated globals
        self._build()

        # Restore widget values
        self._restore_ui_state(ui_state)

        # Restore file info label
        if self._filepath:
            if self.lbl_file is not None:
                self.lbl_file.configure(  # type: ignore[union-attr]
                    text=os.path.basename(self._filepath), text_color=GREEN)
        if self._rpeaks_ok is not None:
            n = len(self._rpeaks_ok)
            if self.lbl_npeaks is not None:
                self.lbl_npeaks.configure(text=f"Peaks detected: {n}", text_color=GREEN)  # type: ignore[union-attr]

        # Restore session / template labels
        # Guard: _wave_template may not exist if _init_state never ran
        # (e.g. theme dialog opened on a fresh launch before file load)
        if self._wave_template is None:
            self._wave_template = WaveTemplate.load()
        if not self._session_dirty:
            self._session_dirty = False
        has_session = bool(
            self._filepath and load_session(self._filepath) is not None)
        self._update_session_ui(has_session=has_session)

        # Enable action buttons if data is ready
        if self._signal_flt is not None:
            if self.btn_save_session is not None:
                self.btn_save_session.configure(state="normal")  # type: ignore[union-attr]
        if self._results is not None:
            for btn_attr in ("btn_run_freq", "btn_run_nonlin", "btn_run_ivl",
                              "btn_save_session"):
                btn = getattr(self, btn_attr, None)
                if btn is not None:
                    btn.configure(state="normal")

        # Repaint signal and results
        apply_theme_config(THEME)
        if self._signal_flt is not None:
            self._draw_detail(self._nav_pos)
        if self._results is not None:
            self.after(80, self._draw_all_results)
            pass  # interpretation removed
        self._update_kpis()
        # Pre-existing gap, not introduced by the Phase 1 top-bar rework:
        # the correlation badge was never repainted after a theme toggle
        # before, so it silently went blank on every _rebuild_ui(). Cheap
        # fix while already touching this exact code path.
        self._update_quality_badge()
        # Same class of gap: the Detection-tab toolbar badges and the right
        # panel's ANNOTATIONS summary both need a refresh after a full
        # rebuild -- both widgets are guaranteed valid here since _build()
        # (called above) has already fully rebuilt every tab and panel.
        self._update_ann_count()
        self._update_pacing_count()

    # ════════════════════════════════════════════════════════
    #  CLIPBOARD / TXT SAVE
    # ════════════════════════════════════════════════════════

    def _copy_summary(self) -> None:
        self.export_ctrl.copy_summary()

    def _save_summary_txt(self) -> None:
        self.export_ctrl.save_summary_txt()

    # ════════════════════════════════════════════════════════
    #  DRAG AND DROP  (tkinterdnd2 optional)
    # ════════════════════════════════════════════════════════

    # ════════════════════════════════════════════════════════
    #  1. KEYBOARD SHORTCUTS
    # ════════════════════════════════════════════════════════

    def _bind_keyboard_shortcuts(self) -> None:
        """Register global keyboard shortcuts."""
        bindings: "list[tuple[str, Any]]" = [
            ("<space>",       lambda e: self._preview()),
            ("<Control-r>",   lambda e: self._run_analysis()),
            ("<Control-o>",   lambda e: self._open_file()),
            ("<Control-s>",   lambda e: self._save_session()),
            ("<Control-e>",   lambda e: self._export_excel()),
            ("<Control-w>",   lambda e: self._export_rr_csv()),
            ("<Control-m>",   lambda e: self._open_compare_segments()),
            ("<F1>",          lambda e: self._show_shortcuts_help()),
        ]
        for seq, cb in bindings:
            try:
                self.bind(seq, cb)
            except Exception as exc:
                log.debug("bind %s: %s", seq, exc)

    def _show_shortcuts_help(self) -> None:
        shortcuts = (
            "Space          Preview Detection\n"
            "Ctrl+R         Run Full Analysis\n"
            "Ctrl+O         Open .mat file\n"
            "Ctrl+S         Save session\n"
            "Ctrl+E         Export Excel\n"
            "Ctrl+W         Export RR intervals CSV\n"
            "Ctrl+M         Compare segments\n"
            "Ctrl+Z / Y     Undo / Redo peak edits\n"
            "←  / →         Navigate tachogram\n"
            "F1             This help\n"
        )
        messagebox.showinfo("Keyboard shortcuts", shortcuts)
    # ════════════════════════════════════════════════════════
    #  3. EXPORT RR INTERVALS AS CSV
    # ════════════════════════════════════════════════════════

    def _export_rr_csv(self) -> None:
        """Export RR intervals to a lightweight CSV (no Excel dependency)."""
        self.export_ctrl.export_rr_csv()

    # ════════════════════════════════════════════════════════
    #  4. SIGNAL QUALITY BADGE (persistent header)
    # ════════════════════════════════════════════════════════

    def _update_quality_badge(self) -> None:
        """Update the persistent quality badge in the KPI bar."""
        self.detection_ctrl.update_quality_badge()

    # ════════════════════════════════════════════════════════
    #  5. DRAG-AND-DROP (fixed wiring)
    # ════════════════════════════════════════════════════════

    def _setup_dnd(self) -> None:
        try:
            self.drop_target_register("DND_Files")   # provided by tkinterdnd2 at runtime  # type: ignore[attr-defined]
            self.dnd_bind("<<Drop>>", self._on_drop)  # provided by tkinterdnd2 at runtime  # type: ignore[attr-defined]
        except Exception:
            pass   # tkinterdnd2 not installed — drag-drop silently disabled

    def _on_drop(self, event) -> None:
        """Handle drag-and-drop .mat file."""
        raw = getattr(event, "data", "").strip()
        # tkinterdnd2 wraps paths with spaces in {braces}
        path = raw.strip("{}")
        if path.lower().endswith(".mat") and os.path.exists(path):
            self._load_path(path)
        else:
            # Try splitting multiple files and load the first .mat
            for part in raw.replace("{", "").replace("}", "").split():
                if part.lower().endswith(".mat") and os.path.exists(part):
                    self._load_path(part)
                    break

    # ════════════════════════════════════════════════════════
    #  7. RECORDING NOTES (per-file, saved in SQLite)
    # ════════════════════════════════════════════════════════

    def _open_notes_dialog(self) -> None:
        """Open a small dialog for free-text notes on the current recording."""
        if not self._filepath:
            messagebox.showinfo("Notes", "Load a file first.")
            return
        current = get_notes(self._filepath) if _DB_AVAILABLE else self._recording_notes

        win = ctk.CTkToplevel(self)
        win.title(f"Notes — {os.path.basename(self._filepath)}")
        win.geometry("500x340")
        win.configure(fg_color=BG)
        win.grab_set(); win.lift()

        ctk.CTkLabel(win, text="📝  Experiment notes",
                     font=FONT_SIDEBAR_HDR, text_color=TEXT,
                     anchor="w").pack(padx=SPACE_L, pady=(SPACE_L, SPACE_S), fill="x")
        ctk.CTkLabel(win, text="Saved per recording (animal ID, drug, dose, time…)",
                     font=FONT_KPI_LABEL, text_color=MUTED, anchor="w").pack(
            padx=SPACE_L, pady=(0, SPACE_S), fill="x")
        txt = ctk.CTkTextbox(win, fg_color=CARD, text_color=TEXT,
                             font=FONT_LABEL, border_width=1,
                             border_color=BORDER2, height=180)
        txt.pack(fill="both", expand=True, padx=SPACE_L, pady=(0, SPACE_M))
        txt.insert("1.0", current)

        def _save():
            notes = txt.get("1.0", "end").strip()
            self._recording_notes = notes
            if _DB_AVAILABLE and self._filepath:
                set_notes(self._filepath, notes)
            self._set_status("Notes saved ✓", GREEN)
            win.destroy()

        btn_row = ctk.CTkFrame(win, fg_color="transparent")
        btn_row.pack(fill="x", padx=SPACE_L, pady=(0, SPACE_L))
        ctk.CTkButton(btn_row, text="✔  Save", command=_save,
                      fg_color=GREEN, hover_color=GREEN_DARK, text_color="white",
                      font=FONT_BTN_PRIMARY, height=30, corner_radius=6).pack(side="left")
        ctk.CTkButton(btn_row, text="✗ Cancel", command=win.destroy,
                      fg_color=BORDER, hover_color=BORDER2, text_color=MUTED,
                      font=FONT_BTN_SEC, height=30, corner_radius=6).pack(
            side="left", padx=(SPACE_M, 0))

    # ════════════════════════════════════════════════════════
    #  9. COMPARE SEGMENTS — statistical test (Wilcoxon)
    # ════════════════════════════════════════════════════════

    def _mannwhitney_test(self, a: "np.ndarray", b: "np.ndarray") -> "tuple[float, str]":
        """Mann-Whitney U test for two independent RR series.

        Returns (p_value, interpretation_string).
        Falls back gracefully if scipy is unavailable.
        """
        return self.analysis_ctrl.mannwhitney_test(a, b)

    # ════════════════════════════════════════════════════════
    #  10. ARRHYTHMIA EPISODE PDF (one strip per episode)
    # ════════════════════════════════════════════════════════

    def _export_arrhythmia_pdf(self) -> None:
        """Export a PDF with one annotated ECG strip per arrhythmia episode."""
        self.export_ctrl.export_arrhythmia_pdf()

    # ════════════════════════════════════════════════════════
    #  11. LIVE THEME TOGGLE (no restart needed)
    # ════════════════════════════════════════════════════════

    def _toggle_dark_live(self) -> None:
        """Switch Light <-> Dark and rebuild the UI without restarting.

        Goes through apply_preset() (not a raw THEME.is_dark flip) so colors
        and is_dark change together -- a bare flip left CTk's native dark
        chrome sitting over the light preset's actual colors.
        """
        from ecg.ui.theme import THEME, apply_theme_config
        THEME.apply_preset("Dark" if not THEME.is_dark else "Light")
        apply_theme_config(THEME)
        THEME.save()
        ui_state = self._snapshot_ui_state()
        self._rebuild_ui()
        self._restore_ui_state(ui_state)
        mode = "Dark" if THEME.is_dark else "Light"
        self._set_status(f"Theme: {mode} mode applied", BLUE)

    # ════════════════════════════════════════════════════════
    #  SIDEBAR WIDGET HELPERS
    # ════════════════════════════════════════════════════════


    def _sidebar_sep(self, parent) -> None:
        ctk.CTkFrame(parent, height=1, fg_color=BORDER).pack(
            fill="x", padx=SPACE_L, pady=SPACE_M)

    def _sidebar_hdr(self, parent, text: str) -> None:
        ctk.CTkLabel(parent, text=text, font=FONT_SIDEBAR_HDR,
                     text_color=RED, anchor="w").pack(padx=SPACE_L, fill="x", pady=(SPACE_M, SPACE_XS))

    def _sidebar_entry(self, parent, label: str, attr: str, default: str, pad: dict) -> None:
        """Add a labelled entry and bind it to self.ent_<attr>."""
        ctk.CTkLabel(parent, text=label, font=FONT_SMALL,
                     text_color=MUTED, anchor="w").pack(**pad, fill="x")
        e = ctk.CTkEntry(parent, font=FONT_LABEL, height=30, fg_color=BG,
                          border_color=BORDER2, text_color=TEXT)
        e.insert(0, default)
        e.pack(**pad, fill="x", pady=(SPACE_XS, SPACE_M))
        setattr(self, f"ent_{attr}", e)

    def _sidebar_entry_row(self, parent, pad: dict, items: list[tuple]) -> None:
        """Render a horizontal row of (label, attr, default) entry pairs."""
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(**pad, fill="x", pady=(0, SPACE_S))
        for i, (label, attr, default) in enumerate(items):
            col = ctk.CTkFrame(row, fg_color="transparent")
            col.pack(side="left", fill="x", expand=True,
                     padx=(0, SPACE_S) if i < len(items) - 1 else 0)
            ctk.CTkLabel(col, text=label, font=FONT_SMALL, text_color=MUTED).pack(anchor="w")
            e = ctk.CTkEntry(col, font=FONT_LABEL, height=28, fg_color=BG,
                              border_color=BORDER2, text_color=TEXT)
            e.insert(0, default)
            e.pack(fill="x")
            setattr(self, f"ent_{attr}", e)

    def _switch(self, parent, label: str, pad: dict, default_on: bool = False,
                command=None) -> ctk.CTkSwitch:
        sw = ctk.CTkSwitch(parent, text=label, font=FONT_LABEL,
                            text_color=MUTED, progress_color=BLUE,
                            button_color=BORDER2, command=command)
        if default_on:
            sw.select()
        sw.pack(**pad, anchor="w", pady=(0, SPACE_S))
        return sw

    def _toolbar_chip(self, parent) -> ctk.CTkFrame:
        """Grouped-control 'chip' background for nav/hdr instrument-toolbar
        clusters. Uses BG (a subtle off-white, distinct from the PANEL/CARD
        pure-white now used for large surfaces) rather than CARD -- since
        PANEL and CARD both became pure white, a CARD-filled chip on a
        PANEL-filled bar was indistinguishable from its background, and
        toolbar button groups lost their visual grouping entirely. BG is a
        gentle enough tint to read as "one control cluster" without
        reintroducing a visible grey block."""
        return ctk.CTkFrame(parent, fg_color=BG, corner_radius=8,
                             border_width=1, border_color=BORDER)

    def _btn(self, parent, text: str, command, pad: dict,
             variant: str = "secondary", h: int = 28, bold: bool = False) -> ctk.CTkButton:
        """Build a sidebar-style secondary/action button in one of 5 named
        variants (primary/success/warning/danger/secondary), matching the
        rationalised colour system in theme.py (COLOR_* semantic tokens).

        Reads the module-level COLOR_* globals fresh on every call (not a
        precomputed lookup table) so buttons rebuilt after a theme switch
        (_rebuild_ui()) always pick up the current theme's colours instead
        of whatever was active when a cached table would have been built.
        """
        if variant == "primary":
            fg, hover, text_color = COLOR_PRIMARY, COLOR_PRIMARY_HOVER, "white"
        elif variant == "success":
            fg, hover, text_color = COLOR_SUCCESS, COLOR_SUCCESS_HOVER, "white"
        elif variant == "warning":
            fg, hover, text_color = COLOR_WARNING, COLOR_WARNING_HOVER, "white"
        elif variant == "danger":
            fg, hover, text_color = COLOR_DANGER, COLOR_DANGER_HOVER, "white"
        else:
            fg, hover, text_color = COLOR_SECONDARY, COLOR_SECONDARY_HOVER, MUTED
        _h = max(24, int(h * THEME.font_scale))
        btn = ctk.CTkButton(
            parent, text=text, command=command,
            fg_color=fg, hover_color=hover, text_color=text_color,
            font=FONT_BTN_PRIMARY if bold else FONT_BTN_SEC, height=_h, corner_radius=8,
        )
        btn.pack(**pad, fill="x", pady=(0, SPACE_S))
        return btn

    def _reset_params(self) -> None:
        """Reset all filter/detection parameters to MouseECG species defaults."""
        fp = FilterParams()   # constructed with all dataclass defaults

        _fields: list[tuple[str, str]] = [
            ("ent_channel", fp.channel),
            ("ent_fs",      str(fp.fs)),
            ("ent_t_start", str(fp.t_start)),
            ("ent_t_end",   str(fp.t_end)),
            ("ent_lp",      str(fp.lp)),
            ("ent_hp",      str(int(fp.hp))),
            ("ent_minrr",   str(int(fp.min_rr_ms))),  # physiological floor
            # peak_distance_ms stored on self._peak_distance_ms (no dedicated widget)
            ("ent_sg_target_fs", str(fp.sg_target_fs)),
            ("ent_sg_window_ms", "20"),
        ]
        for attr, val in _fields:
            w = getattr(self, attr, None)
            if w is None:
                continue
            w.delete(0, "end")
            w.insert(0, val)

        # Threshold slider + entry
        try:
            self.sl_thr.set(fp.thresh)  # type: ignore[union-attr]
            self.lbl_thr.configure(text=f"Sensitivity:  {fp.thresh:.2f}")
        except Exception as e:
            log.debug("sl_thr/lbl_thr restore failed: %s", e)
        try:
            self.ent_thr.delete(0, "end")  # type: ignore[union-attr]
            self.ent_thr.insert(0, str(fp.thresh))  # type: ignore[union-attr]
        except Exception as e:
            log.debug("ent_thr restore failed: %s", e)

        # Epoch entries
        try:
            self.ent_epoch.delete(0, "end")
            self.ent_epoch.insert(0, str(int(MouseECG.EPOCH_DEFAULT_S)))
            self.ent_overlap.delete(0, "end")
            self.ent_overlap.insert(0, "0")
        except Exception as e:
            log.debug("epoch/overlap entry restore failed: %s", e)

        # Switches — match FilterParams defaults (sw_filtering is the
        # inverse of fp.no_filter -- ON means filtering is applied)
        _sw_defaults: list[tuple[str, bool]] = [
            ("sw_notch",         fp.notch_filter),
            ("sw_artifact",      fp.artifact_correction),
            ("sw_filtering",     not fp.no_filter),
            ("sw_epoch",         fp.auto_epochs),
            ("sw_invert_signal", fp.invert_signal),
        ]
        for attr, on in _sw_defaults:
            w = getattr(self, attr, None)
            if w is None:
                continue
            w.select() if on else w.deselect()

        # ComboBox
        try:
            self.cb_clean.set(fp.clean_method)  # type: ignore[union-attr]
        except Exception as e:
            log.debug("cb_clean restore failed: %s", e)
        # Detection method reset
        try:
            if self.cb_det_method is not None:
                self.cb_det_method.set("SG + Derivative (10 kHz)")  # type: ignore[union-attr]
                self._on_det_method_change("SG + Derivative (10 kHz)")
        except Exception as e:
            log.debug("cb_det_method reset failed: %s", e)

        # Resync FILTER SETTINGS greying + the "Processing" summary to the
        # just-reset sw_filtering/notch/cutoff/clean values.
        self._on_filtering_toggle()
        self._update_filter_summary()

        self._set_status("Parameters reset to mouse ECG defaults ✓", GREEN)

    # ════════════════════════════════════════════════════════
    #  GENERAL HELPERS
    # ════════════════════════════════════════════════════════

    def _textbox(self, parent, h: int = 180, padx: int = 0, expand: bool = False):
        kwargs = dict(font=FONT_BODY, fg_color="transparent", text_color=TEXT,
                      border_width=0, scrollbar_button_color=BORDER,
                      scrollbar_button_hover_color=BORDER2)
        if h > 0:
            kwargs["height"] = h
        tb = ctk.CTkTextbox(parent, **kwargs)  # type: ignore[arg-type]
        if h < 0 or expand:
            tb.pack(fill="both", expand=True, padx=padx, pady=(0, SPACE_S))
        else:
            tb.pack(fill="x", padx=padx, pady=(0, SPACE_S))
        return tb

    def _set_textbox(self, widget, text: str,
                     tsv: "str | None" = None) -> None:
        """Update a CTkTextbox with *text* and optionally store *tsv* for
        'Copy as TSV' (clipboard format that pastes into Excel cell-by-cell).
        """
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", text)
        widget.configure(state="disabled")
        wid = id(widget)
        if tsv is not None:
            self._tsv_store[wid] = tsv
            self._bind_copy_menu(widget)
        elif wid in self._tsv_store:
            del self._tsv_store[wid]

    def _bind_copy_menu(self, widget) -> None:
        """Attach a right-click context menu with copy options to *widget*."""
        def _show_menu(event):
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(
                label="📋  Copy as TSV  (paste into Excel)",
                command=lambda: self._copy_tsv(widget))
            menu.add_command(
                label="📄  Copy as plain text",
                command=lambda: self._copy_plain(widget))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
        # Bind to the underlying tk widget inside CTkTextbox
        try:
            inner = widget._textbox  # CTkTextbox wraps a tk.Text
            inner.bind("<Button-3>", _show_menu, add=True)
        except AttributeError:
            widget.bind("<Button-3>", _show_menu, add=True)

    def _copy_tsv(self, widget) -> None:
        """Copy the TSV version of this widget's data to the clipboard."""
        tsv = self._tsv_store.get(id(widget))
        if not tsv:
            self._copy_plain(widget)
            return
        self.clipboard_clear()
        self.clipboard_append(tsv)
        self._set_status("Copied as TSV — paste into Excel ✓", GREEN)

    def _copy_plain(self, widget) -> None:
        """Copy the plain-text content of *widget* to the clipboard."""
        try:
            text = widget.get("1.0", "end")
        except Exception:
            text = ""
        self.clipboard_clear()
        self.clipboard_append(text)
        self._set_status("Copied as plain text ✓", GREEN)

    def _set_status(self, text: str, color: str = MUTED) -> None:
        self.lbl_status.configure(text=text, text_color=color)

    def _bind_hover_tip(self, widget, text: str, color: str = MUTED) -> None:
        """Show *text* in the status bar on hover; clear it on leave.

        Thin wrapper around the existing hand-rolled <Enter>/<Leave> ->
        _set_status(...) hover pattern (previously duplicated once, for
        the Free Placement '?' hint). Reuses the persistent status-bar
        label already on screen -- not a floating tooltip widget.
        """
        widget.bind("<Enter>", lambda _e: self._set_status(text, color))
        widget.bind("<Leave>", lambda _e: self._set_status(""))

    def _widget_float(self, widget: "ctk.CTkEntry", default: float = 0.0) -> float:
        """Read a float from a CTk Entry widget.  Returns *default* on any error.

        Use this when the input is definitely a widget.  Misspelled attribute
        names produce an AttributeError rather than returning a silent 0.0,
        making bugs in callers visible during development.
        """
        try:
            v = float(widget.get())
            if not np.isfinite(v):
                log.debug("_widget_float: non-finite value from %r — using %s", widget, default)
                return default
            return v
        except (ValueError, TypeError, AttributeError) as exc:
            log.debug("_widget_float: could not parse widget %r — using %s: %s",
                      widget, default, exc)
            return default

    def _safe_float(self, widget_or_value: "Any", default: float = 0.0) -> float:
        """Safely extract a float from a CTk widget or a raw scalar value.

        Kept for backward compatibility.  Prefer _widget_float() when the
        caller knows the input is a widget.
        Returns *default* if the value is missing, empty, or non-numeric.
        """
        try:
            raw = widget_or_value.get() if hasattr(widget_or_value, "get") else widget_or_value
            v   = float(raw)
            if not np.isfinite(v):
                log.debug("_safe_float: non-finite value %r — using default %s", raw, default)
                return default
            return v
        except (ValueError, TypeError) as exc:
            log.debug("_safe_float: could not parse %r — using default %s: %s",
                      widget_or_value, default, exc)
            return default

    @staticmethod
    def _safe_df_val(df: "pd.DataFrame | None", col: str, decimals: int = 3) -> str:
        """Return a formatted scalar from a DataFrame cell, or '—' on any error."""
        try:
            if df is None or col not in df.columns:
                return "—"
            v = float(df[col].values[0])
            return f"{v:.{decimals}f}" if np.isfinite(v) else "—"
        except Exception:
            return "—"

    @staticmethod
    def _df_to_tsv(df: "pd.DataFrame | None") -> str:
        """Convert a DataFrame to tab-separated values for Excel paste.

        First row = header (column names), subsequent rows = values.
        Pastes into Excel so each metric is in its own column.
        """
        if df is None or df.empty:
            return ""
        rows = ["	".join(str(c).replace("HRV_", "") for c in df.columns)]
        for _, row in df.iterrows():
            cells = []
            for v in row:
                try:
                    fv = float(v)
                    cells.append(f"{fv:.5g}" if np.isfinite(fv) else "")
                except (TypeError, ValueError):
                    cells.append(str(v) if v is not None else "")
            rows.append("	".join(cells))
        return "\n".join(rows)

    @staticmethod
    def _describe_to_tsv(df: "pd.DataFrame | None") -> str:
        """Convert a describe()-style DataFrame to TSV for Excel.

        Output: stat name in col A, then one column per metric.
        e.g.  stat\tPR_ms\tQRS_ms\tQT_ms
              mean\t42.1\t12.3\t68.4
        """
        if df is None or df.empty:
            return ""
        rows = ["stat	" + "	".join(str(c) for c in df.columns)]
        for stat, row in df.iterrows():
            cells = [str(stat)]
            for v in row:
                try:
                    fv = float(v)
                    cells.append(f"{fv:.5g}" if np.isfinite(fv) else "")
                except (TypeError, ValueError):
                    cells.append(str(v) if v is not None else "")
            rows.append("	".join(cells))
        return "\n".join(rows)

    @staticmethod
    def _kv_to_tsv(pairs: "list[tuple[str, float]]") -> str:
        """Convert a list of (name, value) pairs to two-column TSV."""
        header = "Metric	Value"
        rows   = [f"{k}	{v:.5g}" if np.isfinite(float(v)) else f"{k}	"
                  for k, v in pairs]
        return "\n".join([header] + rows)

    @staticmethod
    def _df_to_text(df: "pd.DataFrame | None") -> str:
        """Render every finite numeric column of a NeuroKit2 HRV DataFrame as readable text.

        Format uses dot-leaders for visual alignment without requiring a fixed-width font:
            SDNN ............  12.35
            RMSSD ...........   8.27
        """
        if df is None or df.empty:
            return "  (not computed)"
        rows = []
        for col in df.columns:
            try:
                v = float(df[col].values[0])
                if np.isfinite(v):
                    name   = col.replace("HRV_", "")
                    # Adaptive precision: integers → 0 dp, small values → 4 dp
                    if abs(v) >= 100:
                        fmt = f"{v:.1f}"
                    elif abs(v) >= 1:
                        fmt = f"{v:.3f}"
                    else:
                        fmt = f"{v:.5f}"
                    # Dot-leader padding to column 30
                    dots = "·" * max(2, 30 - len(name))
                    rows.append(f"  {name} {dots}  {fmt:>10}")
            except Exception as exc:
                log.debug("_df_to_text skip '%s': %s", col, exc)
        return "\n".join(rows) or "  (no finite values)"


# ════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()

    if not NK_AVAILABLE:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "Missing dependency",
            "Install NeuroKit2:\n  pip install neurokit2",
        )
        root.destroy()
    else:
        app = ECGApp()
        app.mainloop()
